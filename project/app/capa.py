import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import os




def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho

def preenche_planilha(planilha, dicionario):

    caminho = pegar_caminho(planilha)
    Plan = planilha
    # carrega a planilha de acordo com o caminho
    workbook = openpyxl.load_workbook(caminho)

   # planilha_preenchida = pegar_caminho('preenchido-' + planilha)
    for nomePlanilha, entradaDados in dicionario.items():
        planilhaAtual = workbook[nomePlanilha]

        for intervaloCelula, entradaCelula in entradaDados:
            if ":" in intervaloCelula:  
                inicioCelula, fimCelula = intervaloCelula.split(":")
                planilhaAtual = workbook[nomePlanilha]
                planilhaAtual[inicioCelula] = entradaCelula
                # planilhaAtual[inicioCelula].fill = color   <--- teste com cores
            else:  
                planilhaAtual[intervaloCelula] = entradaCelula
                # planilhaAtual[intervaloCelula].fill = color  <--- teste com cores


    workbook.save(planilha)

    return planilha


def inserir_round_retangulo(planilha,data1,data2,dicionariofin):


   

    # id_mapeamento               
    # codigo                      
    # nome                        
    # saldo                       
    # data_assinatura             
    # data_vigencia               
    # data_encerramento           
    # tipo_contrato               
    # instituicao_executora       
    # processo                    
    # subprocesso                 
    # cod_proposta                
    # proposta                    
    # objetivos                   
    # valor_aprovado              
    # nome_tp_controle_saldo      
    # grupo_gestores              
    # gestor_resp                 
    # coordenador                 
    # procedimento_compra         
    # tab_frete                   
    # tab_diarias                 
    # custo_op                    
    # nome_financiador            
    # departamento                
    # situacao                    
    # banco                       
    # agencia_bancaria            
    # conta_bancaria              
    # centro_custo                
    # conta_caixa                 
    # categoria_projeto           
    # cod_convenio_conta          
    # cod_status                  
    # ind_sub_projeto             
    # tipo_custo_op               
    # projeto_mae                 
    # id_coordenador              
    # id_financiador              
    # id_instituicao              
    # id_departamento             
    # nome_instituicao            
    # id_instituicao_executora    
    # id_tipo                     

    planilha_local_dados = {
        "Capa Finatec":[
        ("E6",'NOME_FINANCIADOR'),
        ("E9",'NOME'),
        ("E14",'COORDENADOR'),
        ("E17",'PROCESSO'),
        ("E19",'CODIGO'),
        ("K17",'BANCO'),
        ("K19",'CONTA_BANCARIA'),
        ("K21",'AGENCIA_BANCARIA'),
        ("E35",'Suellen Santos Diniz de Carvalho') 
        ]
    }






    # Create an empty dictionary to store the values from the second dictionary
    result_dict = {}

    # Loop pra criar no formato dicionarario por exemplo
    # Capa Finatec : [(E3:COORDENADOR)]
    for key, value in planilha_local_dados["Capa Finatec"]:
        if value in dicionariofin:
            result_dict[key] = dicionariofin[value]
    print(result_dict)
    list_of_tuples = [(key, value) for key, value in result_dict.items()]
   
    novo_dict = {"Capa Finatec":list_of_tuples}
    
    preenche_planilha(planilha, novo_dict)

    caminho = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Capa Finatec'] 

    

    # List of image names
    image_names = [
        'agencia.png',#0
        'agencia_branco.png',#1
        'agente.png',#2
        'agente_branco.png',#3
        'analista.png',#4
        'analista_branco.png',#5
        'assinaturas_responsaveis.png',#6
        'assistente.png',#7
        'assistente_branco.png',#8
        'banco.png',#9
        'banco_branco.png',#10
        'conta_corrente.png',#11
        'conta_corrente_branco.png',#12
        'coordenador.png',#13
        'coordenador_branco.png',#14
        'dadosdoacordo.png',#15
        'gerente.png',#16
        'gerente_branco.png',#17
        'periodo.png',#18
        'periodo_branco.png',#19
        'prestaçãodecontasparcial.png',#20
        'processo.png',#21
        'processo_branco.png',#22
        'projeto.png',#23
        'projeto_branco.png',#24,
        'centrodecusto.png',#25,
        'centrodecusto_branco.png'
    ]

    # Path to the images
    path = '/home/ubuntu/Desktop/entrega/05_PipelineFinatec/project/app/imagensCapa/'
   
    # List to hold Image objects
    images = []

    # Loop through the list of image names and create Image objects
    for i, name in enumerate(image_names):
        image_path = path + name
        pil_image = PILImage.open(image_path)
        pil_image.save(image_path)
        img = Image(image_path)
        images.append(img)

  

 
    #estilo imagem
    worksheet.add_image(images[20], "B2")#prestação
    worksheet.merge_cells('B2:L2')
    worksheet['B2'] = 'PRESTAÇÃO DE CONTAS PARCIAL'
    worksheet['B2'].font = Font(name="Tahoma", size=17, color="204c80",bold=True)
    worksheet['B2'].alignment = Alignment(horizontal="center",vertical="center")

    worksheet.add_image(images[15], "B4")#dados do acordo
    worksheet.merge_cells('B4:L4')
    worksheet['B4'] = 'DADOS DO ACORDO'
    worksheet['B4'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B4'].alignment = Alignment(horizontal="center",vertical="center")

    worksheet.add_image(images[2], "B6")#agente
    worksheet.merge_cells('B6:D7')
    worksheet['B6'] = 'AGENTE FINANCIADOR'
    worksheet['B6'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B6'].alignment = Alignment(horizontal="justify",vertical="center",wrap_text=True)
    worksheet.add_image(images[3], "E6")
    worksheet.merge_cells('E6:M7')
    worksheet['E6'].font = Font(name="Tahoma", size=9, color="204c80",bold=False)
    worksheet['E6'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[23], "B9")#projeto
    worksheet.merge_cells('B9:D13')
    worksheet['B9'] = 'PROJETO'
    worksheet['B9'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B9'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[24], "E9")
    worksheet.merge_cells('E9:M13')
    worksheet['E9'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E9'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[13], "B14")#coordenador
    worksheet.merge_cells('B14:D15')
    worksheet['B14'] = 'COORDENADOR'
    worksheet['B14'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B14'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[14], "E14")
    worksheet.merge_cells('E14:M15')
    worksheet['E14'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E14'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)


    worksheet.add_image(images[21], "B17")#processo
    worksheet.merge_cells('B17:D18')
    worksheet['B17'] = 'PROCESSO'
    worksheet['B17'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B17'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[22], "E17")
    worksheet.merge_cells('E17:G18')
    worksheet['E17'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E17'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[9], "H17")#Banco
    worksheet.merge_cells('H17:J18')
    worksheet['H17'] = 'BANCO'
    worksheet['H17'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['H17'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[10], "K17")
    worksheet.merge_cells('K17:M18')
    worksheet['K17'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['K17'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[11], "H19")#CoOtaCorreOte
    worksheet.merge_cells('H19:J20')
    worksheet['H19'] = 'CONTA CORRENTE'
    worksheet['H19'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['H19'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[12], "K19")
    worksheet.merge_cells('K19:M20')
    worksheet['K19'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['K19'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[0], "H21")#Agencia
    worksheet.merge_cells('H21:J22')
    worksheet['H21'] = 'AGÊNCIA'
    worksheet['H21'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['H21'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[1], "K21")
    worksheet.merge_cells('K21:M22')
    worksheet['K21'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['K21'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[25], "B19")#Centro de custo
    worksheet.merge_cells('B19:D20')
    worksheet['B19'] = 'CENTRO DE CUSTO'
    worksheet['B19'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B19'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[26], "E19")#Centro de custo
    worksheet.merge_cells('E19:G20')
    worksheet['E19'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E19'].alignment = Alignment(horizontal="center",vertical="center")


    worksheet.add_image(images[18], "B21")#Periodo
    worksheet.merge_cells('B21:D22')
    worksheet['B21'] = 'PERIODO'
    worksheet['B21'].font = Font(name="Tahoma", size=9, color="204c80",bold=True)
    worksheet['B21'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[19], "E21")#Periodo
    worksheet.merge_cells('E21:G22')
    worksheet['E21'] = f"{(data1)} a {(data2)}"
    worksheet['E21'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E21'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    worksheet.add_image(images[7], "B30")#Assistente
    worksheet.merge_cells('B30:D32')
    worksheet['B30'] = 'ASSISTENTE'
    worksheet['B30'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B30'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[8], "E30")#Assistente
    worksheet['E30'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E30'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.add_image(images[4], "B26")#Analista
    worksheet.merge_cells('B26:D28')
    worksheet['B26'] = 'ANALISTA'
    worksheet['B26'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B26'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.add_image(images[5], "E26")#Analista
    worksheet.merge_cells('E26:M28')
    worksheet['E26'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E26'].alignment = Alignment(horizontal="center",vertical="center")

    worksheet.add_image(images[16], "B34")#gerengte
    worksheet.merge_cells('B34:D36')
    worksheet['B34'] = 'COORDENADORA DE GESTÃO DE PROJETOS'
    worksheet['B34'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B34'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True,shrink_to_fit=True  )
    worksheet.add_image(images[17], "E34")#gerente
    worksheet['E34'].font = Font(name="Tahoma", size=10, color="204c80",bold=False)
    worksheet['E34'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.add_image(images[6], "B23")#ASSINATURAS
    worksheet.merge_cells('B23:M24')
    worksheet['B23'] = 'ASSINATURAS - RESPONSAVEIS PELA PRESTAÇÃO DE CONTAS'
    worksheet['B23'].font = Font(name="Tahoma", size=10, color="204c80",bold=True)
    worksheet['B23'].alignment = Alignment(horizontal="center",vertical="center")

    # Save the workbook
    workbook.save(planilha)
