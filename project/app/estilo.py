import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle
import os
def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho




def estilo_fub_fisica_juridica(tabela,tamanho):
        # carrega a planilha de acordo com o caminho
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)

    # Create a new workbook
    # workbook = openpyxl.Workbook()

    #workbook.create_sheet(title="Sheet1")

    # You can add one or more worksheets to the workbook (the first one is created by default)
    worksheet = workbook['Pessoa Fisica']  # Replace 'Sheet1' with the name of your worksheet


    # Add data to the worksheet
    #tamanho = 101
    size = tamanho + 10
    #Font(name="Arial", size=12, color="00FF0000",bold=True)
    # Change height of row A1
    #worksheet.row_dimensions[27].height = 50

    cinza = "979CA8"
    azul = "336394"
    
                
                
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
        worksheet.row_dimensions[row[0].row].height = 35


    # Create a custom number format
    custom_number_format = NamedStyle(name='custom_number_format')
    custom_number_format.number_format = 'R$ #,##0.00'
    custom_number_format.font = Font(name="Arial", size=12, color="000000")
    custom_number_format.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)

# Define the value until which you want to apply the format
    value_to_stop = size  # Replace x with your specific value
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'J{row}']
        cell.style = custom_number_format
        
# # Iterate through the column J and apply the custom number format
#     for cell in worksheet['J']:
#         if row < start_row:
#             continue
#         cell.style = custom_number_format
        
#         if cell.value == value_to_stop:
#             break  # Exit the loop once you reach the specified value
#padraocinzaebranco
    for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    


    #subtotal
    celulas_mergidas_subtotal = f"A{size}:I{size}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


    # Specify the row number where you want to sum until (e.g., row 10)

    # Build the formula string
    formula = f"=SUM(J10:J{size-1})"
    celula = f'J{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


    #restituições creditadas
    restituicoes = size + 1
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    #colunas azul
    row_style = NamedStyle(name='row_style')
    row_style.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
    row_style.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style.height = 30

    row_number = size + 2
   
    # # Apply the style to each cell in the row
    # for cell in worksheet[row_number]:
    #     cell.style = row_style
    for column in range(1, 11):  # This will loop through columns 1 to 10
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = row_style


    values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        if coluna == 4:
            coluna = coluna + 1
        coluna = coluna + 1
        

    merge_formula = f'D{row_number}:E{row_number}'
    worksheet.merge_cells(merge_formula)

    #subtotal2
    sub_total2_row = size + 4
    subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    sub_formula_row_celula = f'J{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

     #subtotal12_Formula
    # sub_formula_row = size + 4
    # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
    # sub_formula_row_celula = f'J{sub_formula_row}'
    # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    # worksheet[sub_formula_row_celula] = sub_formula


    #total1-2
    total12_row = size + 5
    total12_merge_cells = f'A{total12_row}:I{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1 -2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

    #total_formula
    total_formula_row = size + 5
    total_formulaa = f'=J{size}'
    total_formula_row_celula = f'J{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa

    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f="='Receita x Despesa'!A42:J42"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f="='Receita x Despesa'!A45"
    diretor_cargo_formula = f="='Receita x Despesa'!A46"
    diretor_cpf_formula = f="='Receita x Despesa'!A47"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f="='Receita x Despesa'!H45"
    coordenadora_cargo_formula = f="='Receita x Despesa'!H46"
    coordenadora_cpf_formula = f="='Receita x Despesa'!H47"
    coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    # #cabeca

    # azul_claro = '1c89b8'
    # row_style_cabecario = NamedStyle(name='row_style_cabecario')
    # row_style_cabecario.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    # row_style_cabecario.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    # row_style_cabecario.alignment = Alignment(horizontal="center",vertical="center")
    # worksheet.row_dimensions[9].height = 50
    # worksheet.column_dimensions['b'].width = 30
    # worksheet.column_dimensions['c'].width = 20
    # worksheet.column_dimensions['d'].width = 30
    # worksheet.column_dimensions['e'].width = 30
    # worksheet.column_dimensions['f'].width = 30
    # worksheet.column_dimensions['g'].width = 30
    # worksheet.column_dimensions['h'].width = 30
    # worksheet.column_dimensions['i'].width = 30
    # worksheet.column_dimensions['j'].width = 15
    # linha_number = 9
    # # Apply the style to each cell in the row
    # for cell in worksheet[linha_number]:
    #     cell.style = row_style_cabecario

    # valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
    # col = 1
    # for a,b in enumerate(valores):
    #     worksheet.cell(row=linha_number, column=col, value=b)
    #     col = col + 1


    # Save the workbook to a file
    workbook.save(tabela)
    workbook.close()


# estilo_fub_fisica_juridica('Modelo_Fub.xlsx')