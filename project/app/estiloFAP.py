import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
import os
import random

# variaveis globais

def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline

def estilo_capa(tabela,tamanho,stringTamanho):
    """

    """

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Capa']

    # fechando e salvando arquivo
    workbook.save(tabela)
    workbook.close()

def estilo_anexoI(tabela,tamanho,stringTamanho):
    """

    """

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO I']

    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    random_number = random.randint(1, 10000)
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False 
    
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=5,max_col=5):
        for cell in row:
            cell.border = borda 
            
    worksheet.column_dimensions['a'].width = 30
    worksheet.column_dimensions['b'].width = 70
    worksheet.column_dimensions['c'].width = 30
    worksheet.column_dimensions['d'].width = 50 # descrição
    worksheet.column_dimensions['e'].width = 50 # descrição

    # fechando e salvando arquivo
    workbook.save(tabela)
    workbook.close()

def estilo_anexoII(tabela,tamanho,stringTamanho):
    """

    """

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO II']

    # fechando e salvando arquivo
    workbook.save(tabela)
    workbook.close()

def estilo_anexoIII(tabela,tamanho,stringTamanho):
    """

    """

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO III']

    # fechando e salvando arquivo
    workbook.save(tabela)
    workbook.close()

def estilo_anexoIV(tabela,tamanho,stringTamanho): 
    """

    """

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO IV']

    # fechando e salvando arquivo
    workbook.save(tabela)
    workbook.close()

def estilo_conciliacao(tabela,tamanho,stringTamanho):
    """

    """

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Conciliação']
    
    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    random_number = random.randint(1, 10000)
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False 
    
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=5,max_col=5):
        for cell in row:
            cell.border = borda
            
    worksheet.column_dimensions['a'].width = 30
    worksheet.column_dimensions['b'].width = 70
    worksheet.column_dimensions['c'].width = 30
    worksheet.column_dimensions['d'].width = 50 # descrição
    worksheet.column_dimensions['e'].width = 50 # descrição

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    # fechando e salvando arquivo
    workbook.save(tabela)
    workbook.close()
