import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import os
import random

def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline

def estiloRelatorioExecFinanceiroA1(tabela,tamanho,stringTamanho):
    '''Estilo da Pagina do Relatorio Execução da Receita e Despesa
      
        Argumentos:
            tabela : recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho : é o tamanho total de linhas que irão ser geradas dinâmicamente. O valor varia dentre o tamanho da quantidade de rubricas diferente que o projeto possui, excluindo Obras e Instalações
            Aplicações Financeira e Equipamento e Material Permanente sendo nacional ou importado.
            stringTamanho : refere-se aonde esta localizado a string brasília na pagina Receita e despesa para a referencias das formulas.
           

    '''

    if tamanho == 0:
         tamanho = 1
    #Carrregar o arquivo da tabela, e inicia a worksheet com o nome da sheet "Relatório de Exec Financ A.1"
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Relatório de Exec Financ A.1']

    #tamanho recebido para alocar o tamanho das linha que e correspondetente ao que vai ser preenchido da consulta do database mais o tamanho do cabecario que e 16 e cores
    size = tamanho + 12;
    size2 = size + 1
  

    azul = 'ccffff'
    cinza = 'bebebf'
    
    #remover as gridlines
    sheet.sheet_view.showGridLines = False
    

    #cabecario RELATÓRIO DE EXECUÇÃO FINANCEIRA
    sheet.merge_cells('A1:I1')
    sheet['A1'] = f'ANEXO 1'
    sheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    

    sheet.merge_cells('A2:I2')
    sheet['A2'] = f'RELATÓRIO DE EXECUÇÃO FINANCEIRA'
    sheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A2'].alignment = Alignment(horizontal="center",vertical="center")
    

    # sheet.merge_cells('B4:F4')
    # sheet['A4'] =  ""
    # sheet['A4'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('B5:F5')
    # sheet['A5'] =  ""
    # sheet['A5'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('B6:F6')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('B7:F7')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('C8:E8')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('B9:E9')
    # sheet['B9'] =  ""
    # sheet['B9'].font = Font(name="Arial", size=12, color="000000")
    # sheet['B9'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('F9:I9')
    # sheet['F9'] =  ""
    # sheet['F9'].font = Font(name="Arial", size=12, color="000000")
    # sheet['F9'].alignment = Alignment(horizontal="left",vertical="center")

    #CATEGORIA ECONÔMICA / RUBRICAS
    sheet.merge_cells('A9:A10')
    sheet['A9'] =  "CATEGORIA ECONÔMICA / RUBRICAS"
    sheet['A9'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    sheet['A9'].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")


    #Concedente,Participe1,Participe 2,total b10 e e10
    values = ["Concedente","Partícipe 1","Partícipe 2",'Total']
    coluna = 2
    for a,b in enumerate(values):
        sheet.cell(row=10, column=coluna, value=b)
        coluna = coluna + 1

    #Concedente,Participe1,Participe 2,total f10 e i10
    values = ["Concedente","Partícipe 1","Partícipe 2",'Total']
    coluna = 6
    for a,b in enumerate(values):
        sheet.cell(row=10, column=coluna, value=b)
        coluna = coluna + 1
    
    #Executado No Periodo
    sheet['B9'] = 'EXECUTADO NO PERÍODO'
    sheet['B9'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['B9'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['B9'].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    sheet['B9'].border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
    
     #EXECUTADO ATÉ O PERÍODO
    sheet['F9'] = 'EXECUTADO ATÉ O PERÍODO'
    sheet['F9'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['F9'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['F9'].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    sheet['F9'].border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

     #DESPESA CORRENTES
    sheet['A11'] = '3.DESPESAS CORRENTES'
    sheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A11'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['A11'].border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

    #pintar inicio da tabela de azul
    for rows in sheet.iter_rows(min_row=9, max_row=10, min_col=1, max_col=9):
            for cell in rows:
                
                cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    

    sheet['E10'] =  "Total"
    sheet['E10'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['E10'].alignment = Alignment(horizontal="left",vertical="center")
    sheet['E10'].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    
    sheet['I10'] =  "Total"
    sheet['I10'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['I10'].alignment = Alignment(horizontal="left",vertical="center")
    sheet['I10'].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")

 

    #4.DESPESAS DE CAPITAL
    sheet[f'A{size2}'] = '4.DESPESAS DE CAPITAL'
    sheet[f'A{size2}'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet[f'A{size2}'].alignment = Alignment(horizontal="left",vertical="center")
    sheet[f'A{size2}'].border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") 
                                ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )          

    sheet.column_dimensions['a'].width = 55
    sheet.column_dimensions['b'].width = 20
    sheet.column_dimensions['c'].width = 20
    sheet.column_dimensions['d'].width = 20
    sheet.column_dimensions['e'].width = 20
    sheet.column_dimensions['f'].width = 20
    sheet.column_dimensions['g'].width = 20
    sheet.column_dimensions['h'].width = 20
    sheet.column_dimensions['i'].width = 20

               
    
    #51- Obras e Instalações
    celulasAzul = f"A{size2+1}"
    cell = sheet[celulasAzul]
    cell.value = f'51 - Obras e Instalações'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
    #deiarlinha de obras e instalções em negrito
    

    #a) Obras
    celulasAzul = f"A{size2+2}"
    cell = sheet[celulasAzul]
    cell.value = f'a) Obras'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        
    #b) Instalações
    celulasAzul = f"A{size2+3}"
    cell = sheet[celulasAzul]
    cell.value = f'b) Instalações'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )



     #Equipamentos e Material Permantente
    celulasAzul = f"A{size2+4}"
    cell = sheet[celulasAzul]
    cell.value = f'52 - Equipamentos e Material Permanente '
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )

     #a) Nacionais
    celulasAzul = f"A{size2+5}"
    cell = sheet[celulasAzul]
    cell.value = f'a) Nacionais'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        
    #b) Importados
    celulasAzul = f"A{size2+6}"
    cell = sheet[celulasAzul]
    cell.value = f'b) Importados'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        
    #Total
    celulasAzul = f"A{size2+7}"
    cell = sheet[celulasAzul]
    cell.value = f'TOTAL'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        
    #pintar celulas de cinza da celula total
    for rows in sheet.iter_rows(min_row=size2+7, max_row=size2+7, min_col=1, max_col=9):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True) 
                cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
     #fazerborda
    for rows in sheet.iter_rows(min_row=9, max_row=size2+7, min_col=1, max_col=9):
            for cell in rows:
                cell.border =  Border(top=Side(border_style="hair") ,left = Side(border_style="hair") 
                                      ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    #bordasNegrito
    for rows in sheet.iter_rows(min_row=1, max_row=size2 + 7, min_col=1, max_col=10):
                for cell in rows:
                    #primeira coluna preta
                    if cell.column == 1 and cell.row > 8:
                        cell.border =  Border(top=Side(border_style="hair") ,left = Side(border_style="hair") 
                                        ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                    
                    #segunda coluna preta
                    if cell.column == 6 and cell.row > 8:
                            cell.border =  Border(top=Side(border_style="hair") ,left = Side(border_style="medium") 
                                            ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                     #linha preta do top       
                    if cell.row == 8:
                        cell.border =  Border(bottom=Side(border_style="medium") )
                    #linha preta da celula 11
                    if cell.row == 11 and cell.column > 1:
                         cell.border =  Border(top=Side(border_style="medium") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

                    if cell.row == 11 and cell.column == 1:
                         cell.border =  Border(top=Side(border_style="medium") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                         
                    if cell.row == 11 and cell.column == 5:
                         cell.border =  Border(top=Side(border_style="medium") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") ) 
                    #linha preta desepsas de capital
                    if cell.row == size2 and cell.column > 1:
                         cell.border =  Border(top=Side(border_style="medium") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") ) 
                    
                    if cell.row == size2 and cell.column == 1:
                         cell.border =  Border(top=Side(border_style="medium") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") ) 
                         
                    if cell.row == size2 and cell.column == 5:
                         cell.border =  Border(top=Side(border_style="medium") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") ) 
                    #linha double total
                         
                    if cell.row == size2+7 and cell.column > 1 :
                         cell.border =  Border(top=Side(border_style="double") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") ) 
                    if cell.row == size2+7 and cell.column == 1 :
                         cell.border =  Border(top=Side(border_style="double") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") ) 
                    if cell.row == size2+7 and cell.column == 5 :
                         cell.border =  Border(top=Side(border_style="double") ,left = Side(border_style="hair") 
                                            ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") ) 
                         
                    if cell.column == 10:
                         cell.border =  Border(left = Side(border_style="medium") ) 

    #
    for rows in sheet.iter_rows(min_row=9, max_row=size2 + 7, min_col=1, max_col=10):
            for cell in rows:
                 if cell.column == 10:
                      cell.border =  Border(left =Side(border_style="medium") ) 
    
    for rows in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=10):
            for cell in rows:
                 if cell.column == 10:
                      cell.border =  Border(left =Side(border_style="thin",color='9e9e9e') ) 
    

   
                      

    #aumentar o tamanho da celula total
    sheet.row_dimensions[size+7].height = 16
     #brasilia
    brasilia_row = size2 + 9
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    sheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = sheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size2 + 10
    diretor_cargo_row = size2 + 11
    diretor_cpf_row = size2 + 12
    
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    sheet.merge_cells(diretor_merge_cells)
    sheet.merge_cells(diretor_cargo_merge_cells)
    sheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = sheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = sheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = sheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size2 + 10
    coordenadora_cargo_row = size2 + 11
    coordenadora_cpf_row = size2 + 12
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:I{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:I{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:I{coordenadora_cpf_row}'
    sheet.merge_cells(coordenadora_merge_cells)
    sheet.merge_cells(coordenadora_cargo_merge_cells)
    sheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = sheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = sheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = sheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    
  
    #adicionar borda fo nim do arquivo
    for row in sheet.iter_rows(min_row=size2 + 8, max_row=coordenadora_cpf_row,min_col=9,max_col=9):
        for cell in row:
         cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e')  )

    #adicionar borda fo nim do arquivo
    for row in sheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=9):
        for cell in row:
            if cell.column == 9:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )


   

    workbook.save(tabela)
    workbook.close()

    return 0

def estiloDEMOSTRRECEITEDESPESAA2(tabela,tamanho):
    '''Estilo da Pagina do Relatorio Receita e Despesa.
      
        Argumentos:
            tabela : recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            
            tamanho : refere-se ao tamanho do dataframe que recebe
    '''
   
       
    caminho = pegar_caminho(tabela)
    #Plan = planilha
    # carrega a planilha de acordo com o caminho
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['DEMOSTR. RECEITA E DESPESA A.2']
    size = tamanho + 13
    #tamanaho contando despes decapital, obras e instalaçoes, e euipamentos
    size2 = size + 8  
    cinza = "d9d9d9"
    azul = 'ccffff'
    
    #remover as gridlines
    sheet.sheet_view.showGridLines = False
    

    #cabecario relação de pagamentos - outro servicoes de terceiros
    sheet.merge_cells('A1:E1')
    sheet['A1'] = f'ANEXO 2'
    sheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    

    sheet.merge_cells('A2:E2')
    sheet['A2'] = f'DEMONSTRATIVO DE RECEITAS E DESPESAS'
    sheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A2'].alignment = Alignment(horizontal="center",vertical="center")
    
    
    # sheet.merge_cells('C3:E3')
    # sheet['A3'] = ""
    # sheet['A3'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A3'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)

    # sheet.merge_cells('C4:E4')
    # sheet['A4'] =  ""
    # sheet['A4'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('C5:E5')
    # sheet['A5'] =  ""
    # sheet['A5'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('C6:E6')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('C7:E7')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('C8:E8')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    # sheet.merge_cells('C9:E9')
    # sheet['A6'] =  ""
    # sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    # sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    

    # sheet.merge_cells('A10:E10')
    # sheet['A10'] = ''
    # sheet['A10'].font = Font(name="Arial", size=12, color="000000",bold=True)
    # sheet['A10'].alignment = Alignment(horizontal="center",vertical="center")
    
    
    #configurando alturas da tabela demonstrativo.receita despesa
    for i in range(1,10):
        sheet.row_dimensions[i].height = 20
        if i ==10:
            sheet.row_dimensions[i].height = 2


    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[3].height = 35
    sheet.row_dimensions[2].height = 25
    sheet.row_dimensions[11].height = 30
    sheet.row_dimensions[12].height = 30
    sheet.column_dimensions['a'].width = 55
    sheet.column_dimensions['b'].width = 25
    sheet.column_dimensions['c'].width = 30
    sheet.column_dimensions['d'].width = 30
    sheet.column_dimensions['e'].width = 30

     #mascara real em toda a tabela
    for rows in sheet.iter_rows(min_row=13, max_row=size2 +13, min_col=2, max_col=5):
        for cell in rows:
                cell.number_format = 'R$ #,##0.00'


    

    #merge das barras azuis da 11 que leva Valor Realizado e Natureza das despesa
    barraMerge1 = f"B{11}:D{11}"
    sheet.merge_cells(barraMerge1)
    #borda Valor realizado
    for row in sheet[barraMerge1]:
        for cell in row:
            cell.border = Border(top=Side(border_style="double"), left=Side(border_style="hair"), bottom=Side(border_style="hair"), right=Side(border_style="double"))
    #Valor Realizado
    celulasAzul = f"B{11}"
    cell = sheet[celulasAzul]
    cell.value = f'Valor Realizado'
    cell.fill = PatternFill(start_color=azul, end_color=azul,
                                            fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="double") )


    #A.Natureza das despesas
    sheet['A11'] = 'NATUREZA DAS DESPESAS'
    sheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A11'].alignment = Alignment(horizontal="left",vertical="center")
    sheet['A11'].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    sheet['A11'].border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
    
    #despesas correntes
    sheet['A13'] = '3.DESPESAS CORRENTES'
    sheet['A13'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A13'].alignment = Alignment(horizontal="left",vertical="center")
    sheet['A13'].border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
    
    for row in sheet.iter_rows(min_row=13,max_row=13,min_col=1,max_col=5):
        for cell in row :
            cell.font = Font(name='Arial',size=12, color="000000",bold=True)

#Grupo/elemenos de despesas/Ate o periodo Anterior/No periodo deteste balancete total acumaladao, previsão na relação de itens
    values = ["Grupos/Elementos de Despesas","Até o período Anterior","No Período deste Balancete",'Total Acumulado',"Previsão na Relação de Itens"]
    coluna = 1
    for a,b in enumerate(values):
        sheet.cell(row=12, column=coluna, value=b)
        coluna = coluna + 1

    for row in sheet.iter_rows(min_row=12, max_row=12,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5 :
                 cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
                 cell.font = Font(name='Arial',bold=True,color='000000')

            else:
                cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
        

        
    for row in sheet.iter_rows(min_row=12, max_row=size2,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5 and cell.row == 12:
                cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
            elif cell.column == 5:
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
            elif cell.row == size2:
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="double") )
        
            else:
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

        
    for row in sheet.iter_rows(min_row=size2, max_row=size2,min_col=5,max_col=5):
        for cell in row:
            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )


    #4.Despesas de Capital
    celulasAzul = f"A{size+1}"
    cell = sheet[celulasAzul]
    cell.value = f'4.DESPESAS DE CAPITAL'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
     #deiarlinha de desoesas de capital em negrito
    for rows in sheet.iter_rows(min_row=size+1, max_row=size+1, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)    


    
    #51- Obras e Instalações
    celulasAzul = f"A{size+2}"
    cell = sheet[celulasAzul]
    cell.value = f'51 - Obras e Instalações'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
    #deiarlinha de obras e instalções em negrito
    for rows in sheet.iter_rows(min_row=size+2, max_row=size+2, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)    

    #a) Obras
    celulasAzul = f"A{size+3}"
    cell = sheet[celulasAzul]
    cell.value = f'a) Obras'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        
    #b) Instalações
    celulasAzul = f"A{size+4}"
    cell = sheet[celulasAzul]
    cell.value = f'b) Instalações'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        

    #Equipamentos e Material Permantente
    celulasAzul = f"A{size+5}"
    cell = sheet[celulasAzul]
    cell.value = f'52 - Equipamentos e Material Permanente '
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )


    #deiarlinha do equipamento e materialpermaneten em negrito
    for rows in sheet.iter_rows(min_row=size+5, max_row=size+5, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                



     #a) Nacionais
    celulasAzul = f"A{size+6}"
    cell = sheet[celulasAzul]
    cell.value = f'a) Nacionais'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        
    #b) Importados
    celulasAzul = f"A{size+7}"
    cell = sheet[celulasAzul]
    cell.value = f'b) Importados'
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
        



    #Total das Receitas (B1+B2)
    celulasAzul = f"A{size2+2}"
    cell = sheet[celulasAzul]
    cell.value = f"B. TOTAL DAS RECEITAS  (B1+B2)"
    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
    #deiarlinha de total das reeitas em negrito
    for rows in sheet.iter_rows(min_row=size2+2, max_row=size2+2, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True) 

    #Barra azull total das receitas
                
    for row in sheet.iter_rows(min_row=size2+2, max_row=size2+2,min_col=2,max_col=4):
        for cell in row:
                if cell.column == 4 :
                    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
                    cell.font = Font(name='Arial',bold=True,color='000000')
                    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="double") )

                else:
                    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
                    cell.font = Font(name='Arial',bold=True,color='000000')
                    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )

    #bordas inferiores
    for row in sheet.iter_rows(min_row=size2+3, max_row=size2+4,min_col=2,max_col=4):
        for cell in row:
                if cell.column == 4 and cell.row == size2+3:
            
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="double") )
                elif cell.column == 4 and cell.row == size2+4:
                    
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="double") )
                elif cell.column != 4 and cell.row == size2+4:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )

                else:
                    
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )




    celulasAzul = f"A{size2+3}"
    cell = sheet[celulasAzul]
    cell.value = f"B.1 Recursos Recebidos"
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
            
    celulasAzul = f"A{size2+4}"
    cell = sheet[celulasAzul]
    cell.value = f"B.2 Rendimento de Aplicações Financeiras"
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )
            
    #saldo
    celulasAzul = f"A{size2+6}"
    cell = sheet[celulasAzul]
    cell.value = f"Saldo"
    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="double") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )
    #deiar negrito o saldo
    for rows in sheet.iter_rows(min_row=size2+6, max_row=size2+6, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True) 


    #saldos disponivel para o perido seguinte
    celulasAzul = f"A{size2+8}"
    cell = sheet[celulasAzul]
    cell.value = f"D. Saldo Disponível para o Período Seguinte"
    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
    
    #Negrito no numero da frente
    celulasAzul = f"B{size2+8}"
    cell = sheet[celulasAzul]
    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="double") )
    cell.number_format = 'R$ #,##0.00'

    #d1 borda
    celulasAzul = f"B{size2+9}"
    cell = sheet[celulasAzul]
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="double") )
    cell.number_format = 'R$ #,##0.00'

    #d2 borda
    celulasAzul = f"B{size2+10}"
    cell = sheet[celulasAzul]
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="double") )
    cell.number_format = 'R$ #,##0.00'

    #d3 borda
    celulasAzul = f"B{size2+11}"
    cell = sheet[celulasAzul]
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="double") )
    cell.number_format = 'R$ #,##0.00'

    #Barra azul saldo

    for row in sheet.iter_rows(min_row=size2+6, max_row=size2+6,min_col=2,max_col=4):
        for cell in row:
            if cell.column == 3:
                cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
                cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )

            elif cell.column == 4 :
                 cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="double") )
            else:    
                cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
                cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )



    #saldo em conta corrente e aplicação financeira
    celulasAzul = f"A{size2+9}"
    cell = sheet[celulasAzul]
    cell.value = f"D.1. Saldo em Conta corrente e Aplicação Financeira"
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
    #restituições não creditadas        
    celulasAzul = f"A{size2+10}"
    cell = sheet[celulasAzul]
    cell.value = f"D.2. Restituições não Creditadas"
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="hair"),right=Side(border_style="hair") )
    #cheques emitido e não descontados
    celulasAzul = f"A{size2+11}"
    cell = sheet[celulasAzul]
    cell.value = f"D.3. Cheques Emitidos e não Descontados"
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )

    #saldos devolvido á Finep
    celulasAzul = f"A{size2+13}"
    cell = sheet[celulasAzul]
    cell.value = f"E. Saldo Devolvido à FINEP"
    cell.fill = PatternFill(start_color=azul, end_color=azul, fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="double") ,bottom=Side(border_style="double"),right=Side(border_style="hair") )


    #Negrito no numero da frente
    celulasAzul = f"B{size2+13}"
    cell = sheet[celulasAzul]
    cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
    cell.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.border = Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="double") )
    cell.number_format = 'R$ #,##0.00'        
   



    #somatorios
    
    #somatorio desepsas correntes

    for col_index in range(2, 6):  
        for row in sheet.iter_rows(min_row=13, max_row=13, min_col=col_index, max_col=col_index):
            column_letter = get_column_letter(col_index)
            cell = sheet[f'{column_letter}13']
            cell.value = f"=SUM({column_letter}14:{column_letter}{size})"
            
    #somatorio linha final
    for col_index in range(2, 6):  
        for row in sheet.iter_rows(min_row=size+8, max_row=size+8, min_col=col_index, max_col=col_index):
            column_letter = get_column_letter(col_index)
            cell = sheet[f'{column_letter}{size+8}']
            cell.value = f"={column_letter}13+{column_letter}{size+1}"         

    #Somatorio total das receitas(B1+B2)
    for col_index in range(2, 5):  
        for row in sheet.iter_rows(min_row=size2+2, max_row=size2+2, min_col=col_index, max_col=col_index):
            column_letter = get_column_letter(col_index)
            cell = sheet[f'{column_letter}{size2+2}']
            cell.value = f"=SUM({column_letter}{size2+3}:{column_letter}{size2+4})"
    
    #somatorio saldo disponivel para o periodo seguinte
    formula = f"=SUM(B{size2+9}:B{size2+11})"
    celula = f'B{size2+8}' 
    sheet[celula] = formula
        
          
    #total acumulado somatorio
     #Adicionar Soma na coluna E
    for rows in sheet.iter_rows(min_row=14, max_row=size+7, min_col=4, max_col=4):
            for cell in rows:
                      cell.value =  f"=SUM(B{cell.row}:C{cell.row})"  



     #brasilia
    brasilia_row = size2 + 16
    brasilia_formula = f"Brasília, 28 de agosto de 2022."
    brasilia_merge_cells = f"A{brasilia_row}:E{brasilia_row}"
    sheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = sheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size2 + 19
    diretor_cargo_row = size2 + 20
    diretor_cpf_row = size2 + 21
    diretor_nome_formula = f"Daniel Monteiro Rosa"
    diretor_cargo_formula = f"Diretor-Financeiro"
    diretor_cpf_formula = f"450.720.272-87"
    diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    sheet.merge_cells(diretor_merge_cells)
    sheet.merge_cells(diretor_cargo_merge_cells)
    sheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = sheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = sheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = sheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12,bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

  #Coordenadora
    coordenadora_row = size2 + 19
    coordenadora_cargo_row = size2 + 20
    coordenadora_cpf_row = size2 + 21
    coordenadora_nome_formula = f"teste"
    coordenadora_cargo_formula = f"Coordenador(a)"
    coordenadora_cpf_formula = f"teste"
    coordenadora_merge_cells = f'C{coordenadora_row}:E{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'C{coordenadora_cargo_row}:E{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'C{coordenadora_cpf_row}:E{coordenadora_cpf_row}'
    sheet.merge_cells(coordenadora_merge_cells)
    sheet.merge_cells(coordenadora_cargo_merge_cells)
    sheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'C{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'C{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'C{coordenadora_cpf_row}'
    top_left_coordenadora_cell = sheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = sheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = sheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    #E melhor fazertodas as bordas no fim pra ter certeza que vao ser aplicadas
    
    
    #borda cabeçario
    for row in sheet.iter_rows(min_row=1, max_row=11,min_col=5,max_col=5):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in sheet.iter_rows(min_row=size2+1, max_row=coordenadora_cpf_row+1,min_col=5,max_col=5):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in sheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )


    workbook.save(tabela)
    workbook.close()

  
    #retorna tamanho de brasilia e de equipamentos
    return size2 + 16

def estiloG(tabela,tamanho,nomeVariavel,nomeTabela,stringTamanho,tamanhoestorno):
    '''Esse estilo e considerado geral por que todas as tabelas que compõe utilizam das mesma colunas.
      
        Argumentos:
            tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho: e o tamanho total de linhas que irao ser geradas dinamicamente correspondente as entradas do respectivo projeto. o valor varia dentre o tamanho das rubricas 
            
            nomeVariavel: variável utilizada para criar o nomes das variaveis dinamicamente para não haver sobreposição de estilos com o mesmo nome. Esses estilos ocorrem nesses codigos:
                                        locals()[input2] = NamedStyle(name=f'{input2}')
                                        locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
                                        locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid"'....
            nomeTabela: variável utilizada para a criação do nome da tabela.Ela deriva das rubricas que são colocadas no input quando essa função e chamada.
            stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referencias das formulas.
            
    '''
    if tamanho == 0:
         tamanho = 1


    nomeSheet=nomeVariavel
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 16
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = 'ccffff'

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:I1')
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO 4'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    #RELAÇÃO DE PAGAMENTOS
    worksheet.merge_cells('A2:I2')
    
    worksheet['A2'] = f'RELAÇÃO DE PAGAMENTOS'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="center",vertical="center")

    #ElementosdeDespesa
   
    
    worksheet['A11'] = f'Elemento de Despesa'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #ElementosdeDespesa
    
    
    worksheet['A14'] = f'DESPESAS REALIZADAS'
    worksheet['A14'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['A14'].alignment = Alignment(horizontal="center",vertical="center")
   
    

    #cabecario que recebe de referencia as celulas A da planilha DEMOSTR. RECEITA E DESPESA A.2
    
   #Convênio nº: 01.14.0032.00
    worksheet['A4'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    #Convenente: 
    worksheet['A5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    #Período de Execução do Convênio:
    worksheet['A6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    #Período Abrangido por este Relatório:
    worksheet['A7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    #Fonte de Recursos:
    worksheet['A8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    #Partícipe (no caso de contrapartida):
    worksheet['A9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A9"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    

    #cabecario que recebe de referencia as celulas C da planilha DEMOSTR. RECEITA E DESPESA A.2
    
    # Fundação de Empreendimentos Científicos e Tecnológicos - FINATEC 		
    worksheet['C5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C5"
    worksheet['C5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C5'].alignment = Alignment(horizontal="left",vertical="center")
    #de 10/02/2014 a 10/06/2023
    worksheet['C6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C6"
    worksheet['C6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C6'].alignment = Alignment(horizontal="left",vertical="center")
    #de 10/02/2014 a 31/07/2022
    worksheet['C7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C7"
    worksheet['C7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C7'].alignment = Alignment(horizontal="left",vertical="center")
    #recursos finep recursos contra partid
    worksheet['C8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C8"
    worksheet['C8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C8'].alignment = Alignment(horizontal="left",vertical="center")
    #xxx
    worksheet['C9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C9"
    worksheet['C9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C9'].alignment = Alignment(horizontal="left",vertical="center")
    
     #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="hair"), right=Side(border_style="hair"),left=Side(border_style='hair') )
    locals()[input2].height = 20

    linha_number = 15
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=9):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 9:
                cell.border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="thin"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')

  
    
    valores = ["Nº DE ORDEM","CREDOR","CNPJ/CPF",'Equivalência na Relação de Itens Apoiados','CHEQUE OU EQUIVALENTE ESTORNADO',"DATA DO CHEQUE",'Nº DO DEPÓSITO','DATA DO DEPÓSITO','VALOR']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=15, max_row=size, min_col=1, max_col=9):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    
    value_to_stop = size  
    start_row = 10

    #estilo mascara de dinheiro
    for row in range(start_row,size+1):
        cell = worksheet[f'i{row}']
        cell.style = locals()[input3]

   #estilocinzasimcinzanao     
    for rows in worksheet.iter_rows(min_row=16, max_row=size, min_col=1, max_col=9):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

               

                if cell.column == 9:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                    if cell.row == size:
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                
     #double em cima double total
                                   
     #bordinha double           
    stringAfinarCelula =size+2
    cellborda = f"I{size+1}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    
    #subtotal
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:H{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="double") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(I16:I{size})"
    celula = f'I{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    input4 = f'row_style_diaria_append{nomeVariavel}'
    #estilo colunas restitucoes creditadas
    locals()[input4] = NamedStyle(name=f'{input4}')
    locals()[input4].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input4].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input4].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input4].height = 30
    locals()[input4].border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"),right=Side(border_style='hair'),left=Side(border_style='hair') )


    row_number = size + 4
   
    for column in range(1, 10):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = locals()[input4]
        if cell.column == 9:
            cell.border = Border(top=Side(border_style="double") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )



    values = ["Nº DE ORDEM","Restituidor","CNPJ/CPF",'Equivalência na Relação de Itens Apoiados',"Cheque equivalente","Data do Cheque",'Nº do Depósito','DATA DO DEPÓSITO','VALOR']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1
        

   
   
    #estorno

    for rows in worksheet.iter_rows(min_row=size+5, max_row=size+4+tamanhoestorno, min_col=1, max_col=9):
        for cell in rows:
            if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))
            if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))

            if cell.column == 9:        
                cell.number_format = 'R$ #,##0.00'
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
                if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
           
           
            
    #bordas,corsimcornao,money
    # Set the height of each row to 60
    for row in worksheet.iter_rows(min_row=size+4, max_row=size+4+tamanhoestorno):
        worksheet.row_dimensions[row[0].row].height = 60

    min_row = size + 4
    max_row = size + 4 + tamanhoestorno

      

    #bordinha double           
    stringAfinarCelula = size + 6 +tamanhoestorno
    cellborda = f"I{size + 5 +tamanhoestorno}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    #subtotal2
    sub_total2_row = size + 6 +tamanhoestorno
    subtotal_merge_cells= f'A{sub_total2_row}:H{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

    sub_formula_row_celula = f'I{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].value = f'=SUM(I{size+5}:I{sub_total2_row-1})'
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="thin") )

      #total1-2
    total12_row = size + 7 + tamanhoestorno
    total12_merge_cells = f'A{total12_row}:H{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="thin") )


    #total_formula
    total_formula_row = size + 7 + tamanhoestorno 
    total_formulaa = f'=I{size+2} - I{sub_total2_row }'
    total_formula_row_celula = f'I{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(bottom=Side(border_style="double"),right=Side(border_style="double") )

    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa

    

    #brasilia
    brasilia_row = size + 8 +tamanhoestorno
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 9 + tamanhoestorno
    diretor_cargo_row = size + 10 + tamanhoestorno
    diretor_cpf_row = size + 11 + tamanhoestorno
    
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:C{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:C{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:C{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 9 + tamanhoestorno 
    coordenadora_cargo_row = size + 10 + tamanhoestorno 
    coordenadora_cpf_row = size + 11 + tamanhoestorno 
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:I{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:I{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:I{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=9):
    #     for cell in row:
    #         cell.border = borda

    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=9, max_col=9):
        for cell in row:
                cell.border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    
    cinzaborda = '9e9e9e'
    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=14,min_col=9,max_col=9):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=brasilia_row, max_row=coordenadora_cpf_row+1,min_col=9,max_col=9):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=9):
        for cell in row:
            if cell.column == 9:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    workbook.save(tabela)
    workbook.close()

    return size+4

def estiloPagamentoPessoal(tabela,tamanho,nomeTabela,stringTamanho,tamanhoestorno):
    '''Esse estilo e considerado geral por que todas as tabelas que compõe utilizam das mesma colunas.
      
        Argumentos:
            tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho: e o tamanho total de linhas que irao ser geradas dinamicamente correspondente as entradas do respectivo projeto. o valor varia dentre o tamanho das rubricas 
            
            nomeVariavel: variável utilizada para criar o nomes das variaveis dinamicamente para não haver sobreposição de estilos com o mesmo nome. Esses estilos ocorrem nesses codigos:
                                        locals()[input2] = NamedStyle(name=f'{input2}')
                                        locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
                                        locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid"'....
            nomeTabela: variável utilizada para a criação do nome da tabela.Ela deriva das rubricas que são colocadas no input quando essa função e chamada.
            stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referencias das formulas.
            
    '''
    if tamanho == 0:
         tamanho = 1


    nomeVariavel = f'PagamentoPessoal'
    nomeSheet=nomeVariavel
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 16
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = 'ccffff'
    

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:J1')
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO 4'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    #RELAÇÃO DE PAGAMENTOS
    worksheet.merge_cells('A2:J2')
    
    worksheet['A2'] = f'RELAÇÃO DE PAGAMENTOS'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="center",vertical="center")

    #ElementosdeDespesa
   
    
    worksheet['A11'] = f'Elemento de Despesa'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #ElementosdeDespesa
    
    
    worksheet['A14'] = f'DESPESAS REALIZADAS'
    worksheet['A14'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['A14'].alignment = Alignment(horizontal="center",vertical="center")
   
    

    #cabecario que recebe de referencia as celulas A da planilha DEMOSTR. RECEITA E DESPESA A.2
    
   
    worksheet['A4'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['A6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A9"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    

    #cabecario que recebe de referencia as celulas C da planilha DEMOSTR. RECEITA E DESPESA A.2
    
    worksheet['C5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C5"
    worksheet['C5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['C6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C6"
    worksheet['C6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C6'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['C7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C7"
    worksheet['C7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['C8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C8"
    worksheet['C8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C8'].alignment = Alignment(horizontal="left",vertical="center")
   
    worksheet['C9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C9"
    worksheet['C9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C9'].alignment = Alignment(horizontal="left",vertical="center")
    
     #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="hair"), right=Side(border_style="hair"),left=Side(border_style='hair') )
    locals()[input2].height = 20

    linha_number = 15
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="thin"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')

  
    
    valores = ["Nº DE ORDEM","BENEFICIÁRIO","CNPJ/CPF",'Vencimento / Tipo de Obrigação Patronal / Tipo de Benefício ','Equivalência na Relação de Itens Apoiados','Nº DO COMPROVANTE OU EQUIVALENTE',"DATA DO COMPROVANTE",'Nº DO CHEQUE OU EQUIVALENTE','DATA DA COMPENSAÇÃO DO CHEQUE','VALOR']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=15, max_row=size, min_col=1, max_col=10):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    
    value_to_stop = size  
    start_row = 10

    #estilo mascara de dinheiro
    for row in range(start_row,size+1):
        cell = worksheet[f'j{row}']
        cell.style = locals()[input3]

   #estilocinzasimcinzanao     
    for rows in worksheet.iter_rows(min_row=16, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

               

                if cell.column == 10:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                    if cell.row == size:
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                
     #double em cima double total
                                   
     #bordinha double           
    stringAfinarCelula =size+2
    cellborda = f"J{size+1}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    
    #subtotal
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="double") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(J16:J{size})"
    celula = f'J{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    input4 = f'row_style_diaria_append{nomeVariavel}'
    #estilo colunas restitucoes creditadas
    locals()[input4] = NamedStyle(name=f'{input4}')
    locals()[input4].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input4].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input4].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input4].height = 30
    locals()[input4].border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"),right=Side(border_style='hair'),left=Side(border_style='hair') )


    row_number = size + 4
   
    for column in range(1, 11):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = locals()[input4]
        if cell.column == 10:
            cell.border = Border(top=Side(border_style="double") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )



    values = ["Nº DE ORDEM","RESTITUIDOR","CNPJ/CPF",'Vencimento / Tipo de Obrigação Patronal / Tipo de Benefício ','Equivalência na Relação de Itens Apoiados','Nº DO COMPROVANTE OU EQUIVALENTE',"DATA DO COMPROVANTE",'Nº DO CHEQUE OU EQUIVALENTE','DATA DA COMPENSAÇÃO DO CHEQUE','VALOR']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1
        

   
   
    #estorno

    for rows in worksheet.iter_rows(min_row=size+5, max_row=size+4+tamanhoestorno, min_col=1, max_col=10):
        for cell in rows:
            if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))
            if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))

            if cell.column == 10:        
                cell.number_format = 'R$ #,##0.00'
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
                if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
           
           
            
    #bordas,corsimcornao,money
    # Set the height of each row to 60
    for row in worksheet.iter_rows(min_row=size+4, max_row=size+4+tamanhoestorno):
        worksheet.row_dimensions[row[0].row].height = 60

    min_row = size + 4
    max_row = size + 4 + tamanhoestorno

      

    #bordinha double           
    stringAfinarCelula = size + 6 +tamanhoestorno
    cellborda = f"J{size + 5 +tamanhoestorno}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    #subtotal2
    sub_total2_row = size + 6 +tamanhoestorno
    subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

    sub_formula_row_celula = f'J{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].value = f'=SUM(J{size+5}:J{sub_total2_row-1})'
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="thin") )

      #total1-2
    total12_row = size + 7 + tamanhoestorno
    total12_merge_cells = f'A{total12_row}:I{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="thin") )


    #total_formula
    total_formula_row = size + 7 + tamanhoestorno 
    total_formulaa = f'=J{size+2} - J{sub_total2_row }'
    total_formula_row_celula = f'J{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(bottom=Side(border_style="double"),right=Side(border_style="double") )

    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa

    

    #brasilia
    brasilia_row = size + 8 +tamanhoestorno
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:J{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 9 + tamanhoestorno
    diretor_cargo_row = size + 10 + tamanhoestorno
    diretor_cpf_row = size + 11 + tamanhoestorno
    
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:C{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:C{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:C{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 9 + tamanhoestorno 
    coordenadora_cargo_row = size + 10 + tamanhoestorno 
    coordenadora_cpf_row = size + 11 + tamanhoestorno 
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=9):
    #     for cell in row:
    #         cell.border = borda

    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=10, max_col=10):
        for cell in row:
                cell.border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    
    cinzaborda = '9e9e9e'
    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=14,min_col=10,max_col=10):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=brasilia_row, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
        for cell in row:
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    workbook.save(tabela)
    workbook.close()

    return size+4

def estiloElementoDeDespesa1415Diarias(tabela,tamanho,nomeTabela,stringTamanho,tamanhoestorno):
    '''Esse estilo e considerado geral por que todas as tabelas que compõe utilizam das mesma colunas.
      
        Argumentos:
            tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho: e o tamanho total de linhas que irao ser geradas dinamicamente correspondente as entradas do respectivo projeto. o valor varia dentre o tamanho das rubricas 
            
            nomeVariavel: variável utilizada para criar o nomes das variaveis dinamicamente para não haver sobreposição de estilos com o mesmo nome. Esses estilos ocorrem nesses codigos:
                                        locals()[input2] = NamedStyle(name=f'{input2}')
                                        locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
                                        locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid"'....
            nomeTabela: variável utilizada para a criação do nome da tabela.Ela deriva das rubricas que são colocadas no input quando essa função e chamada.
            stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referencias das formulas.
            
    '''
    if tamanho == 0:
         tamanho = 1


    nomeVariavel = f'Diaria'
    nomeSheet=nomeVariavel
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 16
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = 'ccffff'
    

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão
    worksheet.column_dimensions['k'].width = 35 #data de emissão
    worksheet.column_dimensions['l'].width = 35 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:L1')
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO 4'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    #RELAÇÃO DE PAGAMENTOS
    worksheet.merge_cells('A2:L2')
    
    worksheet['A2'] = f'RELAÇÃO DE PAGAMENTOS'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="center",vertical="center")

    #ElementosdeDespesa
   
    
    worksheet['A11'] = f'Elemento de Despesa'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #ElementosdeDespesa
    
    
    worksheet['A14'] = f'DESPESAS REALIZADAS'
    worksheet['A14'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['A14'].alignment = Alignment(horizontal="center",vertical="center")
   
    

    #cabecario que recebe de referencia as celulas A da planilha DEMOSTR. RECEITA E DESPESA A.2
    
   
    worksheet['A4'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['A6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A9"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    

    #cabecario que recebe de referencia as celulas C da planilha DEMOSTR. RECEITA E DESPESA A.2
    
    worksheet['C5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C5"
    worksheet['C5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['C6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C6"
    worksheet['C6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C6'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['C7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C7"
    worksheet['C7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['C8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C8"
    worksheet['C8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C8'].alignment = Alignment(horizontal="left",vertical="center")
   
    worksheet['C9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C9"
    worksheet['C9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C9'].alignment = Alignment(horizontal="left",vertical="center")
    
     #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="hair"), right=Side(border_style="hair"),left=Side(border_style='hair') )
    locals()[input2].height = 20

    linha_number = 15
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=12):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="thin"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')

  
    
    valores = ["Nº DE ORDEM","BENEFICIÁRIO","CNPJ/CPF",'Destino','Nº DE DIÁRIAS UTILIZADAS','Evento',"Equivalência na Relação de Itens Apoiados",'Nº DO RECIBO OU EQUIVALENTE','DATA DO RECIBO','Nº DO CHEQUE OU EQUIVALENTE','DATA DA COMPENSAÇÃO DO CHEQUE','VALOR']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=15, max_row=size, min_col=1, max_col=12):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    
    value_to_stop = size  
    start_row = 10

    #estilo mascara de dinheiro
    for row in range(start_row,size+1):
        cell = worksheet[f'L{row}']
        cell.style = locals()[input3]

   #estilocinzasimcinzanao     
    for rows in worksheet.iter_rows(min_row=16, max_row=size, min_col=1, max_col=12):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

               

                if cell.column == 12:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                    if cell.row == size:
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                
     #double em cima double total
                                   
     #bordinha double           
    stringAfinarCelula =size+2
    cellborda = f"L{size+1}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    
    #subtotal
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:K{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="double") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(L16:L{size})"
    celula = f'L{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    input4 = f'row_style_diaria_append{nomeVariavel}'
    #estilo colunas restitucoes creditadas
    locals()[input4] = NamedStyle(name=f'{input4}')
    locals()[input4].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input4].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input4].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input4].height = 30
    locals()[input4].border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"),right=Side(border_style='hair'),left=Side(border_style='hair') )


    row_number = size + 4
   
    for column in range(1, 13):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = locals()[input4]
        if cell.column == 12:
            cell.border = Border(top=Side(border_style="double") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )



    values = ["Nº DE ORDEM","RESTITUIDOR","CNPJ/CPF","Destino",'Nº DE DIÁRIAS UTILIZADAS','Evento','Equivalência na Relação de Itens Apoiados',"CHEQUE OU EQUIVALENTE ESTORNADO",'DATA DO CHEQUE','Nº DO DEPÓSITO','DATA DO DEPÓSITO','VALOR']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1
        

   
   
    #estorno

    for rows in worksheet.iter_rows(min_row=size+5, max_row=size+4+tamanhoestorno, min_col=1, max_col=12):
        for cell in rows:
            if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))
            if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))

            if cell.column == 12:        
                cell.number_format = 'R$ #,##0.00'
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
                if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
           
           
            
    #bordas,corsimcornao,money
    # Set the height of each row to 60
    for row in worksheet.iter_rows(min_row=size+4, max_row=size+4+tamanhoestorno):
        worksheet.row_dimensions[row[0].row].height = 60

    min_row = size + 4
    max_row = size + 4 + tamanhoestorno

      

    #bordinha double           
    stringAfinarCelula = size + 6 +tamanhoestorno
    cellborda = f"L{size + 5 +tamanhoestorno}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    #subtotal2
    sub_total2_row = size + 6 +tamanhoestorno
    subtotal_merge_cells= f'A{sub_total2_row}:K{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

    sub_formula_row_celula = f'L{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].value = f'=SUM(L{size+5}:L{sub_total2_row-1})'
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="thin") )

      #total1-2
    total12_row = size + 7 + tamanhoestorno
    total12_merge_cells = f'A{total12_row}:K{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="thin") )


    #total_formula
    total_formula_row = size + 7 + tamanhoestorno 
    total_formulaa = f'=L{size+2} - L{sub_total2_row }'
    total_formula_row_celula = f'L{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(bottom=Side(border_style="double"),right=Side(border_style="double") )

    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa

    

    #brasilia
    brasilia_row = size + 8 +tamanhoestorno
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:L{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 9 + tamanhoestorno
    diretor_cargo_row = size + 10 + tamanhoestorno
    diretor_cpf_row = size + 11 + tamanhoestorno
    
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:F{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:F{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:F{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 9 + tamanhoestorno 
    coordenadora_cargo_row = size + 10 + tamanhoestorno 
    coordenadora_cpf_row = size + 11 + tamanhoestorno 
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'G{coordenadora_row}:L{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'G{coordenadora_cargo_row}:L{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'G{coordenadora_cpf_row}:L{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'G{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'G{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'G{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=9):
    #     for cell in row:
    #         cell.border = borda

    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=12, max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    
    cinzaborda = '9e9e9e'
    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=14,min_col=12,max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=brasilia_row, max_row=coordenadora_cpf_row+1,min_col=12,max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=12):
        for cell in row:
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    workbook.save(tabela)
    workbook.close()

    return size+4

def estiloElementoDeDespesa33PassagensEDespesa(tabela,tamanho,nomeTabela,stringTamanho,tamanhoestorno):
    '''Esse estilo e considerado geral por que todas as tabelas que compõe utilizam das mesma colunas.
      
        Argumentos:
            tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho: e o tamanho total de linhas que irao ser geradas dinamicamente correspondente as entradas do respectivo projeto. o valor varia dentre o tamanho das rubricas 
            
            nomeVariavel: variável utilizada para criar o nomes das variaveis dinamicamente para não haver sobreposição de estilos com o mesmo nome. Esses estilos ocorrem nesses codigos:
                                        locals()[input2] = NamedStyle(name=f'{input2}')
                                        locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
                                        locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid"'....
            nomeTabela: variável utilizada para a criação do nome da tabela.Ela deriva das rubricas que são colocadas no input quando essa função e chamada.
            stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referencias das formulas.
            
    '''
    if tamanho == 0:
         tamanho = 1

    nomeVariavel = f'Diaria'
    nomeSheet=nomeVariavel
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 16
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = 'ccffff'
    

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão
    worksheet.column_dimensions['k'].width = 35 #data de emissão
    worksheet.column_dimensions['l'].width = 35 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:L1')
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO 4'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    #RELAÇÃO DE PAGAMENTOS
    worksheet.merge_cells('A2:L2')
    
    worksheet['A2'] = f'RELAÇÃO DE PAGAMENTOS'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="center",vertical="center")

    #ElementosdeDespesa
   
    
    worksheet['A11'] = f'Elemento de Despesa'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #ElementosdeDespesa
    
    
    worksheet['A14'] = f'DESPESAS REALIZADAS'
    worksheet['A14'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['A14'].alignment = Alignment(horizontal="center",vertical="center")
   
    

    #cabecario que recebe de referencia as celulas A da planilha DEMOSTR. RECEITA E DESPESA A.2
    
   
    worksheet['A4'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['A6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A9"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    

    #cabecario que recebe de referencia as celulas C da planilha DEMOSTR. RECEITA E DESPESA A.2
    
    worksheet['C5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C5"
    worksheet['C5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['C6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C6"
    worksheet['C6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C6'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['C7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C7"
    worksheet['C7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['C8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C8"
    worksheet['C8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C8'].alignment = Alignment(horizontal="left",vertical="center")
   
    worksheet['C9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C9"
    worksheet['C9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['C9'].alignment = Alignment(horizontal="left",vertical="center")
    
     #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="hair"), right=Side(border_style="hair"),left=Side(border_style='hair') )
    locals()[input2].height = 20

    linha_number = 15
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=12):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="thin"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')

  
    
    valores = ["Nº DE ORDEM","CREDOR","CNPJ/CPF",'BENEFICIÁRIO','CPF DO BENEFICIÁRIO','TRECHO',"Equivalência na Relação de Itens Apoiados",'Nº DA NOTA FISCAL OU EQUIVALENTE','DATA DA NOTA FISCAL','Nº DO CHEQUE OU EQUIVALENTE','DATA DA COMPENSAÇÃO DO CHEQUE','VALOR']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=15, max_row=size, min_col=1, max_col=12):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    
    value_to_stop = size  
    start_row = 10

    #estilo mascara de dinheiro
    for row in range(start_row,size+1):
        cell = worksheet[f'L{row}']
        cell.style = locals()[input3]

   #estilocinzasimcinzanao     
    for rows in worksheet.iter_rows(min_row=16, max_row=size, min_col=1, max_col=12):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

               

                if cell.column == 12:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                    if cell.row == size:
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                
     #double em cima double total
                                   
     #bordinha double           
    stringAfinarCelula =size+2
    cellborda = f"L{size+1}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    
    #subtotal
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:K{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="double") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(L16:L{size})"
    celula = f'L{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    input4 = f'row_style_diaria_append{nomeVariavel}'
    #estilo colunas restitucoes creditadas
    locals()[input4] = NamedStyle(name=f'{input4}')
    locals()[input4].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input4].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input4].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input4].height = 30
    locals()[input4].border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"),right=Side(border_style='hair'),left=Side(border_style='hair') )


    row_number = size + 4
   
    for column in range(1, 13):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = locals()[input4]
        if cell.column == 12:
            cell.border = Border(top=Side(border_style="double") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )


    #            A               B           C           D           E               F                   G           H                                      I                                   J               K                    L
    values = ["Nº DE ORDEM","RESTITUIDOR","CNPJ/CPF","BENEFICIÁRIO",'CPF DO BENEFICIÁRIO','TRECHO',"Equivalência na Relação de Itens Apoiados","CHEQUE OU EQUIVALENTE ESTORNADO",'DATA DO CHEQUE','Nº DO DEPÓSITO','DATA DO DEPÓSITO','VALOR']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1
        

   
   
    #estorno

    for rows in worksheet.iter_rows(min_row=size+5, max_row=size+4+tamanhoestorno, min_col=1, max_col=12):
        for cell in rows:
            if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))
            if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))

            if cell.column == 12:        
                cell.number_format = 'R$ #,##0.00'
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
                if cell.row == size+4+tamanhoestorno:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair"))
           
           
            
    #bordas,corsimcornao,money
    # Set the height of each row to 60
    for row in worksheet.iter_rows(min_row=size+4, max_row=size+4+tamanhoestorno):
        worksheet.row_dimensions[row[0].row].height = 60

    min_row = size + 4
    max_row = size + 4 + tamanhoestorno

      

    #bordinha double           
    stringAfinarCelula = size + 6 +tamanhoestorno
    cellborda = f"L{size + 5 +tamanhoestorno}"
    bordacell = worksheet[cellborda]
    bordacell.border = Border(right =Side(border_style="double"))

    #subtotal2
    sub_total2_row = size + 6 +tamanhoestorno
    subtotal_merge_cells= f'A{sub_total2_row}:K{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

    sub_formula_row_celula = f'L{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].value = f'=SUM(L{size+5}:L{sub_total2_row-1})'
   
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="double") ,bottom=Side(border_style="thin") )

      #total1-2
    total12_row = size + 7 + tamanhoestorno
    total12_merge_cells = f'A{total12_row}:K{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="hair") ,bottom=Side(border_style="double"),right=Side(border_style="thin") )


    #total_formula
    total_formula_row = size + 7 + tamanhoestorno 
    total_formulaa = f'=L{size+2} - L{sub_total2_row }'
    total_formula_row_celula = f'L{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(bottom=Side(border_style="double"),right=Side(border_style="double") )

    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa

    

    #brasilia
    brasilia_row = size + 8 +tamanhoestorno
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:K{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 9 + tamanhoestorno
    diretor_cargo_row = size + 10 + tamanhoestorno
    diretor_cpf_row = size + 11 + tamanhoestorno
    
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:F{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:F{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:F{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 9 + tamanhoestorno 
    coordenadora_cargo_row = size + 10 + tamanhoestorno 
    coordenadora_cpf_row = size + 11 + tamanhoestorno 
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'G{coordenadora_row}:L{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'G{coordenadora_cargo_row}:L{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'G{coordenadora_cpf_row}:L{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'G{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'G{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'G{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=9):
    #     for cell in row:
    #         cell.border = borda

    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=12, max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="double") ,bottom=Side(border_style="hair"), right=Side(border_style="double") )
                cell.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    
    cinzaborda = '9e9e9e'
    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=14,min_col=12,max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=brasilia_row, max_row=coordenadora_cpf_row+1,min_col=12,max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=12):
        for cell in row:
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    workbook.save(tabela)
    workbook.close()

    return size+4

def estilo_conciliacoes_bancaria(tabela,tamanho,tamanho2,stringTamanho):
    """Estilo um pouco diferente pois necessita de dois aspectos dinâmicos que é primeiramente a quantidade de entradas de pagamento de tarifas bancárias e por fim a quantidade de estorno. Sabendo
    esses valores é possivel criar a tabela.
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade de transfêrencia bancárias realizada.
        tamanho2:Corresponde ao tamanho dos estornos.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    if tamanho == 0:
        tamanho = 1
    #pegar o arquivo e carregar ele um worksheet da pagaina Conciliação Bancária
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Conciliação Bancária A.3']
    
    #size e o tamanho da quantidade de arquivos recebido no argumento tamanho mais o tamanho do cabecario que no caso da fub e de 16
    size = tamanho + 18
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"

    azul_claro = 'ccffff'

    #Borda apenas do lado direito da cedula, uma borda mas larga
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
   


    worksheet.column_dimensions['a'].width =69
    worksheet.column_dimensions['b'].width = 20
    worksheet.column_dimensions['c'].width = 20
    

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:C1')
    worksheet['A1'] = f'ANEXO 3'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    
    worksheet.merge_cells('A2:C2')
    worksheet['A2'] = f'CONCILIAÇÃO BANCÁRIA'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="center",vertical="center")
    
  #cabecario que recebe de referencia as celulas A da planilha DEMOSTR. RECEITA E DESPESA A.2
    
   
    worksheet['A4'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['A6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A9"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    

    #cabecario que recebe de referencia as celulas C da planilha DEMOSTR. RECEITA E DESPESA A.2
    
    worksheet['B5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C5"
    worksheet['B5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['B5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['B6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C6"
    worksheet['B6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['B6'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['B7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C7"
    worksheet['B7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['B7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['B8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C8"
    worksheet['B8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['B8'].alignment = Alignment(horizontal="left",vertical="center")
   
    worksheet['B9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!C9"
    worksheet['B9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['B9'].alignment = Alignment(horizontal="left",vertical="center")


    

    worksheet['A11'] = 'A. SALDO CONFORME EXTRATOS BANCÁRIOS NA DATA FINAL DO PERÍODO'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A11'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet['A11'].border = Border(top=Side(border_style="medium"),bottom=Side(border_style="hair"),right=Side(border_style='medium'))
    worksheet.merge_cells('A11:C11')

    
    worksheet['A12'] = 'DISCRIMINAÇÃO'
    worksheet['A12'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A12'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A12'].border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'))
    worksheet.merge_cells('A12:B12')

    worksheet['C12'] = 'VALOR'
    worksheet['C12'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['C12'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['C12'].border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'))
    worksheet['C13'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['C13'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['C13'].border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'))
    worksheet['C14'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['C14'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['C14'].border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'))


    worksheet['A13'] = 'a)Saldo de Conta Corrente(R$)'
    worksheet['A13'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A13'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A13'].border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'))
    worksheet.merge_cells('A13:B13')
    
    
    worksheet['A14'] = 'b)Saldo de Aplicações Financeiras(R$)'
    worksheet['A14'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A14'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A14'].border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'))
    worksheet.merge_cells('A14:B14')
    
    
    worksheet['A15'] = 'c) TOTAL (a+b)'
    worksheet['A15'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A15'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A15'].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='thin'),left=Side(border_style='thin'),top=Side(border_style="thin"))
    worksheet['A15'].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")

    worksheet.merge_cells('A15:B15')

    worksheet['C15'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['C15'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['C15'].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),left=Side(border_style='thin'),top=Side(border_style='thin'))
    worksheet['C15'].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")



    
    worksheet['A16'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A16'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A16'].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'))
    worksheet.merge_cells('A16:C16')    


    
    worksheet['A17'] = 'B. RESTUIÇÕES NÃO CREDITADAS ATÉ A DATA FINAL DESTA PRESTAÇÃO DE CONTAS'
    worksheet['A17'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A17'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A17'].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),left=Side(border_style='thin'),top=Side(border_style='medium'))
    worksheet['A17'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet.merge_cells('A17:C17')

 

    random_number = random.randint(1, 10000)

    custom_number_format_conciliacoes = []
    # MASCARA R$
    if custom_number_format_conciliacoes!= False: 
        custom_number_format_conciliacoes = NamedStyle(name=f'custom_number_format_conciliacoes{random_number}')
        custom_number_format_conciliacoes.number_format = 'R$ #,##0.00'
        custom_number_format_conciliacoes.font = Font(name="Arial", size=12, color="000000")
        custom_number_format_conciliacoes.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
    #stylecinza
    start_row = 19
    for row in range(start_row,size + tamanho2+ 7):
        cell = worksheet[f'C{row}']
        cell.style = custom_number_format_conciliacoes
        
    for rows in worksheet.iter_rows(min_row=18, max_row=size, min_col=1, max_col=3):
            for cell in rows:
                if not cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'),top=Side(border_style='hair'),left=Side(border_style='hair'))


    row_number = 18
    values = ['Descrição',"Data","Valor"]
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        
        coluna = coluna + 1
    
    # FORMULATOTAL
    if size > 18:
        formula = f"=SUM(C19:C{size-1})"
    else:
        formula = f"=SUM(C17:C18)"
    size = size +1
    celula = f'C{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'),left=Side(border_style='medium'))

    #Total
    celula_total_merge = f'A{size}:B{size}'
    celula_total = F'A{size}'
    worksheet.row_dimensions[size].height = 38
    worksheet.row_dimensions[size-1].height = 3
    worksheet[celula_total] = f'd)TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula_total].alignment = Alignment(horizontal="center",vertical="center")
    worksheet[celula_total].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'),left=Side(border_style='medium'))
    worksheet.merge_cells(celula_total_merge)
    #'3. Restituições não creditadas pelo banco até a data final do período'
    string_reituicoes_creditadas = f'A{size+2}:C{size+2}'
    row_creditadas = f'A{size+2}'

    worksheet[row_creditadas] = 'C. VALORES NÃO DEBITADOS ATÉ A DATA FINAL DESTA PRESTAÇÃO DE CONTAS'
    worksheet[row_creditadas].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[row_creditadas].alignment = Alignment(horizontal="left",vertical="center")
    worksheet[row_creditadas].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[row_creditadas].border = Border(top=Side(border_style='medium'))
    worksheet.merge_cells(string_reituicoes_creditadas)

    #data valor documento descrição
    row_number = size+3
    values = ['Descrição',"Data","Valor"]
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1

  


    for rows in worksheet.iter_rows(min_row=18, max_row=18, min_col=1, max_col=3):  
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    for rows in worksheet.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=3):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)



    for row in range(size+4,size+4+tamanho2):
        cell = worksheet[f'C{row}']
        cell.style = custom_number_format_conciliacoes
        
    for rows in worksheet.iter_rows(min_row=size+3, max_row=size+3+tamanho2, min_col=1, max_col=3):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(bottom=Side(border_style="hair"),right=Side(border_style='hair'),top=Side(border_style='hair'),left=Side(border_style='hair'))
        


    # FORMULATOTALrestituição
    formula = f"=SUM(C{size+4}:C{size+tamanho2+3})"
    celula = f'C{size+tamanho2+5}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'),left=Side(border_style='medium'))
    
    #Total
    celula_total = F'A{size+tamanho2+5}'
    string_celula_total= f'A{size+tamanho2+5}:B{size+tamanho2+5}'
    worksheet[celula_total].alignment = Alignment(horizontal="center",vertical="center")
    worksheet[celula_total] = f'e) TOTAL'
    worksheet.row_dimensions[size+tamanho2+5].height = 38
    worksheet.row_dimensions[size+tamanho2+4].height = 3
    worksheet[celula_total].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula_total].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'),left=Side(border_style='medium'))
    worksheet.merge_cells(string_celula_total)

    #Saldo disponível p/ período seguinte (1 +2 - 3)
    string_saldo_disponivel = f'A{size+3+tamanho2+3}:B{size+3+tamanho2+3}'
    celula_string_saldo = f'A{size+tamanho2+6}'
    worksheet.row_dimensions[size+tamanho2+6].height = 30
    worksheet[celula_string_saldo].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[celula_string_saldo].alignment = Alignment(horizontal="left",vertical="center")
    worksheet[celula_string_saldo].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula_string_saldo]= f'D. SALDO CONTÁBIL (c+d-e)'
    worksheet[celula_string_saldo].border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'),left=Side(border_style='medium'))
    worksheet.merge_cells(string_saldo_disponivel)
    
    #total saldo diposnivel
    
    celula_string_total = f'C{size+tamanho2+6}'
    saldodiposnivelformat_conciliacoes = NamedStyle(name=f'saldodiposnivelformat_conciliacoes{random_number}')
    saldodiposnivelformat_conciliacoes.number_format = 'R$ #,##0.00'
    saldodiposnivelformat_conciliacoes.font = Font(name="Arial", size=12, color="000000")
    saldodiposnivelformat_conciliacoes.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    saldodiposnivelformat_conciliacoes.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    saldodiposnivelformat_conciliacoes.border = Border(bottom=Side(border_style="medium"),right=Side(border_style='medium'),top=Side(border_style='medium'),left=Side(border_style='medium'))

    celular = worksheet[celula_string_total]
    celular.style = saldodiposnivelformat_conciliacoes
    celular.value = f'=C15+C{size} - C{size+tamanho2+5}'#corrigir
    # #saldo anterior
    # formula = f"Saldo anterior"
    # celula = f'A16'
    # worksheet[celula] = formula
    # worksheet[celula].font = Font(name="Arial", size=12, color="000000")

    # formula = f"Diversos"
    # celula = f'C16'
    # worksheet[celula] = formula
    # worksheet[celula].font = Font(name="Arial", size=12, color="000000")

    # formula = f"Tarifas Prestações Anteriores"
    # celula = f'D16'
    # worksheet[celula] = formula
    # worksheet[celula].font = Font(name="Arial", size=12, color="000000")

     #brasilia
    brasilia_row = size + tamanho2+ 8
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:C{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 10 + tamanho2
    diretor_cargo_row = size + 11 + tamanho2
    diretor_cpf_row = size + 12 + tamanho2

    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    # diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    # diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    # diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    # worksheet.merge_cells(diretor_merge_cells)
    # worksheet.merge_cells(diretor_cargo_merge_cells)
    # worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = diretor_row
    coordenadora_cargo_row = diretor_cargo_row
    coordenadora_cpf_row = diretor_cpf_row
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'B{coordenadora_row}:C{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'B{coordenadora_cargo_row}:C{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'B{coordenadora_cpf_row}:C{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'B{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'B{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'B{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.border = borda
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")
    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=10,min_col=3,max_col=3):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=brasilia_row-1, max_row=coordenadora_cpf_row+1,min_col=3,max_col=3):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda pagina
    for row in worksheet.iter_rows(min_row=11, max_row=brasilia_row-2,min_col=4,max_col=4):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="none") ,left =Side(border_style="medium") ,bottom=Side(border_style="none") )
           
                
    #
    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=3):
        for cell in row:
            if cell.column == 3:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    for rows in worksheet.iter_rows(min_row=size+3, max_row=size+3, min_col=1, max_col=3):  
        for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)


    workbook.save(tabela)
    workbook.close()

def estilo_rendimento_de_aplicacao(tabela,tamanho,stringTamanho):
    """Estilo da rendimento de aplicação, tabela com as colunas periodo, saldo anterior,valor aplicado no período,valor resgatado no período,rendimento bruto,imposto,rendimento luiquido,saldo.
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade da tabela de rencimentos.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    if tamanho == 0:
        tamanho = 1
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Rendimento de Aplicação']
    
    random_number = random.randint(1, 10000)    
    size = tamanho + 14
    worksheet.row_dimensions[2].height = 48
    worksheet.row_dimensions[1].height = 48
    testeRow = 15;

    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False

    #imagensfinep

    image_names = [
                    'Finep.png',
            
        ]

    # Path to the images
    path = "/imagemFinep/"
   
    # List to hold Image objects
    images = []

    nomePasta = "imagemFinep"
    diretorio = os.path.dirname(__file__)

    # Loop through the list of image names and create Image objects
    for i, name in enumerate(image_names):
        caminhoImage = os.path.join(diretorio, nomePasta, name)
        pil_image = PILImage.open(caminhoImage)
        pil_image.save(caminhoImage)
        img = Image(caminhoImage)
        images.append(img)



    worksheet.add_image(images[0], "A1")# FINEP
   


    worksheet.column_dimensions['a'].width = 20
    worksheet.column_dimensions['b'].width = 20
    worksheet.column_dimensions['c'].width = 20
    worksheet.column_dimensions['d'].width = 20
    worksheet.column_dimensions['e'].width = 20
    worksheet.column_dimensions['f'].width = 20
    worksheet.column_dimensions['g'].width = 20
    worksheet.column_dimensions['h'].width = 20
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:H2')
    worksheet['A1'] = f'Demonstrativo dos Ganhos Auferidos com Aplicações '
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
 
    #unidade executora
  
    worksheet['A3'] = "Unidade Executora:"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A3'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet['A3'].border = Border(top=Side(border_style="double"))
    worksheet.merge_cells('A3:F3')
    
    #convenio
    worksheet['G3'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A4"
    worksheet['G3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['G3'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['G3'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet['G3'].border = Border(right=Side(border_style="double"),top=Side(border_style="double"))
    worksheet.merge_cells('G3:H3')

    #projeto
    worksheet['A5'] = "Projeto"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A5'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet.merge_cells('A5:F5')

    #Período de Execução do Convênio:
    worksheet['G5'] = "Período de Execução do Convênio"
    worksheet['G5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['G5'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['G5'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet['G5'].border = Border(right=Side(border_style="double"))
    worksheet.merge_cells('G5:H5')

    #linha 4
    worksheet['A4'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet.merge_cells('A4:F4')
    worksheet['G4'].border = Border(right=Side(border_style="double"))
    worksheet['G4'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet.merge_cells('G4:H4')
    #linha 6
    worksheet['G6'].border = Border(right=Side(border_style="double"))
    worksheet['G6'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet.merge_cells('G6:H6')

    #vaireceberinputnopreencher
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A6'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet.merge_cells('A6:F7')

    #Período Abrangido por este Relatório:
    worksheet['G7'] = "Período Abrangido por este Relatório"
    worksheet['G7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['G7'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['G7'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet['G7'].border = Border(right=Side(border_style="double"))
    worksheet.merge_cells('G7:H7')


    #barrvazia
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A8'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet['A8'].border = Border(bottom=Side(border_style="double"))
    worksheet.merge_cells('A8:F8')
     #Período Abrangido por este Relatório:
    
    worksheet['G8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['G8'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['G8'].fill = openpyxl.styles.PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type='solid')
    worksheet['G8'].border = Border(bottom=Side(border_style="double"),right=Side(border_style="double"))
    worksheet.merge_cells('G8:H8')


    worksheet.row_dimensions[9].height = 20
    worksheet.row_dimensions[10].height = 20

    worksheet.merge_cells('A9:H10')
    worksheet['A9'] = f'APLICAÇÃO FINANCEIRA'
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A9'].alignment = Alignment(horizontal="center",vertical="center")

    worksheet.row_dimensions[11].height = 20
    worksheet.row_dimensions[12].height = 20

   
    worksheet['A11'] = f'APLICAÇÃO FINANCEIRA - RF REF DI PLUS ÁGIL'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A11'].border = Border(bottom=Side(border_style="double"),right=Side(border_style="double"),left=Side(border_style="double"),top=Side(border_style="double"))
    worksheet.merge_cells('A11:H12')
   


    worksheet.row_dimensions[13].height = 20
    worksheet.row_dimensions[14].height = 20

    # #stylecinza
    start_row = 11
    for rows in worksheet.iter_rows(min_row=13, max_row=14, min_col=1, max_col=8):
            for cell in rows:
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")

    #cabcerario abaixo de aplicação financeira
    row_number = 13
    values = ["Período","Saldo Anterior","Valor Aplicado no período",'Valor Resgatado no Período','Rendimento Bruto','Imposto de Renda / IOF','Rendimento Líquido','Saldo']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1

    for i in range(1,9):
        worksheet.merge_cells(start_row=13,end_row=14,start_column=i,end_column=i)


    # #RENDIMENTO LIQUIDO
    # # print(size)
    for row in worksheet.iter_rows(min_row=testeRow, max_row=size, min_col=7, max_col=7):
        for cell in row:
                stringSaldo = f"=E{cell.row} - F{cell.row}"
                cell.value = stringSaldo
          

                

    #BARRAS DE DADOS
    start_row = 14
    for rows in worksheet.iter_rows(min_row=testeRow, max_row=size, min_col=1, max_col=8):
            for cell in rows:
                if cell.row % 2==0:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #MASCARA VERMELHO
    for rows in worksheet.iter_rows(min_row=testeRow, max_row=size-1, min_col=6, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="f90000")
                cell.number_format ='#,##0.00'
   
    #MASCARANEGRITO
    for rows in worksheet.iter_rows(min_row=testeRow, max_row=size-1, min_col=1, max_col=1):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                
   
    # # #MASCARA AZUL
    # # for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=6, max_col=6):
    # #     for cell in rows:
    # #         cell.font = Font(name="Arial", size=12, color="141fca")
    # #         cell.number_format ='#,##0.00'

    # # for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=7, max_col=7):
    # #         for cell in rows:
    # #             cell.font = Font(name="Arial", size=12, color="141fca",bold=True)
    # #             cell.number_format ='#,##0.00'
   

    
    #barra de totais
    formula = f"Saldo anterior"
    celula = f'A{testeRow}'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    
    

    # FORMULATOTAL
    #B
    size = size + 1
    formula = f"=SUM(B{testeRow}:B{size-1})"
    celula = f'B{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
     #C
    formula = f"=SUM(C{testeRow}:C{size-1})"
    celula = f'C{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #D
    formula = f"=SUM(D{testeRow}:D{size-1})"
    celula = f'D{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #E
    formula = f"=SUM(E{testeRow}:E{size-1})"
    celula = f'E{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #F
    formula = f"=SUM(F{testeRow}:F{size-1})"
    celula = f'F{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #G
    formula = f"=SUM(G{testeRow}:G{size-1})"
    celula = f'G{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #H
    formula = f"=SUM(H{testeRow}:H{size-1})"
    celula = f'H{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)

    #Total
    celula_total = F'A{size}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)    
    #borda total



    
    #brasilia
    brasilia_row = size + 2
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:H{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 6 
    diretor_cargo_row = size + 7 
    diretor_cpf_row = size + 8
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:C{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:C{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:C{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size  + 6
    coordenadora_cargo_row = size + 7 
    coordenadora_cpf_row = size + 8
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:H{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:H{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:H{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

 #borda cabeçario
    for row in worksheet.iter_rows(min_row=13, max_row=brasilia_row-2,min_col=8,max_col=8):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="double") ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=brasilia_row-1, max_row=coordenadora_cpf_row+1,min_col=8,max_col=8):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=8):
        for cell in row:
            if cell.column == 8:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    
     #borda da barra total
    for row in worksheet.iter_rows(min_row=size, max_row=size,min_col=1,max_col=8):
        for cell in row:
                cell.border = Border(top=Side(border_style="double") ,right = Side(border_style="none") ,left =Side(border_style="none") ,bottom=Side(border_style="double") )
        if cell.column == 8:
                 cell.border = Border(top=Side(border_style="double") ,right = Side(border_style="double") ,left =Side(border_style="none") ,bottom=Side(border_style="double") )


    workbook.save(tabela)
    workbook.close()

def estiloRelacaoBens(tabela,tamanho,stringTamanho):
    """Estilo da tabela de bens, consulta no banco sap
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade de bens.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    if tamanho == 0:
         tamanho = 1

    random_number = random.randint(1, 10000)
    nomeTabela = f'relacaodebens'
    nomeVariavel = f'material{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Relação Bens Adquiridos A.5']
    size = tamanho + 13
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = 'ccffff'
    
 
    worksheet.sheet_view.showGridLines = False

            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão
    worksheet.column_dimensions['k'].width = 35 #data de emissão


   #cabecario 
    worksheet.merge_cells('A1:K1')
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO 5'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    #RELAÇÃO DE PAGAMENTOS
    worksheet.merge_cells('A2:K2')
    nomeTabela = nomeTabela.upper()
    worksheet['A2'] = f'RELAÇÃO DE BENS ADQUIRIDOS OU PRODUZIDOS'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="center",vertical="center")
    
        #cabecario que recebe de referencia as celulas A da planilha DEMOSTR. RECEITA E DESPESA A.2
    

    worksheet['A5'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet['A6'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A7'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A8'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet['A9'] = "='DEMOSTR. RECEITA E DESPESA A.2'!A9"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")



    #declaração
   
    worksheet['A11'] = f'Declaramos que os bens abaixo especificados, adquiridos ou produzidos com os recursos do concedente, foram  inventariados e encontram-se localizados nas instalações do Convenente ou dos Executores conforme relacionado abaixo. Relacionamos, também, os responsáveis pela guarda dos bens.'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="left",vertical="center")

  
    input2=f'rowStyle{nomeVariavel}'
   
    borda = Border(
    left=Side(border_style='hair'),  
    right=Side(border_style='hair'),  
    top=Side(border_style='double'), 
    bottom=Side(border_style='hair')  
)
    borda2 = Border(
    left=Side(border_style='hair', color='000000'),  
    right=Side(border_style='hair', color='000000'),  
    top=Side(border_style='hair', color='000000'), 
    bottom=Side(border_style='hair', color='000000')  
)

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="double")  ,bottom=Side(border_style="hair") )
    locals()[input2].height = 20
    linha_number = 13
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=11):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 11:
                cell.border = Border(left=Side(border_style="hair")  ,bottom=Side(border_style="hair"), right=Side(border_style="double") )


            cell.border = borda       

##CABECARIO
    
    worksheet["A13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("A13:A14")
    worksheet["A13"] ="Nº DO ITEM"
    

    worksheet["B13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("B13:B14")
    worksheet["B13"] = "DESCRIÇÃO DO BEM"

    worksheet["C13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("C13:C14")
    worksheet["C13"] = "NÚMERO PATRIMONIAL DO BEM"

    worksheet["D13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("D13:E13")
    worksheet["D13"] = "DOCUMENTAÇÃO FISCAL"

    worksheet["D14"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet["D14"] = "DATA"
    worksheet["D14"].border = borda2   

    worksheet["E14"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet["E14"] = "Nº "
    worksheet["E14"].border = borda2   

    worksheet["F13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("F13:F14")
    worksheet["F13"] = "LOCALIZAÇÃO"

    worksheet["G13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("G13:G14")
    worksheet["G13"] = "Equivalência na Relação de Itens Apoiados"

    worksheet["H13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("H13:H14")
    worksheet["H13"] = "QUANTIDADE"

    worksheet["I13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("I13:J13")
    worksheet["I13"] = "VALOR (R$)"

    worksheet["I14"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet["I14"] = "Unitário"
    worksheet["I14"].border = borda2   
    worksheet["J14"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet["J14"] = "Total"
    worksheet["J14"].border = borda2   

    worksheet["K13"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet["K13"].border = openpyxl.styles.Border(top=Side(border_style="double") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
    worksheet.merge_cells("K13:K14")
    worksheet["K13"] = "RESPONSÁVEL PELA GUARDA DO BEM"

    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=13, max_row=size, min_col=1, max_col=11):
        worksheet.row_dimensions[row[0].row].height = 35
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(15,size+1):
        cell = worksheet[f'I{row}']
        cell.style = locals()[input3]
    
    for row in range(15,size+1):
        cell = worksheet[f'J{row}']
        cell.style = locals()[input3]
        
    for rows in worksheet.iter_rows(min_row=15, max_row=size+1, min_col=1, max_col=11):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
               
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

                if cell.row == size+1:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="double") )

                if cell.column == 11:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="hair") )
                    if cell.row == size+1:
                            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )

    for rows in worksheet.iter_rows(min_row=size+1, max_row=size+1, min_col=1, max_col=11):
            for cell in rows:            
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="double") )
                if cell.column == 11:
                           cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="double") ,bottom=Side(border_style="double") )

    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "TOTAL"
    top_left_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(J13:J{size})"
    celula = f'J{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border( left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'


    #brasilia
    brasilia_row = size + 7
    brasilia_formula =  f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:K{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:F{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:F{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:F{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+3}"
    coordenadora_cargo_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+4}"
    coordenadora_cpf_formula = f"='DEMOSTR. RECEITA E DESPESA A.2'!C{stringTamanho+5}"
    coordenadora_merge_cells = f'G{coordenadora_row}:K{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'G{coordenadora_cargo_row}:K{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'G{coordenadora_cpf_row}:K{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'G{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'G{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'G{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    

    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=12,min_col=11,max_col=11):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=size+2, max_row=coordenadora_cpf_row+1,min_col=11,max_col=11):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=11):
        for cell in row:
            if cell.column == 11:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    


    workbook.save(tabela)
    workbook.close()

def estilo_demonstrativoDeReceita(tabela,tamanho,stringTamanho):
    """Estilo da demonstrativo de receita que inclui entradas de receita ISS 2%, ISS 5%.
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade de bens.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """

    if tamanho == 0:
         tamanho = 1


    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Demonstrativo de Receita']
    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    random_number = random.randint(1, 10000)
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=5,max_col=5):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 30
    worksheet.column_dimensions['b'].width = 70
    worksheet.column_dimensions['c'].width = 30
    worksheet.column_dimensions['d'].width = 50#descrição
    worksheet.column_dimensions['e'].width = 50#descrição
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:E2')
    worksheet['A1'] = f'D E M O N S T R A T I V O   D E   R E C E I T A  E    ISS 5% E ISS 2%'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:E3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:E4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:E5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:E6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:E7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")

    #colunas azul cabecario
    row_style_demonstrativo = NamedStyle(name=f'row_style_demonstrativo{random_number}')
    row_style_demonstrativo.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style_demonstrativo.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    row_style_demonstrativo.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style_demonstrativo.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    row_style_demonstrativo.height = 20
    linha_number = 9
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=5):
        for cell in row:
            cell.style = row_style_demonstrativo
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

    valores = ["NomeFavorecido","Histórico","Documento","Data de Entrada",'Valor']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=4):
        worksheet.row_dimensions[row[0].row].height = 80

    custom_number_format_demonstrativo = []
    # MASCARA R$
    if custom_number_format_demonstrativo!= False: 
        custom_number_format_demonstrativo = NamedStyle(name=f'custom_number_format_demonstrativo{random_number}')
        custom_number_format_demonstrativo.number_format = 'R$ #,##0.00'
        custom_number_format_demonstrativo.font = Font(name="Arial", size=12, color="000000")
        custom_number_format_demonstrativo.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'E{row}']
        cell.style = custom_number_format_demonstrativo
        
    for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=5):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 5:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:D{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(E10:E{size})"
    celula = f'E{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "Estorno de Mensalidades"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30

    #estilo colunas restitucoes creditadas
    row_style_demonstrativo_append = NamedStyle(name=f'row_style_demonstrativo_append{random_number}')
    row_style_demonstrativo_append.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style_demonstrativo_append.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    row_style_demonstrativo_append.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style_demonstrativo_append.height = 30
    row_style_demonstrativo_append.border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


    row_number = size + 4
   
    for column in range(1, 6):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = row_style_demonstrativo_append
        if cell.column == 5:
            cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



    values = ["NomeFavorecido","Histórico","Documento","Data de Entrada",'Valor']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        if coluna == 5:
            coluna = coluna + 1
        coluna = coluna + 1
        



    
    #subtotal2
    sub_total2_row = size + 5
    subtotal_merge_cells= f'A{sub_total2_row}:D{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

    sub_formula_row_celula = f'E{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

      #total1-2
    total12_row = size + 6
    total12_merge_cells = f'A{total12_row}:D{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


    #total_formula
    total_formula_row = size + 6
    total_formulaa = f'=E{size}'
    total_formula_row_celula = f'E{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )
    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa


    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:J{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:E{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='Receita x Despesa'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='Receita x Despesa'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='Receita x Despesa'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'C{coordenadora_row}:E{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'C{coordenadora_cargo_row}:E{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'C{coordenadora_cpf_row}:E{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'C{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'C{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'C{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")


    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=4):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()



# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\ModeloFINEP.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# sheet = workbook.create_sheet(title="DEMOSTR. RECEITA E DESPESA A.2")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# sheet = workbook.create_sheet(title="Relatório de Exec Financ A.1")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# nomeTabela = "Kek"
# sheet = workbook.create_sheet(title=f"{nomeTabela}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# nomeTabelaa = "Keka"
# sheet = workbook.create_sheet(title=f"{nomeTabelaa}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# nomeTabelaaa = "PAgamentoPEssoal"
# sheet = workbook.create_sheet(title=f"{nomeTabelaaa}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# nomeTabelaaaa = "Elemento1415"
# sheet = workbook.create_sheet(title=f"{nomeTabelaaaa}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()


# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# nomeTabela33 = "Elemento33"
# sheet = workbook.create_sheet(title=f"{nomeTabela33}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# BENS = "Relação Bens Adquiridos A.5"
# sheet = workbook.create_sheet(title=f"{BENS}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# conc = "Conciliação Bancária A.3"
# sheet = workbook.create_sheet(title=f"{conc}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook = openpyxl.load_workbook(tabela)
# rend = "Rendimento de Aplicação"
# sheet = workbook.create_sheet(title=f"{rend}")
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelaEstilizada.xlsx")
# workbook.close()


# tamanho = 0
# nomeVariavel = "Quea"

# tamanhoestorno = 0

# tamanhoBrasilia = estiloDEMOSTRRECEITEDESPESAA2(tabela,20)

# stringTamanho = tamanhoBrasilia
# estiloRelatorioExecFinanceiroA1(tabela,20,tamanhoBrasilia)
# estiloG(tabela,tamanho,nomeVariavel,nomeTabela,stringTamanho,tamanhoestorno)
# estiloPagamentoPessoal(tabela,tamanho,nomeTabelaaa,stringTamanho,tamanhoestorno)
# estiloElementoDeDespesa1415Diarias(tabela,tamanho,nomeTabelaaaa,stringTamanho,tamanhoestorno)
# estiloElementoDeDespesa33PassagensEDespesa(tabela,tamanho,nomeTabela33,stringTamanho,tamanhoestorno)
# estiloRelacaoBens(tabela,tamanho,stringTamanho)
# estilo_conciliacoes_bancaria(tabela,tamanho,tamanhoestorno,stringTamanho)
# estilo_rendimento_de_aplicacao(tabela,tamanho,stringTamanho)