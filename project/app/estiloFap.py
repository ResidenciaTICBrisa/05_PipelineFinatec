import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.comments import Comment
from PIL import Image as PILImage
from openpyxl.utils import units
import os
import random
from openpyxl.drawing.geometry import GeomRect

def pegar_caminho(subdiretorio):

    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline


def estiloAnexoDois(tabela,tamanho):
    '''Esse estilo e considerado geral por que todas as tabelas que compõe utilizam das mesma colunas.
      
        Argumentos:
            tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho: e o tamanho total de linhas que irao ser geradas dinamicamente correspondente as entradas do respectivo projeto. o valor varia dentre o tamanho das rubricas 
            
            nomeVariavel: variável utilizada para criar o nomes das variaveis dinamicamente para não haver sobreposição de estilos com o mesmo nome. Esses estilos ocorrem nesses codigos:
                                        locals()[input2] = NamedStyle(name=f'{input2}')
                                        locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
                                        locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid"'....
            nomeTabela: variável utilizada para a criação do nome da tabela.Ela deriva das rubricas que são colocadas no input quando essa função e chamada.
            stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referencias das formulas.
            
    '''
    nomeVariavel = f'Anexo2'
    
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO II']
    size = tamanho + 10
    cinza = "f3f3f2"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = 'ccffff'
    

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão
    worksheet.column_dimensions['k'].width = 35 #data de emissão


    #ANEXO II
   
   
    worksheet['A1'] = f'ANEXO 2'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet.merge_cells('A1:K1')

    #Fapdf
    worksheet['A2'] = f'FAPDF'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="left",vertical="center")
   
    #FUNDAÇÃO DE APOIO A PESQUISA DO DISTRITO FEDERAL
    worksheet['A3'] = f'FUNDAÇÃO DE APOIO A PESQUISA DO DISTRITO FEDERAL'
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A3'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    

    #RELAÇÃO DE PAGAMENTOS
    worksheet['F3'] = f'RELAÇÃO DE PAGAMENTOS'
    worksheet['F3'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['F3'].alignment = Alignment(horizontal="center",vertical="center")
   
    
    #N TOA/Processo
    worksheet['I2'] = f'Nº TOA / Processo'
    worksheet['I2'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['I2'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['I2'].border = Border(top=Side(border_style="none")  ,bottom=Side(border_style="none"), right=Side(border_style="thin"),left=Side(border_style='thin') )
    worksheet.merge_cells('I2:J2')
    
    worksheet['I3'].border = Border(top=Side(border_style="none")  ,bottom=Side(border_style="thin"), right=Side(border_style="thin"),left=Side(border_style='thin') )
    worksheet.merge_cells('I3:J3')


    #outorgado

    worksheet['A5'] = "Outorgado:"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    #Título do Projeto:
    worksheet['A6'] = "Título do Projeto:"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    #Instituição Gestora:
    worksheet['A7'] = "Instituição Gestora:Fundação de Empreendimentos Científicos e Tecnológicos - FINATEC"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    #Instituição Executora:
    worksheet['A8'] = "Instituição Executora: "
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")







    #merges
    worksheet.merge_cells('A5:K5')
    worksheet.merge_cells('A6:K6')
    worksheet.merge_cells('A7:K7')
    worksheet.merge_cells('A8:K8')

 
    
     #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #abecario #Item	Rubrica	Nº Cheque ou F. de caixa	Data	Nº Fatura	Favorecido	Descrição do Bem ou Serviço                                                                                  (nome, marca, tipo, modelo)	Qtde.	 Unitário 	 Custeio 	 Capital 

    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="hair"), right=Side(border_style="hair"),left=Side(border_style='hair') )
    locals()[input2].height = 20

    linha_number = 10
    

  
   
                                                                                 	 

    valores = ['Item','Rubrica','Nº Cheque ou F. de caixa','Data','Nº Fatura','Favorecido','Descrição do Bem ou Serviço','Qtde','Unitário','Custeio','Capital']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=11):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    
    value_to_stop = size  
    start_row = 10

    #estilo mascara de dinheiro lina I
    for row in range(start_row,size+1):
        cell = worksheet[f'I{row}']
        cell.style = locals()[input3]
    
    #estilo mascara de dinheiro lina J
    for row in range(start_row,size+1):
        cell = worksheet[f'J{row}']
        cell.style = locals()[input3]

   #estilocinzasimcinzanao     
    for rows in worksheet.iter_rows(min_row=11, max_row=size, min_col=1, max_col=11):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                if cell.row == size:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

                     
               

                if cell.column == 11:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                    if cell.row == size:
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                
 
    #subtotal
    #worksheet.row_dimensions[size+1].height = 6
    celulas_mergidas_subtotal = f"A{size+1}:H{size+1}"
    
    left_celula_cell = f"A{size+1}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    worksheet.merge_cells(celulas_mergidas_subtotal)
    worksheet.row_dimensions[size+1].height = 56.25

     # FORMULASOMATORIOTOTAL
    formula = f"=SUM(J11:J{size})"
    celula = f'J{size+1}'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    # CEDULATOTAL
    celulaTotal = size + 1
    celula_Total=f'I{celulaTotal}'
    worksheet[celula_Total].value = "Total"
    worksheet[celula_Total].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula_Total].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    worksheet.row_dimensions[celulaTotal].height = 30
    

    #capital
    formula = f"=SUM(K11:K{size})"
    celula = f'K{size+1}'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'


   
    

    # #brasilia
    brasilia_row = size + 3
    brasilia_formula =  f"Brasilia"
    brasilia_merge_cells = f'A{brasilia_row}:H{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_brasilia_cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    top_left_brasilia_cell.value = brasilia_formula

    

    # #DiretorFinanceiro
    diretor_row = size + 5
    diretor_cargo_row = size + 6
    diretor_cpf_row = size + 7
    
    diretor_nome_formula = f"Daniel Monteiro Rosa"
    diretor_cargo_formula = f"Diretor-Financeiro"
    diretor_cpf_formula = f"450.720.272-87"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    # #Coordenadora
    coordenadora_row = size + 5 
    coordenadora_cargo_row = size + 6 
    coordenadora_cpf_row = size + 7 
    coordenadora_nome_formula = f"nomeFormula"
    coordenadora_cargo_formula = f"cargoformula"
    coordenadora_cpf_formula = f"cpformula"
    coordenadora_merge_cells = f'E{coordenadora_row}:H{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'E{coordenadora_cargo_row}:H{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'E{coordenadora_cpf_row}:H{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'E{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'E{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'E{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font = Font(name="Arial", size=12, color="000000")

    

    #custeio
    # capital
    # Bolsa
    # total
   

    #QUADRADO CINZA
    for row in worksheet.iter_rows(min_row=size + 2, max_row=size + 7, min_col=9, max_col=11):
        for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.number_format = 'R$ #,##0.00'
                if cell.column == 9:
                        cell.border = Border(left=Side(border_style="thin") )
                        cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                if cell.column == 11:
                        cell.border = Border(right=Side(border_style="thin") )
                if cell.row == size + 7:
                    cell.border = Border(bottom=Side(border_style="thin") )
                    if cell.row == size + 7 and cell.column == 11:
                            cell.border = Border(bottom=Side(border_style="thin"), right=Side(border_style="thin") )
                    if cell.row == size + 7 and cell.column == 9:
                            cell.border = Border(bottom=Side(border_style="thin"), left=Side(border_style="thin") )

                if cell.row == size + 7 and cell.column == 10:
                           cell.font = Font(name="Arial", size=12, color="000000",bold=True)

            
    #TOTAL QUADRADO CINZA
    #CUSTEIO
    custeio_formula =  f"CUSTEIO"
    top_left_custeio_cell_formula = f'I{size+3}'
    top_left_custeio_cell = worksheet[top_left_custeio_cell_formula]
    top_left_custeio_cell.value = custeio_formula
    top_left_custeio_cell.alignment = Alignment(horizontal="center",vertical="center")                       
    #CAPITAL
    custeio_formula =  f"CAPITAL"
    top_left_custeio_cell_formula = f'I{size+4}'
    top_left_custeio_cell = worksheet[top_left_custeio_cell_formula]
    top_left_custeio_cell.value = custeio_formula
    top_left_custeio_cell.alignment = Alignment(horizontal="center",vertical="center")                       
    #BOLSA
    custeio_formula =  f"BOLSA"
    top_left_custeio_cell_formula = f'I{size+5}'
    top_left_custeio_cell = worksheet[top_left_custeio_cell_formula]
    top_left_custeio_cell.value = custeio_formula
    top_left_custeio_cell.alignment = Alignment(horizontal="center",vertical="center")                       
    #TOTAL
    custeio_formula =  f"TOTAL"
    top_left_custeio_cell_formula = f'I{size+6}'
    top_left_custeio_cell = worksheet[top_left_custeio_cell_formula]
    top_left_custeio_cell.value = custeio_formula
    top_left_custeio_cell.alignment = Alignment(horizontal="center",vertical="center")  
    #FORMULA TOTAL        
    custeio_formula =  f"=SUM(J{size+3}:J{size+5})"
    top_left_custeio_cell_formula = f'J{size+6}'
    top_left_custeio_cell = worksheet[top_left_custeio_cell_formula]
    top_left_custeio_cell.value = custeio_formula
    top_left_custeio_cell.alignment = Alignment(horizontal="center",vertical="center")               
    top_left_custeio_cell.font = Font(name="Arial", size=12, color="000000",bold=True)


    cinzaborda = '9e9e9e'

    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=11):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 11:
                cell.border = Border(top=Side(border_style="medium")  , right=Side(border_style="medium") )


    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=9,min_col=11,max_col=11):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    # #borda fim pagina
    # for row in worksheet.iter_rows(min_row=size + 3, max_row=coordenadora_cpf_row+1,min_col=11,max_col=11):
    #     for cell in row:
    #             cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

                

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row, max_row=coordenadora_cpf_row,min_col=1,max_col=8):
        for cell in row:
            if cell.column == 11:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    workbook.save(tabela)
    workbook.close()

    return brasilia_row

def estiloAnexoTres(tabela,tamanho,stringTamanho):
    if tamanho == 0:
         tamanho = 1

    nomeVariavel = f'Anexo3'
    nomeTabela = f'ANEXO III'
    
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO III']
    size = tamanho + 10
    cinza = "f3f3f2"
    cinza_escuro = "d9d9d9"
   

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            

    worksheet.column_dimensions['a'].width = 45
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 45#descrição
    worksheet.column_dimensions['e'].width = 45 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 35 #data de emissão
    worksheet.column_dimensions['g'].width = 35 #data de emissão
    worksheet.column_dimensions['h'].width = 35 #data de emissão
    worksheet.column_dimensions['i'].width = 35 #data de emissão
    worksheet.column_dimensions['j'].width = 35 #data de emissão
    worksheet.column_dimensions['k'].width = 35 #data de emissão
    worksheet.column_dimensions['l'].width = 35 #data de emissão


    #ANEXO II
   
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO 3'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet.merge_cells('A1:L1')

    #Fapdf
    worksheet['A2'] = f'FAPDF'
    worksheet['A2'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A2'].alignment = Alignment(horizontal="left",vertical="center")
   
    #FUNDAÇÃO DE APOIO A PESQUISA DO DISTRITO FEDERAL
    worksheet['A3'] = f'FUNDAÇÃO DE APOIO A PESQUISA DO DISTRITO FEDERAL'
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A3'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    

    #RELAÇÃO DE PAGAMENTOS
    worksheet['F3'] = f'LISTA DE BENS ADQUIRIDOS NO PROJETO'
    worksheet['F3'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['F3'].alignment = Alignment(horizontal="center",vertical="center")
   
    
      #N TOA/Processo
    worksheet['J2'] = f'Nº TOA / Processo'
    worksheet['J2'].font = Font(name="Arial", size=12, color="000000",bold=True,italic=True)
    worksheet['J2'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['J2'].border = Border(top=Side(border_style="none")  ,bottom=Side(border_style="none"), right=Side(border_style="thin"),left=Side(border_style='thin') )
    worksheet.merge_cells('J2:K2')
    
    worksheet['J3'] = "='ANEXO II'!I3"
    worksheet['J3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['J3'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['J3'].border = Border(top=Side(border_style="none")  ,bottom=Side(border_style="thin"), right=Side(border_style="thin"),left=Side(border_style='thin') )
    worksheet.merge_cells('J3:K3')

    #outorgado

    worksheet['A5'] = "Outorgado:"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    #Título do Projeto:
    worksheet['A6'] = "='ANEXO II'!A6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    #Instituição Gestora:
    worksheet['A7'] = "='ANEXO II'!A7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    #Instituição Executora:
    worksheet['A8'] = "='ANEXO II'!A8"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")







    #merges
    worksheet.merge_cells('A5:L5')
    worksheet.merge_cells('A6:L6')
    worksheet.merge_cells('A7:L7')
    worksheet.merge_cells('A8:L8')

 
    
     #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #abecario #Item	Rubrica	Nº Cheque ou F. de caixa	Data	Nº Fatura	Favorecido	Descrição do Bem ou Serviço                                                                                  (nome, marca, tipo, modelo)	Qtde.	 Unitário 	 Custeio 	 Capital 
    #estilo cabeçario


    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="thin"),left=Side(border_style='thin') )
    locals()[input2].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type = "solid")
    locals()[input2].height = 20

    linha_number = 10
    

  
    
                                                                                 	 

    valores = ['Item','Nº Cheque ou F. de caixa','Data','Nº Fatura','Favorecido','Descrição do Bem ou Serviço','Qtde','Unitário','Capital','Origem / Plaqueta','Localização','Responsável']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=12):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    
    value_to_stop = size  
    start_row = 10

    #estilo mascara de dinheiro lina I
    for row in range(start_row,size+1):
        cell = worksheet[f'H{row}']
        cell.style = locals()[input3]
    
    #estilo mascara de dinheiro lina J
    for row in range(start_row,size+1):
        cell = worksheet[f'I{row}']
        cell.style = locals()[input3]

   #estilocinzasimcinzanao     
    for rows in worksheet.iter_rows(min_row=11, max_row=size, min_col=1, max_col=12):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                    
                
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                if cell.row == size:
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )

                     
               

                if cell.column == 12:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                    if cell.row == size:
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                
 
    #subtotal
    #worksheet.row_dimensions[size+1].height = 6
    #subtotal barra
    celulas_mergidas_subtotal = f"A{size+1}:G{size+1}"
    
    left_celula_cell = f"A{size+1}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    top_left_cell.value = f'Declaro que as despesas relacionadas acima foram pagas e que os materiais e equipamentos foram recebidos.'
    top_left_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
    worksheet.merge_cells(celulas_mergidas_subtotal)
    worksheet.row_dimensions[size+1].height = 56.25
    #barra 2
    celulas_mergidas_subtotal = f"J{size+1}:L{size+1}"
    left_celula_cell = f"J{size+1}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    top_left_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
    worksheet.merge_cells(celulas_mergidas_subtotal)
    



     # FORMULASOMATORIOTOTAL
    formula = f"=SUM(I11:I{size})"
    celula = f'I{size+1}'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
    # CEDULATOTAL
    celulaTotal = size + 1
    celula_Total=f'H{celulaTotal}'
    worksheet[celula_Total].value = "Total"
    worksheet[celula_Total].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula_Total].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    worksheet.row_dimensions[celulaTotal].height = 30
    worksheet[celula_Total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
    


  #brasilia
    brasilia_row = size + 3
    brasilia_formula =  f"='ANEXO II'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:L{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 5
    diretor_cargo_row = size + 6
    diretor_cpf_row = size + 7
    
    diretor_nome_formula = f"='ANEXO II'!A{stringTamanho+2}:D{stringTamanho+2}"
    diretor_cargo_formula = f"='ANEXO II'!A{stringTamanho+3}:D{stringTamanho+3}"
    diretor_cpf_formula = f"='ANEXO II'!A{stringTamanho+4}:D{stringTamanho+4}"
    diretor_merge_cells = f'A{diretor_row}:F{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:F{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:F{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    # #Coordenadora
    coordenadora_row = size + 5 
    coordenadora_cargo_row = size + 6 
    coordenadora_cpf_row = size + 7 
    coordenadora_nome_formula = f"='ANEXO II'!E{stringTamanho+2}"
    coordenadora_cargo_formula = f"='ANEXO II'!E{stringTamanho+3}"
    coordenadora_cpf_formula = f"='ANEXO II'!E{stringTamanho+4}"
    coordenadora_merge_cells = f'G{coordenadora_row}:L{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'G{coordenadora_cargo_row}:L{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'G{coordenadora_cpf_row}:L{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'G{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'G{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'G{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font = Font(name="Arial", size=12, color="000000")

    

    cinzaborda = '9e9e9e'

    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=12):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="medium")  , right=Side(border_style="medium"),bottom=Side(border_style='thin') )


    
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=9,min_col=12,max_col=12):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    # #borda fim pagina
    # for row in worksheet.iter_rows(min_row=size + 3, max_row=coordenadora_cpf_row+1,min_col=11,max_col=11):
    #     for cell in row:
    #             cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    for row in worksheet.iter_rows(min_row=size+2, max_row=coordenadora_cpf_row-1,min_col=12,max_col=12):
        for cell in row:
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="none")  )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,bottom =Side(border_style="none") ,right=Side(border_style="thin",color='9e9e9e') )


    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row, max_row=coordenadora_cpf_row,min_col=1,max_col=12):
        for cell in row:
            if cell.column == 12:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )

    workbook.save(tabela)
    workbook.close()

def estiloAnexoQuatro(tabela,tamanho,stringTamanho):

    if tamanho == 0:
         tamanho = 1


    nomeVariavel = f'Anexo4'
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['ANEXO IV']
    size = tamanho + 10
    cinza = "f3f3f2"
    cinza_escuro = "d9d9d9"
    testeRow = 17;
   

    # borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 

    #imagens

    image_names = [
                'fapdf.png',
               
    ]
    images = []

    nomePasta = "../../imagemFap"
    diretorio = os.path.dirname(__file__)

    # Loop through the list of image names and create Image objects
    for i, name in enumerate(image_names):
            caminhoImage = os.path.join(diretorio, nomePasta, name)
            pil_image = PILImage.open(caminhoImage)
            pil_image.save(caminhoImage)
            img = Image(caminhoImage)
            images.append(img)


    worksheet.add_image(images[0], "A1")#fap
       





    # for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    worksheet.row_dimensions[1].height = 25
    worksheet.row_dimensions[2].height = 25
    worksheet.column_dimensions['a'].width = 45 #Período
    worksheet.column_dimensions['b'].width = 35 #Saldo anterior
    worksheet.column_dimensions['c'].width = 35 #Valor Aplicado no período
    worksheet.column_dimensions['d'].width = 45 #Valor Resgatado no Período 
    worksheet.column_dimensions['e'].width = 45 #Imposto de Renda / IOF
    worksheet.column_dimensions['f'].width = 35 #Rendimento Bruto
    worksheet.column_dimensions['g'].width = 35 #Saldo



    #ANEXO II
    nomeTabela = 'ANEXO IV'
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'ANEXO IV'
    worksheet['A1'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet.merge_cells('A1:G1')

    #Fapdf
    worksheet['A4'] = f'FAPDF'
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet.merge_cells('A4:G4')
   
    #FUNDAÇÃO DE APOIO A PESQUISA DO DISTRITO FEDERAL
    worksheet['A5'] = f'FUNDAÇÃO DE APOIO A PESQUISA DO DISTRITO FEDERAL'
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A5'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.merge_cells('A5:G5')

    #Demonstrativo dos Ganhos Auferidos com Aplicações Financeiras
    worksheet['A7'] = f'Demonstrativo dos Ganhos Auferidos com Aplicações Financeiras'
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A7'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet.merge_cells('A7:G7')

    
      #N TOA/Processo
    worksheet['F9'] = f'Nº TOA / Processo'
    worksheet['F9'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['F9'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet.merge_cells('F9:G9')
      #N TOA/Processo
    worksheet['F10'] = f"='ANEXO II'!I3"
    worksheet['F10'].font = Font(name="Arial", size=12, color="000000")
    worksheet['F10'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet.merge_cells('F10:G10')


    #outorgado
    #Unidade Gestora:
    worksheet['A9'] = "Unidade Gestora:"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    worksheet.merge_cells('A9:E9')
    #Unidade Gestora:
    worksheet['A10'] = "FINATEC - Fundação de Empreendimentos Científicos e Tecnológicos"
    worksheet['A10'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A10'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    worksheet.merge_cells('A10:E10')
    #Projeto:
    worksheet['A11'] = "Projeto"
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    worksheet.merge_cells('A11:E11')
    #ProjetorEFERENCIA:
    worksheet['A12'] = "='ANEXO II'!A6"
    worksheet['A12'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A12'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    worksheet.merge_cells('A12:E13')
    #Período abrangido:
    worksheet['F11'] = "Período abrangido:"
    worksheet['F11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['F11'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.merge_cells('F11:G11')

    #APLICAÇÃO FINANCEIRA - CURTO PRAZO
    worksheet['A14'] = "APLICAÇÃO FINANCEIRA - "
    worksheet['A14'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A14'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    worksheet.merge_cells('A14:G14')

 
    

    worksheet.row_dimensions[15].height = 20
    worksheet.row_dimensions[16].height = 20

    #
    input2=f'rowStyle{nomeVariavel}'
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="000000",bold=True)
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro, fill_type = "solid")
    locals()[input2].border = Border(top=Side(border_style="medium")   )
    locals()[input2].height = 20


    linha_number = 15




    # #stylecinza
    start_row = 17
    for rows in worksheet.iter_rows(min_row=15, max_row=16, min_col=1, max_col=7):
            for cell in rows:
                if cell.row % 2 == 0:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")

    #cabcerario abaixo de aplicação financeira
    row_number = 15
    values = ["Período","Saldo Anterior","Valor Aplicado no período",'Valor Resgatado no Período','Imposto de Renda / IOF','Rendimento Bruto','Saldo']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1


   

    # #RENDIMENTO LIQUIDO
    # # # print(size)
    # for row in worksheet.iter_rows(min_row=testeRow, max_row=size, min_col=7, max_col=7):
    #     for cell in row:
    #             stringSaldo = f"=E{cell.row} - F{cell.row}"
    #             cell.value = stringSaldo
          

                

    #BARRAS DE DADOS
    start_row = 14
    for rows in worksheet.iter_rows(min_row=testeRow, max_row=size, min_col=1, max_col=7):
            for cell in rows:
                if cell.row % 2==0:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    # #MASCARA VERMELHO
    # for rows in worksheet.iter_rows(min_row=testeRow, max_row=size-1, min_col=6, max_col=6):
    #         for cell in rows:
    #             cell.font = Font(name="Arial", size=12, color="f90000")
    #             cell.number_format ='#,##0.00'
   
    #MASCARANEGRITO
    for rows in worksheet.iter_rows(min_row=testeRow, max_row=size-1, min_col=1, max_col=1):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                
   
    # # #MASCARA AZUL
    # # for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=6, max_col=6):
    # #     for cell in rows:
    # #         cell.font = Font(name="Arial", size=12, color="141fca")
    # #         cell.number_format ='#,##0.00'

    # # for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=7, max_col=7):
    # #         for cell in rows:
    # #             cell.font = Font(name="Arial", size=12, color="141fca",bold=True)
    # #             cell.number_format ='#,##0.00'
   

    
    #barra de totais
    formula = f"Saldo anterior"
    celula = f'A{testeRow}'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    
    




    # FORMULATOTAL
    #B
    size = size + 1
    formula = f"=SUM(B{testeRow}:B{size-1})"
    celula = f'B{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

     #C
    formula = f"=SUM(C{testeRow}:C{size-1})"
    celula = f'C{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    #D
    formula = f"=SUM(D{testeRow}:D{size-1})"
    celula = f'D{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #E
    formula = f"=SUM(E{testeRow}:E{size-1})"
    celula = f'E{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #F
    formula = f"=SUM(F{testeRow}:F{size-1})"
    celula = f'F{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #G
    formula = f"=SUM(G{testeRow}:G{size-1})"
    celula = f'G{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
   

    #Total
    celula_total = F'A{size}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)    
    worksheet[celula].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #borda total



      #brasilia
    brasilia_row = size + 3
    brasilia_formula =  f"='ANEXO II'!A{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:G{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 5
    diretor_cargo_row = size + 6
    diretor_cpf_row = size + 7
    
    diretor_nome_formula = f"='ANEXO II'!A{stringTamanho+2}:C{stringTamanho+2}"
    diretor_cargo_formula = f"='ANEXO II'!A{stringTamanho+3}:C{stringTamanho+3}"
    diretor_cpf_formula = f"='ANEXO II'!A{stringTamanho+4}:C{stringTamanho+4}"
    diretor_merge_cells = f'A{diretor_row}:C{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:C{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:C{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    # #Coordenadora
    coordenadora_row = size + 5 
    coordenadora_cargo_row = size + 6 
    coordenadora_cpf_row = size + 7 
    coordenadora_nome_formula = f"='ANEXO II'!E{stringTamanho+2}"
    coordenadora_cargo_formula = f"='ANEXO II'!E{stringTamanho+3}"
    coordenadora_cpf_formula = f"='ANEXO II'!E{stringTamanho+4}"
    coordenadora_merge_cells = f'E{coordenadora_row}:G{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'E{coordenadora_cargo_row}:G{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'E{coordenadora_cpf_row}:G{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'E{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'E{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'E{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.font = Font(name="Arial", size=12, color="000000")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.font= Font(name="Arial", size=12, color="000000")

    

    cinzaborda = '9e9e9e'

   #estiloccinza cabeçario
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=7):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 7:
                cell.border = Border(top=Side(border_style="medium")  , right=Side(border_style="medium"),bottom=Side(border_style='thin') )

    for i in range(1,7):
        worksheet.merge_cells(start_row=15,end_row=16,start_column=i,end_column=i)

    #estilo dinheiro tabela toda
    for row in worksheet.iter_rows(min_row=17, max_row=size, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'
          
        
        

   
    #borda cabeçario
    for row in worksheet.iter_rows(min_row=1, max_row=14,min_col=7,max_col=7):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="thin",color='9e9e9e') ,left =Side(border_style="none") ,bottom=Side(border_style="none") )

    #borda fim pagina
    for row in worksheet.iter_rows(min_row=15, max_row=size,min_col=1,max_col=7):
        for cell in row:
                if cell.column == 7 and cell.row == 15 :
                    cell.border = Border(top=Side(border_style="medium") ,right = Side(border_style="medium") ,left =Side(border_style="none") ,bottom=Side(border_style="none") )
                if cell.column == 7 and cell.row != 15:
                    cell.border = Border(top=Side(border_style="none") ,right = Side(border_style="medium") ,left =Side(border_style="none") ,bottom=Side(border_style="none") )
                if cell.row == size and cell.column == 7:
                    cell.border = Border(top=Side(border_style="none") ,bottom = Side(border_style="medium") ,left =Side(border_style="none") ,right=Side(border_style="medium") )
                if cell.row == size and cell.column != 7:
                    cell.border = Border(top=Side(border_style="none") ,bottom = Side(border_style="medium") ,left =Side(border_style="none") ,right=Side(border_style="none") )


    for row in worksheet.iter_rows(min_row=size+1, max_row=coordenadora_cpf_row-1,min_col=7,max_col=7):
        for cell in row:
            if cell.column == 7:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="none")  )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,bottom =Side(border_style="none") ,right=Side(border_style="thin",color='9e9e9e') )


    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row, max_row=coordenadora_cpf_row,min_col=1,max_col=7):
        for cell in row:
            if cell.column == 7:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="thin",color='9e9e9e') ,bottom=Side(border_style="thin",color='9e9e9e') )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="thin",color='9e9e9e') )



    workbook.save(tabela)
    workbook.close()

# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\modeloFap.xlsx")
# # Load the workbook and create a new sheet
# workbook = openpyxl.load_workbook(tabela)
# sheet = workbook.create_sheet(title="ANEXO II")

# # Save the workbook
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# workbook.close()






# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# tamanho = 20
# nomeTabela = "ANEXO II"
# stringTamanho = 0
# tamanhoestorno = 0
# rowBrasilia = estiloAnexoDois(tabela,tamanho,nomeTabela,stringTamanho,tamanhoestorno)



# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# # Load the workbook and create a new sheet
# workbook = openpyxl.load_workbook(tabela)
# sheet = workbook.create_sheet(title="ANEXO III")

# # Save the workbook
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# workbook.close()



# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# tamanho = 20
# nomeTabela = "ANEXO III"
# stringTamanho = 0
# tamanhoestorno = 0
# estiloAnexoTres(tabela,tamanho,nomeTabela,rowBrasilia,tamanhoestorno)

#  # anexo4
# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# # Load the workbook and create a new sheet
# workbook = openpyxl.load_workbook(tabela)
# sheet = workbook.create_sheet(title="ANEXO IV")

# # Save the workbook
# workbook.save("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# workbook.close()



# tabela = pegar_caminho("C:\\Users\\hemanoel.brito\\Desktop\\estilos\\tabelafap.xlsx")
# tamanho = 20
# nomeTabela = "ANEXO IV"
# stringTamanho = 0
# tamanhoestorno = 0
# estiloAnexoQuatro(tabela,tamanho,nomeTabela,rowBrasilia,tamanhoestorno)





