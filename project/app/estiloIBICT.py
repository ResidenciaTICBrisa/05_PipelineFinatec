import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
import os
import random
from PIL import Image as PILImage
from openpyxl.drawing.image import Image
#pegar o caminho do arquivo
# def pegar_caminho(nome_arquivo):

#     # Obter o caminho absoluto do arquivo Python em execução
#     caminho_script = os.path.abspath(__file__)

#     # Obter o diretório da pasta onde o script está localizado
#     pasta_script = os.path.dirname(caminho_script)

#     # Combinar o caminho da pasta com o nome do arquivo Excel
#     caminho = os.path.join(pasta_script, nome_arquivo)

#     return caminho

def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline

def imagemIbict(tabela,nomeSheet):
    '''Estilo da Pagina do Relatorio Execução da Receita e Despesa
      
        Argumentos:
            tabela : recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            nomeSheet = recebe o nome da tabela que tem que preencher com as imagens
           

    '''
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook[nomeSheet]

    image_names = [
                'ibict.png',
                'finatec.png'
               
    ]
    images = []

    nomePasta = "../../imagensIBICIT"
    diretorio = os.path.dirname(__file__)

    # Loop through the list of image names and create Image objects
    for i, name in enumerate(image_names):
            caminhoImage = os.path.join(diretorio, nomePasta, name)
            pil_image = PILImage.open(caminhoImage)
            pil_image.save(caminhoImage)
            img = Image(caminhoImage)
            images.append(img)


    sheet.add_image(images[0], "A1")#ibict
    sheet.add_image(images[0], "D1")#finatec
       
    # for row in sheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[2].height = 25
    
    workbook.save(tabela)
    workbook.close()

def estiloExecReceitaDespesa(tabela,tamanho,stringTamanho):
    '''Estilo da Pagina do Relatorio Execução da Receita e Despesa
      
        Argumentos:
            tabela : recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho : é o tamanho total de linhas que irão ser geradas dinâmicamente. O valor varia dentre o tamanho da quantidade de rubricas diferente que o projeto possui, excluindo Obras e Instalações
            Aplicações Financeira e Equipamento e Material Permanente sendo nacional ou importado.
            stringTamanho : refere-se aonde esta localizado a string brasília na pagina Receita e despesa para a referencias das formulas.
           

    '''
    
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Exec. Receita e Despesa']

    size = tamanho + 16;
    size2 = size + 1
    cinza = "f1f1f1"
    azul_claro = '1c8cbc'
    # sheet.row_dimensions[3] = 28
    numRowInicial = 7
    numRowInicialMerge = 8
    variavelA = f'A{numRowInicial}'
    

    #cabecario 
    sheet.merge_cells('A{numRowInicial}:I{numRowInicialMerge}')#7
    sheet['A{numRowInicial}'] = f'E X E C U Ç Ã O   D A  R E C E I T A   E  D E S P E S A '
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['A{numRowInicial}'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    numRowInicial = numRowInicial + 1

    sheet.merge_cells('A{numRowInicial}:I{numRowInicial}')#8
    sheet['A{numRowInicial}'] = "='Receita x Despesa'!A{numRowInicial}:I{numRowInicial}"
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="000000")
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="left",vertical="center")
    numRowInicial = numRowInicial + 1
   

    sheet.merge_cells('A{numRowInicial}:I{numRowInicial}')#9
    sheet['A{numRowInicial}'] = "='Receita x Despesa'!A{numRowInicial}:I{numRowInicial}"
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="000000")
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="left",vertical="center")
    numRowInicial = numRowInicial + 1
   
    
    sheet.merge_cells('A{numRowInicial}:I{numRowInicial}')#10
    sheet['A{numRowInicial}'] = "='Receita x Despesa'!A{numRowInicial}:I{numRowInicial}"
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="000000")
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="left",vertical="center")
    numRowInicial = numRowInicial + 1
   
    
    sheet.merge_cells('A{numRowInicial}:I{numRowInicial}')#11
    sheet['A{numRowInicial}'] = "='Receita x Despesa'!A{numRowInicial}:I{numRowInicial}"
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="000000")
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="left",vertical="center")
    numRowInicial = numRowInicial + 1
   
    
    sheet.merge_cells('A{numRowInicial}:I{numRowInicial}')#12
    sheet['A{numRowInicial}'] = "='Receita x Despesa'!A{numRowInicial}:I{numRowInicial}"
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="000000")
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="left",vertical="center")
    numRowInicial = numRowInicial + 1
   
    
    sheet.merge_cells('A{numRowInicial}:I{numRowInicial}')#13
    sheet['A{numRowInicial}'] = 'E X E C U Ç Ã O    D A   R E C E I T A    E    D E S P E S A'
    sheet['A{numRowInicial}'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    sheet['A{numRowInicial}'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['A{numRowInicial}'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    numRowInicial = numRowInicial + 1
   
    
    #aumentao altura
    for i in range(1,10):
        sheet.row_dimensions[i].height = 20

    sheet.row_dimensions[7].height = 25
    sheet.row_dimensions[9].height = 35
    sheet.row_dimensions[8].height = 25
    sheet.row_dimensions[15].height = 30


 
    #esticando coluna A e pintando ela
    numRowInicial = numRowInicial + 1 #ta na 15
    sheet.merge_cells('A{numRowInicial}:A16')
    sheet['A{numRowInicial}'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    sheet.column_dimensions['a'].width = 55
    sheet.column_dimensions['b'].width = 20
    sheet.column_dimensions['c'].width = 20
    sheet.column_dimensions['d'].width = 20
    sheet.column_dimensions['e'].width = 20
    sheet.column_dimensions['f'].width = 20
    sheet.column_dimensions['g'].width = 20
    sheet.column_dimensions['h'].width = 20
    sheet.column_dimensions['i'].width = 20
    sheet.column_dimensions['j'].width = 1
    #resto do cabecario
    sheet.merge_cells('B12:E13')
    sheet['B12'] = 'EXECUTADO NO PERÍODO\n(Valores em R$)'
    sheet['B12'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    sheet['B12'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['B12'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    sheet.merge_cells('F12:I13')
    sheet['F12'] = 'ACUMULADO ATÉ O PERÍODO\n(Valores em R$)'
    sheet['F12'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    sheet['F12'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['F12'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")




   


    #%porcentagem Despesas correntes
    #IFERROR (C16/B16;0)
    #%porcentagem Despesas de Capital
    #IFERROR (C16/B16;0)
    percentage_style = NamedStyle(name='percentage', number_format='0%')
    #diferença e calculo da porcentagem
    for row in sheet.iter_rows(min_row=15, max_row=size2+9, min_col=1, max_col=9):
        for cell in row:
            if cell.column == 4:
                stringSaldo = f"=B{cell.row} - C{cell.row}"
                cell.value = stringSaldo
            if cell.column == 8:
                stringSaldo = f"=F{cell.row} - G{cell.row}"
                cell.value = stringSaldo
              

            if cell.column == 5:
                stringPorcentagem = f"=IFERROR(C{cell.row}/B{cell.row}, 0)"
                cell.value = stringPorcentagem
                cell.style = percentage_style
            if cell.column == 9:
                stringPorcentagem = f"=IFERROR(G{cell.row}/F{cell.row}, 0)"
                cell.value = stringPorcentagem
                cell.style = percentage_style
            
    #adicionara mascara do numero
    for row in sheet.iter_rows(min_row=15, max_row=size2+9, min_col=1, max_col=9):
        for cell in row:
            if cell.column != 5 and cell.column != 9:
                cell.number_format ='#,##0.00'
   
                
        





    #adicionar borda em tudo
    borda = Border(right=Side(border_style="medium"))
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=1, max_row=size2+14,min_col=9,max_col=9):
        for cell in row:
            cell.border = borda

    font = Font(name='Arial', size=12)
    #Colocar tudo arial 12
    for row in sheet.iter_rows(min_row=16):
        for cell in row:
            cell.font = font
    #preenche a coluna A e        F com cinza
    for row in sheet.iter_rows(min_row=16, max_row=size2+9,min_col=1,max_col=9):
        for cell in row:
            if cell.column == 1 or cell.column == 6 or cell.column == 2:
                cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
            if cell.column == 6 or cell.column == 2:
                cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = cell.font = Font(name='Arial', size=12,bold= True)
                
###pinta de azul a linha desejada
    for row in sheet.iter_rows(min_row=15, max_row=sheet.max_row, min_col=sheet.min_row, max_col= 9):
        for cell in row:
            if cell.row == 15:
                cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,
                                            fill_type = "solid")
                cell.font = Font(name='Arial', size=12,bold= True)

            if cell.row == size2 or cell.row == size2 +5 or cell.row == size2 + 7  or cell.row== size2+9:
                cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,
                                            fill_type = "solid")
                cell.font = Font(name='Arial', size=12,bold= True)    
    
    

    
    #cabecario negrito sem fundo
    
    valores = ["RUBRICA","PREVISTO","REALIZADO","SALDO",'(%)\nREALIZADO',"PREVISTO","REALIZADO","SALDO","(%)\nREALIZADO"]
    col = 1
    for a,b in enumerate(valores):
        sheet.cell(row=14, column=col, value=b)
        col = col + 1
    #coloca em negrito
    font = Font(name='Arial', size=12,bold=True)
    #Colocar tudo arial 12
    for row in sheet.iter_rows(min_row=14,max_row=14):
        for cell in row:
            cell.font = font


    #depsesas e capital
    despesasDeCapitalString = f"I. DESPESAS CORRENTES"
    despesasDeCapitalStringCelula = f'A{15}'
    sheet[despesasDeCapitalStringCelula] = despesasDeCapitalString
    cell=sheet[despesasDeCapitalStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    



     #depsesas e capital
    despesasDeCapitalString = f"II. DESPESAS DE CAPITAL"
    despesasDeCapitalStringCelula = f'A{size2}'
    sheet[despesasDeCapitalStringCelula] = despesasDeCapitalString
    cell=sheet[despesasDeCapitalStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    
    
    

    obrasEInstalacoeslString = f"Obras e Instalações"
    obrasEInstalacoeslStringCelula = f'A{size2+1}'
    sheet[obrasEInstalacoeslStringCelula] = obrasEInstalacoeslString
 

    equipamentoseMaterialPermanenteString = f"Equipamentos e Material Permanente"
    equipamentoseMaterialPermanenteStringCelula = f'A{size2+2}'
    sheet[equipamentoseMaterialPermanenteStringCelula] = equipamentoseMaterialPermanenteString
    cell=sheet[equipamentoseMaterialPermanenteStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell=sheet[f'B{size2+2}']
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    
    

    nacionalString = f"     a) Nacional"
    nacionalStringCelula = f'A{size2+3}'
    sheet[nacionalStringCelula] = nacionalString

    

    importadoString = f"    b) Importado"
    importadoStringCelula = f'A{size2+4}'
    sheet[importadoStringCelula] = importadoString
    
    

    #depsesas e capital
    utilRendimentosString = f"III.UTILIZAÇÃO DE RENDIMENTOS"
    utilRendimentosStringCelula = f'A{size2+7}'
    sheet[utilRendimentosStringCelula] = utilRendimentosString
    cell=sheet[utilRendimentosStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell=sheet[f'B{size2+1}']
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    

    
    apliFinString = f"Aplicação Financeira"
    apliFinStringCelula = f'A{size2+8}'
    sheet[apliFinStringCelula] = apliFinString


    #barra de total nova 22/02
    totalString = f"TOTAL"
    totalStringCelula = f'A{size2+5}'
    sheet[totalStringCelula] = totalString
    cell=sheet[totalStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.fill = PatternFill(start_color='9c9c9c', end_color='9c9c9c',fill_type = "solid")
    cell.alignment = Alignment(horizontal="center",vertical="center")

    
    totalString = f"TOTAL"
    totalStringCelula = f'A{size2+9}'
    sheet[totalStringCelula] = totalString
    cell=sheet[totalStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.fill = PatternFill(start_color='9c9c9c', end_color='9c9c9c',fill_type = "solid")
    cell.alignment = Alignment(horizontal="center",vertical="center")

    #barra total nova 

    formula = f"=SUM(B{size2}, B15)"
    sheet[f'B{size2+5}'] = formula

    formula = f"=SUM(C{size2}, C15)"
    sheet[f'C{size2+5}'] = formula

    formula = f"=SUM(D{size2}, D15)"
    sheet[f'D{size2+5}'] = formula

    formula = f"=SUM(F{size2}, F15)"
    sheet[f'F{size2+5}'] = formula

    formula = f"=SUM(G{size2}, G15)"
    sheet[f'G{size2+5}'] = formula

    formula = f"=SUM(H{size2},H15)"
    sheet[f'H{size2+5}'] = formula


    #sheet[f'H{size2+5}']
    sheet.row_dimensions[size2 + 6].height = 1
    


    #3.UTILIZAÇÂO DE RENDIMENTOS
    # formula = f"=SUM(A{size2+8})"
    # sheet[f'A{size2+7}'] = formula

    formula = f"=SUM(B{size2+8})"
    sheet[f'B{size2+7}'] = formula

    formula = f"=SUM(C{size2+8})"
    sheet[f'C{size2+7}'] = formula

    formula = f"=SUM(D{size2+8})"
    sheet[f'D{size2+7}'] = formula

    formula = f"=SUM(F{size2+8})"
    sheet[f'F{size2+7}'] = formula

    formula = f"=SUM(G{size2+8})"
    sheet[f'G{size2+7}'] = formula

    formula = f"=SUM(H{size2+8})"
    sheet[f'H{size2+7}'] = formula

     ##SOMATORIO NEGATIVO C
    Formula = f'=SUMIF(D16:D{size}, "<0")'
    sheet[f'C{size2+8}']= Formula
    ##############

    ##SOMATORIO NEGATIVO G
    Formula = f'=SUMIF(H16:H{size}, "<0")'
    sheet[f'G{size2+8}'] = Formula
    ##############


    
 
        #totais
    #total espesas correntes linha 15

    formula = f"=SUM(B16:B{size})"
    sheet['B15'] = formula
    formula = f"=SUM(C16:C{size})"
    sheet['C15'] = formula
    formula = f'=SUMIF(D16:D{size}, ">0")'
    sheet['D15'] = formula

    formula = f"=SUM(F16:F{size})"
    sheet['F15'] = formula
    formula = f"=SUM(G16:G{size})"
    sheet['G15'] = formula
    formula = f'=SUMIF(H16:H{size}, ">0")'
    sheet['H15'] = formula

    #Total Despesas de Capital
    formula = f"=SUM(B{size2+2}:B{size2+4})"
    sheet[f'B{size2}'] = formula

    formula = f"=SUM(C{size2+2}:C{size2+4})"
    sheet[f'C{size2}'] = formula

    formula = f"=SUM(D{size2+2}:D{size2+4})"
    sheet[f'D{size2}'] = formula

    formula = f"=SUM(F{size2+2}:F{size2+4})"
    sheet[f'F{size2}'] = formula

    formula = f"=SUM(G{size2+2}:G{size2+4})"
    sheet[f'G{size2}'] = formula

    formula = f"=SUM(H{size2+2}:H{size2+4})"
    sheet[f'H{size2}'] = formula


    #Total  abaixo de Utilização de rendimentos

    formula = f"=B{size2+8}"
    sheet[f'B{size2+9}'] = formula

    formula = f"=C{size2+8}"
    sheet[f'C{size2+9}'] = formula

    formula = f"=D{size2+8}"
    sheet[f'D{size2+9}'] = formula

    formula = f"=F{size2+8}"
    sheet[f'F{size2+9}'] = formula

    formula = f"=G{size2+8}"
    sheet[f'G{size2+9}'] = formula

    formula = f"=H{size2+8}"
    sheet[f'H{size2+9}'] = formula

   

    #somasaldo h
    stringSaldo = f"=SUM(F{size2+8}+G{size2+8})"
   
    sheet[f"H{size2+8}"] = stringSaldo
    

     #brasilia
    brasilia_row = size2 + 11
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:I{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    sheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = sheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size2 + 12
    diretor_cargo_row = size2 + 13
    diretor_cpf_row = size2 + 14
    
    diretor_nome_formula = f"Daniel Monteiro Rosa"
    diretor_cargo_formula = f"Diretor-Financeiro"
    diretor_cpf_formula = f"450.720.272-87"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    sheet.merge_cells(diretor_merge_cells)
    sheet.merge_cells(diretor_cargo_merge_cells)
    sheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = sheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = sheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = sheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size2 + 12
    coordenadora_cargo_row = size2 + 13
    coordenadora_cpf_row = size2 + 14
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:I{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:I{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:I{coordenadora_cpf_row}'
    sheet.merge_cells(coordenadora_merge_cells)
    sheet.merge_cells(coordenadora_cargo_merge_cells)
    sheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = sheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = sheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = sheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # sheet.sheet_view.showGridLines = False
    # # 
    # for row in sheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in sheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=9):
        for cell in row:
            if cell.column == 9:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )


    
     #bordinha branca

    for row in sheet.iter_rows(min_row=1,max_row=coordenadora_cpf_row,min_col=1,max_col=10):
        for cell in row:
           
                if cell.row == 15:
                    cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='FFFFFF'))
                    if cell.column == 9:
                        cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                        right=Side(border_style='medium', color='000000'),
                        top=Side(border_style='thin', color='FFFFFF'),
                        bottom=Side(border_style='thin', color='FFFFFF'))
                if cell.row == size2:
                    cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='FFFFFF'))
                    if cell.column == 9:
                        cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                        right=Side(border_style='medium', color='000000'),
                        top=Side(border_style='thin', color='FFFFFF'),
                        bottom=Side(border_style='thin', color='FFFFFF'))
                    
                if cell.row == size2+5:
                    cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='FFFFFF'))
                    if cell.column == 9:
                        cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                        right=Side(border_style='medium', color='000000'),
                        top=Side(border_style='thin', color='FFFFFF'),
                        bottom=Side(border_style='thin', color='FFFFFF'))
                if cell.row == size2+7:
                    cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='FFFFFF'))
                    if cell.column == 9:
                        cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                        right=Side(border_style='medium', color='000000'),
                        top=Side(border_style='thin', color='FFFFFF'),
                        bottom=Side(border_style='thin', color='FFFFFF'))
                if cell.row == size2+9:
                    cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                    right=Side(border_style='thin', color='FFFFFF'),
                    top=Side(border_style='thin', color='FFFFFF'),
                    bottom=Side(border_style='thin', color='FFFFFF'))
                    if cell.column == 9:
                        cell.border = Border(left=Side(border_style='thin', color='FFFFFF'),  # white
                        right=Side(border_style='medium', color='000000'),
                        top=Side(border_style='thin', color='FFFFFF'),
                        bottom=Side(border_style='thin', color='FFFFFF'))

    for row in sheet.iter_rows(min_row=1,max_row=coordenadora_cpf_row,min_col=10,max_col=10):
        cell.border = Border(left=Side(border_style='medium', color='000000'))

    workbook.save(tabela)
    workbook.close()


    


    
    return 0

def estiloReceitaXDespesa(tabela,stringTamanho):
    '''Estilo da Pagina do Relatorio Receita e Despesa.
      
        Argumentos:
            tabela : recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            
            stringTamanho : refere-se aonde esta localizado a string Brasilia nessa pagina, ela pega o  valor  entre o tamanho quantidade de rubricas que o projeto possui salvo algumas exeções
            e a entrada de receitas/iss, oque for maior esse ditará tamanho.
    '''
       
    caminho = pegar_caminho(tabela)
    #Plan = planilha
    # carrega a planilha de acordo com o caminho
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Receita x Despesa']
    size = stringTamanho + 16;
    size2 = size + 5
    cinza = "d9d9d9"
    azul_claro = '1c8cbc'
    # sheet.row_dimensions[3] = 28

    

    #cabecario relação de pagamentos - outro servicoes de terceiros
    sheet.merge_cells('A1:J2')
    sheet['A1'] = f'D E M O N S T R A T I V O   D E   R E C E I T A   E   D E S P E S A'
    sheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    sheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    sheet.merge_cells('A3:J3')
    sheet['A3'] = ""
    sheet['A3'].font = Font(name="Arial", size=12, color="000000")
    sheet['A3'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)

    sheet.merge_cells('A4:J4')
    sheet['A4'] =  ""
    sheet['A4'].font = Font(name="Arial", size=12, color="000000")
    sheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    sheet.merge_cells('A5:J5')
    sheet['A5'] =  ""
    sheet['A5'].font = Font(name="Arial", size=12, color="000000")
    sheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    sheet.merge_cells('A6:J6')
    sheet['A6'] =  ""
    sheet['A6'].font = Font(name="Arial", size=12, color="000000")
    sheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    sheet.merge_cells('A7:J7')
    sheet['A7'] =  ""
    sheet['A7'].font = Font(name="Arial", size=12, color="000000")
    sheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    sheet.merge_cells('A9:J9')
    sheet['A9'] = 'R E C E I T A   E   D E S P E S A'
    sheet['A9'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    sheet['A9'].alignment = Alignment(horizontal="center",vertical="center")
    sheet['A9'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    for i in range(1,10):
        sheet.row_dimensions[i].height = 20

    sheet.row_dimensions[1].height = 25
    sheet.row_dimensions[3].height = 35
    sheet.row_dimensions[2].height = 25
    sheet.row_dimensions[9].height = 30
    sheet.column_dimensions['b'].width = 40
    sheet.column_dimensions['c'].width = 30
    sheet.column_dimensions['d'].width = 30
    sheet.column_dimensions['e'].width = 30
    sheet.column_dimensions['f'].width = 30
    sheet.column_dimensions['i'].width = 30
    sheet.column_dimensions['j'].width = 30
    #barra cinza

    barrazinzaMerge = f"A{11}:J{11}"
    sheet.merge_cells(barrazinzaMerge)
    barrazinzaStringCelula = f"A{11}"
    cell = sheet[barrazinzaStringCelula]
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)




    #adicionar borda em tudo
    borda = Border(right=Side(border_style="medium"))
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows(min_row=1, max_row=size2+21,min_col=10,max_col=10):
        for cell in row:
            cell.border = borda


    


    #merge das receitas
    for i in range(16,size):
        sttring = f"C{i}:D{i}"
        sheet.merge_cells(sttring)


    font = Font(name='Arial', size=12)
    #Colocar tudo arial 12
    for row in sheet.iter_rows(min_row=16):
        for cell in row:
            cell.font = font
    
    #mascara de dinheiro despesas correntes
    for row in sheet.iter_rows(min_row=16,max_row=size,min_col=5,max_col=5):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    #mascara de dinehrio despesas realizadas
    for row in sheet.iter_rows(min_row=16,max_row=size,min_col=9,max_col=9):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'


    #preenche qualquer linha com cinza
    for row in sheet.iter_rows(min_row=16, max_row=sheet.max_row, min_col=sheet.min_row, max_col= sheet.max_column):
        for cell in row:
            if cell.row == size2+2 or cell.row == size2 +12:
                cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
            



    #Receita

    rendimentoDeAplicacaoFinanceiraMerge = f"A{size2+2}:E{size2+2}"
    sheet.merge_cells(rendimentoDeAplicacaoFinanceiraMerge)
    rendimentoDeAplicacaoFinanceiraString= f"Rendimento de Aplicação financeira"
    rendimentoDeAplicacaoFinanceiraStringCelula = f"A{size2+2}"
    sheet[rendimentoDeAplicacaoFinanceiraStringCelula] = rendimentoDeAplicacaoFinanceiraString
    cell=sheet[f'A{size2+2}']
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)


    #Rendimento de aplicação financeira formula
    formula = f"=SUM(E{size2+3}:E{size2+6})"
    celula = f'F{size2+2}'
    sheet[celula] = formula
    cell=sheet[f'F{size2+2}']
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'

    #soma despesas realizadas
    formula = f"=SUM(I{16}:I{size+5})"
    celula = f'J13'
    sheet[celula] = formula

    #soma saldo conciliado
    formula = f'=SUM(I{size2+4}+I{size2+5}+I{size2+6}+I{size2+9})'
    celula = f'J{size2+2}'
    sheet[celula] = formula

    #valores recebidos no periodo
    formula = f"=SUM(E{14}:E{size2+1})"
    celula = f'F13'
    sheet[celula] = formula

    #SUM total receita

    formula = f'=SUM(F{size2+2}+F13)'
    celula = f'F{size2+12}'
    sheet[celula] = formula

    #SUM total despesa
    formula = f'=SUM(J{size2+2}+J13)'
    celula = f'J{size2+12}'
    sheet[celula] = formula

    #sum II. DESPESAS DE CAPITAl
    formula = f'=SUM(I{size+2}:I{size+5})'
    celula = f'I{size+1}'
    sheet[celula] = formula

   #Despesas realizadas
    #depsesas e capital
    despesasDeCapitalString = f"II. DESPESAS DE CAPITAL"
    despesasDeCapitalStringCelula = f'H{size+1}'
    sheet[despesasDeCapitalStringCelula] = despesasDeCapitalString
    cell=sheet[despesasDeCapitalStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell=sheet[f'I{size+1}']
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'
    
    

    obrasEInstalacoeslString = f"Obras e Instalações"
    obrasEInstalacoeslStringCelula = f'H{size+2}'
    sheet[obrasEInstalacoeslStringCelula] = obrasEInstalacoeslString
    cell=sheet[f'I{size+2}']
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.number_format = 'R$ #,##0.00'
    

    equipamentoseMaterialPermanenteString = f"Equipamentos e Material Permanente"
    equipamentoseMaterialPermanenteStringCelula = f'H{size+3}'
    sheet[equipamentoseMaterialPermanenteStringCelula] = equipamentoseMaterialPermanenteString
    cell=sheet[equipamentoseMaterialPermanenteStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell=sheet[f'I{size+3}']
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'
    

    nacionalString = f"     a) Nacional"
    nacionalStringCelula = f'H{size+4}'
    sheet[nacionalStringCelula] = nacionalString
    cell=sheet[f'I{size+4}']
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.number_format = 'R$ #,##0.00'
    

    importadoString = f"    b) Importado"
    importadoStringCelula = f'H{size+5}'
    sheet[importadoStringCelula] = importadoString
    cell=sheet[f'I{size+5}']
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.number_format = 'R$ #,##0.00'

    #saldo conciliado em:

    saldoConciliadoEmString= f"Saldo Conciliado em: "
    saldoConciliadoEmStringCelula = f"H{size2+2}"
    sheet[saldoConciliadoEmStringCelula] = saldoConciliadoEmString
    cell=sheet[saldoConciliadoEmStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
   
   ##saldo conciliado formula
    cell=sheet[f'J{size2+2}']
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'

    #mergecelulasabaixo de Rendimento de Aplicação financeira
    for i in range(size2+3,size2+6):
        sttring = f"A{i}:D{i}"
        sheet.merge_cells(sttring)
    
  
    
    contaCorrenteEmString= f"Conta Corrente"
    contaCorrenteEmStringCelula = f"H{size2+4}"
    sheet[contaCorrenteEmStringCelula] = contaCorrenteEmString
    cell=sheet[f'I{size2+4}']
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.number_format = 'R$ #,##0.00'
    
    devolucaoRecursosEmString= f"Devolução de recursos - GRU SIMPLES"
    devolucaoRecursosEmStringCelula = f"H{size2+5}"
    sheet[devolucaoRecursosEmStringCelula] = devolucaoRecursosEmString
    cell=sheet[f'I{size2+5}']
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.number_format = 'R$ #,##0.00'
    
    tarifaBancariaSaldoEmString= f"Tarifa Bancária - Saldo"
    tarifaBancariaSaldoEmStringCelula = f"H{size2+6}"
    sheet[tarifaBancariaSaldoEmStringCelula] = tarifaBancariaSaldoEmString
    cell=sheet[saldoConciliadoEmStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell=sheet[f'I{size2+6}']
    cell.value= f"=I{size2+7}-I{size2+8}"
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'
    
    tarifaBancariaDespesaEmString= f"Tarifa Bancária - Despesa (-)"
    tarifaBancariaDespesaEmStringCelula = f"H{size2+7}"
    sheet[tarifaBancariaDespesaEmStringCelula] = tarifaBancariaDespesaEmString
    cell=sheet[f'I{size2+7}']
    cell.font = Font(name="Arial", size=12, color="000000")
    cell.number_format = 'R$ #,##0.00'
    
    TarifaBancariaRestituicaoEmString= f"Tarifa Bancária - Restituição (+)"
    TarifaBancariaRestituicaoEmStringCelula = f"H{size2+8}"
    sheet[TarifaBancariaRestituicaoEmStringCelula] = TarifaBancariaRestituicaoEmString
    cell=sheet[f'I{size2+8}']
    cell.number_format = 'R$ #,##0.00'
    
    aplicacaoFinanceiraEmString= f"Aplicação Financeira"
    aplicacaoFinanceiraEmStringCelula = f"H{size2+9}"
    sheet[aplicacaoFinanceiraEmStringCelula] = aplicacaoFinanceiraEmString
    cell=sheet[aplicacaoFinanceiraEmStringCelula]
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell=sheet[f'I{size2+9}']
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'
    
    rendimentoDeaplicacaoEmString= f"Rendimento de aplicação financeira"
    rendimentoDeaplicacaoEmStringCelula = f"H{size2+10}"
    sheet[rendimentoDeaplicacaoEmStringCelula] = rendimentoDeaplicacaoEmString

    cell=sheet[f'I{size2+10}']
    cell.number_format = 'R$ #,##0.00'

    #Total Receita

    totalMerge = f"A{size2+12}:E{size2+12}"
    sheet.merge_cells(totalMerge)
    totalString= f"TOTAL"
    totalStringCelula = f"A{size2+12}"
    sheet[totalStringCelula] = totalString
    
    cell = sheet[totalStringCelula]
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    #TotalReceitaFormula
    cell=sheet[f'F{size2+12}']
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'
     #Total Despesa

    totalMerge = f"H{size2+12}:I{size2+12}"
    sheet.merge_cells(totalMerge)
    totalString= f"TOTAL"
    totalStringCelula = f"H{size2+12}"
    sheet[totalStringCelula] = totalString
    cell = sheet[totalStringCelula]
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #Total Despesa Formula
    cell=sheet[f'J{size2+12}']
    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
    cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    cell.number_format = 'R$ #,##0.00'


    #Barraazul

    barrazulMerge = f"A{size2+14}:J{size2+14}"
    sheet.merge_cells(barrazulMerge)
    barrazulStringCelula = f"A{size2+14}"
    cell = sheet[barrazulStringCelula]
    
    cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,
                                            fill_type = "solid")
    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

   

     #brasilia
    brasilia_row = size2 + 16
    brasilia_formula = f""
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    sheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = sheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size2 + 19
    diretor_cargo_row = size2 + 20
    diretor_cpf_row = size2 + 21
    diretor_nome_formula = f"Daniel Monteiro Rosa"
    diretor_cargo_formula = f"Diretor-Financeiro"
    diretor_cpf_formula = f"450.720.272-87"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    sheet.merge_cells(diretor_merge_cells)
    sheet.merge_cells(diretor_cargo_merge_cells)
    sheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = sheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = sheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = sheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12,bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

  #Coordenadora
    coordenadora_row = size2 + 19
    coordenadora_cargo_row = size2 + 20
    coordenadora_cpf_row = size2 + 21
    coordenadora_nome_formula = f"teste"
    coordenadora_cargo_formula = f"Coordenador(a)"
    coordenadora_cpf_formula = f"teste"
    coordenadora_merge_cells = f'H{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'H{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'H{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    sheet.merge_cells(coordenadora_merge_cells)
    sheet.merge_cells(coordenadora_cargo_merge_cells)
    sheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'H{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'H{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'H{coordenadora_cpf_row}'
    top_left_coordenadora_cell = sheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = sheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = sheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    for row in sheet.iter_rows(min_row=1,max_row=coordenadora_cpf_row,min_col=10,max_col=10):
        for cell in row:
            cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="none") )
   
    for row in sheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
        for cell in row:
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )
    
    # #deixar azul a letra de receita
    # cell=sheet['E16']
    # cell.font = Font(name="Arial", size=12,color="0000FF")
    # #deixar VERMELHO a letra de receita
    # cell=sheet[f'E{17}']
    # cell.font = Font(name="Arial", size=12,color="FF0000") 

    # cell=sheet[f'E{18}']
    # cell.font = Font(name="Arial", size=12,color="FF0000") 

    workbook.save(tabela)
    workbook.close()

    # print(f'printandosize2')
    # print(size2)
    #retorna tamanho de brasilia e de equipamentos
    return size2 + 16,size+3

def estiloGeral(tabela,tamanho,nomeVariavel,nomeTabela,stringTamanho,tamanhoestorno):
    '''Esse estilo e considerado geral por que todas as tabelas que compõe utilizam das mesma colunas.
      
        Argumentos:
            tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo ja foi iniciado e passou pela preencher fub mas ainda esta sem o estilo que sera aplicado nessa função.
            tamanho: e o tamanho total de linhas que irao ser geradas dinamicamente correspondente as entradas do respectivo projeto. o valor varia dentre o tamanho das rubricas 
            
            nomeVariavel: variável utilizada para criar o nomes das variaveis dinamicamente para não haver sobreposição de estilos com o mesmo nome. Esses estilos ocorrem nesses codigos:
                                        locals()[input2] = NamedStyle(name=f'{input2}')
                                        locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
                                        locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid"'....
            nomeTabela: variável utilizada para a criação do nome da tabela.Ela deriva das rubricas que são colocadas no input quando essa função e chamada.
            stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referencias das formulas.
            tamanhoEstorno = Correspondente ao tamanho do estrono
    '''
    
    nomeSheet=nomeVariavel
    random_number = random.randint(1, 10000)
    nomeVariavel = f'{nomeVariavel}{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'

    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+12,min_col=10,max_col=10):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 25
    worksheet.column_dimensions['b'].width = 25
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35#descrição
    worksheet.column_dimensions['e'].width = 65 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 25 #data de emissão
    worksheet.column_dimensions['g'].width = 25 #data de emissão
    worksheet.column_dimensions['h'].width = 25 #data de emissão
    worksheet.column_dimensions['i'].width = 25 #data de emissão
    worksheet.column_dimensions['j'].width = 25 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:J2')
    nomeTabela = nomeTabela.upper()
    worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - {nomeTabela}'

    # if nomeSheet == "diarias":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - DIÁRIAS'
    # elif nomeSheet == "pessoaFisica":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S -  OUTROS SERVIÇOS DE TERCEIROS - PF'
    # elif nomeSheet == "pessoaJuridica":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - OUTROS SERVIÇOS DE TERCEIROS - PESSOA JURÍDICA'
    # elif nomeSheet == "passagenDespLocomo":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - PASSAGENS E DESPESAS COM LOCOMOÇÃO'
    # elif nomeSheet == "outrosServiçosTerceiros":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - O U T R O S  S E R V I Ç O S D E T E R C E I R O S - C E L E T I S T A S'
    # elif nomeSheet == "auxilio":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S -  AUXÍLIO FINANCEIRO A ESTUDANTE E PESQUISADOR'
    # elif nomeSheet == "bolsaExtensao":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S -  BOLSA DE EXTENSÃO'
    # elif nomeSheet == "estagiario":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S -  ESTAGIÁRIO'
    # elif nomeSheet == "custosIndiretos":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - CUSTOS INDIRETOS - FUB'
    # elif nomeSheet == "materialDeConsumo":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S -  MATERIAL DE CONSUMO'
    # elif nomeSheet == "equipamentoMaterialPermanente":
    #     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S -  EQUIPAMENTO E MATERIAL PERMANENTE'
    # elif nomeSheet == "isss":
    #     worksheet['A1'] = f'ISS 5% e ISS 2%'

  

    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:J3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:J4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:J5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:J6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:J7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    locals()[input2].height = 20
    linha_number = 9
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

    valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
        worksheet.row_dimensions[row[0].row].height = 75
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'J{row}']
        cell.style = locals()[input3]
        
    for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 10:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(J10:J{size})"
    celula = f'J{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    input4 = f'row_style_diaria_append{nomeVariavel}'
    #estilo colunas restitucoes creditadas
    locals()[input4] = NamedStyle(name=f'{input4}')
    locals()[input4].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    locals()[input4].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input4].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input4].height = 30
    locals()[input4].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


    row_number = size + 4
   
    for column in range(1, 11):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = locals()[input4]
        if cell.column == 10:
            cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



    values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        if coluna == 4:
            coluna = coluna + 1
        coluna = coluna + 1
        

    merge_formula = f'D{row_number}:E{row_number}'
    worksheet.merge_cells(merge_formula)

    #mergecells
    for x in range(size+4,size+5+tamanhoestorno):
        worksheet.merge_cells(start_row=x ,start_column=4, end_row=x, end_column=5)

   
    #estorno

    for rows in worksheet.iter_rows(min_row=size+5, max_row=size+4+tamanhoestorno, min_col=1, max_col=10):
        for cell in rows:
            if cell.row % 2:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
            if cell.column == 10:        
                cell.number_format = 'R$ #,##0.00'
                cell.font = Font(name="Arial", size=12, color="000000")

            cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair"))
            cell.font = Font(name="Arial", size=12, color="000000")


    #bordas,corsimcornao,money
    # Set the height of each row to 60
    for row in worksheet.iter_rows(min_row=size+4, max_row=size+4+tamanhoestorno):
        worksheet.row_dimensions[row[0].row].height = 75

    min_row = size + 4
    max_row = size + 4 + tamanhoestorno

      

    #subtotal2
    sub_total2_row = size + 6 +tamanhoestorno
    subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

    sub_formula_row_celula = f'J{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].value = f'=SUM(J{size+5}:J{sub_total2_row-1})'
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

      #total1-2
    total12_row = size + 7 + tamanhoestorno
    total12_merge_cells = f'A{total12_row}:I{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


    #total_formula
    total_formula_row = size + 7 + tamanhoestorno
    total_formulaa = f'=J{size+2} - J{sub_total2_row }'
    total_formula_row_celula = f'J{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )

    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa

    

    #brasilia
    brasilia_row = size + 8 +tamanhoestorno
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:J{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 9 + tamanhoestorno
    diretor_cargo_row = size + 10 + tamanhoestorno
    diretor_cpf_row = size + 11 + tamanhoestorno
    
    diretor_nome_formula = f"='Receita x Despesa'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='Receita x Despesa'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='Receita x Despesa'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 9 + tamanhoestorno 
    coordenadora_cargo_row = size + 10 + tamanhoestorno  
    coordenadora_cpf_row = diretor_cpf_row
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=11,max_col=11):
        for cell in row:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="medium") ,right =Side(border_style="none") ,bottom=Side(border_style="none") )

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
        for cell in row:
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()

    return size+4

def estilo_conciliacoes_bancaria(tabela,tamanho,tamanho2,stringTamanho):
    """Estilo um pouco diferente pois necessita de dois aspectos dinâmicos que é primeiramente a quantidade de entradas de pagamento de tarifas bancárias e por fim a quantidade de estorno. Sabendo
    esses valores é possivel criar a tabela.
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade de transfêrencia bancárias realizada.
        tamanho2:Corresponde ao tamanho dos estornos.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    
    #pegar o arquivo e carregar ele um worksheet da pagaina Conciliação Bancária
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Conciliação Bancária']
    
    #size e o tamanho da quantidade de arquivos recebido no argumento tamanho mais o tamanho do cabecario que no caso da fub e de 16
    size = tamanho + 17
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'

    #Borda apenas do lado direito da cedula, uma borda mas larga
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
   


    worksheet.column_dimensions['a'].width = 25
    worksheet.column_dimensions['b'].width = 25
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35
    worksheet.column_dimensions['e'].width = 20
    worksheet.column_dimensions['f'].width = 20
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:F2')
    worksheet['A1'] = f'C O N C I L I A Ç Ã O   B A N C Á R I A'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:F3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:F4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:F5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:F6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:F7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A9:F9')
    worksheet['A9'] = '1.Saldo conforme extratos bancários na data final do período'
    worksheet['A9'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A9'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    worksheet.merge_cells('A10:E10')
    worksheet['A10'] = 'Saldo de Conta Corrente(R$)'
    worksheet['A10'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A10'].alignment = Alignment(horizontal="right",vertical="center")

    worksheet.merge_cells('A11:E11')
    worksheet['A11'] = 'Saldo de Aplicações Financeiras(R$)'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="right",vertical="center")

    worksheet.merge_cells('A13:F13')
    worksheet['A13'] = '2. Restituições não creditadas pelo banco até a data final do período'
    worksheet['A13'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A13'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A13'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    for i in range(15,size):
        sttring = f"D{i}:F{i}"
        worksheet.merge_cells(sttring)
        
    for i in range(size+3,size+4+tamanho2+1):
        sttring = f"D{i}:F{i}"
        worksheet.merge_cells(sttring)

    random_number = random.randint(1, 10000)

    custom_number_format_conciliacoes = []
    # MASCARA R$
    if custom_number_format_conciliacoes!= False: 
        custom_number_format_conciliacoes = NamedStyle(name=f'custom_number_format_conciliacoes{random_number}')
        custom_number_format_conciliacoes.number_format = 'R$ #,##0.00'
        custom_number_format_conciliacoes.font = Font(name="Arial", size=12, color="000000")
        custom_number_format_conciliacoes.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
    #stylecinza
    start_row = 15
    for row in range(start_row,size+1):
        cell = worksheet[f'B{row}']
        cell.style = custom_number_format_conciliacoes
        
    for rows in worksheet.iter_rows(min_row=15, max_row=size, min_col=1, max_col=6):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)


    row_number = 15
    values = ["Data","Valor(R$)","Documento",'Descrição']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        
        coluna = coluna + 1

    # FORMULATOTAL
    formula = f"=SUM(B16:B{size-1})"
    celula = f'B{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #Total
    celula_total = F'A{size}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)
    #'3. Restituições não creditadas pelo banco até a data final do período'
    string_reituicoes_creditadas = f'A{size+2}:F{size+2}'
    row_creditadas = f'A{size+2}'
    worksheet.merge_cells(string_reituicoes_creditadas)
    
    worksheet[row_creditadas] = '3. Restituições creditadas pelo banco até a data final do período'
    worksheet[row_creditadas].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet[row_creditadas].alignment = Alignment(horizontal="left",vertical="center")
    worksheet[row_creditadas].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    #data valor documento descrição
    row_number = size+3
    values = ["Data","Valor(R$)","Documento",'Descrição']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1

    for rows in worksheet.iter_rows(min_row=15, max_row=15, min_col=1, max_col=6):  
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    for rows in worksheet.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                cell.alignment = Alignment(horizontal="center",vertical="center")

    #saldo anterior restituição
    row_restituicao = f'A{size+4}'         
    worksheet[row_restituicao] = 'Saldo anterior'
    worksheet[row_restituicao].font = Font(name="Arial", size=12, color="000000")
    worksheet[row_restituicao].alignment = Alignment(horizontal="left",vertical="center")
    row_restituicao = f'b{size+4}'         
    worksheet[row_restituicao].font = Font(name="Arial", size=12, color="000000")
    worksheet[row_restituicao].alignment = Alignment(horizontal="left",vertical="center")
    row_restituicao = f'c{size+4}'         
    worksheet[row_restituicao] = 'Diversos'
    worksheet[row_restituicao].font = Font(name="Arial", size=12, color="000000")
    worksheet[row_restituicao].alignment = Alignment(horizontal="left",vertical="center")
    row_restituicao = f'd{size+4}'         
    worksheet[row_restituicao] = 'Restituição Prestações Anteriores'
    worksheet[row_restituicao].font = Font(name="Arial", size=12, color="000000")
    worksheet[row_restituicao].alignment = Alignment(horizontal="left",vertical="center")



    size = size + 1

    for row in range(size+3,size+4+tamanho2):
        cell = worksheet[f'B{row}']
        cell.style = custom_number_format_conciliacoes
        
    for rows in worksheet.iter_rows(min_row=size+3, max_row=size+3+tamanho2, min_col=1, max_col=6):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                if cell.column == 6: 
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )


    # FORMULATOTALrestituição
    formula = f"=SUM(B{size+3}:B{size+tamanho2+3})"
    celula = f'B{size+tamanho2+5}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #Total
    celula_total = F'A{size+tamanho2+5}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)
    #Saldo disponível p/ período seguinte (1 +2 - 3)
    string_saldo_disponivel = f'A{size+3+tamanho2+3}:D{size+3+tamanho2+3}'
    celula_string_saldo = f'A{size+tamanho2+6}'
    worksheet[celula_string_saldo].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet.merge_cells(string_saldo_disponivel)
    worksheet[celula_string_saldo]= f'Saldo disponível p/ período seguinte (1 + 2 - 3)'
    #total saldo diposnivel
    string_merge_saldo_disponivel = f'E{size+3+tamanho2+3}:F{size+3+tamanho2+3}'
    celula_string_total = f'E{size+tamanho2+6}'
    worksheet.merge_cells(string_merge_saldo_disponivel)
    saldodiposnivelformat_conciliacoes = NamedStyle(name=f'saldodiposnivelformat_conciliacoes{random_number}')
    saldodiposnivelformat_conciliacoes.number_format = 'R$ #,##0.00'
    saldodiposnivelformat_conciliacoes.font = Font(name="Arial", size=12, color="000000")
    saldodiposnivelformat_conciliacoes.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    saldodiposnivelformat_conciliacoes.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    celular = worksheet[celula_string_total]
    celular.style = saldodiposnivelformat_conciliacoes
    celular.value = f'=F10+F11+B{size-1} -B{size+tamanho2+5}'
    #saldo anterior
    formula = f"Saldo anterior"
    celula = f'A16'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000")

    formula = f"Diversos"
    celula = f'C16'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000")

    formula = f"Tarifas Prestações Anteriores"
    celula = f'D16'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000")

     #brasilia
    brasilia_row = size + tamanho2+ 8
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:J{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:F{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 10 + tamanho2
    diretor_cargo_row = size + 11 + tamanho2
    diretor_cpf_row = size + 12 + tamanho2
    diretor_nome_formula = f"='Receita x Despesa'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='Receita x Despesa'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='Receita x Despesa'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = diretor_row
    coordenadora_cargo_row = diretor_cargo_row
    coordenadora_cpf_row = diretor_cpf_row
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'D{coordenadora_row}:F{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'D{coordenadora_cargo_row}:F{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'D{coordenadora_cpf_row}:F{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'D{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'D{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'D{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.border = borda
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.border = borda
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.border = borda

    for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=6,max_col=6):
        for cell in row:
            cell.border = borda

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=6):
        for cell in row:
            if cell.column == 6:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()

def estilo_rendimento_de_aplicacao(tabela,tamanho,stringTamanho):
    """Estilo da rendimento de aplicação, tabela com as colunas periodo, saldo anterior,valor aplicado no período,valor resgatado no período,rendimento bruto,imposto,rendimento luiquido,saldo.
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade da tabela de rencimentos.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Rendimento de Aplicação']
    
    random_number = random.randint(1, 10000)    
    size = tamanho + 15
    worksheet.row_dimensions[10].height = 2
    worksheet.row_dimensions[9].height = 20

    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+9,min_col=8,max_col=8):
        for cell in row:
            cell.border = borda

    

    worksheet.column_dimensions['a'].width = 20
    worksheet.column_dimensions['b'].width = 20
    worksheet.column_dimensions['c'].width = 20
    worksheet.column_dimensions['d'].width = 20
    worksheet.column_dimensions['e'].width = 20
    worksheet.column_dimensions['f'].width = 20
    worksheet.column_dimensions['g'].width = 20
    worksheet.column_dimensions['h'].width = 20
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:H2')
    worksheet['A1'] = f'D E M O N S T R A T I V O   D E   R E N D I M E N T O   D E   A P L I C A Ç Ã O   F I N A N C E I R A'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:H3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:H4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:H5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:H6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:H7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A9:H9')
    worksheet['A9'] = 'RF Ref DI Plus Ágil - CNP JRF REF DI PLUS ÁGIL'
    worksheet['A9'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A9'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A9'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    

    

    #stylecinza
    start_row = 11
    for rows in worksheet.iter_rows(min_row=start_row, max_row=13, min_col=1, max_col=8):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)


    row_number = 11
    values = ["Período","Saldo Anterior","Valor Aplicado no período",'Valor Resgatado no Período','Rendimento Bruto','Imposto de Renda / IOF','Rendimento Líquido','Saldo']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        
        coluna = coluna + 1

    for i in range(1,9):
        worksheet.merge_cells(start_row=11,end_row=13,start_column=i,end_column=i)


    #RENDIMENTO LIQUIDO
    # print(size)
    for row in worksheet.iter_rows(min_row=14, max_row=size, min_col=7, max_col=7):
        for cell in row:
                stringSaldo = f"=E{cell.row} - F{cell.row}"
                cell.value = stringSaldo
          

                

    #BARRAS DE DADOS
    start_row = 14
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size, min_col=1, max_col=8):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #MASCARA VERMELHO
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=6, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="f90000")
                cell.number_format ='#,##0.00'
   
    #MASCARANEGRITO
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=1, max_col=1):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                
   
    # #MASCARA AZUL
    # for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=6, max_col=6):
    #     for cell in rows:
    #         cell.font = Font(name="Arial", size=12, color="141fca")
    #         cell.number_format ='#,##0.00'

    # for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=7, max_col=7):
    #         for cell in rows:
    #             cell.font = Font(name="Arial", size=12, color="141fca",bold=True)
    #             cell.number_format ='#,##0.00'
   
    #barra de totais
    formula = f"Saldo anterior"
    celula = f'A14'
    worksheet[celula] = formula
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    
    #barra de totais
    # FORMULATOTAL
    #B
    formula = f"=SUM(B15:B{size-1})"
    celula = f'B{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
     #C
    formula = f"=SUM(C15:C{size-1})"
    celula = f'C{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #D
    formula = f"=SUM(D15:D{size-1})"
    celula = f'D{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #E
    formula = f"=SUM(E15:E{size-1})"
    celula = f'E{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #F
    formula = f"=SUM(F15:F{size-1})"
    celula = f'F{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #G
    formula = f"=SUM(G15:G{size-1})"
    celula = f'G{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #H
    formula = f"=SUM(H15:H{size-1})"
    celula = f'H{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)

    #Total
    celula_total = F'A{size}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)    

    #brasilia
    brasilia_row = size + 2
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:J{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:F{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 6 
    diretor_cargo_row = size + 7 
    diretor_cpf_row = size + 8
    diretor_nome_formula = f"='Receita x Despesa'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='Receita x Despesa'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='Receita x Despesa'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:C{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:C{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:C{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size  + 6
    coordenadora_cargo_row = size + 7 
    coordenadora_cpf_row = size + 8
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'E{coordenadora_row}:G{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'E{coordenadora_cargo_row}:G{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'E{coordenadora_cpf_row}:G{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'E{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'E{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'E{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")


    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=8):
        for cell in row:
            if cell.column == 8:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )


    workbook.save(tabela)
    workbook.close()

def estiloRelacaoBens(tabela,tamanho,nomeVariavel,nomeTabela,stringTamanho):
    """Estilo da tabela de bens, consulta no banco sap
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade de bens.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    random_number = random.randint(1, 10000)
    
    nomeVariavel = f'material{random_number}'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 13
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 25
    worksheet.column_dimensions['b'].width = 50
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35#descrição
    worksheet.column_dimensions['e'].width = 40 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 70 #data de emissão
    worksheet.column_dimensions['g'].width = 25 #data de emissão
    worksheet.column_dimensions['h'].width = 25 #data de emissão
    worksheet.column_dimensions['i'].width = 25 #data de emissão
    worksheet.column_dimensions['j'].width = 25 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:J2')
    worksheet['A1'] = f'RELAÇÃO DE BENS'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    worksheet.merge_cells('A3:J4')
    worksheet['A3'] = f'(ADQUIRIDOS, PRODUZIDOS OU CONSTRUÍDOS COM RECURSOS)'
    worksheet['A3'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A3'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A3'].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    
    worksheet.merge_cells('A5:F5')
    worksheet['A5'] = "='Receita x Despesa'!A3:J3"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A6:F6')
    worksheet['A6'] = "='Receita x Despesa'!A4:J4"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:F7')
    worksheet['A7'] = "='Receita x Despesa'!A5:J5"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A8:F8')
    worksheet['A8'] = "='Receita x Despesa'!A6:J6"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A9:F9')
    worksheet['A9'] = "='Receita x Despesa'!A7:J7"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   
    borda = Border(
    left=Side(border_style='thin', color='FFFFFF'),  
    right=Side(border_style='thin', color='FFFFFF'),  
    top=Side(border_style='thin', color='FFFFFF'), 
    bottom=Side(border_style='thin', color='FFFFFF')  
)
    borda2 = Border(
    left=Side(border_style='hair', color='000000'),  
    right=Side(border_style='hair', color='000000'),  
    top=Side(border_style='hair', color='000000'), 
    bottom=Side(border_style='hair', color='000000')  
)

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    locals()[input2].height = 20
    linha_number = 11
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 10:
                cell.border = Border(left=Side(border_style="thin", color='FFFFFF')  ,bottom=Side(border_style="thin", color='FFFFFF'), right=Side(border_style="medium") )


            cell.border = borda       

##CABECARIO
    
    worksheet["A11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("A11:A12")
    worksheet["A11"] ="Nº DO ITEM"
    

    worksheet["B11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("B11:B12")
    worksheet["B11"] = "DESCRIÇÃO DO BEM"

    worksheet["C11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("C11:C12")
    worksheet["C11"] = "NÚMERO PATRIMONIAL DO BEM"

    worksheet["D11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("D11:E11")
    worksheet["D11"] = "DOCUMENTAÇÃO FISCAL"

    worksheet["D12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["D12"] = "DATA"
    worksheet["D12"].border = borda2   

    worksheet["E12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["E12"] = "Nº "
    worksheet["E12"].border = borda2   

    worksheet["F11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("F11:F12")
    worksheet["F11"] = "LOCALIZAÇÃO"

    worksheet["G11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("G11:G12")
    worksheet["G11"] = "QTD."

    worksheet["H11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("H11:I11")
    worksheet["H11"] = "VALOR (R$)"

    worksheet["H12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["H12"] = "Unitário"
    worksheet["H12"].border = borda2   
    worksheet["I12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["I12"] = "Total"
    worksheet["I12"].border = borda2   

    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
        worksheet.row_dimensions[row[0].row].height = 35
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(13,size+1):
        cell = worksheet[f'H{row}']
        cell.style = locals()[input3]
    
    for row in range(13,size+1):
        cell = worksheet[f'I{row}']
        cell.style = locals()[input3]
        
    for rows in worksheet.iter_rows(min_row=13, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 10:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "TOTAL"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(I13:I{size})"
    celula = f'J{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'


    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:J{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='Receita x Despesa'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='Receita x Despesa'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='Receita x Despesa'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
        for cell in row:
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    worksheet["J11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("J11:J12")
    worksheet["J11"] = "RESPONSÁVEL PELA GUARDA DO BEM"
    worksheet["J11"].border = Border(left=Side(border_style="thin", color='FFFFFF')  ,bottom=Side(border_style="thin", color='FFFFFF'), right=Side(border_style="medium") )
    worksheet["J11"].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet["J10"].border = Border(right =Side(border_style="medium"))
    worksheet.row_dimensions[10].height = 2


    workbook.save(tabela)
    workbook.close()

def estilo_demonstrativoDeReceita(tabela,tamanho,stringTamanho):
    """Estilo da demonstrativo de receita que inclui entradas de receita ISS 2%, ISS 5%.
    
        Argumentos:
        tabela: recebe o arquivo correspondente a tabela Fub extensão xlsx. Esse arquivo foi iniciado e passou pela preencher fub mas ainda está sem o estilo que será aplicado nessa função.
        tamanho:Corresponde ao tamanho das quantidade de bens.
        stringTamanho: refere-se aonde esta localizado a string brasilia na pagina Receita e despesa para a referências das formulas.
    """
    caminho = pegar_caminho(tabela)
    
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Demonstrativo de Receita']
    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    random_number = random.randint(1, 10000)
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=5,max_col=5):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 30
    worksheet.column_dimensions['b'].width = 70
    worksheet.column_dimensions['c'].width = 30
    worksheet.column_dimensions['d'].width = 50#descrição
    worksheet.column_dimensions['e'].width = 50#descrição
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:E2')
    worksheet['A1'] = f'D E M O N S T R A T I V O   D E   R E C E I T A  E    ISS 5% E ISS 2%'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:E3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:E4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:E5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:E6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:E7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")

    #colunas azul cabecario
    row_style_demonstrativo = NamedStyle(name=f'row_style_demonstrativo{random_number}')
    row_style_demonstrativo.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style_demonstrativo.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    row_style_demonstrativo.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style_demonstrativo.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    row_style_demonstrativo.height = 20
    linha_number = 9
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=5):
        for cell in row:
            cell.style = row_style_demonstrativo
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

    valores = ["NomeFavorecido","Histórico","Documento","Data de Entrada",'Valor']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=4):
        worksheet.row_dimensions[row[0].row].height = 80

    custom_number_format_demonstrativo = []
    # MASCARA R$
    if custom_number_format_demonstrativo!= False: 
        custom_number_format_demonstrativo = NamedStyle(name=f'custom_number_format_demonstrativo{random_number}')
        custom_number_format_demonstrativo.number_format = 'R$ #,##0.00'
        custom_number_format_demonstrativo.font = Font(name="Arial", size=12, color="000000")
        custom_number_format_demonstrativo.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'E{row}']
        cell.style = custom_number_format_demonstrativo
        
    for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=5):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 5:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:D{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(E10:E{size})"
    celula = f'E{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "Estorno de Mensalidades"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30

    #estilo colunas restitucoes creditadas
    row_style_demonstrativo_append = NamedStyle(name=f'row_style_demonstrativo_append{random_number}')
    row_style_demonstrativo_append.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style_demonstrativo_append.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    row_style_demonstrativo_append.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style_demonstrativo_append.height = 30
    row_style_demonstrativo_append.border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


    row_number = size + 4
   
    for column in range(1, 6):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = row_style_demonstrativo_append
        if cell.column == 5:
            cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



    values = ["NomeFavorecido","Histórico","Documento","Data de Entrada",'Valor']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        if coluna == 5:
            coluna = coluna + 1
        coluna = coluna + 1
        



    
    #subtotal2
    sub_total2_row = size + 5
    subtotal_merge_cells= f'A{sub_total2_row}:D{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

    sub_formula_row_celula = f'E{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

      #total1-2
    total12_row = size + 6
    total12_merge_cells = f'A{total12_row}:D{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


    #total_formula
    total_formula_row = size + 6
    total_formulaa = f'=E{size}'
    total_formula_row_celula = f'E{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )
    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa


    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f"='Receita x Despesa'!A{stringTamanho}:J{stringTamanho}"
    brasilia_merge_cells = f'A{brasilia_row}:E{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='Receita x Despesa'!A{stringTamanho+3}"
    diretor_cargo_formula = f"='Receita x Despesa'!A{stringTamanho+4}"
    diretor_cpf_formula = f"='Receita x Despesa'!A{stringTamanho+5}"
    diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(name="Arial", size=12, color="000000",bold = True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='Receita x Despesa'!H{stringTamanho+3}"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H{stringTamanho+4}"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H{stringTamanho+5}"
    coordenadora_merge_cells = f'C{coordenadora_row}:E{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'C{coordenadora_cargo_row}:E{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'C{coordenadora_cpf_row}:E{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'C{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'C{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'C{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(name="Arial", size=12, color="000000",bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=4):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()
