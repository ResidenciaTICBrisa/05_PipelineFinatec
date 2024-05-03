import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
import os
#pegar o caminho do arquivo
def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho

def estiloOpas(tabela,tamanho,nomeVariavel,nomeTabela):
    nomeSheet=nomeVariavel
    print(tabela)
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 10
    print(size)
    cinza = "d9d9d9"
    cinza_escuro = "d8e0f2"
   

    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=12, max_row=size+10,min_col=5,max_col=5):
        for cell in row:
            cell.border = borda


    worksheet.column_dimensions['a'].width = 20 #N
    worksheet.column_dimensions['b'].width = 65 #Fornecedor
    worksheet.column_dimensions['c'].width = 35 # ateogira de despesa
    worksheet.column_dimensions['d'].width = 35 #data
    worksheet.column_dimensions['e'].width = 35 #montante


        #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=11, max_row=size, min_col=1, max_col=5):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)

    for row in range(12,size+1):
        cell = worksheet[f'E{row}']
        cell.style = locals()[input3]
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 12
#
   
        
    for rows in worksheet.iter_rows(min_row=12, max_row=size, min_col=1, max_col=5):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 5:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )



    #subtotal
    stringAfinarCelula =size+1
    worksheet.row_dimensions[size+1].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:B{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    left_celula_cell2 = f"B{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Total"
    top_left_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )
    worksheet[left_celula_cell2].border = Border(top=Side(border_style="medium")  ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )


    
    worksheet.row_dimensions[size+2].height = 56.25
    

     # FORMULATOTAL
    celulas_mergidas_total = f"C{size+2}:E{size+2}"
    worksheet.merge_cells(celulas_mergidas_total)
    formula = f"=SUM(E10:E{size})"
    celula = f'C{size+2}'
    celula2 =f'D{size+2}'
    celula3 =f'E{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula2].border = Border(top=Side(border_style="medium") , bottom=Side(border_style="medium") )
    worksheet[celula3].border = Border(top=Side(border_style="medium") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

    worksheet[celula].number_format = 'R$ #,##0.00'


    #Local
    brasilia_row = size + 4
    brasilia_formula = f"Local:"
    top_left_brasilia_cell_formula = f'B{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)

    #data
    data_row = size + 5
    data_formula = f"Data:"
    top_left_data_cell_formula = f'B{data_row}'
    top_left_data_cell = worksheet[top_left_data_cell_formula]
    top_left_data_cell.value = data_formula
    top_left_data_cell.alignment = Alignment(horizontal="center",vertical="center")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)

    #Representante Legal da Instituição Beneficiária:
    repre_row = size + 7
    repre_formula = f"Representante Legal da Instituição Beneficiária::"
    top_left_repre_cell_formula = f'B{repre_row}'
    top_left_repre_cell = worksheet[top_left_repre_cell_formula]
    top_left_repre_cell.value = repre_formula
    top_left_repre_cell.alignment = Alignment(horizontal="center",vertical="center")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


    #diretor preseitente
    diretor_row = size + 9
    diretor = f"Diretor-Presidente"
    top_left_diretor_cell_formula = f'C{diretor_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell.value = diretor
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")

    for row in worksheet.iter_rows(min_row=diretor_row+1, max_row=diretor_row+1,min_col=1,max_col=5):
        for cell in row:
            if cell.column == 5:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )
    
    workbook.save(tabela)
    workbook.close()




tabela = pegar_caminho('ModeloOPAS.xlsx')
workbook = openpyxl.load_workbook(tabela)
nomeTabela ="Relatório Detalhado"
tituloStyle = "aff"
workbook.save("tabelapreenchida.xlsx")
workbook.close()
maior = 20
tabela2 = pegar_caminho('tabelapreenchida.xlsx')
print(tabela2)
estiloOpas(tabela2,maior,tituloStyle,nomeTabela)