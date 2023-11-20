import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
import os
#pegar o caminho do arquivo
def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho

def create_variable(name1, name2):
    # Create a dictionary to store values with keys based on inputs
    variables = {}
    variable_name = f"{name1}_{name2}"  # Create a variable name based on
    variables[variable_name] = []
    
       # Assign the value to the dynamically created variable name
    return variables

def estiloGeral(tabela,tamanho,nomeVariavel,nomeTabela):
    nomeSheet=nomeVariavel
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'

    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 25
    worksheet.column_dimensions['b'].width = 25
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35#descrição
    worksheet.column_dimensions['e'].width = 65 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 25 #data de emissão
    worksheet.column_dimensions['g'].width = 25 #data de emissão
    worksheet.column_dimensions['h'].width = 25 #data de emissão
    worksheet.column_dimensions['i'].width = 25 #data de emissão
    worksheet.column_dimensions['j'].width = 25 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:J2')
    worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - DIÁRIAS'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:F3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:F4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:F5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:F6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:F7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    locals()[input2].height = 20
    linha_number = 9
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

    valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
        worksheet.row_dimensions[row[0].row].height = 60
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'J{row}']
        cell.style = locals()[input3]
        
    for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 10:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(J10:J{size})"
    celula = f'J{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30


    input4 = f'row_style_diaria_append{nomeVariavel}'
    #estilo colunas restitucoes creditadas
    locals()[input4] = NamedStyle(name=f'{input4}')
    locals()[input4].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    locals()[input4].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input4].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input4].height = 30
    locals()[input4].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


    row_number = size + 4
   
    for column in range(1, 11):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = locals()[input4]
        if cell.column == 10:
            cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



    values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        if coluna == 4:
            coluna = coluna + 1
        coluna = coluna + 1
        

    merge_formula = f'D{row_number}:E{row_number}'
    worksheet.merge_cells(merge_formula)

    
    #subtotal2
    sub_total2_row = size + 5
    subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

    sub_formula_row_celula = f'J{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

      #total1-2
    total12_row = size + 6
    total12_merge_cells = f'A{total12_row}:I{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


    #total_formula
    total_formula_row = size + 6
    total_formulaa = f'=J{size}'
    total_formula_row_celula = f'J{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )

    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa


    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f"='Receita x Despesa'!A42:J42"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='Receita x Despesa'!A45"
    diretor_cargo_formula = f"='Receita x Despesa'!A46"
    diretor_cpf_formula = f"='Receita x Despesa'!A47"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='Receita x Despesa'!H45"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
    coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
        for cell in row:
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()

def estilo_conciliacoes_bancaria(tabela,tamanho,tamanho2):
    
  
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Conciliação Bancária']

   
    size = tamanho + 16
    #worksheet.row_dimensions[27].height = 50
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'

    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
   


    worksheet.column_dimensions['a'].width = 25
    worksheet.column_dimensions['b'].width = 25
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35
    worksheet.column_dimensions['e'].width = 20
    worksheet.column_dimensions['f'].width = 20
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:F2')
    worksheet['A1'] = f'C O N C I L I A Ç Ã O   B A N C Á R I A'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:F3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:F4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:F5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:F6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:F7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A9:F9')
    worksheet['A9'] = '1.Saldo conforme extratos bancários na data final do período'
    worksheet['A9'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A9'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    worksheet.merge_cells('A10:E10')
    worksheet['A10'] = 'Saldo de Conta Corrente(R$)'
    worksheet['A10'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A10'].alignment = Alignment(horizontal="right",vertical="center")

    worksheet.merge_cells('A11:E11')
    worksheet['A11'] = 'Saldo de Aplicações Financeiras(R$)'
    worksheet['A11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet['A11'].alignment = Alignment(horizontal="right",vertical="center")

    worksheet.merge_cells('A13:F13')
    worksheet['A13'] = '2. Restituições não creditadas pelo banco até a data final do período'
    worksheet['A13'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A13'].alignment = Alignment(horizontal="left",vertical="center")
    worksheet['A13'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    for i in range(15,size):
        sttring = f"D{i}:F{i}"
        worksheet.merge_cells(sttring)
        
    for i in range(size+3,size+3+tamanho2):
        sttring = f"D{i}:F{i}"
        worksheet.merge_cells(sttring)

    custom_number_format_conciliacoes = []
    # MASCARA R$
    if custom_number_format_conciliacoes!= False: 
        custom_number_format_conciliacoes = NamedStyle(name='custom_number_format_conciliacoes')
        custom_number_format_conciliacoes.number_format = 'R$ #,##0.00'
        custom_number_format_conciliacoes.font = Font(name="Arial", size=12, color="000000")
        custom_number_format_conciliacoes.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
    #stylecinza
    start_row = 15
    for row in range(start_row,size+1):
        cell = worksheet[f'B{row}']
        cell.style = custom_number_format_conciliacoes
        
    for rows in worksheet.iter_rows(min_row=15, max_row=size, min_col=1, max_col=6):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)


    row_number = 15
    values = ["Data","Valor(R$)","Documento",'Descrição']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        
        coluna = coluna + 1

    # FORMULATOTAL
    formula = f"=SUM(B16:B{size-1})"
    celula = f'B{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #Total
    celula_total = F'A{size}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)
    #'3. Restituições não creditadas pelo banco até a data final do período'
    string_reituicoes_creditadas = f'A{size+2}:F{size+2}'
    row_creditadas = f'A{size+2}'
    worksheet.merge_cells(string_reituicoes_creditadas)
    
    worksheet[row_creditadas] = '3. Restituições creditadas pelo banco até a data final do período'
    worksheet[row_creditadas].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet[row_creditadas].alignment = Alignment(horizontal="left",vertical="center")
    worksheet[row_creditadas].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    #data valor documento descrição
    row_number = size+3
    values = ["Data","Valor(R$)","Documento",'Descrição']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        coluna = coluna + 1

    for rows in worksheet.iter_rows(min_row=15, max_row=15, min_col=1, max_col=6):  
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    for rows in worksheet.iter_rows(min_row=row_number, max_row=row_number, min_col=1, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                


    for row in range(size+4,size+4+tamanho2):
        cell = worksheet[f'B{row}']
        cell.style = custom_number_format_conciliacoes
        
    for rows in worksheet.iter_rows(min_row=size+3, max_row=size+3+tamanho, min_col=1, max_col=6):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                if cell.column == 6: 
                        cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )


    # FORMULATOTALrestituição
    formula = f"=SUM(B{size+4}:B{size+tamanho2+3})"
    celula = f'B{size+tamanho2+4}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #Total
    celula_total = F'A{size+tamanho2+4}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)
    #Saldo disponível p/ período seguinte (1 +2 - 3)
    string_saldo_disponivel = f'A{size+3+tamanho2+3}:D{size+3+tamanho2+3}'
    celula_string_saldo = f'A{size+tamanho2+6}'
    worksheet[celula_string_saldo].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet.merge_cells(string_saldo_disponivel)
    worksheet[celula_string_saldo]= f'Saldo disponível p/ período seguinte (1 + 2 - 3)'
    #total saldo diposnivel
    string_merge_saldo_disponivel = f'E{size+3+tamanho2+3}:F{size+3+tamanho2+3}'
    celula_string_total = f'E{size+tamanho2+6}'
    worksheet.merge_cells(string_merge_saldo_disponivel)
    saldodiposnivelformat_conciliacoes = NamedStyle(name='saldodiposnivelformat_conciliacoes')
    saldodiposnivelformat_conciliacoes.number_format = 'R$ #,##0.00'
    saldodiposnivelformat_conciliacoes.font = Font(name="Arial", size=12, color="000000")
    saldodiposnivelformat_conciliacoes.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    saldodiposnivelformat_conciliacoes.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    celular = worksheet[celula_string_total]
    celular.style = saldodiposnivelformat_conciliacoes
    celular.value = f'=F10+F11+B{size} -B{size+tamanho2+4}'

     #brasilia
    brasilia_row = size + tamanho2+ 8
    brasilia_formula = f"='Receita x Despesa'!A42:F42"
    brasilia_merge_cells = f'A{brasilia_row}:F{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 10 + tamanho2
    diretor_cargo_row = size + 11 + tamanho2
    diretor_cpf_row = size + 12 + tamanho2
    diretor_nome_formula = f"='Receita x Despesa'!A45"
    diretor_cargo_formula = f"='Receita x Despesa'!A46"
    diretor_cpf_formula = f"='Receita x Despesa'!A47"
    diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + tamanho2 + 10
    coordenadora_cargo_row = size + 11 + tamanho2
    coordenadora_cpf_row = size + 12+ tamanho2
    coordenadora_nome_formula = f"='Receita x Despesa'!H45"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
    coordenadora_merge_cells = f'D{coordenadora_row}:F{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'D{coordenadora_cargo_row}:F{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'D{coordenadora_cpf_row}:F{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'D{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'D{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'D{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.border = borda
    top_left_coordenadora_cell.font= Font(bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cargo_formula.border = borda
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.border = borda

    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=6,max_col=6):
        for cell in row:
            cell.border = borda

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=6):
        for cell in row:
            if cell.column == 6:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()

def estilo_rendimento_de_aplicacao(tabela,tamanho):
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Rendimento de Aplicação']

   
    size = tamanho + 16
    worksheet.row_dimensions[10].height = 2
    worksheet.row_dimensions[9].height = 20

    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'
    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+9,min_col=8,max_col=8):
        for cell in row:
            cell.border = borda

    

    worksheet.column_dimensions['a'].width = 20
    worksheet.column_dimensions['b'].width = 20
    worksheet.column_dimensions['c'].width = 20
    worksheet.column_dimensions['d'].width = 20
    worksheet.column_dimensions['e'].width = 20
    worksheet.column_dimensions['f'].width = 20
    worksheet.column_dimensions['g'].width = 20
    worksheet.column_dimensions['h'].width = 20
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:H2')
    worksheet['A1'] = f'D E M O N S T R A T I V O   D E   R E N D I M E N T O   D E   A P L I C A Ç Ã O   F I N A N C E I R A'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:H3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:H4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:H5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:H6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:H7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A9:H9')
    worksheet['A9'] = 'RF Ref DI Plus Ágil - CNP JRF REF DI PLUS ÁGIL'
    worksheet['A9'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A9'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A9'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    

    

    #stylecinza
    start_row = 11
    for rows in worksheet.iter_rows(min_row=start_row, max_row=13, min_col=1, max_col=8):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)


    row_number = 11
    values = ["Período","Saldo Anterior","Valor Aplicado no período",'Valor Resgatado no Período','Rendimento Bruto','Imposto de Renda / IOF','Rendimento Líquido','Saldo']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        
        coluna = coluna + 1

    for i in range(1,9):
        worksheet.merge_cells(start_row=11,end_row=13,start_column=i,end_column=i)
    #BARRAS DE DADOS
    start_row = 14
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size, min_col=1, max_col=8):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,
                                            fill_type = "solid")
                cell.font = Font(name="Arial", size=12, color="000000")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    #MASCARA VERMELHO
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=6, max_col=6):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="f90000")
    #MASCARANEGRITO
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=1, max_col=1):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    #MASCARA AZUL
    for rows in worksheet.iter_rows(min_row=start_row, max_row=size-1, min_col=7, max_col=7):
            for cell in rows:
                cell.font = Font(name="Arial", size=12, color="141fca",bold=True)

    
    #barra de totais
    # FORMULATOTAL
     #C
    formula = f"=SUM(C14:C{size-1})"
    celula = f'C{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #D
    formula = f"=SUM(D14:D{size-1})"
    celula = f'D{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #E
    formula = f"=SUM(E14:E{size-1})"
    celula = f'E{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #F
    formula = f"=SUM(F14:F{size-1})"
    celula = f'F{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #G
    formula = f"=SUM(G14:G{size-1})"
    celula = f'G{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    #H
    formula = f"=SUM(H14:H{size-1})"
    celula = f'H{size}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)

    #Total
    celula_total = F'A{size}'
    worksheet[celula_total] = f'TOTAL'
    worksheet[celula_total].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[celula_total].font = Font(name="Arial", size=12, color="000000",bold=True)    

    #brasilia
    brasilia_row = size + 2
    brasilia_formula = f"='Receita x Despesa'!A42:F42"
    brasilia_merge_cells = f'A{brasilia_row}:F{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    # #DiretorFinanceiro
    diretor_row = size + 6 
    diretor_cargo_row = size + 7 
    diretor_cpf_row = size + 8
    diretor_nome_formula = f"='Receita x Despesa'!A45"
    diretor_cargo_formula = f"='Receita x Despesa'!A46"
    diretor_cpf_formula = f"='Receita x Despesa'!A47"
    diretor_merge_cells = f'A{diretor_row}:C{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:C{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:C{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size  + 6
    coordenadora_cargo_row = size + 7 
    coordenadora_cpf_row = size + 8
    coordenadora_nome_formula = f"='Receita x Despesa'!H45"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
    coordenadora_merge_cells = f'E{coordenadora_row}:G{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'E{coordenadora_cargo_row}:G{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'E{coordenadora_cpf_row}:G{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'E{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'E{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'E{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")


    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=8):
        for cell in row:
            if cell.column == 4:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )


    workbook.save(tabela)
    workbook.close()

def estiloRelacaoBens(tabela,tamanho,nomeVariavel,nomeTabela):
    nomeVariavel = f'material'
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook[nomeTabela]
    size = tamanho + 1
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'

    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 25
    worksheet.column_dimensions['b'].width = 25
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35#descrição
    worksheet.column_dimensions['e'].width = 40 #n do recibo ou qeuivalente
    worksheet.column_dimensions['f'].width = 25 #data de emissão
    worksheet.column_dimensions['g'].width = 25 #data de emissão
    worksheet.column_dimensions['h'].width = 25 #data de emissão
    worksheet.column_dimensions['i'].width = 25 #data de emissão
    worksheet.column_dimensions['j'].width = 25 #data de emissão


    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:J2')
    worksheet['A1'] = f'RELAÇÃO DE BENS'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")

    worksheet.merge_cells('A3:J4')
    worksheet['A3'] = f'(ADQUIRIDOS, PRODUZIDOS OU CONSTRUÍDOS COM RECURSOS)'
    worksheet['A3'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A3'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A3'].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    
    worksheet.merge_cells('A5:F5')
    worksheet['A5'] = "='Receita x Despesa'!A3:J3"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A6:F6')
    worksheet['A6'] = "='Receita x Despesa'!A4:J4"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:F7')
    worksheet['A7'] = "='Receita x Despesa'!A5:J5"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A8:F8')
    worksheet['A8'] = "='Receita x Despesa'!A6:J6"
    worksheet['A8'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A8'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A9:F9')
    worksheet['A9'] = "='Receita x Despesa'!A7:J7"
    worksheet['A9'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A9'].alignment = Alignment(horizontal="left",vertical="center")
    
    #variavel
  
    input2=f'rowStyle{nomeVariavel}'
   
    borda = Border(
    left=Side(border_style='thin', color='FFFFFF'),  
    right=Side(border_style='thin', color='FFFFFF'),  
    top=Side(border_style='thin', color='FFFFFF'), 
    bottom=Side(border_style='thin', color='FFFFFF')  
)
    borda2 = Border(
    left=Side(border_style='hair', color='000000'),  
    right=Side(border_style='hair', color='000000'),  
    top=Side(border_style='hair', color='000000'), 
    bottom=Side(border_style='hair', color='000000')  
)

    #colunas azul cabecario
    locals()[input2] = NamedStyle(name=f'{input2}')
    locals()[input2].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    locals()[input2].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    locals()[input2].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    locals()[input2].border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    locals()[input2].height = 20
    linha_number = 11
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
        for cell in row:
            cell.style = locals()[input2]
            if cell.column == 10:
                cell.border = Border(left=Side(border_style="thin", color='FFFFFF')  ,bottom=Side(border_style="thin", color='FFFFFF'), right=Side(border_style="medium") )


            cell.border = borda       

    
  


##CABECARIO
    
    worksheet["A11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("A11:A12")
    worksheet["A11"] ="Nº DO ITEM"
    

    worksheet["B11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("B11:B12")
    worksheet["B11"] = "DESCRIÇÃO DO BEM"

    worksheet["C11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("C11:C12")
    worksheet["C11"] = "NÚMERO PATRIMONIAL DO BEM"

    worksheet["D11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("D11:E11")
    worksheet["D11"] = "DOCUMENTAÇÃO FISCAL"

    worksheet["D12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["D12"] = "DATA"
    worksheet["D12"].border = borda2   

    worksheet["E12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["E12"] = "Nº "
    worksheet["E12"].border = borda2   

    worksheet["F11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("F11:F12")
    worksheet["F11"] = "LOCALIZAÇÃO"

    worksheet["G11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("G11:G12")
    worksheet["G11"] = "QTD."

    worksheet["H11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("H11:I11")
    worksheet["H11"] = "VALOR (R$)"

    worksheet["H12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["H12"] = "Unitário"
    worksheet["H12"].border = borda2   
    worksheet["I12"].fill = openpyxl.styles.PatternFill(start_color=cinza, end_color=cinza, fill_type='solid')
    worksheet["I12"] = "Total"
    worksheet["I12"].border = borda2   

  
   

   












    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
        worksheet.row_dimensions[row[0].row].height = 35
    input3 = f'customNumber{nomeVariavel}'
    
    # MASCARA R$
   
    locals()[input3] = NamedStyle(name=f'{input3}')
    locals()[input3].number_format = 'R$ #,##0.00'
    locals()[input3].font = Font(name="Arial", size=12, color="000000")
    locals()[input3].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'J{row}']
        cell.style = locals()[input3]
        
    for rows in worksheet.iter_rows(min_row=13, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 10:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "TOTAL"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(J10:J{size})"
    celula = f'J{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'


    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f"='Receita x Despesa'!A42:J42"
    brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='Receita x Despesa'!A45"
    diretor_cargo_formula = f"='Receita x Despesa'!A46"
    diretor_cpf_formula = f"='Receita x Despesa'!A47"
    diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='Receita x Despesa'!H45"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
    coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
        for cell in row:
            if cell.column == 10:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    worksheet["J11"].fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    worksheet.merge_cells("J11:J12")
    worksheet["J11"] = "RESPONSÁVEL PELA GUARDA DO BEM"
    worksheet["J11"].border = Border(left=Side(border_style="thin", color='FFFFFF')  ,bottom=Side(border_style="thin", color='FFFFFF'), right=Side(border_style="medium") )
    worksheet["J11"].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet["J10"].border = Border(right =Side(border_style="medium"))
    worksheet.row_dimensions[10].height = 2


    workbook.save(tabela)
    workbook.close()


def estilo_demonstrativoDeReceita(tabela,tamanho):
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    worksheet = workbook['Demonstrativo de Receita']
    size = tamanho + 10
    cinza = "d9d9d9"
    cinza_escuro = "bfbfbf"
    azul = "336394"
    azul_claro = '1c8cbc'

    borda = Border(right=Side(border_style="medium"))
    worksheet.sheet_view.showGridLines = False
    # 
    for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=4,max_col=4):
        for cell in row:
            cell.border = borda
            

    worksheet.column_dimensions['a'].width = 35
    worksheet.column_dimensions['b'].width = 35
    worksheet.column_dimensions['c'].width = 35
    worksheet.column_dimensions['d'].width = 35#descrição
   

    #cabecario relação de pagamentos - outro servicoes de terceiros
    worksheet.merge_cells('A1:D2')
    worksheet['A1'] = f'D E M O N S T R A T I V O   D E   R E C E I T A'
    worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
    worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
    worksheet.merge_cells('A3:D3')
    worksheet['A3'] = "='Receita x Despesa'!A3:J3"
    worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

    worksheet.merge_cells('A4:D4')
    worksheet['A4'] = "='Receita x Despesa'!A4:J4"
    worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A5:D5')
    worksheet['A5'] = "='Receita x Despesa'!A5:J5"
    worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A6:D6')
    worksheet['A6'] = "='Receita x Despesa'!A6:J6"
    worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
    worksheet.merge_cells('A7:D7')
    worksheet['A7'] = "='Receita x Despesa'!A7:J7"
    worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
    worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")

    #colunas azul cabecario
    row_style_demonstrativo = NamedStyle(name='row_style_demonstrativo')
    row_style_demonstrativo.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style_demonstrativo.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    row_style_demonstrativo.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style_demonstrativo.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
    row_style_demonstrativo.height = 20
    linha_number = 9
    for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=4):
        for cell in row:
            cell.style = row_style_demonstrativo
            if cell.column == 4:
                cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

    valores = ["Data de Entrada","Cod.BB_Histórico","Documento",'Valor']
    col = 1
    for a,b in enumerate(valores):
        worksheet.cell(row=linha_number, column=col, value=b)
        col = col + 1


    #Aumentar  a altura das celulas 
    for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=4):
        worksheet.row_dimensions[row[0].row].height = 35

    custom_number_format_demonstrativo = []
    # MASCARA R$
    if custom_number_format_demonstrativo!= False: 
        custom_number_format_demonstrativo = NamedStyle(name='custom_number_format_demonstrativo')
        custom_number_format_demonstrativo.number_format = 'R$ #,##0.00'
        custom_number_format_demonstrativo.font = Font(name="Arial", size=12, color="000000")
        custom_number_format_demonstrativo.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    #estilocinzasimcinzanao
    value_to_stop = size  
    start_row = 10
#
    for row in range(start_row,size+1):
        cell = worksheet[f'D{row}']
        cell.style = custom_number_format_demonstrativo
        
    for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=4):
            for cell in rows:
                if cell.row % 2:
                    cell.fill = PatternFill(start_color=cinza, end_color=cinza,
                                            fill_type = "solid")
                if cell.column == 4:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
                else:
                    cell.font = Font(name="Arial", size=12, color="000000")
                    cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                    cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
    #subtotal
    stringAfinarCelula =size+2
    worksheet.row_dimensions[size+2].height = 6
    celulas_mergidas_subtotal = f"A{size+2}:C{size+2}"
    worksheet.merge_cells(celulas_mergidas_subtotal)
    left_celula_cell = f"A{size+2}"
    top_left_cell = worksheet[left_celula_cell]
    top_left_cell.value = "Sub Total1"
    top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

    worksheet.row_dimensions[size+2].height = 56.25

     # FORMULATOTAL
    formula = f"=SUM(D10:D{size})"
    celula = f'D{size+2}'
    worksheet[celula] = formula
    worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
    worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
    worksheet[celula].number_format = 'R$ #,##0.00'
    #restituições creditadas
    restituicoes = size + 3
    celula_restituicoes=f'A{restituicoes}'
    worksheet[celula_restituicoes].value = "Estorno de Mensalidades"
    worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet.row_dimensions[restituicoes].height = 30

    #estilo colunas restitucoes creditadas
    row_style_demonstrativo_append = NamedStyle(name='row_style_demonstrativo_append')
    row_style_demonstrativo_append.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
    row_style_demonstrativo_append.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
    row_style_demonstrativo_append.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    row_style_demonstrativo_append.height = 30
    row_style_demonstrativo_append.border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


    row_number = size + 4
   
    for column in range(1, 5):  
        cell = worksheet.cell(row=row_number, column=column)
        cell.style = row_style_demonstrativo_append
        if cell.column == 4:
            cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



    values = ["Data de Entrada","Cod.BB_Histórico","Documento",'Valor']
    coluna = 1
    for a,b in enumerate(values):
        worksheet.cell(row=row_number, column=coluna, value=b)
        if coluna == 4:
            coluna = coluna + 1
        coluna = coluna + 1
        



    
    #subtotal2
    sub_total2_row = size + 5
    subtotal_merge_cells= f'A{sub_total2_row}:C{sub_total2_row}'
    worksheet.merge_cells(subtotal_merge_cells)
    top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
    top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
    top_left_subtotal2_cell.value = "Sub Total 2"
    top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

    sub_formula_row_celula = f'D{sub_total2_row}'
    worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
    worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

      #total1-2
    total12_row = size + 6
    total12_merge_cells = f'A{total12_row}:C{total12_row}'
    worksheet.merge_cells(total12_merge_cells)
    top_left_total12_cell_formula = f'A{total12_row}'
    top_left_total12_cell = worksheet[top_left_total12_cell_formula]
    top_left_total12_cell.value = "Total(1-2)"
    top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
    top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


    #total_formula
    total_formula_row = size + 6
    total_formulaa = f'=D{size}'
    total_formula_row_celula = f'D{total_formula_row}'
    worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
    worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
    worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )
    worksheet.row_dimensions[total_formula_row].height = 30
    worksheet[total_formula_row_celula] = total_formulaa


    #brasilia
    brasilia_row = size + 7
    brasilia_formula = f"='Receita x Despesa'!A42:D42"
    brasilia_merge_cells = f'A{brasilia_row}:D{brasilia_row}'
    worksheet.merge_cells(brasilia_merge_cells)
    top_left_brasilia_cell_formula = f'A{brasilia_row}'
    top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
    top_left_brasilia_cell.value = brasilia_formula
    top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

    #DiretorFinanceiro
    diretor_row = size + 8
    diretor_cargo_row = size + 9
    diretor_cpf_row = size + 10
    diretor_nome_formula = f"='Receita x Despesa'!A45"
    diretor_cargo_formula = f"='Receita x Despesa'!A46"
    diretor_cpf_formula = f"='Receita x Despesa'!A47"
    diretor_merge_cells = f'A{diretor_row}:B{diretor_row}'
    diretor_cargo_merge_cells = f'A{diretor_cargo_row}:B{diretor_cargo_row}'
    diretor_cpf_merge_cells = f'A{diretor_cpf_row}:B{diretor_cpf_row}'
    worksheet.merge_cells(diretor_merge_cells)
    worksheet.merge_cells(diretor_cargo_merge_cells)
    worksheet.merge_cells(diretor_cpf_merge_cells)
    top_left_diretor_cell_formula = f'A{diretor_row}'
    top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
    top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
    top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
    top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
    top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
    top_left_diretor_cell.value = diretor_nome_formula
    top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
    top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
    top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell.font = Font(bold=True)
    top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
    #Coordenadora
    coordenadora_row = size + 8
    coordenadora_cargo_row = size + 9
    coordenadora_cpf_row = size + 10
    coordenadora_nome_formula = f"='Receita x Despesa'!H45"
    coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
    coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
    coordenadora_merge_cells = f'C{coordenadora_row}:D{coordenadora_row}'
    coordenadora_cargo_merge_cells = f'C{coordenadora_cargo_row}:D{coordenadora_cargo_row}'
    coordenadora_cpf_merge_cells = f'C{coordenadora_cpf_row}:D{coordenadora_cpf_row}'
    worksheet.merge_cells(coordenadora_merge_cells)
    worksheet.merge_cells(coordenadora_cargo_merge_cells)
    worksheet.merge_cells(coordenadora_cpf_merge_cells)
    top_left_coordenadora_cell_formula = f'C{coordenadora_row}'
    top_left_coordenadora_cell_cargo_formula = f'C{coordenadora_cargo_row}'
    top_left_coordenadora_cell_cpf_formula = f'C{coordenadora_cpf_row}'
    top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
    top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
    top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
    top_left_coordenadora_cell.value = coordenadora_nome_formula
    top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
    top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
    top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell.font= Font(bold = True)
    top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
    top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
    # borda = Border(right=Side(border_style="medium"))
    # worksheet.sheet_view.showGridLines = False
    # # 
    # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=4):
    #     for cell in row:
    #         cell.border = borda
            
    

    for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=4):
        for cell in row:
            if cell.column == 8:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
            else:
                cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

    workbook.save(tabela)
    workbook.close()

# def estilo_fub_fisica(tabela,tamanho):
    
#     # caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(tabela)
#     worksheet = workbook['Pessoa Fisica']

   
#     size = tamanho + 10
#     #worksheet.row_dimensions[27].height = 50

#     cinza = "979CA8"
#     azul = "336394"
    
                
                
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format_fisica = []
#     # MASCARA R$
#     if custom_number_format_fisica!= False: 
#         custom_number_format_fisica = NamedStyle(name='custom_number_format_fisica')
#         custom_number_format_fisica.number_format = 'R$ #,##0.00'
#         custom_number_format_fisica.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format_fisica.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
#     #stylecinza
    

#     value_to_stop = size  
#     start_row = 10

#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format_fisica
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 cell.font = Font(name="Arial", size=12, color="000000")
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                 cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    


#     #subtotal
#     celulas_mergidas_subtotal = f"A{size}:I{size}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


  

#     # FORMULATOTAL
#     formula = f"=SUM(J10:J{size-1})"
#     celula = f'J{size}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


#     #restituições creditadas
#     restituicoes = size + 1
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30


#     #colunas azul
#     row_style_fisica = NamedStyle(name='row_style_fisica')
#     row_style_fisica.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_fisica.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
#     row_style_fisica.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_fisica.height = 30

#     row_number = size + 2
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style_fisica


#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

#     #subtotal2
#     sub_total2_row = size + 4
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

#      #subtotal12_Formula
#     # sub_formula_row = size + 4
#     # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
#     # sub_formula_row_celula = f'J{sub_formula_row}'
#     # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     # worksheet[sub_formula_row_celula] = sub_formula


#     #total1-2
#     total12_row = size + 5
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1 -2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #total_formula
#     total_formula_row = size + 5
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa

#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

#     #nome dos indices

#     azul_claro = '1c89b8'
#     row_style_fisica_cabecario = NamedStyle(name='row_style_fisica_cabecario')
#     row_style_fisica_cabecario.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_fisica_cabecario.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_fisica_cabecario.alignment = Alignment(horizontal="center",vertical="center",wrap_text = True)
#     worksheet.row_dimensions[9].height = 50
#     worksheet.column_dimensions['b'].width = 35
#     worksheet.column_dimensions['c'].width = 20
#     worksheet.column_dimensions['d'].width = 35
#     worksheet.column_dimensions['e'].width = 60
#     worksheet.column_dimensions['f'].width = 35
#     worksheet.column_dimensions['g'].width = 30
#     worksheet.column_dimensions['h'].width = 30
#     worksheet.column_dimensions['i'].width = 30
#     worksheet.column_dimensions['j'].width = 30
#     worksheet.column_dimensions['e'].height = 20
  
  
#     linha_number = 9
#     # Apply the style to each cell in the row
#     for cell in worksheet[linha_number]:
#         cell.style = row_style_fisica_cabecario

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - O U T R O S   S E R V I Ç O S   T E R C E I R O S   -   P E S S O A   F Í S I C A '
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet.merge_cells('A3:I3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:I4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:I5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:I6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:I7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
   
#     altura = 45
#     linha_inicio = 9
#     for row_number in range(linha_inicio, linha_inicio + tamanho+1):
#         worksheet.row_dimensions[row_number].height = altura

#     workbook.save('output.xlsx')
#     workbook.close()

# def estilo_fub_juridica(tabela,tamanho):
    
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['Pessoa Fisica']

   
#     size = tamanho + 10
#     #worksheet.row_dimensions[27].height = 50

#     cinza = "979CA8"
#     azul = "336394"
    
                
                
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format_juridica = []
#     # MASCARA R$
#     if custom_number_format_juridica!= False: 
#         custom_number_format_juridica = NamedStyle(name='custom_number_format_juridica')
#         custom_number_format_juridica.number_format = 'R$ #,##0.00'
#         custom_number_format_juridica.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format_juridica.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
#     #stylecinza
    

#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format_juridica
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 cell.font = Font(name="Arial", size=12, color="000000")
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                 cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    


#     #subtotal
#     celulas_mergidas_subtotal = f"A{size}:I{size}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


  

#     # FORMULATOTAL
#     formula = f"=SUM(J10:J{size-1})"
#     celula = f'J{size}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


#     #restituições creditadas
#     restituicoes = size + 1
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30


#     #colunas azul
#     row_style_juridica = NamedStyle(name='row_style_juridica')
#     row_style_juridica.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_juridica.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
#     row_style_juridica.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_juridica.height = 30

#     row_number = size + 2
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style_juridica


#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

#     #subtotal2
#     sub_total2_row = size + 4
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

#      #subtotal12_Formula
#     # sub_formula_row = size + 4
#     # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
#     # sub_formula_row_celula = f'J{sub_formula_row}'
#     # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     # worksheet[sub_formula_row_celula] = sub_formula


#     #total1-2
#     total12_row = size + 5
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1 -2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #total_formula
#     total_formula_row = size + 5
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa

#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

#     #nome dos indices

#     azul_claro = '1c89b8'
#     row_style_cabecario_juridica = NamedStyle(name='row_style_cabecario_juridica')
#     row_style_cabecario_juridica.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_cabecario_juridica.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_cabecario_juridica.alignment = Alignment(horizontal="center",vertical="center",wrap_text = True)
#     worksheet.row_dimensions[9].height = 50
#     worksheet.column_dimensions['b'].width = 35
#     worksheet.column_dimensions['c'].width = 20
#     worksheet.column_dimensions['d'].width = 35
#     worksheet.column_dimensions['e'].width = 60
#     worksheet.column_dimensions['f'].width = 35
#     worksheet.column_dimensions['g'].width = 30
#     worksheet.column_dimensions['h'].width = 30
#     worksheet.column_dimensions['i'].width = 30
#     worksheet.column_dimensions['j'].width = 30
#     worksheet.column_dimensions['e'].height = 20
  
  
#     linha_number = 9
#     # Apply the style to each cell in the row
#     for cell in worksheet[linha_number]:
#         cell.style = row_style_cabecario_juridica

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - O U T R O S   S E R V I Ç O S   T E R C E I R O S   -   P E S S O A   F Í S I C A '
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet.merge_cells('A3:I3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:I4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:I5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:I6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:I7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
   
#     altura = 45
#     linha_inicio = 9
#     for row_number in range(linha_inicio, linha_inicio + tamanho+1):
#         worksheet.row_dimensions[row_number].height = altura
    

#     # Save the workbook to a file
#     workbook.save(tabela)
#     workbook.close()

# def estilo_fub_juridica_juridica(tabela,tamanho):
    
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['Pessoa Jurídica']

   
#     size = tamanho + 10
#     #worksheet.row_dimensions[27].height = 50

#     cinza = "979CA8"
#     azul = "336394"
    
                
                
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_formatjuridica_juridica = []
#     # MASCARA R$
#     if custom_number_formatjuridica_juridica!= False: 
#         custom_number_formatjuridica_juridica = NamedStyle(name='custom_number_formatjuridica_juridica')
#         custom_number_formatjuridica_juridica.number_format = 'R$ #,##0.00'
#         custom_number_formatjuridica_juridica.font = Font(name="Arial", size=12, color="000000")
#         custom_number_formatjuridica_juridica.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
#     #stylecinza
    

#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_formatjuridica_juridica
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 cell.font = Font(name="Arial", size=12, color="000000")
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                 cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    


#     #subtotal
#     celulas_mergidas_subtotal = f"A{size}:I{size}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


  

#     # FORMULATOTAL
#     formula = f"=SUM(J10:J{size-1})"
#     celula = f'J{size}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


#     #restituições creditadas
#     restituicoes = size + 1
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30


#     #colunas azul
#     row_stylejuridica_juridica = NamedStyle(name='row_stylejuridica_juridica')
#     row_stylejuridica_juridica.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_stylejuridica_juridica.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
#     row_stylejuridica_juridica.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_stylejuridica_juridica.height = 30

#     row_number = size + 2
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_stylejuridica_juridica


#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

#     #subtotal2
#     sub_total2_row = size + 4
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

#      #subtotal12_Formula
#     # sub_formula_row = size + 4
#     # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
#     # sub_formula_row_celula = f'J{sub_formula_row}'
#     # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     # worksheet[sub_formula_row_celula] = sub_formula


#     #total1-2
#     total12_row = size + 5
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1 -2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #total_formula
#     total_formula_row = size + 5
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa

#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

#     #nome dos indices

#     azul_claro = '1c89b8'
#     row_style_cabecariojuridica_juridica = NamedStyle(name='row_style_cabecariojuridica_juridica')
#     row_style_cabecariojuridica_juridica.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_cabecariojuridica_juridica.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_cabecariojuridica_juridica.alignment = Alignment(horizontal="center",vertical="center",wrap_text = True)
#     worksheet.row_dimensions[9].height = 50
#     worksheet.column_dimensions['b'].width = 35
#     worksheet.column_dimensions['c'].width = 20
#     worksheet.column_dimensions['d'].width = 35
#     worksheet.column_dimensions['e'].width = 60
#     worksheet.column_dimensions['f'].width = 35
#     worksheet.column_dimensions['g'].width = 30
#     worksheet.column_dimensions['h'].width = 30
#     worksheet.column_dimensions['i'].width = 30
#     worksheet.column_dimensions['j'].width = 30
#     worksheet.column_dimensions['e'].height = 20
  
  
#     linha_number = 9
#     # Apply the style to each cell in the row
#     for cell in worksheet[linha_number]:
#         cell.style = row_style_cabecariojuridica_juridica

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - O U T R O S   S E R V I Ç O S   D E   T E R C E I R O S   -   P E S S O A   J U R Í D I C A '
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet.merge_cells('A3:I3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:I4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:I5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:I6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:I7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
   
#     altura = 45
#     linha_inicio = 9
#     for row_number in range(linha_inicio, linha_inicio + tamanho+1):
#         worksheet.row_dimensions[row_number].height = altura
    

#     # Save the workbook to a file
#     workbook.save(tabela)
#     workbook.close()

# def estilo_serv_terceiro(tabela,tamanho):
    
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['Outros Serviços Terceiros - PF']

   
#     size = tamanho + 10
#     #worksheet.row_dimensions[27].height = 50

#     cinza = "979CA8"
#     azul = "336394"
    
                
                
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format = []
#     # MASCARA R$
#     if custom_number_format!= False: 
#         custom_number_format = NamedStyle(name='custom_number_format')
#         custom_number_format.number_format = 'R$ #,##0.00'
#         custom_number_format.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
#     #stylecinza
    

#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 cell.font = Font(name="Arial", size=12, color="000000")
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                 cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    


#     #subtotal
#     celulas_mergidas_subtotal = f"A{size}:I{size}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


  

#     # FORMULATOTAL
#     formula = f"=SUM(J10:J{size-1})"
#     celula = f'J{size}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


#     #restituições creditadas
#     restituicoes = size + 1
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30


#     #colunas azul
#     row_style = NamedStyle(name='row_style')
#     row_style.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
#     row_style.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style.height = 30

#     row_number = size + 2
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style


#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

#     #subtotal2
#     sub_total2_row = size + 4
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

#      #subtotal12_Formula
#     # sub_formula_row = size + 4
#     # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
#     # sub_formula_row_celula = f'J{sub_formula_row}'
#     # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     # worksheet[sub_formula_row_celula] = sub_formula


#     #total1-2
#     total12_row = size + 5
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1 -2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #total_formula
#     total_formula_row = size + 5
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa

#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

#     #nome dos indices

#     azul_claro = '1c89b8'
#     row_style_cabecario = NamedStyle(name='row_style_cabecario')
#     row_style_cabecario.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_cabecario.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_cabecario.alignment = Alignment(horizontal="center",vertical="center",wrap_text = True)
#     worksheet.row_dimensions[9].height = 50
#     worksheet.column_dimensions['b'].width = 35
#     worksheet.column_dimensions['c'].width = 20
#     worksheet.column_dimensions['d'].width = 35
#     worksheet.column_dimensions['e'].width = 60
#     worksheet.column_dimensions['f'].width = 35
#     worksheet.column_dimensions['g'].width = 30
#     worksheet.column_dimensions['h'].width = 30
#     worksheet.column_dimensions['i'].width = 30
#     worksheet.column_dimensions['j'].width = 30
#     worksheet.column_dimensions['e'].height = 20
  
  
#     linha_number = 9
#     # Apply the style to each cell in the row
#     for cell in worksheet[linha_number]:
#         cell.style = row_style_cabecario

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - O U T R O S  S E R V I Ç O S D E T E R C E I R O S - C E L E T I S T A S'
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet.merge_cells('A3:I3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:I4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:I5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:I6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:I7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
   
#     altura = 45
#     linha_inicio = 9
#     for row_number in range(linha_inicio, linha_inicio + tamanho+1):
#         worksheet.row_dimensions[row_number].height = altura
    

#     # Save the workbook to a file
#     workbook.save(tabela)
#     workbook.close()

# def estilo_Iss(tabela,tamanho):
    
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['ISS']

   
#     size = tamanho + 10
#     #worksheet.row_dimensions[27].height = 50

#     cinza = "979CA8"
#     azul = "336394"
    
                
                
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format_iss = []
#     # MASCARA R$
#     if custom_number_format_iss!= False: 
#         custom_number_format_iss = NamedStyle(name='custom_number_format_iss')
#         custom_number_format_iss.number_format = 'R$ #,##0.00'
#         custom_number_format_iss.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format_iss.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
#     #stylecinza
    

#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format_iss
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 cell.font = Font(name="Arial", size=12, color="000000")
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                 cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    


#     #subtotal
#     celulas_mergidas_subtotal = f"A{size}:I{size}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


  

#     # FORMULATOTAL
#     formula = f"=SUM(J10:J{size-1})"
#     celula = f'J{size}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


#     #restituições creditadas
#     restituicoes = size + 1
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30


#     #colunas azul
#     row_style_iss = NamedStyle(name='row_style_iss')
#     row_style_iss.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_iss.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
#     row_style_iss.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_iss.height = 30

#     row_number = size + 2
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style_iss


#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

#     #subtotal2
#     sub_total2_row = size + 4
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

#      #subtotal12_Formula
#     # sub_formula_row = size + 4
#     # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
#     # sub_formula_row_celula = f'J{sub_formula_row}'
#     # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     # worksheet[sub_formula_row_celula] = sub_formula


#     #total1-2
#     total12_row = size + 5
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1 -2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #total_formula
#     total_formula_row = size + 5
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa

#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

#     #nome dos indices

#     azul_claro = '1c89b8'
#     row_style_cabecario_iss = NamedStyle(name='row_style_cabecario_iss')
#     row_style_cabecario_iss.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_cabecario_iss.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_cabecario_iss.alignment = Alignment(horizontal="center",vertical="center",wrap_text = True)
#     worksheet.row_dimensions[9].height = 50
#     worksheet.column_dimensions['b'].width = 35
#     worksheet.column_dimensions['c'].width = 20
#     worksheet.column_dimensions['d'].width = 35
#     worksheet.column_dimensions['e'].width = 60
#     worksheet.column_dimensions['f'].width = 35
#     worksheet.column_dimensions['g'].width = 30
#     worksheet.column_dimensions['h'].width = 30
#     worksheet.column_dimensions['i'].width = 30
#     worksheet.column_dimensions['j'].width = 30
#     worksheet.column_dimensions['e'].height = 20
  
  
#     linha_number = 9
#     # Apply the style to each cell in the row
#     for cell in worksheet[linha_number]:
#         cell.style = row_style_cabecario_iss

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - A U X Í L I O   F I N A N C E I R O  A  E S T U D A N T E '
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet.merge_cells('A3:I3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:I4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:I5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:I6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:I7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
   
#     altura = 45
#     linha_inicio = 9
#     for row_number in range(linha_inicio, linha_inicio + tamanho+1):
#         worksheet.row_dimensions[row_number].height = altura
    

#     # Save the workbook to a file
#     workbook.save(tabela)
#     workbook.close()

# def estilo_passagens(tabela,tamanho):
    
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['Passagens e Desp. Locomoção']
#     size = tamanho + 10
#     cinza = "d9d9d9"
#     cinza_escuro = "bfbfbf"
#     azul = "336394"
#     azul_claro = '1c8cbc'

#     borda = Border(right=Side(border_style="medium"))
#     worksheet.sheet_view.showGridLines = False
#     # 
#     for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
#         for cell in row:
#             cell.border = borda
            

#     worksheet.column_dimensions['a'].width = 25
#     worksheet.column_dimensions['b'].width = 25
#     worksheet.column_dimensions['c'].width = 35
#     worksheet.column_dimensions['d'].width = 40#descrição
#     worksheet.column_dimensions['e'].width = 20 #n do recibo ou qeuivalente
#     worksheet.column_dimensions['f'].width = 25 #data de emissão
#     worksheet.column_dimensions['g'].width = 25 #data de emissão
#     worksheet.column_dimensions['h'].width = 25 #data de emissão
#     worksheet.column_dimensions['i'].width = 25 #data de emissão
#     worksheet.column_dimensions['j'].width = 25 #data de emissão


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - PASSAGENS E DESPESAS COM LOCOMOÇÃO'
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
#     worksheet.merge_cells('A3:F3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:F4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:F5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:F6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:F7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")

#     #colunas azul cabecario
#     row_style_passagens = NamedStyle(name='row_style_passagens')
#     row_style_passagens.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_passagens.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_passagens.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_passagens.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
#     row_style_passagens.height = 20
#     linha_number = 9
#     for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
#         for cell in row:
#             cell.style = row_style_passagens
#             if cell.column == 10:
#                 cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #Aumentar  a altura das celulas 
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format_passagens = []
#     # MASCARA R$
#     if custom_number_format_passagens!= False: 
#         custom_number_format_passagens = NamedStyle(name='custom_number_format_passagens')
#         custom_number_format_passagens.number_format = 'R$ #,##0.00'
#         custom_number_format_passagens.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format_passagens.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
#     #estilocinzasimcinzanao
#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format_passagens
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 if cell.column == 10:
#                     cell.font = Font(name="Arial", size=12, color="000000")
#                     cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                     cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
#                 else:
#                     cell.font = Font(name="Arial", size=12, color="000000")
#                     cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                     cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
#     #subtotal
#     stringAfinarCelula =size+2
#     worksheet.row_dimensions[size+2].height = 6
#     celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size+2}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total1"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

#     worksheet.row_dimensions[size+2].height = 56.25

#      # FORMULATOTAL
#     formula = f"=SUM(J10:J{size})"
#     celula = f'J{size+2}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
#     worksheet[celula].number_format = 'R$ #,##0.00'
#     #restituições creditadas
#     restituicoes = size + 3
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30

#     #estilo colunas restitucoes creditadas
#     row_style_passagens_append = NamedStyle(name='row_style_passagens_append')
#     row_style_passagens_append.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_passagens_append.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_passagens_append.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_passagens_append.height = 30
#     row_style_passagens_append.border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


#     row_number = size + 4
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style_passagens_append
#         if cell.column == 10:
#           cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

    
#     #subtotal2
#     sub_total2_row = size + 5
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

#       #total1-2
#     total12_row = size + 6
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1-2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


#     #total_formula
#     total_formula_row = size + 6
#     total_formulaa = f'=J{size+2}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )

#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa


#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
#     # borda = Border(right=Side(border_style="medium"))
#     # worksheet.sheet_view.showGridLines = False
#     # # 
#     # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
#     #     for cell in row:
#     #         cell.border = borda
            
    

#     for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
#         for cell in row:
#             if cell.column == 10:
#                 cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
#             else:
#                 cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

#     workbook.save(tabela)
#     workbook.close()

# def estilo_obrigacoes_tributarias(tabela,tamanho):
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['Obrigações Trib. - Encargos 20%']

   
#     size = tamanho + 10
#     #worksheet.row_dimensions[27].height = 50

#     cinza = "979CA8"
#     azul = "336394"

                
                
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format_obrig = []
#     # MASCARA R$
#     if custom_number_format_obrig!= False: 
#         custom_number_format_obrig = NamedStyle(name='custom_number_format_obrig')
#         custom_number_format_obrig.number_format = 'R$ #,##0.00'
#         custom_number_format_obrig.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format_obrig.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
#     #stylecinza
    

#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format_obrig
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 cell.font = Font(name="Arial", size=12, color="000000")
#                 cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                 cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                

    


#     #subtotal
#     celulas_mergidas_subtotal = f"A{size}:I{size}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")


  

#     # FORMULATOTAL
#     formula = f"=SUM(J10:J{size-1})"
#     celula = f'J{size}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)


#     #restituições creditadas
#     restituicoes = size + 1
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30


#     #colunas azul
#     row_style_obrig = NamedStyle(name='row_style_obrig')
#     row_style_obrig.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_obrig.fill = openpyxl.styles.PatternFill(start_color=azul, end_color=azul, fill_type='solid')
#     row_style_obrig.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_obrig.height = 30

#     row_number = size + 2
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style_obrig


#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

#     #subtotal2
#     sub_total2_row = size + 4
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)

#      #subtotal12_Formula
#     # sub_formula_row = size + 4
#     # sub_formula = f'=SOMA(J{sub_formula_row}:J{sub_formula_row})'
#     # sub_formula_row_celula = f'J{sub_formula_row}'
#     # worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     # worksheet[sub_formula_row_celula] = sub_formula


#     #total1-2
#     total12_row = size + 5
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1 -2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #total_formula
#     total_formula_row = size + 5
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul, end_color=azul,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa

#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

#     #nome dos indices

#     azul_claro = '1c89b8'
#     row_style_cabecario_obrig = NamedStyle(name='row_style_cabecario_obrig')
#     row_style_cabecario_obrig.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_cabecario_obrig.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_cabecario_obrig.alignment = Alignment(horizontal="center",vertical="center",wrap_text = True)
#     worksheet.row_dimensions[9].height = 50
#     worksheet.column_dimensions['b'].width = 35
#     worksheet.column_dimensions['c'].width = 20
#     worksheet.column_dimensions['d'].width = 35
#     worksheet.column_dimensions['e'].width = 60
#     worksheet.column_dimensions['f'].width = 35
#     worksheet.column_dimensions['g'].width = 30
#     worksheet.column_dimensions['h'].width = 30
#     worksheet.column_dimensions['i'].width = 30
#     worksheet.column_dimensions['j'].width = 30
#     worksheet.column_dimensions['e'].height = 20
  
  
#     linha_number = 9
#     # Apply the style to each cell in the row
#     for cell in worksheet[linha_number]:
#         cell.style = row_style_cabecario_obrig

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - P A S S A G E N S  E  D E S P E S A S   C O M  L O C O M O Ç Ã O'
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet.merge_cells('A3:I3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:I4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:I5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:I6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:I7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")
   
#     altura = 45
#     linha_inicio = 9
#     for row_number in range(linha_inicio, linha_inicio + tamanho+1):
#         worksheet.row_dimensions[row_number].height = altura
    

#     # Save the workbook to a file
#     workbook.save(tabela)
#     workbook.close()

# def estilo_diarias(tabela,tamanho):
#     caminho = pegar_caminho(tabela)
#     workbook = openpyxl.load_workbook(caminho)
#     worksheet = workbook['Diárias']
#     size = tamanho + 10
#     cinza = "d9d9d9"
#     cinza_escuro = "bfbfbf"
#     azul = "336394"
#     azul_claro = '1c8cbc'

#     borda = Border(right=Side(border_style="medium"))
#     worksheet.sheet_view.showGridLines = False
#     # 
#     for row in worksheet.iter_rows(min_row=1, max_row=size+11,min_col=10,max_col=10):
#         for cell in row:
#             cell.border = borda
            

#     worksheet.column_dimensions['a'].width = 25
#     worksheet.column_dimensions['b'].width = 25
#     worksheet.column_dimensions['c'].width = 35
#     worksheet.column_dimensions['d'].width = 40#descrição
#     worksheet.column_dimensions['e'].width = 20 #n do recibo ou qeuivalente
#     worksheet.column_dimensions['f'].width = 25 #data de emissão
#     worksheet.column_dimensions['g'].width = 25 #data de emissão
#     worksheet.column_dimensions['h'].width = 25 #data de emissão
#     worksheet.column_dimensions['i'].width = 25 #data de emissão
#     worksheet.column_dimensions['j'].width = 25 #data de emissão


#     #cabecario relação de pagamentos - outro servicoes de terceiros
#     worksheet.merge_cells('A1:J2')
#     worksheet['A1'] = f'R E L A Ç Ã O   D E   P A G A M E N T O S - DIÁRIAS'
#     worksheet['A1'].font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     worksheet['A1'].alignment = Alignment(horizontal="center",vertical="center")
#     worksheet['A1'].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
    
#     worksheet.merge_cells('A3:F3')
#     worksheet['A3'] = "='Receita x Despesa'!A3:J3"
#     worksheet['A3'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A3'].alignment = Alignment(horizontal="left",vertical="center")

#     worksheet.merge_cells('A4:F4')
#     worksheet['A4'] = "='Receita x Despesa'!A4:J4"
#     worksheet['A4'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A4'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A5:F5')
#     worksheet['A5'] = "='Receita x Despesa'!A5:J5"
#     worksheet['A5'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A5'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A6:F6')
#     worksheet['A6'] = "='Receita x Despesa'!A6:J6"
#     worksheet['A6'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A6'].alignment = Alignment(horizontal="left",vertical="center")
    
#     worksheet.merge_cells('A7:F7')
#     worksheet['A7'] = "='Receita x Despesa'!A7:J7"
#     worksheet['A7'].font = Font(name="Arial", size=12, color="000000")
#     worksheet['A7'].alignment = Alignment(horizontal="left",vertical="center")

#     #colunas azul cabecario
#     row_style_diaria = NamedStyle(name='row_style_diaria')
#     row_style_diaria.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_diaria.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_diaria.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_diaria.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin") )
#     row_style_diaria.height = 20
#     linha_number = 9
#     for row in worksheet.iter_rows(min_row=linha_number, max_row=linha_number, min_col=1, max_col=10):
#         for cell in row:
#             cell.style = row_style_diaria
#             if cell.column == 10:
#                 cell.border = Border(top=Side(border_style="medium")  ,bottom=Side(border_style="thin"), right=Side(border_style="medium") )

#     valores = ["ITEM","NOME","CNPJ/CPF",'ESPECIFICAÇÃO DA DESPESA','DESCRIÇÃO',"Nº DO RECIBO OU EQUIVALENTE","DATA DE EMISSÃO",'CHEQUE / ORDEM BANCÁRIA','DATA DE PGTO','Valor']
#     col = 1
#     for a,b in enumerate(valores):
#         worksheet.cell(row=linha_number, column=col, value=b)
#         col = col + 1


#     #Aumentar  a altura das celulas 
#     for row in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#         worksheet.row_dimensions[row[0].row].height = 35

#     custom_number_format_diaria = []
#     # MASCARA R$
#     if custom_number_format_diaria!= False: 
#         custom_number_format_diaria = NamedStyle(name='custom_number_format_diaria')
#         custom_number_format_diaria.number_format = 'R$ #,##0.00'
#         custom_number_format_diaria.font = Font(name="Arial", size=12, color="000000")
#         custom_number_format_diaria.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
#     #estilocinzasimcinzanao
#     value_to_stop = size  
#     start_row = 10
# #
#     for row in range(start_row,size+1):
#         cell = worksheet[f'J{row}']
#         cell.style = custom_number_format_diaria
        
#     for rows in worksheet.iter_rows(min_row=10, max_row=size, min_col=1, max_col=10):
#             for cell in rows:
#                 if cell.row % 2:
#                     cell.fill = PatternFill(start_color=cinza, end_color=cinza,
#                                             fill_type = "solid")
#                 if cell.column == 10:
#                     cell.font = Font(name="Arial", size=12, color="000000")
#                     cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                     cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="medium") ,bottom=Side(border_style="hair") )
#                 else:
#                     cell.font = Font(name="Arial", size=12, color="000000")
#                     cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#                     cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="hair") ,right =Side(border_style="hair") ,bottom=Side(border_style="hair") )
                
                
#     #subtotal
#     stringAfinarCelula =size+2
#     worksheet.row_dimensions[size+2].height = 6
#     celulas_mergidas_subtotal = f"A{size+2}:I{size+2}"
#     worksheet.merge_cells(celulas_mergidas_subtotal)
#     left_celula_cell = f"A{size+2}"
#     top_left_cell = worksheet[left_celula_cell]
#     top_left_cell.value = "Sub Total1"
#     top_left_cell.fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     top_left_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="medium") ,right =Side(border_style="thin") ,bottom=Side(border_style="medium") )

#     worksheet.row_dimensions[size+2].height = 56.25

#      # FORMULATOTAL
#     formula = f"=SUM(J10:J{size})"
#     celula = f'J{size+2}'
#     worksheet[celula] = formula
#     worksheet[celula].fill = PatternFill(start_color=cinza, end_color=cinza,fill_type = "solid")
#     worksheet[celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
#     worksheet[celula].number_format = 'R$ #,##0.00'
#     #restituições creditadas
#     restituicoes = size + 3
#     celula_restituicoes=f'A{restituicoes}'
#     worksheet[celula_restituicoes].value = "RESTITUIÇÕES CREDITADAS"
#     worksheet[celula_restituicoes].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet.row_dimensions[restituicoes].height = 30

#     #estilo colunas restitucoes creditadas
#     row_style_diaria_append = NamedStyle(name='row_style_diaria_append')
#     row_style_diaria_append.font = Font(name="Arial", size=12, color="FFFFFF",bold=True)
#     row_style_diaria_append.fill = openpyxl.styles.PatternFill(start_color=azul_claro, end_color=azul_claro, fill_type='solid')
#     row_style_diaria_append.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
#     row_style_diaria_append.height = 30
#     row_style_diaria_append.border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium") )


#     row_number = size + 4
   
#     for column in range(1, 11):  
#         cell = worksheet.cell(row=row_number, column=column)
#         cell.style = row_style_diaria_append
#         if cell.column == 10:
#             cell.border = Border(top=Side(border_style="medium") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )



#     values = ["Item","Restituidor","CNPJ/CPF",'Descrição',"Cheque equivalente","Data do Cheque",'Nº do Depósito','Data da Devolução','Valor']
#     coluna = 1
#     for a,b in enumerate(values):
#         worksheet.cell(row=row_number, column=coluna, value=b)
#         if coluna == 4:
#             coluna = coluna + 1
#         coluna = coluna + 1
        

#     merge_formula = f'D{row_number}:E{row_number}'
#     worksheet.merge_cells(merge_formula)

    
#     #subtotal2
#     sub_total2_row = size + 5
#     subtotal_merge_cells= f'A{sub_total2_row}:I{sub_total2_row}'
#     worksheet.merge_cells(subtotal_merge_cells)
#     top_left_subtotal2_cell_formula = f'A{sub_total2_row}'
#     top_left_subtotal2_cell = worksheet[top_left_subtotal2_cell_formula]
#     top_left_subtotal2_cell.value = "Sub Total 2"
#     top_left_subtotal2_cell.fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
#     top_left_subtotal2_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_subtotal2_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_subtotal2_cell.border = Border(top=Side(border_style="hair") ,left = Side(border_style="medium") ,right =Side(border_style="hair") ,bottom=Side(border_style="medium") )

#     sub_formula_row_celula = f'J{sub_total2_row}'
#     worksheet[sub_formula_row_celula].fill = PatternFill(start_color=cinza_escuro, end_color=cinza_escuro,fill_type = "solid")
#     worksheet[sub_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[sub_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet[sub_formula_row_celula].border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )

#       #total1-2
#     total12_row = size + 6
#     total12_merge_cells = f'A{total12_row}:I{total12_row}'
#     worksheet.merge_cells(total12_merge_cells)
#     top_left_total12_cell_formula = f'A{total12_row}'
#     top_left_total12_cell = worksheet[top_left_total12_cell_formula]
#     top_left_total12_cell.value = "Total(1-2)"
#     top_left_total12_cell.fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     top_left_total12_cell.font = Font(name="Arial", size=12, color="000000",bold=True)
#     top_left_total12_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_total12_cell.border = Border(top=Side(border_style="medium") ,left = Side(border_style="medium") ,bottom=Side(border_style="medium") )


#     #total_formula
#     total_formula_row = size + 6
#     total_formulaa = f'=J{size}'
#     total_formula_row_celula = f'J{total_formula_row}'
#     worksheet[total_formula_row_celula].fill = PatternFill(start_color=azul_claro, end_color=azul_claro,fill_type = "solid")
#     worksheet[total_formula_row_celula].font = Font(name="Arial", size=12, color="000000",bold=True)
#     worksheet[total_formula_row_celula].number_format = 'R$ #,##0.00'
#     worksheet[total_formula_row_celula].border = Border(top=Side(border_style="medium") ,bottom=Side(border_style="medium"),right=Side(border_style="medium") )

#     worksheet.row_dimensions[total_formula_row].height = 30
#     worksheet[total_formula_row_celula] = total_formulaa


#     #brasilia
#     brasilia_row = size + 7
#     brasilia_formula = f"='Receita x Despesa'!A42:J42"
#     brasilia_merge_cells = f'A{brasilia_row}:I{brasilia_row}'
#     worksheet.merge_cells(brasilia_merge_cells)
#     top_left_brasilia_cell_formula = f'A{brasilia_row}'
#     top_left_brasilia_cell = worksheet[top_left_brasilia_cell_formula]
#     top_left_brasilia_cell.value = brasilia_formula
#     top_left_brasilia_cell.alignment = Alignment(horizontal="center",vertical="center")

#     #DiretorFinanceiro
#     diretor_row = size + 8
#     diretor_cargo_row = size + 9
#     diretor_cpf_row = size + 10
#     diretor_nome_formula = f"='Receita x Despesa'!A45"
#     diretor_cargo_formula = f"='Receita x Despesa'!A46"
#     diretor_cpf_formula = f"='Receita x Despesa'!A47"
#     diretor_merge_cells = f'A{diretor_row}:D{diretor_row}'
#     diretor_cargo_merge_cells = f'A{diretor_cargo_row}:D{diretor_cargo_row}'
#     diretor_cpf_merge_cells = f'A{diretor_cpf_row}:D{diretor_cpf_row}'
#     worksheet.merge_cells(diretor_merge_cells)
#     worksheet.merge_cells(diretor_cargo_merge_cells)
#     worksheet.merge_cells(diretor_cpf_merge_cells)
#     top_left_diretor_cell_formula = f'A{diretor_row}'
#     top_left_diretor_cell_cargo_formula = f'A{diretor_cargo_row}'
#     top_left_diretor_cell_cpf_formula = f'A{diretor_cpf_row}'
#     top_left_diretor_cell = worksheet[top_left_diretor_cell_formula]
#     top_left_diretor_cell_cargo_formula = worksheet[top_left_diretor_cell_cargo_formula]
#     top_left_diretor_cell_cpf_formula = worksheet[top_left_diretor_cell_cpf_formula]
#     top_left_diretor_cell.value = diretor_nome_formula
#     top_left_diretor_cell_cargo_formula.value = diretor_cargo_formula
#     top_left_diretor_cell_cpf_formula.value = diretor_cpf_formula
#     top_left_diretor_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell.font = Font(bold=True)
#     top_left_diretor_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_diretor_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")
#     #Coordenadora
#     coordenadora_row = size + 8
#     coordenadora_cargo_row = size + 9
#     coordenadora_cpf_row = size + 10
#     coordenadora_nome_formula = f"='Receita x Despesa'!H45"
#     coordenadora_cargo_formula = f"='Receita x Despesa'!H46"
#     coordenadora_cpf_formula = f"='Receita x Despesa'!H47"
#     coordenadora_merge_cells = f'F{coordenadora_row}:J{coordenadora_row}'
#     coordenadora_cargo_merge_cells = f'F{coordenadora_cargo_row}:J{coordenadora_cargo_row}'
#     coordenadora_cpf_merge_cells = f'F{coordenadora_cpf_row}:J{coordenadora_cpf_row}'
#     worksheet.merge_cells(coordenadora_merge_cells)
#     worksheet.merge_cells(coordenadora_cargo_merge_cells)
#     worksheet.merge_cells(coordenadora_cpf_merge_cells)
#     top_left_coordenadora_cell_formula = f'F{coordenadora_row}'
#     top_left_coordenadora_cell_cargo_formula = f'F{coordenadora_cargo_row}'
#     top_left_coordenadora_cell_cpf_formula = f'F{coordenadora_cpf_row}'
#     top_left_coordenadora_cell = worksheet[top_left_coordenadora_cell_formula]
#     top_left_coordenadora_cell_cargo_formula = worksheet[top_left_coordenadora_cell_cargo_formula]
#     top_left_coordenadora_cell_cpf_formula = worksheet[top_left_coordenadora_cell_cpf_formula]
#     top_left_coordenadora_cell.value = coordenadora_nome_formula
#     top_left_coordenadora_cell_cargo_formula.value = coordenadora_cargo_formula
#     top_left_coordenadora_cell_cpf_formula.value = coordenadora_cpf_formula
#     top_left_coordenadora_cell.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell.font= Font(bold = True)
#     top_left_coordenadora_cell_cargo_formula.alignment = Alignment(horizontal="center",vertical="center")
#     top_left_coordenadora_cell_cpf_formula.alignment = Alignment(horizontal="center",vertical="center")

    
#     # borda = Border(right=Side(border_style="medium"))
#     # worksheet.sheet_view.showGridLines = False
#     # # 
#     # for row in worksheet.iter_rows(min_row=1, max_row=coordenadora_cpf_row+1,min_col=10,max_col=10):
#     #     for cell in row:
#     #         cell.border = borda
            
    

#     for row in worksheet.iter_rows(min_row=coordenadora_cpf_row+1, max_row=coordenadora_cpf_row+1,min_col=1,max_col=10):
#         for cell in row:
#             if cell.column == 10:
#                 cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="medium") ,bottom=Side(border_style="medium") )
#             else:
#                 cell.border = Border(top=Side(border_style="none") ,left = Side(border_style="none") ,right =Side(border_style="none") ,bottom=Side(border_style="medium") )

#     workbook.save(tabela)
#     workbook.close()


