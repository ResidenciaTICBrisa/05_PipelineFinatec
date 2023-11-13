import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment,NamedStyle,Border, Side
import os
from datetime import datetime,date

def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho
def estilo_fundep(tabela,tamanho):
    
    # caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(tabela)
    worksheet = workbook['Relação de despesas']
    ######periodo de prestação de contas
    

    #corpo
    size = tamanho + 6
    print(size)

    custom_number_format = []
    # MASCARA R$
    if custom_number_format!= False: 
        custom_number_format = NamedStyle(name='custom_number_format')
        custom_number_format.number_format = 'R$ #,##0.00'
        custom_number_format.font = Font(name="Calibri", size=11, color="000000")
        custom_number_format.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
    
    value_to_stop = size  
    start_row = 7
    for row in range(start_row,size+2):
        cell = worksheet[f'J{row}']
        cell.style = custom_number_format

    #Bordas
    for rows in worksheet.iter_rows(min_row=7, max_row=size, min_col=1, max_col=10):
            for cell in rows:
                cell.font = Font(name="Calibri", size=11, color="000000")
                cell.alignment = Alignment(horizontal="general",vertical="bottom",wrap_text=True)
                cell.border = Border(top=Side(border_style="thin") ,left = Side(border_style="thin") ,right =Side(border_style="thin") ,bottom=Side(border_style="thin") )
                

    #total despesas nesta
    total_despesa_string_LOCATION_merge = f'A{size+1}:I{size+1}'
    worksheet.merge_cells(total_despesa_string_LOCATION_merge)
    total_despesa_string_LOCATION = f'A{size+1}'
    worksheet[total_despesa_string_LOCATION] = f'TOTAL DE DESPESAS NESTA'
    worksheet[total_despesa_string_LOCATION].alignment = Alignment(horizontal="right",vertical="center",wrap_text = True)
    worksheet[total_despesa_string_LOCATION].fill= PatternFill(start_color='d9d9d9',end_color='d9d9d9',fill_type='solid')

    #Total Nessa
    formula = f"=SUM(J7:J{size})"
    formulat_total_location = f'J{size+1}'
    worksheet[formulat_total_location]= formula = f"=SUM(J10:J{size-1})"
    worksheet[formulat_total_location].font=Font(bold=True)

    #total de despesas Acumuladas
    total_despesa_string_LOCATION_merge_acumuladas = f'A{size+2}:I{size+2}'
    worksheet.merge_cells(total_despesa_string_LOCATION_merge_acumuladas)
    total_despesa_string_LOCATION_acumuladas = f'A{size+2}'
    worksheet[total_despesa_string_LOCATION_acumuladas] = f'TOTAL DE DESPESAS ACUMULADAS'
    worksheet[total_despesa_string_LOCATION_acumuladas].alignment = Alignment(horizontal="right",vertical="center",wrap_text = True)
    worksheet[total_despesa_string_LOCATION_acumuladas].fill= PatternFill(start_color='d9d9d9',end_color='d9d9d9',fill_type='solid')

        #Total de despesas Acumuladas
    
    total_despesa_string_LOCATION_acumuladas_valor = f'J{size+2}'
    worksheet[total_despesa_string_LOCATION_acumuladas_valor].font=Font(bold=True)

    workbook.save(tabela)
    workbook.close()


