import openpyxl as op
import datetime
import os


# def pegar_caminho(nome_arquivo):

#     # Obter o caminho absoluto do arquivo Python em execução
#     caminho_script = os.path.abspath(__file__)

#     # Obter o diretório da pasta onde o script está localizado
#     pasta_script = os.path.dirname(caminho_script)

#     # Combinar o caminho da pasta com o nome do arquivo Excel
#     caminho = os.path.join(pasta_script, nome_arquivo)

#     return caminho

# def pegar_caminho(nome_arquivo, diretorio=''):
#     # Obter o caminho absoluto do diretório onde este script está localizado
#     pasta_script = os.path.dirname(os.path.abspath(__file__))

#     # Navegar para o diretório do projeto
#     pasta_projeto = os.path.dirname(os.path.dirname(pasta_script))

#     # Combinar o caminho do diretório fornecido com o nome do arquivo Excel
#     caminho = os.path.join(pasta_projeto, diretorio, nome_arquivo)

#     return caminho

# caminho2 = pegar_caminho("planilhas_preenchidas.txt")
# print(caminho2)

import os

def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline

# caminho = pegar_caminho("planilhas")
# print(caminho)


# def preenche_planilha(planilha, dicionario):

#     caminho = pegar_caminho(planilha)
    
#     # carrega a planilha de acordo com o caminho
#     workbook = op.load_workbook(caminho)

#    # planilha_preenchida = pegar_caminho('preenchido-' + planilha)
#     for nomePlanilha, entradaDados in dicionario.items():
#         planilhaAtual = workbook[nomePlanilha]

#         for intervaloCelula, entradaCelula in entradaDados:
#             if ":" in intervaloCelula:  
#                 inicioCelula, fimCelula = intervaloCelula.split(":")
#                 planilhaAtual = workbook[nomePlanilha]
#                 planilhaAtual[inicioCelula] = entradaCelula
#                 # planilhaAtual[inicioCelula].fill = color   <--- teste com cores
#             else:  
#                 planilhaAtual[intervaloCelula] = entradaCelula
#                 # planilhaAtual[intervaloCelula].fill = color  <--- teste com cores

#     caminho_planilha = os.path.join(caminho, planilha)
#     workbook.save(caminho_planilha)

#     print(f"arquivo salvo como  planilhas_preenchidas/{planilha}")
#     return caminho_planilha

def preenche_planilha(planilha, dicionario,codigo,template_id,consultaInicio,consultaFim,stringNomeFinanciador):


    filename = os.path.basename(planilha)
    # carrega a planilha de acordo com o caminho
    workbook = op.load_workbook(planilha)
    #print("Nomes das Planilhas no Excel:", workbook.sheetnames)

   # planilha_preenchida = pegar_caminho('preenchido-' + planilha)
    for nomePlanilha, entradaDados in dicionario.items():
        planilhaAtual = workbook[nomePlanilha]

        for intervaloCelula, entradaCelula in entradaDados:
            if ":" in intervaloCelula:  
                inicioCelula, fimCelula = intervaloCelula.split(":")
                planilhaAtual = workbook[nomePlanilha]
                planilhaAtual[inicioCelula] = entradaCelula
                # planilhaAtual[inicioCelula].fill = color   <--- teste com cores
            else:  
                planilhaAtual[intervaloCelula] = entradaCelula
                # planilhaAtual[intervaloCelula].fill = color  <--- teste com cores


    caminho_pasta_planilhas = "../../planilhas_preenchidas/"

    # Obtém o diretório atual do script
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))

    # Combina o diretório atual com o caminho para a pasta "planilhas_preenchidas" e o nome do arquivo
    #salvar = os.path.join(diretorio_atual, caminho_pasta_planilhas, f"planilhaPreenchida{filename}")
    salvar = os.path.join(diretorio_atual, caminho_pasta_planilhas, f"PC - {stringNomeFinanciador} - {codigo} - {consultaInicio} a {consultaFim}.xlsx")


    workbook.save(salvar)
    workbook.close()
    print(f"arquivo salvo como {salvar}")

def extrair(text_list):
    start_delimiter = "@@"
    end_delimiter = "@@"
    extracted_texts = []

    for text in text_list:
        start = text.find(start_delimiter)
        if start != -1:
            start += len(start_delimiter)
            end = text.find(end_delimiter, start)
            if end != -1:
                extracted_texts.append(text[start:end])
    
    return extracted_texts



planilha_local_dados = {
    "Receita x Despesa": [
        ("A3:J3", "Título do Projeto:String 1 A3:J3"),
        ("A4:J4", "Executora:  String 2 A4:J4"),
        ("A5:J5", "Partícipe: String 3 A5:J5"),
        ("A6:J6", "Período de Execução Físico-Financeiro: String 4 A6:J6"),
        ("A7:J7", "Período que abrange esta prestação:  String 5 A7:J7"),
        ("A16:A25",datetime.datetime.strptime("2014-06-23", "%Y-%m-%d")),
        ("B16:B25","STRINGB16B25"),# intervalo nao interfere
        ("C16:C25","STRINGC16C25"),#intervalo nao interfere
        ("E16:E25",200),
        #("I16:I21",223),
        ("I16", 23),
        ("I17", 213),
        ("I18", 223),
        ("I19", 233),
        ("I20", 243),
        ("I21", 253),
        ("I24", 263),
        ("I26", 23787),
        ("I27", 10),
        ("I32", 100),
        ("I33", 1000),
        ("I38", 10000),
        ("H45", "COORDENADORA_TESTE")
     ],
    "Exec. Receita e Despesa":[
        ("B16",200),
        ("B17",201),
        ("B18",202),
        ("B19",203),
        ("B20",204),
        ("B21",205),
        ("B22",206),
        ("B23",207),
        ("C16",205),
        ("C17",206),
        ("C18",207),
        ("C19",208),
        ("C20",209),
        ("C21",210),
        ("C22",211),
        ("C23",212),
        ("C24",213),
        ("C25",214),
        ("F16",2051),
        ("F17",2061),
        ("F18",2071),
        ("F19",2081),
        ("F20",2091),
        ("F21",2101),
        ("F22",2111),
        ("F23",2121),
        ("G16",20351),
        ("G17",20361),
        ("G18",20371),
        ("G19",20381),
        ("G20",20391),
        ("G21",21301),
        ("G22",21311),
        ("G23",21321),
        ("I26",2011),
        ("I28",2001),
        ("I29",2001),
        ("B26",3011),
        ("B28",3001),
        ("B29",3001),
        ("B31",3011),
        ("C26",4011),
        ("C29",4011),
        ("C31",4011),
        ("F26",5011),
        ("F28",5001),
        ("F29",5001),
        ("F31",5011),
        ("G26",6011),
        ("G28",6001),
        ("G29",6001),
        ("G31",6011)
    ],
    "Pessoa Jurídica":[
        ("B11","TESTE_NOME"),
        ("C11","TESTE_CPF"),
        ("D11","TESTE_ESPECIFICACAO"),
        ("E11","TESTE_DESCRICAO"),
        ("F11","TESTE_DESCRICAO"),
        ("F11","TESTE_RECIBO"),
        ("G11",110101),
        ("H11","TESTE_CHEQUE"),
        ("I11",98765431),
        ("J11",3000)
       
    ],
    "Conciliação Bancária":[
        ("F10",5000),
        ("F11",5000),
        ("A15",120623),
        ("A15",120623),
        ("B15",9777),
        ("C15","DOCUMENTO_TESTE"),
        ("D15","DESCRIÇÃO_TESTE"),
        ("B38",9777),
        ("B39",7878),
        ("C38","TESTESTRALEATORIOC38"),
        ("C39",'TESTESTRALEATORIOC39'),
        ("D38","TESTESTRALEATORIOD38"),
        ("D39",'TESTESTRALEATORIOD39')
       
    ],
    "Rendimento de Aplicação":[
        ("B12",2023),
        ("C12",2024),
        ("D12",2025),
        ("E12",2026),
        ("F12",2027),
        ("G12",2028),
        ("H12",2029)
    ]


}

modelo_fundep = {
	"Relação e despesas" : [
        # campos cadastrais
		("C3", "Instituição Gestora"),
		("F3", "0000*XX"),
		("I3", "n_acordo"),
		("C4", "Titulo_Projeto"),
		("C5", "Nome_Coordenador"),

        # item beneficiarios
		("I5", "Dt_PeriodoPrestacao"),
		("C7", "Nome_Beneficiario"),
		("D7", "CNPJ-CPF-Baneficiario"),
		("E7", "Rubrica-Beneficiario"),
		("F7", "CH-OB-Beneficiario"),
		("G7", "dt-inicial-Beneficiario"),
		("H7", "NF-DOC.FISCAL-Beneficiario"),
		("I7", "dt-final-Beneficiario"),
		("J7", "valor-Beneficiario"),
	]
}

# for nomePlanilha, entradaDados in planilha_local_dados.items():
#     planilhaAtual = workbook[nomePlanilha]

#     for intervaloCelula, entradaCelula in entradaDados:
#         if ":" in intervaloCelula:  
#             inicioCelula, fimCelula = intervaloCelula.split(":")
#             planilhaAtual = workbook[nomePlanilha]
#             planilhaAtual[inicioCelula] = entradaCelula
#         else:  
#             planilhaAtual[intervaloCelula] = entradaCelula


# workbook.save("modified_ModeloFub.xlsx")

#preenche_planilha('planilhas/FUNDEP.xlsx', planilha_local_dados)