import pyodbc
from datetime import datetime,date
import openpyxl
from openpyxl.styles import Font
import os
from collections import defaultdict
from .estiloFap import *
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.engine import URL
import numpy as np  

def convert_datetime_to_string(value):
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return value

def convert_datetime_to_stringdt(dt):
    # Check if the value is a pandas Timestamp
    if isinstance(dt, pd.Timestamp):
        # Convert the Timestamp to a string using strftime
        return dt.strftime('%d/%m/%Y')  # You can customize the format as needed
    else:
        # If it's not a Timestamp, return the original value
        return dt

def formatar_data(row):
    """ Formata a data com o mes abreviado transformando 01 em jan por exemplo
    """
    dia = row.day
    mes = row.month
    ano = row.year

    # Mapear o número do mês para o nome abreviado
    meses = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'}

    # Obter o nome abreviado do mês
    mes_abreviado = meses.get(mes, mes)

    # Criar a string formatada
    data_formatada = f'{dia}-{mes_abreviado}-{ano}'
    
    return data_formatada

def formatarDataSemDia(row):
    """ Formata a data com o mes abreviado transformando 01 em jan por exemplo
    """
    dia = row.day
    mes = row.month
    ano = row.year

    # Mapear o número do mês para o nome abreviado
    meses = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'}

    # Obter o nome abreviado do mês
    mes_abreviado = meses.get(mes, mes)

    # Criar a string formatada
    data_formatada = f'{mes_abreviado}-{ano}'
    
    return data_formatada

def formatar_cpf(cpf):
    cpf_formatado = f'{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}'
    return cpf_formatado

def check_format(time_data, format='%Y-%m-%d'):
    try:
        # Try to parse the time_data using the specified format
        datetime.strptime(time_data, format)
        return True  # The time_data matches the format
    except ValueError:
        return False  # The time_data does not match the format
    
def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline

def pegar_pass(chave):
    arq_atual = os.path.abspath(__file__)
    app = os.path.dirname(arq_atual)
    project = os.path.dirname(app)
    pipeline = os.path.dirname(project)
    desktop = os.path.dirname(pipeline)
    caminho_pipeline = os.path.join(desktop, chave)
    
    return caminho_pipeline


    # return records

#todas as consultas em sql
def consultaCabecarioAnexoDois(IDPROJETO,DATA1,DATA2):
    """ Informa o titulo do projeto
        a instiuição gestora
        a instuição executora
        NTOA/ e processo
    
NomeConvenio	Processo	SubProcesso	ValorAprovado
FAP - FUB/IQ - Edital 05/2016 -Micropoluentes emergentes e o uso da água na Bacia do Paranoá: diagnóstico, identificação de fontes, efeitos tóxicos métodos de remoção e de detecção in situ	

Processo

0193.000714/2016

SubProcesso

Projeto TOA n° 503/2016 - Edital 05/2016

ValorAprovado

699000.0000
    
    """
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO,)]
    queryNomeCabecario = f"SELECT [LisConvenio].NomeConvenio ,[LisConvenio].Processo,SubProcesso,ValorAprovado,NomePessoa  FROM [Conveniar].[dbo].[LisConvenio] WHERE CodConvenio = ? "
    dfCabecarios = pd.read_sql(queryNomeCabecario, engine, params=parametros)

    return dfCabecarios

def consultaID(IDPROJETO):

   #file_path = "/home/ubuntu/Desktop/devfront/devfull/pass.txt"
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    conn = None
    
    conn = pyodbc.connect(conStr)
    cursor = conn.cursor()
    
   
    consulta = {}
   

    # SQL querys
    
    sql = f"SELECT [LisConvenio].* , [LisPessoa].[CPFCNPJ] as 'CPFCoordenador' FROM [Conveniar].[dbo].[LisConvenio] INNER JOIN  [Conveniar].[dbo].[LisUsuario] ON [LisConvenio].[CodUsuarioResponsavel] = [LisUsuario].[CodUsuario] INNER JOIN  [Conveniar].[dbo].[LisPessoa] ON [LisUsuario].[CodPessoa] = [LisPessoa].[CodPessoa] WHERE CodConvenio = ? "

    # Execute the query
    cursor.execute(sql, IDPROJETO)


    records = cursor.fetchall()
    
    collums = cursor.description
    

    for i in range(len(collums)):
        consulta[collums[i][0]] = records[0][i]
  
    cursor.close()
    conn.close()
    
 
    
    # return records
    return consulta

def consultaAnexoUm(IDPROJETO):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO,)]
    queryNomeConvenioProcessoSubProcesso = f"SELECT [LisConvenio].NomeConvenio ,[LisConvenio].Processo,SubProcesso,ValorAprovado,NomePessoaResponsavel FROM [Conveniar].[dbo].[LisConvenio] WHERE CodConvenio = ? "
    dfConvenioProcessoSubProcessos = pd.read_sql(queryNomeConvenioProcessoSubProcesso, engine, params=parametros)

    
    return dfConvenioProcessoSubProcessos

def consultaAnexoDois(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO,DATA1,DATA2)]
    queryAnexoDois = f"""SELECT [NomeRubrica]
		,NumChequeDeposito
		,CONVERT(varchar(10), DataPagamento, 103) AS FormattedDate
		,NumDocPago
		,[NomeFavorecido]
		,HisLancamento
		,ValorPago
        ,ValorPago
        FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     WHERE [LisLancamentoConvenio].CodConvenio = ? AND [LisLancamentoConvenio].CodStatus = 27
     AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ? and [LisLancamentoConvenio].CodRubrica not in (2,3,9,67,88,0) order by DataPagamento"""
    dfConvenioAnexoDois = pd.read_sql(queryAnexoDois, engine, params=parametros)

    
    
    return dfConvenioAnexoDois

def consultaAnexoTres(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passss.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    idprojetoComZero = f"0{IDPROJETO}"
    parametros = [(IDPROJETO,idprojetoComZero, DATA1, DATA2)]
    
    queryConsultaComRubrica = f"""SELECT 
    [Data de Aquisição][dataAqui],
	[Nº Nota][nota],
    [Descrição][descri],
	[Valor de Aquisição][valorAqui],
    [Valor de Aquisição][valorAqui2],
    [Patrimônio][patri],
    [Localização][localiza],
    [Responsável][responsavel]
    FROM [SBO_FINATEC].[dbo].[VW_BENS_ADQUIRIDOS] 
    WHERE ([Cod Projeto] = ? or [Cod Projeto] = ? ) 
    AND [Status] = 'Imobilizado' 
    AND [Data de Aquisição] BETWEEN ? AND ? 
    Order by [Data de Aquisição]"""
    dfConsultaBens = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)

    
    return dfConsultaBens  
    return 0

def consultaRendimentosIRRFConciliacao(IDPROJETO,DATA1,DATA2):
    """consulta fora do comum utilizado para pegar o total de rendimentos para aparecer na pasta conciliação 
    """
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    
    consultaComPeriodo =f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 0 AND DataPagamento BETWEEN  ? AND ?"

    consultaRendimentoAplicacao = f"""
    SELECT 
        NomeTipoLancamento,
        SUM(CASE WHEN NomeTipoLancamento = 'IRRF Pessoa Jurídica' THEN ValorPago ELSE 0 END) AS IRRF,
        SUM(CASE WHEN NomeTipoLancamento = 'Aplicação Financeira' THEN ValorPago ELSE 0 END) AS Aplicação
    FROM 
        [Conveniar].[dbo].[LisLancamentoConvenio] 
    WHERE 
        CodConvenio = ? 
        AND CodStatus = 27 
        AND CodRubrica = 3 
        AND DataPagamento BETWEEN ? AND ?
    GROUP BY 
        NomeTipoLancamento;
    """


    Soma = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
   
    return Soma

def consultaDevolucaoRecursosConciliacao(IDPROJETO,DATA1,DATA2):
    """Consulta pra informar o total de devolução de recurso dos projetos
    """
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    
    consultaComPeriodo = f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 0 AND DataPagamento BETWEEN  ? AND ?"


    Soma = pd.read_sql(consultaComPeriodo, engine, params=parametros)
   
    return Soma

def consultaRendimentosAplicacao(IDPROJETO,DATA1,DATA2):
    """Imprime as colunas de imposto e rendimento bruto.

    Args:
      
      IDPROJETO: numero do projeto

      DATA1: Data inicio prestação de contas

      DATA2: Data fim prestação de contas


    Returns:

      Retorna tres dataframes, um contem o imposto o outro contem o rendimento bruto
    """

    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    consultaRendimentoAplicacao = f"SELECT  DataPagamento,ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and NomeTipoLancamento = 'Aplicação Financeira'  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaRendimentoAplicacao = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
     
    consultaRendimentoEImposto =  f"SELECT DataPagamento,ValorPago,NomeTipoLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and (NomeTipoLancamento = 'Aplicação Financeira' or NomeTipoLancamento = 'IRRF Pessoa Jurídica')  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaRendimentoEImposto = pd.read_sql(consultaRendimentoEImposto, engine, params=parametros)
    
    consultaImposto =  f"SELECT  DataPagamento,ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and NomeTipoLancamento = 'IRRF Pessoa Jurídica'  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaImposto = pd.read_sql(consultaImposto, engine, params=parametros)
     

    return dfConsultaRendimentoAplicacao,dfConsultaImposto,dfConsultaRendimentoEImposto

def consultaConciliacao(IDPROJETO,DATA1,DATA2):
    """ Informa o nome do convenio,Processo,SubProcesso e Valor Aprovado do projeto
    
NomeConvenio	Processo	SubProcesso	ValorAprovado
FAP - FUB/IQ - Edital 05/2016 -Micropoluentes emergentes e o uso da água na Bacia do Paranoá: diagnóstico, identificação de fontes, efeitos tóxicos métodos de remoção e de detecção in situ	

Processo

0193.000714/2016

SubProcesso

Projeto TOA n° 503/2016 - Edital 05/2016

ValorAprovado

699000.0000
    
    """
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO,)]
    queryNomeConvenioProcessoSubProcesso = f"""
    SELECT [LisConvenio].NomeConvenio ,
    [LisConvenio].Processo,
    SubProcesso,
    ValorAprovado,NomePessoaResponsavel 
    FROM [Conveniar].[dbo].[LisConvenio] WHERE CodConvenio = ? """
    dfConvenioProcessoSubProcessos = pd.read_sql(queryNomeConvenioProcessoSubProcesso, engine, params=parametros)

    return dfConvenioProcessoSubProcessos


#preencher 

def anexoUm(tabela,codigo,data1,data2):
    dfAnexoUm = consultaAnexoUm(codigo)
    dfAnexoRendimento = consultaRendimentosIRRFConciliacao(codigo,data1,data2)
    pd.set_option('display.max_colwidth', None)   
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['ANEXO I']
    # Soma = dfSoma["Aplicação"] + dfSoma["IRRF"]
    stringNome = dfAnexoUm['NomeConvenio']
    stringValorAprovado = dfAnexoUm['ValorAprovado']
    stringProcesso = dfAnexoUm['Processo']
    stringSubProcesso = dfAnexoUm['SubProcesso']
    stringCoordenador = dfAnexoUm['NomePessoaResponsavel']
    stringRendApli= dfAnexoRendimento.loc[dfAnexoRendimento['NomeTipoLancamento'] == 'Aplicação Financeira', 'Aplicação'].values[0]
    sheet['A20'] = stringNome.to_string(index=False)
    sheet['A20'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A20'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    sheet['A22'] = f'Valor Global R$:{stringValorAprovado.to_string(index=False)}'
    sheet['A22'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['A22'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    sheet['E7'] = f'{stringSubProcesso.to_string(index=False)}     /     {stringProcesso.to_string(index=False)}'
    sheet['E7'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['E7'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    sheet['C29'] = stringRendApli
    sheet['C29'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['C29'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
    sheet['D46'] = stringCoordenador.to_string(index=False)
    sheet['D46'].font = Font(name="Arial", size=12, color="000000")
    sheet['D46'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)



    
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None


   #periodo prestação de contas
    stringPeriodoAbrangido = f'Período da Prestação de Contas: {output_date_str} a {output_date_str2}'
    sheet['C22'] = stringPeriodoAbrangido
    sheet['C22'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['C22'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    

    meses_dict = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}   

    stringTamanhoBrasilia = f'A{40}' # retorna lugar de brasilia
    hoje = date.today()
    data_formatada = f"{hoje.day} de {meses_dict[hoje.month]} de {hoje.year}"
    sheet[stringTamanhoBrasilia] = f'Brasilia, {data_formatada}'
    sheet[stringTamanhoBrasilia].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet[stringTamanhoBrasilia].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)







    workbook.save(tabela)
    workbook.close()

def anexoDois(tabela,codigo,data1,data2):
    consultaAnexoDois(codigo,data1,data2)
    df= consultaCabecarioAnexoDois(codigo,data1,data2)
    
    tabela = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="ANEXO II")
    workbook.save(tabela)
    workbook.close()


    #carregar a tabela
    
    tamanho = consultaAnexoDois(codigo,data1,data2)
    tamanho = len(tamanho)
    
    brasiliaRow = estiloAnexoDois(tabela,tamanho)
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['ANEXO II']
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None


   
    
    consulta_coordenador = consultaID(codigo)
    stringCoordenador= f'E{brasiliaRow+2}' # retorna lugar do coordanor
    stringCoordanadorCargo = f'E{brasiliaRow+3}'
    sheet[stringCoordanadorCargo] = f"Coordenador(a)"
    stringTamanhoCPF = f'E{brasiliaRow+4}' # retorna lugar do coordanor
    sheet[stringCoordenador] = consulta_coordenador['NomePessoaResponsavel']
    sheet[stringTamanhoCPF] = formatar_cpf(consulta_coordenador['CPFCoordenador'])
    string_titulo = f"Título do Projeto: {consulta_coordenador['NomeConvenio']}"
    string_executora = f"Instituição Executora: {consulta_coordenador['NomePessoa']}"

    #string toa
    stringTOA = f"{consulta_coordenador['SubProcesso']}  /  {consulta_coordenador['Processo']}"
   
   # Convert 'DataAssinatura' to "dd/mm/YYYY" format
    datetime_obj1 = consulta_coordenador['DataAssinatura']

    if datetime_obj1 is not None:
        formatted_date1 = datetime_obj1.strftime("%d/%m/%Y")
    
    # Convert 'DataVigencia' to "dd/mm/YYYY" format
    datetime_obj2 = consulta_coordenador['DataVigencia']
    
    if datetime_obj2 is not None:
         formatted_date2 = datetime_obj2.strftime("%d/%m/%Y")
#    # Convert 'DataAssinatura' to "dd/mm/YYYY" format
#     datetime_obj1 = consulta_coordenador['DataAssinatura']

#     if datetime_obj1 is not None:
#         formatted_date1 = datetime_obj1.strftime("%d/%m/%Y")
    
#     # Convert 'DataVigencia' to "dd/mm/YYYY" format
#     datetime_obj2 = consulta_coordenador['DataVigencia']
    
#     if datetime_obj2 is not None:
#          formatted_date2 = datetime_obj2.strftime("%d/%m/%Y")

# Create the string representing the period of execution
    
    sheet['A6'] = string_titulo
    sheet['A8'] = string_executora
    sheet['I3'] = stringTOA
 
    #dadosquefaltam = getAnalistaDoProjetoECpfCoordenador(codigo)
    #sheet['H47'] = formatar_cpf(dadosquefaltam["CPF_COORDENADOR"])
    meses_dict = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}   

    stringTamanhoBrasilia = f'A{brasiliaRow}' # retorna lugar de brasilia
    hoje = date.today()
    data_formatada = f"{hoje.day} de {meses_dict[hoje.month]} de {hoje.year}"
    sheet[stringTamanhoBrasilia] = f'Brasilia, {data_formatada}'

    dfAnexoDois = consultaAnexoDois(codigo,data1,data2)

    dfAnexoDois.insert(6, "col1", 1)
    
   

   
    for row_num, row_data in enumerate(dfAnexoDois.itertuples(), start=11): #inicio linha
        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                    value = convert_datetime_to_stringdt(value)
                    sheet.cell(row=row_num, column=col_num, value=value)  





    workbook.save(tabela)
    workbook.close()


    return brasiliaRow 

def anexoTres(tabela,codigo,data1,data2,rowBrasilia):
    tabela = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="ANEXO III")
    workbook.save(tabela)
    workbook.close()

    dfAnexoTres = consultaAnexoTres(codigo,data1,data2)
    tamanho = len(dfAnexoTres)
    estiloAnexoTres(tabela,tamanho,rowBrasilia)
    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['ANEXO III']

   


    dfAnexoTres.insert(0, "col1",None)
    dfAnexoTres.insert(3, "col2",None)
    dfAnexoTres.insert(5, "col3",1)
    
    
   

   
    for row_num, row_data in enumerate(dfAnexoTres.itertuples(), start=11): #inicio linha
        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                    value = convert_datetime_to_stringdt(value)
                    sheet.cell(row=row_num, column=col_num, value=value)  



    workbook.save(tabela)
    workbook.close()

def anexoQuatro(tabela,codigo,data1,data2,rowBrasilia):

    tabela = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="ANEXO IV")
    workbook.save(tabela)
    workbook.close()
   
    dfConsultaRendimentoAplicacao,dfConsultaImposto,dfImpostoERendimento = consultaRendimentosAplicacao(codigo,data1,data2)
    merged_df = pd.merge(dfConsultaRendimentoAplicacao, dfConsultaImposto, on='DataPagamento') 
    tamanho = 0
    # tamanhoMaior = lambda tamanho: len(dfConsultaRendimentoAplicacao) if len(dfConsultaRendimentoAplicacao) > len(dfConsultaImposto) else len(dfConsultaImposto)
    # tamanhoMaiorParaOEstilo = tamanhoMaior(None)

    tamanhoMaiorParaOEstilo = len(merged_df)
     # tem q somra oito algum erro do estilo
    tamanhoMaiorParaOEstilo = tamanhoMaiorParaOEstilo + 8
    estiloAnexoQuatro(tabela,tamanhoMaiorParaOEstilo,rowBrasilia)

    workbook = openpyxl.load_workbook(tabela)
    sheet = workbook['ANEXO IV']

    merged_df['data_formatada'] = merged_df['DataPagamento'].apply(formatarDataSemDia)
    merged_df['DataPagamento'] = merged_df['data_formatada']
    merged_df = merged_df.drop('data_formatada', axis=1)
    
   
    for row_num, row_data in enumerate(merged_df.itertuples(index=False), start=18):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna 
            
            if col_num == 2:
                col_num = 5
            if col_num == 3:
                col_num = 6
            sheet.cell(row=row_num, column=col_num, value=value)
        

    #datas

    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None


   
    stringPeriodoAbrangido = f'{output_date_str} a {output_date_str2}'
    
    sheet['F12'] = stringPeriodoAbrangido
    sheet['F12'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['F12'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    
    sheet.merge_cells('F12:G12')






   
    workbook.save(tabela)
    workbook.close()
     
def Conciliacao(tabela,codigo,data1,data2):

    caminho = pegar_caminho(tabela)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Conciliação ']
    dfRendimentoConciliacao = consultaRendimentosIRRFConciliacao(codigo,data1,data2)
    dfConciliacao = consultaConciliacao(codigo,data1,data2)
    dfDevolucao = consultaDevolucaoRecursosConciliacao(codigo,data1,data2)
    stringProcesso = dfConciliacao['Processo']
    stringValorAprovado = dfConciliacao['ValorAprovado']
    stringSubProcesso = dfConciliacao['SubProcesso']
    stringCoordenador = dfConciliacao['NomePessoaResponsavel']
    stringDevolucaoTotal = dfDevolucao['TotalPago']
    stringRendApli= dfRendimentoConciliacao.loc[dfRendimentoConciliacao['NomeTipoLancamento'] == 'Aplicação Financeira', 'Aplicação'].values[0]
    
    
    cell_is_null = dfDevolucao.isnull().iloc[0, 0]
    
    
   
    
    #toa
    sheet['D5'] = f'{stringSubProcesso.to_string(index=False)}     /     {stringProcesso.to_string(index=False)}'
    sheet['D5'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['D5'].alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
   
   
    #valorparovado
    if  not stringValorAprovado.empty:
        sheet['B15'] = int(stringValorAprovado.iloc[0])
      
    else:
        sheet['B15'] = 0
        
    sheet['B15'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['B15'].alignment = Alignment(horizontal="right",vertical="center",wrap_text=True)
   #rendimento
    sheet['B21'] = stringRendApli
    sheet['B21'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['B21'].alignment = Alignment(horizontal="right",vertical="center",wrap_text=True)
    sheet['B21'].number_format = 'R$ #,##0.00'
    #devolução
    if  not cell_is_null:
        sheet['D19'] = int(stringDevolucaoTotal.iloc[0])
       
    else:
        sheet['D19'] = 0

    sheet['D19'].font = Font(name="Arial", size=12, color="000000",bold=True)
    sheet['D19'].alignment = Alignment(horizontal="right",vertical="center",wrap_text=True)
    sheet['D19'].number_format = 'R$ #,##0.00'
    #COORDENADOR
    sheet['C48'] = stringCoordenador.to_string(index=False)
    sheet['C48'].font = Font(name="Arial", size=12, color="000000")
    sheet['C48'].alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
    sheet['C48'].number_format = 'R$ #,##0.00'


    for row in sheet.iter_rows(min_row=12, max_row=24, min_col=1, max_col=4):
         for cell in row:
              cell.number_format = 'R$ #,##0.00'

    workbook.save(tabela)
    workbook.close() 

def preencheFap(codigo,data1,data2,tabela):
    '''Preenche a planilha fap

        Argumentos:
            codigo = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            tabela = tabela a ser preenchida  extensão xlsx


   '''

    rowBrasilia = anexoDois(tabela,codigo,data1,data2)
    anexoUm(tabela,codigo,data1,data2)
    anexoTres(tabela,codigo,data1,data2,rowBrasilia)
    anexoQuatro(tabela,codigo,data1,data2,rowBrasilia)
    Conciliacao(tabela,codigo,data1,data2)




