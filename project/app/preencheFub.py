import pyodbc
from datetime import datetime,date
import openpyxl
from openpyxl.styles import Font
import os
from .estilo_fub import *
from collections import defaultdict
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.engine import URL
import numpy as np  

def convert_datetime_to_string(value):
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return value

def convert_datetime_to_stringdt(dt):
    # Check if the value is a pandas Timestamp
    if isinstance(dt, pd.Timestamp):
        # Convert the Timestamp to a string using strftime
        return dt.strftime('%d/%m/%Y')  # You can customize the format as needed
    else:
        # If it's not a Timestamp, return the original value
        return dt

def formatar_data(row):
    """ Formata a data com o mes abreviado transformando 01 em jan por exemplo
    """
    dia = row.day
    mes = row.month
    ano = row.year

    # Mapear o número do mês para o nome abreviado
    meses = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'}

    # Obter o nome abreviado do mês
    mes_abreviado = meses.get(mes, mes)

    # Criar a string formatada
    data_formatada = f'{dia}-{mes_abreviado}-{ano}'
    
    return data_formatada

def formatarDataSemDia(row):
    """ Formata a data com o mes abreviado transformando 01 em jan por exemplo
    """
    dia = row.day
    mes = row.month
    ano = row.year

    # Mapear o número do mês para o nome abreviado
    meses = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'}

    # Obter o nome abreviado do mês
    mes_abreviado = meses.get(mes, mes)

    # Criar a string formatada
    data_formatada = f'{mes_abreviado}-{ano}'
    
    return data_formatada

def formatar_cpf(cpf):
    cpf_formatado = f'{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}'
    return cpf_formatado

def check_format(time_data, format='%Y-%m-%d'):
    try:
        # Try to parse the time_data using the specified format
        datetime.strptime(time_data, format)
        return True  # The time_data matches the format
    except ValueError:
        return False  # The time_data does not match the format
    
def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline

def pegar_pass(chave):
    arq_atual = os.path.abspath(__file__)
    app = os.path.dirname(arq_atual)
    project = os.path.dirname(app)
    pipeline = os.path.dirname(project)
    desktop = os.path.dirname(pipeline)
    caminho_pipeline = os.path.join(desktop, chave)
    
    return caminho_pipeline


    # return records

#todas as consultas em sql
def consultaRendimentosTodosAteOPeriodo(IDPROJETO,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA2)]
    
    consultaRendimentoAplicacao = f"""
    SELECT 
        NomeTipoLancamento,
        SUM(CASE WHEN NomeTipoLancamento = 'IRRF Pessoa Jurídica' THEN ValorPago ELSE 0 END) AS IRRF,
        SUM(CASE WHEN NomeTipoLancamento = 'Aplicação Financeira' THEN ValorPago ELSE 0 END) AS Aplicação
    FROM 
        [Conveniar].[dbo].[LisLancamentoConvenio] 
    WHERE 
        CodConvenio = ? 
        AND CodStatus = 27 
        AND CodRubrica = 3 
        AND DataPagamento <= ?
    GROUP BY 
        NomeTipoLancamento;
    """


    Soma = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
    
   
    return Soma

def consultaRendimentosIRRF(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    
    consultaComPeriodo =f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 0 AND DataPagamento BETWEEN  ? AND ?"

    consultaRendimentoAplicacao = f"""
    SELECT 
        NomeTipoLancamento,
        SUM(CASE WHEN NomeTipoLancamento = 'IRRF Pessoa Jurídica' THEN ValorPago ELSE 0 END) AS IRRF,
        SUM(CASE WHEN NomeTipoLancamento = 'Aplicação Financeira' THEN ValorPago ELSE 0 END) AS Aplicação
    FROM 
        [Conveniar].[dbo].[LisLancamentoConvenio] 
    WHERE 
        CodConvenio = ? 
        AND CodStatus = 27 
        AND CodRubrica = 3 
        AND DataPagamento BETWEEN ? AND ?
    GROUP BY 
        NomeTipoLancamento;
    """


    Soma = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
   
    return Soma

def consultaDevolucaoRecursos(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    
    consultaComPeriodo = f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 0 AND DataPagamento BETWEEN  ? AND ?"


    Soma = pd.read_sql(consultaComPeriodo, engine, params=parametros)
   
    return Soma

def consultaRendimentosAplicacao(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    consultaRendimentoAplicacao = f"SELECT  DataPagamento,ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and NomeTipoLancamento = 'Aplicação Financeira'  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaRendimentoAplicacao = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
     
    consultaRendimentoEImposto =  f"SELECT DataPagamento,ValorPago,NomeTipoLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and (NomeTipoLancamento = 'Aplicação Financeira' or NomeTipoLancamento = 'IRRF Pessoa Jurídica')  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaRendimentoEImposto = pd.read_sql(consultaRendimentoEImposto, engine, params=parametros)
    
    consultaImposto =  f"SELECT  DataPagamento,ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and NomeTipoLancamento = 'IRRF Pessoa Jurídica'  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaImposto = pd.read_sql(consultaImposto, engine, params=parametros)
     

    return dfConsultaRendimentoAplicacao,dfConsultaImposto,dfConsultaRendimentoEImposto

def consultaConciliacaoBancaria(IDPROJETO, DATA1, DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    #consultaSemEstorno = f"SELECT DISTINCT DataPagamento,ValorPago,NumChequeDeposito,HisLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) NOT LIKE '%estorno%' order by DataPagamento"
    #consultaComEstorno =  f"SELECT DISTINCT DataPagamento,ValorPago,NumChequeDeposito,HisLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) LIKE '%estorno%' order by DataPagamento"
    
    consultaSemEstorno = f"""
    SELECT [LisLancamentoConvenio].DataPagamento,
    [LisPagamentoDespesaConvenioAdministrativa].Valor ,
    [LisLancamentoConvenio].NumChequeDeposito,[LisLancamentoConvenio].HisLancamento
    FROM [Conveniar].[dbo].[LisLancamentoConvenio]
    INNER JOIN  [Conveniar].[dbo].[LisDocumentoConvenio] ON [LisLancamentoConvenio].[CodDocFinConvenio] = [LisDocumentoConvenio].[CodDocFinConvenio]
    INNER JOIN  [Conveniar].[dbo].[DocFinConvPagDespesa] ON [LisDocumentoConvenio].[CodDocFinConvenio] = [DocFinConvPagDespesa].[CodDocFinConvenio]
    INNER JOIN  [Conveniar].[dbo].[LisPagamentoDespesaConvenio] ON [DocFinConvPagDespesa].[CodPedido] = [LisPagamentoDespesaConvenio].[CodPedido]
    INNER JOIN  [Conveniar].[dbo].[LisPagamentoDespesaConvenioAdministrativa] ON [LisPagamentoDespesaConvenio].CodDespesaConvenio = [LisPagamentoDespesaConvenioAdministrativa].CodDespesaConvenio
    AND [LisPagamentoDespesaConvenio].CodConvenio = [LisPagamentoDespesaConvenioAdministrativa].CodConvenio
    WHERE [LisLancamentoConvenio].CodConvenio = ? AND [LisLancamentoConvenio].CodStatus = 27 AND [LisLancamentoConvenio].CodRubrica = 9 AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ?
    AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' order by [LisLancamentoConvenio].DataPagamento"""

    consultaComEstorno = f"SELECT DISTINCT DataPagamento,ValorPago,NumChequeDeposito,HisLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento)  LIKE '%estorno%' order by DataPagamento"
    
    dfSemEstorno = pd.read_sql(consultaSemEstorno, engine, params=parametros)
    dfComEstorno = pd.read_sql(consultaComEstorno, engine, params=parametros)
   

    return dfSemEstorno,dfComEstorno

def consultaConciliaoBancarioSaldoTotal(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1)]
    PARAMET = [(IDPROJETO, DATA1, DATA2)]
    consultaSumTotal= f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento <= ? AND LOWER(HisLancamento) NOT LIKE '%estorno%'"
    sumTotalEstorno = f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN  ? AND ? AND LOWER(HisLancamento)  LIKE '%estorno%'  "
    sumTotalSemEstorno = f"SELECT SUM(ValorPago) AS TotalPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN  ? AND ? AND LOWER(HisLancamento) NOT LIKE '%estorno%'  "

    
    consultaSaldoTotal= pd.read_sql(consultaSumTotal, engine, params=parametros)
    consultaSaldoTotalEstorno= pd.read_sql(sumTotalEstorno, engine, params=PARAMET)
    consultaSaldoTotalSemEstorno= pd.read_sql(sumTotalSemEstorno, engine, params=PARAMET)
    
   

    return consultaSaldoTotal,consultaSaldoTotalEstorno,consultaSaldoTotalSemEstorno

def consultaID(IDPROJETO):

   #file_path = "/home/ubuntu/Desktop/devfront/devfull/pass.txt"
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    conn = None
    
    conn = pyodbc.connect(conStr)
    cursor = conn.cursor()
    
   
    consulta = {}
   

    # SQL querys
    
    sql = f"SELECT [LisConvenio].* , [LisPessoa].[CPFCNPJ] as 'CPFCoordenador' FROM [Conveniar].[dbo].[LisConvenio] INNER JOIN  [Conveniar].[dbo].[LisUsuario] ON [LisConvenio].[CodUsuarioResponsavel] = [LisUsuario].[CodUsuario] INNER JOIN  [Conveniar].[dbo].[LisPessoa] ON [LisUsuario].[CodPessoa] = [LisPessoa].[CodPessoa] WHERE CodConvenio = ? "

    # Execute the query
    cursor.execute(sql, IDPROJETO)


    records = cursor.fetchall()
    
    collums = cursor.description
    

    for i in range(len(collums)):
        consulta[collums[i][0]] = records[0][i]
  
    cursor.close()
    conn.close()
    
 
    
    # return records
    return consulta


    values = [d.get(key) for d in list_of_dicts for key in keys]
    
   
    return values

def consultaNomeRubricaCodRubrica(IDPROJETO, DATA1, DATA2,):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    queryNomeRubrica = f"SELECT CodRubrica,NomeRubrica FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? GROUP BY NomeRubrica, CodRubrica"
    dfNomeRubricaCodRubrica = pd.read_sql(queryNomeRubrica, engine, params=parametros)


    return dfNomeRubricaCodRubrica

def consultaProjeto(IDPROJETO, DATA1, DATA2,codigoRubrica):
    ''' Consulta dinamica do SQL relacionado a rubrica correspondente,cada pagina tem sua própria consulta correspondente a rubrica
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            codigoRubrica = código da rubrica 
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2,codigoRubrica)]
    parametrosComRubricaEstorno  = [(IDPROJETO, DATA1, DATA2,codigoRubrica,IDPROJETO,DATA1, DATA2,codigoRubrica,)]
    parametrosPJ=[(IDPROJETO, DATA1, DATA2)]
    parametrosPJestorno=[(IDPROJETO, DATA1, DATA2,IDPROJETO, DATA1, DATA2)]
    queryConsultaComRubrica = f"""SELECT NomeFavorecido,
     CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     HisLancamento,
     NumDocPago,
     DataEmissao,
     NumChequeDeposito,
     DataPagamento,
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
     WHERE [LisLancamentoConvenio].CodConvenio = ? AND [LisLancamentoConvenio].CodStatus = 27
     AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ? AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' and [LisLancamentoConvenio].CodRubrica = ? order by DataPagamento"""
    
    queryConsultaComRubricaEstorno = f"""SELECT NomeFavorecido
    ,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-')
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     HisLancamento,
     NumChequeDeposito,
     DataPagamento, 
     ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado]
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado]
     WHERE 
     [LisLancamentoConvenio].CodConvenio = ? 
     AND CodStatus = 27 
     AND NomeTipoCreditoDebito = 'C' 
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento) LIKE '%estorno%' 
     and [LisLancamentoConvenio].CodRubrica = ? 
     OR CodStatus = 27
     AND [LisLancamentoConvenio].CodConvenio = ?  
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento)  LIKE '%estorno%'
     AND [LisLancamentoConvenio].CodRubrica = ? 
     
     order by DataPagamento """
    
    queryConsultaPJDOA = f"""SELECT NomeFavorecido,
     CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     HisLancamento,
     NumDocPago,
     DataEmissao,
     NumChequeDeposito,
     DataPagamento,
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
     WHERE [LisLancamentoConvenio].CodConvenio = ? 
     AND CodStatus = 27
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento) NOT LIKE '%estorno%' 
     AND [LisLancamentoConvenio].CodRubrica IN (57,75,26) 
     order by DataPagamento """

    queryConsultaPJDOAEstorno = f"""
    SELECT NomeFavorecido,
    CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
    WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
    HisLancamento,
    NumChequeDeposito,
    DataPagamento,
    ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
    
    LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
    LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
    WHERE 
    [LisLancamentoConvenio].CodConvenio = ?
    AND CodStatus = 27 
    AND DataPagamento BETWEEN ? AND ? 
    AND LOWER(HisLancamento) LIKE '%estorno%' 
    AND [LisLancamentoConvenio].CodRubrica IN (57,75,26)
    OR CodStatus = 27  
    AND [LisLancamentoConvenio].CodConvenio = ? 
    AND DataPagamento BETWEEN ? AND ? 
    AND LOWER(HisLancamento)  LIKE '%estorno%' 
    AND [LisLancamentoConvenio].CodRubrica IN (57,75,26)
    AND NomeTipoCreditoDebito = 'C'
    order by DataPagamento 
      """
    dfconsultaDadosPorRubrica = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)
    # queryConsultaComRubricaEstorno = f"SELECT NomeFavorecido,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,HisLancamento,NumChequeDeposito,DataPagamento, ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) LIKE '%estorno%' and CodRubrica = ? order by DataPagamento "
    # queryConsultaPJDOA = f"SELECT NomeFavorecido,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,NomeTipoLancamento,HisLancamento,NumDocPago,DataEmissao,NumChequeDeposito,DataPagamento, ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) NOT LIKE '%estorno%' AND CodRubrica IN (57,75,26) order by DataPagamento "
    # queryConsultaPJDOAEstorno = f"SELECT NomeFavorecido,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,HisLancamento,NumChequeDeposito,DataPagamento, ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) LIKE '%estorno%' AND CodRubrica IN (57,75,26) order by DataPagamento "
    # dfconsultaDadosPorRubrica = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)
    dfconsultaDadosPorRubricaComEstorno = pd.read_sql(queryConsultaComRubricaEstorno,engine, params=parametrosComRubricaEstorno)
    dfPJDOA = pd.read_sql(queryConsultaPJDOA, engine, params=parametrosPJ)
    dfPJDOAESTORNO = pd.read_sql(queryConsultaPJDOAEstorno,engine, params=parametrosPJestorno)
    
    return dfPJDOA,dfPJDOAESTORNO,dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno

def consultaEntradaReceitas(IDPROJETO, DATA1, DATA2):
  
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2,)]
    consultaEntradaReceita = f"SELECT DataPagamento,NumChequeDeposito,NomeFavorecido,ValorPago,CodRubrica FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND  (CodRubrica = 2 OR CodRubrica = 67 OR CodRubrica = 88) AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? ORDER BY  CodDocFinConvenio,CodRubrica"
    consultaDemonstrativoReceita = f"SELECT NomeFavorecido,HisLancamento,NumChequeDeposito,DataPagamento,ValorPago,CodRubrica FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND (CodRubrica = 2 OR CodRubrica = 67 OR CodRubrica = 88) AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? ORDER BY  DataPagamento,NumChequeDeposito"
    dfReceitas = pd.read_sql(consultaEntradaReceita, engine, params=parametros)
    dfDemonstrativoReceitas = pd.read_sql(consultaDemonstrativoReceita, engine, params=parametros)

    
    


    return dfReceitas,dfDemonstrativoReceitas

def consultaReceitaEExecReceita(IDPROJETO, DATA1, DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametros2 = [(IDPROJETO, DATA2)]
    parametros3 = IDPROJETO
    consultaComPeriodo =f"SELECT NomeRubrica, SUM(ValorPago) AS VALOR_TOTAL_PERIODO FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica"
    consultaAteAData =  f"SELECT NomeRubrica, SUM(ValorPago) AS VALOR_TOTAL_DATA FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento <= ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica "
    consultaPrevisto = f"SELECT NomeRubrica, SUM(VALOR*Quantidade) AS VALOR_TOTAL_PREVISTO FROM [Conveniar].[dbo].[LisConvenioItemAprovado] WHERE CodConvenio = ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica"
    
    dfComPeriodo= pd.read_sql(consultaComPeriodo, engine, params=parametros)
    dfAteAData = pd.read_sql(consultaAteAData, engine, params=parametros2)
    dfPrevisto = pd.read_sql(consultaPrevisto, engine, params=(IDPROJETO,))
   

    return dfComPeriodo,dfAteAData,dfPrevisto    

def consultaBens(IDPROJETO,DATA1,DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passss.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    idprojetoComZero = f"0{IDPROJETO}"
    parametros = [(IDPROJETO,idprojetoComZero, DATA1, DATA2)]
    
    queryConsultaComRubrica = f"SELECT [Descrição][descri],[Patrimônio][patri],[Data de Aquisição][dataAqui],[Nº Nota][nota],[Localização][localiza],[telefone],[Valor de Aquisição][valorAqui],[Valor de Aquisição][valorAqui2],[Responsável][responsavel] FROM [SBO_FINATEC].[dbo].[VW_BENS_ADQUIRIDOS] WHERE ([Cod Projeto] = ? or [Cod Projeto] = ? ) AND [Status] = 'Imobilizado' AND [Data de Aquisição] BETWEEN ? AND ? Order by [Data de Aquisição]"
    dfConsultaBens = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)


    return dfConsultaBens   
    
#preenche planilha 

def conciliacaoBancaria(codigo,data1,data2,planilha,stringTamanho):

    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Conciliação Bancária")
    workbook.save(tabela)
    workbook.close()
    tamanho = []
    # categorized_data= separarporrubrica(codigo,data1,data2)

    dataframeSemEstorno,dataframeComEstorno = consultaConciliacaoBancaria(codigo,data1,data2)
   
    #####pergar os dados do db e separar por mes e ano###################3
    
    grupos_por_ano_mes = defaultdict(list)
    if  dataframeSemEstorno.empty and dataframeComEstorno.empty:
                print("Data not available or empty.")
                maior = 5
                maior2= 5
                tabela = pegar_caminho(planilha)
                estilo_conciliacoes_bancaria(tabela,maior,maior2,stringTamanho)
                return None  # or handle the case accordingly
    else:
        
        tamanho = len(dataframeSemEstorno)
        tamanho2 = len(dataframeComEstorno)

        tabela = pegar_caminho(planilha)
        estilo_conciliacoes_bancaria(tabela,tamanho,tamanho2,stringTamanho)
       

        workb = openpyxl.load_workbook(tabela)
        worksheet333 = workb["Conciliação Bancária"]
        i = 16
        j=0
        estorno_valor = 0
        
        dataframeSemEstorno['data_formatada'] = dataframeSemEstorno['DataPagamento'].apply(formatar_data)
        dataframeComEstorno['data_formatada'] = dataframeComEstorno['DataPagamento'].apply(formatar_data)
        dataframeSemEstorno['DataPagamento'] = dataframeSemEstorno['data_formatada']
        dataframeComEstorno['DataPagamento'] = dataframeComEstorno['data_formatada']

        dataframeSemEstorno = dataframeSemEstorno.drop('data_formatada', axis=1)
        dataframeComEstorno = dataframeComEstorno.drop('data_formatada', axis=1)

        #for row in worksheet333.iter_rows(min_row=16, max_row=tamanho, min_col=1, max_col=4):

      
        for row_num, row_data in enumerate(dataframeSemEstorno.itertuples(index=False), start=17):#inicio linha
            for col_num, value in enumerate(row_data, start=1):#inicio coluna
                worksheet333.cell(row=row_num, column=col_num, value=value)
               
       
        linha2 = 17+4+tamanho


        for row_num, row_data in enumerate(dataframeComEstorno.itertuples(index=False), start=linha2):#inicio linha
            for col_num, value in enumerate(row_data, start=1):#inicio coluna
                worksheet333.cell(row=row_num, column=col_num, value=value)
         
       #saldo anterior
                
        consultaSaldoTotal,consultaSaldoTotalEstorno,consultaSaldoTotalSemEstorno = consultaConciliaoBancarioSaldoTotal(codigo,data1,data2)
        
        a = consultaSaldoTotal.iloc[0].item()
        
        worksheet333['B16'] = a
        workb.save(tabela)
        workb.close

def demonstrativo(codigo,data1,data2,planilha,rowBrasilia,dataframeDemonstrativoReceita,dfReceitas):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Demonstrativo de Receita")
    workbook.save(tabela)
    workbook.close()
    tamanho = len(dataframeDemonstrativoReceita)
    estilo_demonstrativoDeReceita(tabela,tamanho,rowBrasilia)

    workbook = openpyxl.load_workbook(tabela)
    worksheet = workbook['Demonstrativo de Receita']
    
    for row_num, row_data in enumerate(dataframeDemonstrativoReceita.itertuples(index = False), start=10):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
            if col_num != 6:
                worksheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)) 
                cod_rubrica_value = row_data[5]
                if cod_rubrica_value == 67 or cod_rubrica_value == 88:
                    worksheet.cell(row=row_num, column=5).font = Font(name='Arial', size=12, bold=True, italic=False, color='FF0000')  
                else:
                    worksheet.cell(row=row_num, column=5).font = Font(name='Arial', size=12, bold=True, italic=False, color='0000FF')  

                     




    
    workbook.save(tabela)
    workbook.close()

def relacaodeBens(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Relação de Bens"
    tituloStyle = "relacaoBEns"
    sheet2 = workbook.create_sheet(title="Relação de Bens")
    workbook.save(tabela)
    workbook.close()
    dfBens = consultaBens(codigo,data1,data2)
    
    tamanho = len(dfBens)
   
    estiloRelacaoBens(tabela,tamanho,tituloStyle,nomeTabela,rowBrasilia)

    workbook = openpyxl.load_workbook(tabela)
    sheet = workbook['Relação de Bens']
    for row_num, row_data in enumerate(dfBens.itertuples(), start=13):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
            value = convert_datetime_to_stringdt(value)
            if col_num == 7:
                  value = 1
            sheet.cell(row=row_num, column=col_num, value=value)
    

    workbook.save(tabela)
    workbook.close()
    
def rendimentoDeAplicacao(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Rendimento de Aplicação")
    workbook.save(tabela)
    workbook.close()
   
    dfConsultaRendimentoAplicacao,dfConsultaImposto,dfImpostoERendimento = consultaRendimentosAplicacao(codigo,data1,data2)
    merged_df = pd.merge(dfConsultaRendimentoAplicacao, dfConsultaImposto, on='DataPagamento') 
    tamanho = 0
    # tamanhoMaior = lambda tamanho: len(dfConsultaRendimentoAplicacao) if len(dfConsultaRendimentoAplicacao) > len(dfConsultaImposto) else len(dfConsultaImposto)
    # tamanhoMaiorParaOEstilo = tamanhoMaior(None)
    tamanhoMaiorParaOEstilo = len(merged_df)
    estilo_rendimento_de_aplicacao(tabela,tamanhoMaiorParaOEstilo,rowBrasilia)

    workbook = openpyxl.load_workbook(tabela)
    sheet = workbook['Rendimento de Aplicação']

    merged_df['data_formatada'] = merged_df['DataPagamento'].apply(formatarDataSemDia)
    merged_df['DataPagamento'] = merged_df['data_formatada']
    merged_df = merged_df.drop('data_formatada', axis=1)
   
    for row_num, row_data in enumerate(merged_df.itertuples(index=False), start=15):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna 
            if col_num == 2:
                col_num = 5
            if col_num == 3:
                col_num = 6
            sheet.cell(row=row_num, column=col_num, value=value)
        
   
    workbook.save(tabela)
    workbook.close()
    return (len(merged_df))+15

def rubricaGeral(codigo,data1,data2,planilha,rowBrasilia):
     #consulta todas rubricas
    tabela = pegar_caminho(planilha)
    dfNomeRubricaCodigoRubrica = consultaNomeRubricaCodRubrica(codigo, data1, data2)
    
    for index, values in dfNomeRubricaCodigoRubrica.iterrows():
       
        dfPJDOA,dfPJDOAESTORNO,dfConsultaProjeto ,dfconsultaDadosPorRubricaComEstorno= consultaProjeto(codigo, data1, data2,values['CodRubrica'])
        
            #remove as rubricas nao desejadas
        # values_to_remove = ["Receitas", "Rendimentos de Aplicações Financeiras", "Despesas Financeiras",'Devolução de Recursos','Outros Serviços de Terceiros - Pessoa Jurídica ','Despesas Operacionais e Administrativas - Finatec']
        # dfConsultaProjeto = dfConsultaProjeto[~dfConsultaProjeto['NomeRubrica'].isin(values_to_remove)]
        # dfconsultaDadosPorRubricaComEstorno = dfconsultaDadosPorRubricaComEstorno[~dfconsultaDadosPorRubricaComEstorno['NomeRubrica'].isin(values_to_remove)]

       
        if values['NomeRubrica'] == "Obrigações Tributárias e contributivas":
            values['NomeRubrica'] = "Obrigações Tributárias"
        if values['NomeRubrica'] == "Serviços de Terceiros Pessoa Física":
            values['NomeRubrica'] = "Serviços de Terceiros PF"
        if values['NomeRubrica'] == f"Obrigações Tributárias e Contributivas - 20% de OST " :
            values['NomeRubrica'] = f"Obrigações Trib. - Encargos 20%"
        if values['NomeRubrica'] == f"Obrigações Tributárias e contributivas " :
            values['NomeRubrica'] = f"Obrigações Tributárias"
        if values['NomeRubrica'] == f"Outros Serviços de Terceiros - Pessoa Física" :
            values['NomeRubrica'] = f"Outros Serviços Terceiros - PF"
        if values['NomeRubrica'] == f"Outros Serviços de Terceiros - Pessoa Jurídica " :
            values['NomeRubrica'] = f"Outros Serviços Terceiros - PJ"
        if values['NomeRubrica'] == f"Passagens e Despesas com Locomoção" :
            values['NomeRubrica'] = f"Passagens e Desp. Locomoção"
        if values['NomeRubrica'] == f"Despesas Operacionais e Administrativas - Finatec" :
            values['NomeRubrica'] = f"Despesas Operacionais"
        if values['NomeRubrica'] == f"Despesas Operacionais e Administrativas - Finatec" :
            values['NomeRubrica'] = f"Despesas Operacionais"
        if values['NomeRubrica'] == f"Auxílio Financeiro a Pesquisador" :
            values['NomeRubrica'] = f"AuxFinanceiro Pesquisador"
        if values['NomeRubrica'] == f"Equipamentos e Material Permanente" :
            values['NomeRubrica'] = f"Equip e Mat Permanente"


    

        if values['NomeRubrica'] == "Outros Serviços Terceiros - PJ" or values['NomeRubrica'] == "Serviços de Terceiros Pessoa Jurídica":
            values['NomeRubrica'] = "Outros Serviços Terceiros -PJ"
            nomeTabela = values['NomeRubrica']
            tituloStyle = values['NomeRubrica']
            workbook = openpyxl.load_workbook(tabela)
            sheet2 = workbook.create_sheet(title=f"{values['NomeRubrica']}")
            workbook.save(tabela)
            workbook.close()

            
            tamanho = len(dfPJDOA)
            tamanhoRetorno = len(dfPJDOAESTORNO)
            
            rownovo = estiloGeral(tabela,tamanho,tituloStyle,nomeTabela,rowBrasilia,tamanhoRetorno)
            workbook = openpyxl.load_workbook(tabela)
            sheet2 = workbook[values['NomeRubrica']]
            dfPJDOA.index = dfPJDOA.index + 1
            for row_num, row_data in enumerate(dfPJDOA.itertuples(), start=10):#inicio linha
                for col_num, value in enumerate(row_data, start=1):#inicio coluna
                    value = convert_datetime_to_stringdt(value)
                    sheet2.cell(row=row_num, column=col_num, value=value)
            dfPJDOAESTORNO.index = dfPJDOAESTORNO.index + 1


            for row_num, row_data in enumerate(dfPJDOAESTORNO.itertuples(), start=rownovo): #inicio linha
                for col_num, value in enumerate(row_data, start=1): #inicio coluna
                    if col_num == 5:
                        continue
                    value = convert_datetime_to_stringdt(value)
                    sheet2.cell(row=row_num, column=col_num, value=value)    
            
            workbook.save(tabela)
            workbook.close()
             
     
        else:

            if values['NomeRubrica'] != "Rendimentos de Aplicações Financeiras" and values['NomeRubrica'] != "Despesas Financeiras" and values['NomeRubrica'] != "Receitas" and values['NomeRubrica'] != "Devolução de Recursos" and values['NomeRubrica'] != "Outros Serviços Terceiros - PJ"and values['NomeRubrica'] != "Despesas Operacionais":
                    nomeTabela = values['NomeRubrica']
                    tituloStyle = values['NomeRubrica']
                    workbook = openpyxl.load_workbook(tabela)
                    sheet2 = workbook.create_sheet(title=f"{values['NomeRubrica']}")
                    workbook.save(tabela)
                    workbook.close()

                    tamanho = len(dfConsultaProjeto)
                    tamanhoRetorno = len(dfconsultaDadosPorRubricaComEstorno)
                    
                  
                    
                    rowEstorno = estiloGeral(tabela,tamanho,tituloStyle,nomeTabela,rowBrasilia,tamanhoRetorno)
                    workbook = openpyxl.load_workbook(tabela)
                    sheet2 = workbook[values['NomeRubrica']]
                    dfConsultaProjeto.index = dfConsultaProjeto.index + 1
                    for row_num, row_data in enumerate(dfConsultaProjeto.itertuples(), start=10):#inicio linha
                        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                            value = convert_datetime_to_stringdt(value)
                            sheet2.cell(row=row_num, column=col_num, value=value)
                    
                    # dfconsultaDadosPorRubricaComEstorno.index = dfconsultaDadosPorRubricaComEstorno.index + 1
                    rowEstorno = rowEstorno + 1
                    #
                    tamanhoDf = len(dfconsultaDadosPorRubricaComEstorno)
                    dfconsultaDadosPorRubricaComEstorno.insert(0, "col1", None)
                    dfconsultaDadosPorRubricaComEstorno.insert(4, 'Col2', None)
                    dfconsultaDadosPorRubricaComEstorno.insert(4, 'Col3', None)
                    dfconsultaDadosPorRubricaComEstorno.insert(5, 'Col4', None)
                    
                    
                    
                    
                    for row_num, row_data in enumerate(dfconsultaDadosPorRubricaComEstorno.itertuples(index=False), start=rowEstorno): #inicio linha
                        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                        
                            if col_num == 5:
                                continue
                            value = convert_datetime_to_stringdt(value)
                            sheet2.cell(row=row_num, column=col_num, value=value)    
                    
                    workbook.save(tabela)
                    workbook.close()

       

     #consulta individual
    return 0

def Receita(planilha,codigo,data1,data2,tamanhoResumo,dataframe):
   
    #dfReceitas,dfDemonstrativoReceitas,dfIss2,dfIss5 = consultaEntradaReceitas(codigo,data1,data2)
    dfReceitas,dfDemonstrativoReceitas = consultaEntradaReceitas(codigo,data1,data2)
    
    if len(dfReceitas) > tamanhoResumo:
        tamanhoResumo = len(dfReceitas)
    elif len(dataframe) > tamanhoResumo:
         tamanhoResumo = len(dataframe)
    
    
    caminho = pegar_caminho(planilha)
    #tamanho equipamento ja e valido pois recebeu o tamanho maior e retornou o tamanho equipamentos da string
    tamanho,tamanhoequipamentos = estiloReceitaXDespesa(caminho,tamanhoResumo)

    
    
    #o tamanho na verdade tem que ser do data frame se ele for maior
   
    #Plan = planilha
    # carrega a planilha de acordo com o caminho
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Receita x Despesa']
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None
    
    string_periodo = f"Período que abrange esta prestação: {output_date_str} a {output_date_str2}"
    sheet['A7'] = string_periodo
    consulta_coordenador = consultaID(codigo)

    stringCoordenador= f'H{tamanho+3}' # retorna lugar do coordanor
    stringTamanhoCPF = f'H{tamanho+5}' # retorna lugar do coordanor
    sheet[stringCoordenador] = consulta_coordenador['NomePessoaResponsavel']
    sheet[stringTamanhoCPF] = formatar_cpf(consulta_coordenador['CPFCoordenador'])
    string_titulo = f"Título do Projeto: {consulta_coordenador['NomeConvenio']}"
    string_executora = f"Executora: {consulta_coordenador['NomePessoaFinanciador']}"
    string_participe = f"Partícipe: Fundação de Empreendimentos Científicos e Tecnológicos - FINATEC"
   # Convert 'DataAssinatura' to "dd/mm/YYYY" format
    datetime_obj1 = consulta_coordenador['DataAssinatura']
    formatted_date1 = datetime_obj1.strftime("%d/%m/%Y")

    # Convert 'DataVigencia' to "dd/mm/YYYY" format
    datetime_obj2 = consulta_coordenador['DataVigencia']
    formatted_date2 = datetime_obj2.strftime("%d/%m/%Y")

# Create the string representing the period of execution
    string_periodo = f"Período de Execução Físico-Financeiro: {formatted_date1} a {formatted_date2}"
    sheet['A3'] = string_titulo
    sheet['A4'] = string_executora
    sheet['A5'] = string_participe
    sheet['A6'] = string_periodo
 
    #dadosquefaltam = getAnalistaDoProjetoECpfCoordenador(codigo)
   
    #sheet['H47'] = formatar_cpf(dadosquefaltam["CPF_COORDENADOR"])
    meses_dict = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
    }   
    strintT = tamanho
    stringTamanho = f'A{tamanho}' # retorna lugar de brasilia
    hoje = date.today()
    data_formatada = f"{hoje.day} de {meses_dict[hoje.month]} de {hoje.year}"
    sheet[stringTamanho] = f'Brasilia, {data_formatada}'
    # tamanhoequipamentos = tamanhoequipamentos + 16
     #Obras e Instalações
    #previsto
    string_exists = dataframe['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
        
        #periodo
        stringObras = f'I{tamanhoequipamentos -1}'
        sheet[stringObras] = dataframe.loc[dataframe['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PERIODO'].values[0]
        

    string_exists = dataframe['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
    #Materiais Equipamentos e Material Permanente   
        
        #periodo
        stringObras = f'I{tamanhoequipamentos}'
        sheet[stringObras] = dataframe.loc[dataframe['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PERIODO'].values[0]

    #Materiais Equipamentos e Material nACIONAL
    string_exists = dataframe['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
      
        #periodo
        stringObras = f'I{tamanhoequipamentos +1}'
        sheet[stringObras] = dataframe.loc[dataframe['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PERIODO'].values[0]
       

    #Materiais Equipamentos e Material iMPORTADO
    string_exists = dataframe['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
      
        #periodo
        stringObras = f'I{tamanhoequipamentos + 2}'
        sheet[stringObras] = dataframe.loc[dataframe['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PERIODO'].values[0]
       

    #remove as rubricas nao desejadas
    values_to_remove = ["Receitas", "Rendimentos de Aplicações Financeiras", "Despesas Financeiras",'Material Permanente e Equipamento Nacional','Material Permanente e Equipamento Importado','Equipamentos e Material Permanente','Devolução de Recursos','Encargos - ISS 5% ']
    dataframe = dataframe[~dataframe['NomeRubrica'].isin(values_to_remove)]
  
    string_exists = dataframe['NomeRubrica'].isin(["Despesas Operacionais e Administrativas - Finatec"]).any()
    if string_exists:
    # Extract the value from "Despesas Operacionais e Administrativas - Finatec"
     value_to_add = dataframe.loc[dataframe['NomeRubrica'] == 'Despesas Operacionais e Administrativas - Finatec', 'VALOR_TOTAL_PERIODO'].values[0]

    string_exists = dataframe['NomeRubrica'].isin(["Outros Serviços de Terceiros - Pessoa Jurídica "]).any()
    string_exists2 = dataframe['NomeRubrica'].isin(["Serviços de Terceiros Pessoa Jurídica"]).any()
    if string_exists or string_exists2:
        if string_exists:
        # Find the index of "Outros Serviços de Terceiros - Pessoa Jurídica"
            index_to_update = dataframe.loc[dataframe['NomeRubrica'] == 'Outros Serviços de Terceiros - Pessoa Jurídica '].index[0]

            # Add the value to "Outros Serviços de Terceiros - Pessoa Jurídica"
            dataframe.at[index_to_update, 'VALOR_TOTAL_PERIODO'] += value_to_add

            # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
            dataframe = dataframe[dataframe['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']
        if string_exists2:
        # Find the index of "Outros Serviços de Terceiros - Pessoa Jurídica"
            index_to_update = dataframe.loc[dataframe['NomeRubrica'] == 'Serviços de Terceiros Pessoa Jurídica'].index[0]

            # Add the value to "Outros Serviços de Terceiros - Pessoa Jurídica"
            dataframe.at[index_to_update, 'VALOR_TOTAL_PERIODO'] += value_to_add

            # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
            dataframe = dataframe[dataframe['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']

    
    
    #dataframe.loc[['Outros Serviços de Terceiros - Pessoa Jurídica ']] += dataframe.loc[['Despesas Operacionais e Administrativas - Finatec']]
   
    for row_num, row_data in enumerate(dataframe.itertuples(index = False), start=16):#inicio linha
        for col_num, value in enumerate(row_data, start=8):#inicio coluna
                sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value))



    #vai preencher o lado esquerdo da tabela a parte de receita
    for row_num, row_data in enumerate(dfReceitas.itertuples(index = False), start=16):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
        
            
            if col_num != 6:
                if col_num == 4:
                    col_num = 5
                sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)) 
                cod_rubrica_value = row_data[4]
                if col_num ==5:
                    kek = row_data[3]
                    if cod_rubrica_value == 88 or cod_rubrica_value == 67:
                         kek = -row_data[3]
                     
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(kek)) 


                if cod_rubrica_value == 88:
                    if col_num == 3:
                        
                         val = f'ISS 2% {value}'
                         sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(val)) 

                  
                    
                    sheet.cell(row=row_num, column=5).font = Font(name='Arial', size=12, bold=True, italic=False, color='FF0000')  

                elif cod_rubrica_value == 67:
                    if col_num == 3:
                         
                         vala = f'ISS 5% {value}'
                         sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(vala)) 
                   
                    sheet.cell(row=row_num, column=5).font = Font(name='Arial', size=12, bold=True, italic=False, color='FF0000')  
                else:
                    sheet.cell(row=row_num, column=5).font = Font(name='Arial', size=12, bold=True, italic=False, color='0000FF')  

    #rendimentosdeapliacação
    dfSoma = consultaRendimentosIRRF(codigo,data1,data2)
    dfcComPeriodo = consultaDevolucaoRecursos(codigo,data1,data2)
    Soma = dfSoma["Aplicação"] + dfSoma["IRRF"]
    
    all_null = Soma.isnull().all()
    if all_null != True :
            if len(Soma) == 1:
                result = Soma.iloc[0]
            else:
                result = Soma.iloc[0] - Soma.iloc[1]
                stringRendimento = f'Rendimento de Aplicação'
                stringRendimentoValor = f'E{tamanhoequipamentos + 6}'
                sheet[stringRendimentoValor] = result
                sheet[f'A{tamanhoequipamentos + 6}'] = stringRendimento

    
    #tarifasbancarias
                
    consultaSaldoTotal,consultaSaldoTotalEstorno,consultaSaldoTotalSemEstorno=consultaConciliaoBancarioSaldoTotal(codigo,data1,data2)

    c = consultaSaldoTotal.iloc[0].item()
    resultEstorno = consultaSaldoTotalSemEstorno.iloc[0].item()
    resultSemEstorno = consultaSaldoTotalEstorno.iloc[0].item()
    

   
    if c == None:
         c = 0

    if resultEstorno == None:
         resultEstorno = 0
    if resultSemEstorno == None:
         resultSemEstorno = 0

    

    stringRendimentoValor = f'I{tamanhoequipamentos + 9}'
    sheet[stringRendimentoValor] = resultEstorno + c
   
   
    
    stringRendimentoValor = f'I{tamanhoequipamentos + 10}'
    sheet[stringRendimentoValor] = resultSemEstorno
    

    #devoluçãoderecursos

    stringRendimentoValor = f'I{tamanhoequipamentos + 7}'
    sheet[stringRendimentoValor] = dfcComPeriodo.iloc[0,0]
    

    workbook.save(planilha)
    workbook.close()
    return strintT,dfReceitas,dfDemonstrativoReceitas

def ExeReceitaDespesa(planilha,codigo,data1,data2,stringTamanho):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Exec. Receita e Despesa")
    workbook.save(tabela)
    workbook.close()
    stringTamanho = 15 #esqueci o pq deve ser  =o tamanho na tabela

    #dataframe com os dados
    dfComPeriodo,dfAteAData,dfPrevisto = consultaReceitaEExecReceita(codigo,data1,data2)
    # Merge with an outer join
    merged_df = pd.merge(dfPrevisto, dfComPeriodo, on='NomeRubrica', how='outer')
    dfMerged = pd.merge(merged_df,dfAteAData, on = 'NomeRubrica', how = 'outer')
    tamanho = len(dfMerged)#tamanho para deixar dinamico para imprimir sa rubricas
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
         tamanho = tamanho - 1
    
    tamanho = tamanho - 3

    stringTamanho = tamanho + 16
    estiloExecReceitaDespesa(tabela,tamanho,stringTamanho)
    #preencher
    workbook = openpyxl.load_workbook(planilha)
    sheet = workbook['Exec. Receita e Despesa']

    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None


    string_periodo = f"Período que abrange esta prestação: {output_date_str} a {output_date_str2}"
    sheet['A7'] = string_periodo
    consulta_coordenador = consultaID(codigo)
    stringCoordenador= f'F{stringTamanho+13}' # retorna lugar do coordanor
    stringCoordanadorCargo = f'F{stringTamanho+14}'
    sheet[stringCoordanadorCargo] = f"Coordenador(a)"
    stringTamanhoCPF = f'F{stringTamanho+15}' # retorna lugar do coordanor
    sheet[stringCoordenador] = consulta_coordenador['NomePessoaResponsavel']
    sheet[stringTamanhoCPF] = formatar_cpf(consulta_coordenador['CPFCoordenador'])
    string_titulo = f"Título do Projeto: {consulta_coordenador['NomeConvenio']}"
    string_executora = f"Executora: {consulta_coordenador['NomePessoaFinanciador']}"
    string_participe = f"Partícipe: Fundação de Empreendimentos Científicos e Tecnológicos - FINATEC"
   # Convert 'DataAssinatura' to "dd/mm/YYYY" format
    datetime_obj1 = consulta_coordenador['DataAssinatura']
    formatted_date1 = datetime_obj1.strftime("%d/%m/%Y")

    # Convert 'DataVigencia' to "dd/mm/YYYY" format
    datetime_obj2 = consulta_coordenador['DataVigencia']
    formatted_date2 = datetime_obj2.strftime("%d/%m/%Y")

# Create the string representing the period of execution
    string_periodo = f"Período de Execução Físico-Financeiro: {formatted_date1} a {formatted_date2}"
    sheet['A3'] = string_titulo
    sheet['A4'] = string_executora
    sheet['A5'] = string_participe
    sheet['A6'] = string_periodo
 
    #dadosquefaltam = getAnalistaDoProjetoECpfCoordenador(codigo)
    #sheet['H47'] = formatar_cpf(dadosquefaltam["CPF_COORDENADOR"])
    meses_dict = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}   

    stringTamanhoBrasilia = f'A{stringTamanho+12}' # retorna lugar de brasilia
    hoje = date.today()
    data_formatada = f"{hoje.day} de {meses_dict[hoje.month]} de {hoje.year}"
    sheet[stringTamanhoBrasilia] = f'Brasilia, {data_formatada}'

    #despesas correntes 
    imprimirResumoLinha = 16

    #Obras e Instalações
    #previsto
    string_exists = dfMerged['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
        
        stringObras = f'B{stringTamanho + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PREVISTO'].values[0]
        stringObras = f'F{stringTamanho + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PREVISTO'].values[0]

        #periodo
        stringObras = f'C{stringTamanho + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'G{stringTamanho + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_DATA'].values[0]

    string_exists = dfMerged['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
    #Materiais Equipamentos e Material Permanente   
        stringObras = f'B{stringTamanho + 3}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PREVISTO'].values[0]
        stringObras = f'F{stringTamanho + 3}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PREVISTO'].values[0]
        #periodo
        stringObras = f'C{stringTamanho + 3}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'G{stringTamanho + 3}'
        #sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_DATA'].values[0]

    #Materiais Equipamentos e Material nACIONAL
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
        stringObras = f'B{stringTamanho + 4}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PREVISTO'].values[0]
        stringObras = f'F{stringTamanho + 4}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PREVISTO'].values[0]
        #periodo
        stringObras = f'C{stringTamanho + 4}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'G{stringTamanho + 4}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_DATA'].values[0]

    #Materiais Equipamentos e Material iMPORTADO
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
        stringObras = f'B{stringTamanho + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PREVISTO'].values[0]
        stringObras = f'F{stringTamanho + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PREVISTO'].values[0]
        #periodo
        stringObras = f'C{stringTamanho + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'G{stringTamanho + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_DATA'].values[0]

    dfSoma = consultaRendimentosIRRF(codigo,data1,data2)
    dfRendimentoAteOPeriodo = consultaRendimentosTodosAteOPeriodo(codigo,data2)
    
    Soma = dfSoma["Aplicação"] + dfSoma["IRRF"]
    all_null = Soma.isnull().all()
    
    if all_null != True :
        
        if len(Soma) == 1:
             result = Soma.iloc[0]
        else:
            result = Soma.iloc[0] - Soma.iloc[1]
            stringObras = f'B{stringTamanho + 9}'
            sheet[stringObras] = result

            SomaRendimentoAteoperido =  dfRendimentoAteOPeriodo["Aplicação"] + dfRendimentoAteOPeriodo["IRRF"]
            resultado = SomaRendimentoAteoperido.iloc[0] - SomaRendimentoAteoperido.iloc[1]
            stringObras = f'F{stringTamanho + 9}'
            sheet[stringObras] = resultado
    

   


    #remover essas linhas da tabela
    values_to_remove = ["Receitas", "Rendimentos de Aplicações Financeiras", "Despesas Financeiras",'Material Permanente e Equipamento Nacional','Material Permanente e Equipamento Importado','Devolução de Recursos','Obras e Instalações','Equipamentos e Material Permanente']

    # Use boolean indexing to drop rows based on the values in the first column
    dfMerged = dfMerged[~dfMerged['NomeRubrica'].isin(values_to_remove)]
   

    string_exists = dfMerged['NomeRubrica'].isin(["Despesas Operacionais e Administrativas - Finatec"]).any()
    if string_exists:
    # Extract the row for "Despesas Operacionais e Administrativas - Finatec"
            row_to_add = dfMerged.loc[dfMerged['NomeRubrica'] == 'Despesas Operacionais e Administrativas - Finatec'].iloc[0]
    else:   
            row_to_add = 0


    string_exists = dfMerged['NomeRubrica'].isin(["Outros Serviços de Terceiros - Pessoa Jurídica "]).any()
    string_exists2 = dfMerged['NomeRubrica'].isin(["Serviços de Terceiros Pessoa Jurídica"]).any()
    if string_exists or string_exists2:
        if string_exists:
        # Find the index of "Outros Serviços de Terceiros - Pessoa Jurídica"
            index_to_update = dfMerged.loc[dfMerged['NomeRubrica'] == 'Outros Serviços de Terceiros - Pessoa Jurídica '].index[0]

            # Update the values in "Outros Serviços de Terceiros - Pessoa Jurídica" row with the values from "Despesas Operacionais e Administrativas - Finatec"
            dfMerged.iloc[index_to_update] += row_to_add

            # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
            dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']
        if string_exists2:
        # Find the index of "Outros Serviços de Terceiros - Pessoa Jurídica"
            index_to_update = dfMerged.loc[dfMerged['NomeRubrica'] == "Serviços de Terceiros Pessoa Jurídica"].index[0]

            # Update the values in "Outros Serviços de Terceiros - Pessoa Jurídica" row with the values from "Despesas Operacionais e Administrativas - Finatec"
            dfMerged.iloc[index_to_update] += row_to_add

            # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
            dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']


    

    for row_num, row_data in enumerate(dfMerged.itertuples(index = False), start=16):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                if col_num == 2:
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value))
                    col_num = 6
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value))
                    col_num=2

                if col_num == 4:
                    col_num = 7
                sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value))
                   
            

            


    workbook.save(planilha)
    workbook.close()
    return tamanho,dfComPeriodo

def preencheFub(codigo,data1,data2,tabela):
    '''Preencher fub legado
        dadoRubrica = transforma em dicionario a consulta feita por sql e separa por rubricas.
        variaveisResumo= preenchimento da tabela Exec, são definidas em tres tipos:
            -variavelResumoComPeriodo = consulta no sql com datas1 e data2
            -variavelResumoComPeriodo = consulta no sql ate o periodo da data2
            -variavelResumoComPeriodo = consulta no sql com valor total previsto
        execReceitaDespesa = preenche a sheet Exec.Receita e Despesa da planilha,
        e retorna 3 argumentos:
            -tamanho = o tamanho necessario para fazer o estilo, ele varia de acordo
            com a quantidade de rubricas
            -dictPraCalcularTamanho = o dicionario utilizado para calcular o tamanho, ele
            contem a rubricas utilizadas no projeto para poder preencher a sheet ReceitaxDespesa
            -merged_dict = dicionario que contem todas as rubricas inclusive equipamento e 
            material permanente, utilizado para colocar esses valores na ReceitaxDespesa
        

        Argumentos:
            codigo = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            KEYS = Lista responsavel por filtrar no dicionario quais dados irão ser preenchidos nas
            planilhas, exemplo cpf,data etc
            tabela = tabela a ser preenchida  extensão xlsx


    '''
    tamanho,dataframe = ExeReceitaDespesa(tabela,codigo,data1,data2,15)
    tamanhoPosicaoBrasilia,dfReceitas,dfDemonstrativoReceitas = Receita(tabela,codigo,data1,data2,tamanho,dataframe)
    # demonstrativo(codigo,data1,data2,tabela,tamanhoPosicaoBrasilia,dfDemonstrativoReceitas,dfReceitas)
    rubricaGeral(codigo,data1,data2,tabela,tamanhoPosicaoBrasilia)
    # conciliacaoBancaria(codigo,data1,data2,tabela,tamanhoPosicaoBrasilia)
    # rowRendimento= rendimentoDeAplicacao(codigo,data1,data2,tabela,tamanhoPosicaoBrasilia)
    # relacaodeBens(codigo,data1,data2,tabela,tamanhoPosicaoBrasilia)
    
