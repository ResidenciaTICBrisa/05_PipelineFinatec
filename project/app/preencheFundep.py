import pyodbc
from datetime import datetime,date
import openpyxl
import os
from .estilo_fundep import estilo_fundep
from openpyxl.styles import Font,Alignment
def check_format(time_data, format='%Y-%m-%d'):
    try:
        # Try to parse the time_data using the specified format
        datetime.strptime(time_data, format)
        return True  # The time_data matches the format
    except ValueError:
        return False  # The time_data does not match the format

def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho

def convert_datetime_to_string(value):
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return value
#connection string in the format
#<username>/<password>@<dBhostAddress>:<dbPort>/<dbServiceNam
# def getCollumNames(IDPROJETO):
def pegar_pass(chave):
    arq_atual = os.path.abspath(__file__)
    app = os.path.dirname(arq_atual)
    project = os.path.dirname(app)
    pipeline = os.path.dirname(project)
    desktop = os.path.dirname(pipeline)
    caminho_pipeline = os.path.join(desktop, chave)
    
    return caminho_pipeline


def consultaID(IDPROJETO):

   #file_path = "/home/ubuntu/Desktop/devfront/devfull/pass.txt"
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    conn = None
    
    conn = pyodbc.connect(conStr)
    cursor = conn.cursor()
    
   
    consulta = {}
   

    # SQL querys
    
    sql = f"SELECT [LisConvenio].* , [LisPessoa].[CPFCNPJ] as 'CPFCoordenador' FROM [Conveniar].[dbo].[LisConvenio] INNER JOIN  [Conveniar].[dbo].[LisUsuario] ON [LisConvenio].[CodUsuarioResponsavel] = [LisUsuario].[CodUsuario] INNER JOIN  [Conveniar].[dbo].[LisPessoa] ON [LisUsuario].[CodPessoa] = [LisPessoa].[CodPessoa] WHERE CodConvenio = ? "

    # Execute the query
    cursor.execute(sql, IDPROJETO)


    records = cursor.fetchall()
    
    collums = cursor.description


    for i in range(len(collums)):
        consulta[collums[i][0]] = records[0][i]

           
    cursor.close()
    conn.close()
    print("The connection is closed")
    
    # return records
    return consulta

def getCollumNames(IDPROJETO, DATA1, DATA2):
    
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    conn = None
    
    conn = pyodbc.connect(conStr)
    cursor = conn.cursor()
    

  
    sql = f"SELECT * FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? AND CodRubrica NOT IN(2,3)  ORDER BY DataPagamento"


    cursor.execute(sql, IDPROJETO, DATA1, DATA2)

    return cursor
#retorna todos os valores dos dicionarios
def get_values_from_dict(codigo,data1,data2):
  
    gete = getCollumNames(codigo,data1,data2)

    collums = []
    for i in gete.description:
        collums.append(i[0])
    
    #print(collums)

    value = []
    for i in gete:
        val = tuple(convert_datetime_to_string(item) for item in i)
        value.append(val)
    #print(value)
    list_of_dicts = [dict(zip(collums, values)) for values in value]

    #print(list_of_dicts)
    return list_of_dicts
#retorna os valores dado uma chave, por exmeplo se for VALOR_PAGO = 4,50
def retornavalores(list_of_dicts,keys):
    values = [d.get(key) for d in list_of_dicts for key in keys]
    
    #print(values)
    return values
#separa  os dics por rubrica, por exemplo caso queira acessar a da rubrica 87 a= separarporrubrica() - > a[87]


def preenche_fundep(codigo,data1,data2,keys,planilha):
    dados_db = get_values_from_dict(codigo,data1,data2)
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None
    
    string_periodo = f"{output_date_str} a {output_date_str2}"
    
    

    #dados_gerais = retornavalores(dados_db,keys)
    tamanho = []
    for j in keys:
            lj = [j]
            valores_dboracle = retornavalores(dados_db,lj)
            size = len(valores_dboracle)
            tamanho.append(size)
    maior = max(tamanho)
    # print(tamanho)
    tabela = pegar_caminho(planilha)
    estilo_fundep(tabela,maior)
    
    workb = openpyxl.load_workbook(tabela)
    worksheet5 = workb['Relação de despesas']
    consultaCabecario = consultaID(codigo)
    
    worksheet5['C3'] = consultaCabecario['NomePessoaFinanciador'] #gestora
    worksheet5['C4'] = consultaCabecario['NomeConvenio'] #titulo do projeto
    worksheet5['C5'] = consultaCabecario['NomePessoaResponsavel'] #coordenador
    worksheet5['I3'] = consultaCabecario['SubProcesso']
    worksheet5['I5'] = string_periodo
    worksheet5['I5'].font= Font(name="Calibri", size=10, color="000000")
    worksheet5['I5'].alignment = Alignment(horizontal="left",vertical="bottom",wrap_text=True)

    for i in range(1,maior+1):
        valor_coluna = 6 + i
        worksheet5.cell(row=valor_coluna, column=2, value=i)  # column index starts from 1
    
    for i in range(1,maior+1):
        valor_coluna = 6 + i
        worksheet5.cell(row=valor_coluna, column=1, value=1)  # column index starts from 1
    
    coluna = 3
    for i in keys:
        li = [i]
        valores_preenchimento = retornavalores(dados_db,li) 
        for rowkek, cell_data in enumerate(valores_preenchimento, start=7):
            worksheet5.cell(row=rowkek, column=coluna, value=cell_data)
            # print(cell_data)

        coluna = coluna + 1
        # print(coluna)
    


    workb.save(tabela)
    workb.close()
    