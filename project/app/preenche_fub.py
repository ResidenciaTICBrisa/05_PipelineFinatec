import oracledb
from datetime import datetime,date
import openpyxl
import os
#from .estilo_fub import (estilo_conciliacoes_bancaria,estilo_rendimento_de_aplicacao,
#estilo_demonstrativoDeReceita,estiloGeral,estiloRelacaoBens)
from .estilo_fub import *
from collections import defaultdict
from .oracle_cruds import consultaPorID,getAnalistaDoProjetoECpfCoordenador

def formatar_cpf(cpf):
    cpf_formatado = f'{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}'
    return cpf_formatado
def check_format(time_data, format='%Y-%m-%d'):
    try:
        # Try to parse the time_data using the specified format
        datetime.strptime(time_data, format)
        return True  # The time_data matches the format
    except ValueError:
        return False  # The time_data does not match the format

# def pegar_caminho(nome_arquivo):

#     # Obter o caminho absoluto do arquivo Python em execução
#     caminho_script = os.path.abspath(__file__)

#     # Obter o diretório da pasta onde o script está localizado
#     pasta_script = os.path.dirname(caminho_script)

#     # Combinar o caminho da pasta com o nome do arquivo Excel
#     caminho = os.path.join(pasta_script, nome_arquivo)

#     return caminho

def pegar_caminho(subdiretorio):
    # Obtém o caminho do script atual
    arq_atual = os.path.abspath(__file__)
    
    # Obtém o diretório do script
    app = os.path.dirname(arq_atual)
    
    # Obtém o diretório pai do script
    project = os.path.dirname(app)
    
    # Obtém o diretório pai do projeto
    pipeline = os.path.dirname(project)
    
    # Junta o diretório pai do projeto com o subdiretório desejado
    caminho_pipeline = os.path.join(pipeline, subdiretorio)
    
    return caminho_pipeline


def convert_datetime_to_string(value):
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y')
    return value
#connection string in the format
#<username>/<password>@<dBhostAddress>:<dbPort>/<dbServiceNam
# def getCollumNames(IDPROJETO):

def pegar_pass(chave):
    arq_atual = os.path.abspath(__file__)
    app = os.path.dirname(arq_atual)
    project = os.path.dirname(app)
    pipeline = os.path.dirname(project)
    desktop = os.path.dirname(pipeline)
    caminho_pipeline = os.path.join(desktop, chave)
    
    return caminho_pipeline

def getCollumNames(IDPROJETO, DATA1, DATA2):

    #file_path = "/home/ubuntu/Desktop/devfront/devfull/pass.txt"
    file_path = pegar_pass("pass.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    conn = None
    conn = oracledb.connect(conStr)
    cur = conn.cursor()

    # sql = """SELECT DISTINCT * FROM IDEA.FAT_LANCAMENTO_CONVENIAR 
    #          WHERE ID_PROJETO = :IDPROJETO 
    #          AND ID_STATUS = 27  
    #          ORDER BY NUM_DOC_FIN"""
    sql = """SELECT DISTINCT * FROM IDEA.FAT_LANCAMENTO_CONVENIAR 
             WHERE ID_PROJETO = :IDPROJETO 
             AND ID_STATUS = 27 
             AND DATA_PAGAMENTO BETWEEN TO_DATE(:DATA1, 'YYYY-MM-DD') 
             AND TO_DATE(:DATA2, 'YYYY-MM-DD') 
             ORDER BY NUM_DOC_FIN"""

    # cur.execute(sql, {
    #     'IDPROJETO': IDPROJETO
    # })
    cur.execute(sql, {
        'IDPROJETO': IDPROJETO,
        'DATA1': DATA1,
        'DATA2': DATA2
    })
    return cur
#retorna todos os valores dos dicionarios
def get_values_from_dict(codigo,data1,data2):
  
    gete = getCollumNames(codigo,data1,data2)

    collums = []
    for i in gete.description:
        collums.append(i[0])
    
    #print(collums)

    value = []
    for i in gete:
        val = tuple(convert_datetime_to_string(item) for item in i)
        value.append(val)
    #print(value)
    list_of_dicts = [dict(zip(collums, values)) for values in value]

    #print(list_of_dicts)
    return list_of_dicts
#retorna os valores dado uma chave, por exmeplo se for VALOR_PAGO = 4,50
def retornavalores(list_of_dicts,keys):
    values = [d.get(key) for d in list_of_dicts for key in keys]
    
    #print(values)
    return values
#separa  os dics por rubrica, por exemplo caso queira acessar a da rubrica 87 a= separarporrubrica() - > a[87]
def separarporrubrica(codigo,data1,data2):
    valor = get_values_from_dict(codigo,data1,data2)


    # # Step 1: Extract unique values from the 'ID_RUBRICA' key
    unique_id_rubrica_values = set(item['ID_RUBRICA'] for item in valor)

    # # Step 2: Create separate lists of dictionaries for each unique 'ID_RUBRICA' value
    categorized_data = {value: [] for value in unique_id_rubrica_values}
    for item in valor:
        categorized_data[item['ID_RUBRICA']].append(item)
    
    return categorized_data

#separa por tipo de favorecido as rubricas 87 e 9
def tipodefavorecido(codigo,data1,data2):
    data_categorizada = separarporrubrica(codigo,data1,data2)
    #print(data_categorizada)
    if 87 not in data_categorizada or not data_categorizada[87]:
        print("Data not available or empty.")
        return None  # or handle the case accordingly
    separarportipodefavorecido = set(item['TIPO_FAVORECIDO'] for item in data_categorizada[87])
    #print(separarportipodefavorecido)

    # # Step 2: Create separate lists of dictionaries for each unique 'ID_RUBRICA' value
    dict_favorecido_fisica_e_juridica = {value: [] for value in separarportipodefavorecido}
    for item in data_categorizada[87]:
        dict_favorecido_fisica_e_juridica[item['TIPO_FAVORECIDO']].append(item)

    #print(dict_favorecido_fisica_e_juridica)
    return dict_favorecido_fisica_e_juridica

def preencherCapa(codigo,planilha):
    analista = getAnalistaDoProjetoECpfCoordenador(codigo)
    caminho = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Capa Finatec']
    sheet['E26'] = analista['NOME_ANALISTA']
    workbook.save(planilha)
    workbook.close()

#preenche planilha de referencia pra nome do coordenador e diretor
def criaout(planilha,codigo,data1,data2):
    
    caminho = pegar_caminho(planilha)
    Plan = planilha
    # carrega a planilha de acordo com o caminho
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Receita x Despesa']
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None
    
    
    
    string_periodo = f"Período que abrange esta prestação: {output_date_str} a {output_date_str2}"
    sheet['A7'] = string_periodo
    consulta_coordenador = consultaPorID(codigo)
    sheet['H45'] = consulta_coordenador['COORDENADOR']
    dadosquefaltam = getAnalistaDoProjetoECpfCoordenador(codigo)
    #print(f'aaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa{dadosquefaltam}')
    sheet['H47'] = formatar_cpf(dadosquefaltam["CPF_COORDENADOR"])
    meses_dict = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
}

    hoje = date.today()
    data_formatada = f"{hoje.day} de {meses_dict[hoje.month]} de {hoje.year}"
    sheet["A42"] = f'Brasilia,{data_formatada}'
    workbook.save(planilha)
    workbook.close()

# ##########################################Pessoa Fisica######################################### CANCELADO
def pessoa_fisica(codigo,data1,data2,keys,planilha):
    tabela = pegar_caminho(planilha)
    nomeTabela ="Outros Serviços Terceiros - PF"
    tituloStyle = "pessoaFisica"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Outros Serviços Terceiros - PF")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)
    
    if 87 not in categorized_data or not categorized_data[87]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[87])
    
    print(len(categorized_data[87]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)

    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet5 = workb['Outros Serviços Terceiros - PF']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet5.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 87 not in categorized_data or not categorized_data[87]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[87],li)
        
        n = len(valores_preenchimento)  
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet5.cell(row=rowkek, column=coluna, value=cell_data)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1

    workb.save(tabela)
    workb.close()



def pessoaFisica(codigo,data1,data2,keys,planilha):
   
    tabela = pegar_caminho(planilha)
    nomeTabela ="Outros Serviços Terceiros - PF"
    tituloStyle = "pessoaFisica"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Outros Serviços Terceiros - PF")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)

    tamanho = []
   
    if 87 in categorized_data and 25 in categorized_data:
        categorized_data[87].extend(categorized_data[25])
    elif 87 not in categorized_data and 25 in categorized_data:
       categorized_data[87] = categorized_data[25]
    else:
        print("Data not available or empty.")
        maior = 1
        tabela = pegar_caminho(planilha)
        estiloGeral(tabela,maior,tituloStyle,nomeTabela)
        return None  # or handle the case accordingly
    
    if 87 not in categorized_data or not categorized_data[87]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[87])
    print(maior)
    print(len(categorized_data[87]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)


    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet555 = workb['Outros Serviços Terceiros - PF']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet555.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 87 not in categorized_data or not categorized_data[87]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[87],li)
        
     
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet555.cell(row=rowkek, column=coluna, value=cell_data)
            print(cell_data)
            print(f'row :')
            print(rowkek)
            print(f'coluna :')
            print(coluna)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1

    workb.save(tabela)
    workb.close()


# ##########################################Pessoa Juridica#########################################
def pessoa_juridica(codigo,data1,data2,keys,planilha):
   
   

    tabela = pegar_caminho(planilha)
    nomeTabela ="Pessoa Jurídica"
    tituloStyle = "pessoaJuridica"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Pessoa Jurídica")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)

    tamanho = []
   
    if 75 in categorized_data and 57 in categorized_data:
        categorized_data[75].extend(categorized_data[57])
    elif 75 not in categorized_data and 57 in categorized_data:
       categorized_data[75] = categorized_data[57]
    else:
        print("Data not available or empty.")
        maior = 1
        tabela = pegar_caminho(planilha)
        estiloGeral(tabela,maior,tituloStyle,nomeTabela)
        return None  # or handle the case accordingly
    
    if 75 not in categorized_data or not categorized_data[75]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[75])
    print(maior)
    print(len(categorized_data[75]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)


    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet5 = workb['Pessoa Jurídica']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet5.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 75 not in categorized_data or not categorized_data[75]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[75],li)
        
        n = len(valores_preenchimento)  
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet5.cell(row=rowkek, column=coluna, value=cell_data)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1

    workb.save(tabela)
    workb.close()

# ##########################################ISS#########################################CANCELADO
def iss(codigo,data1,data2,keys,planilha):
    tabela = pegar_caminho(planilha)
    nomeTabela ="ISS"
    tituloStyle = "isss"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="ISS")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)
    
    if 67 not in categorized_data or not categorized_data[67]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[67])
    print(maior)
    print(len(categorized_data[67]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)

    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet5 = workb['ISS']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet5.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 67 not in categorized_data or not categorized_data[67]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[67],li)
        
        n = len(valores_preenchimento)  
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet5.cell(row=rowkek, column=coluna, value=cell_data)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1

    workb.save(tabela)
    workb.close()

# ##########################################Passagem Locomoção#########################################
def passagem_locomção(codigo,data1,data2,keys,planilha):
    tabela = pegar_caminho(planilha)
    nomeTabela ="Passagens e Desp. Locomoção"
    tituloStyle = "passagenDespLocomo"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Passagens e Desp. Locomoção")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)
    
    if 7 not in categorized_data or not categorized_data[7]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[7])
    print(maior)
    print(len(categorized_data[7]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)


    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet3 = workb["Passagens e Desp. Locomoção"]

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet3.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 7 not in categorized_data or not categorized_data[7]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[7],li)
        
        n = len(valores_preenchimento)  
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet3.cell(row=rowkek, column=coluna, value=cell_data)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1
    

    workb.save(tabela)
    workb.close()

# ##########################################Serv.Terceiro CLTa#########################################CANCELADO
def terclt(codigo,data1,data2,keys,planilha):
    tabela = pegar_caminho(planilha)
    nomeTabela ="Serv. Terceiro CLT"
    tituloStyle = "outrosServiçosTerceiros"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Serv. Terceiro CLT")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)
    
    if 25 not in categorized_data or not categorized_data[25]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[25])
    print(maior)
    print(len(categorized_data[25]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)

    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet4 = workb['Serv. Terceiro CLT']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet4.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 25 not in categorized_data or not categorized_data[25]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[25],li)
        
        n = len(valores_preenchimento)  
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet4.cell(row=rowkek, column=coluna, value=cell_data)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1

    workb.save(planilha)
    workb.close()

# ##########################################Obrigaçoes tributárias #########################################
def obricacao_tributaria(codigo,data1,data2,keys,planilha):
    tabela = pegar_caminho(planilha)
    nomeTabela ="Obrigações Trib. - Encargos 20%"
    tituloStyle = "obrigacoesTribu"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Obrigações Trib. - Encargos 20%")
    workbook.save(tabela)
    workbook.close()
    categorized_data= separarporrubrica(codigo,data1,data2)
    
    if 66 not in categorized_data or not categorized_data[66]:
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)
                print("Data not available or empty.")
                return None  # or handle the case accordingly
    maior = len(categorized_data[66])
    print(maior)
    print(len(categorized_data[66]))
    tabela = pegar_caminho(planilha)

    estiloGeral(tabela,maior,tituloStyle,nomeTabela)
    coluna = 2
    # caminho = pegar_caminho(planilha)

    workb = openpyxl.load_workbook(tabela)
    worksheet4 = workb['Obrigações Trib. - Encargos 20%']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet4.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1


    for i in keys:
        li = [i]
        if 66 not in categorized_data or not categorized_data[66]:
                print("Data not available or empty.")
                maior = 1
                tabela = pegar_caminho(planilha)
                estiloGeral(tabela,maior,tituloStyle,nomeTabela)

                return None  # or handle the case accordingly
        valores_preenchimento = retornavalores(categorized_data[66],li)
        
        n = len(valores_preenchimento)  
        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
            worksheet4.cell(row=rowkek, column=coluna, value=cell_data)  
        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1

    workb.save(planilha)
    workb.close()

# ##########################################Conciliação Bancária #########################################
def conciliacao_bancaria(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Conciliação Bancária")
    workbook.save(tabela)
    workbook.close()
    tamanho = []
    categorized_data= separarporrubrica(codigo,data1,data2)
    #####pergar os dados do db e separar por mes e ano###################3
    
    grupos_por_ano_mes = defaultdict(list)
    if 9 not in categorized_data or not categorized_data[9]:
                print("Data not available or empty.")
                maior = 1
                maior2= 2
                tabela = pegar_caminho(planilha)
                estilo_conciliacoes_bancaria(tabela,maior,maior2)
                return None  # or handle the case accordingly
    else:
        
        for item in categorized_data[9]:
            data_criacao_str = item['DATA_CRIACAO']
            
            # Converter a string de data para um objeto datetime
            data_criacao = datetime.strptime(data_criacao_str, '%d/%m/%Y')
            # Extrair o componente do ano e do mês
            ano = data_criacao.year
            mes = data_criacao.month
            dia = data_criacao.day
            # Adicionar o item ao grupo correspondente ao ano e mês
                
            grupos_por_ano_mes[(ano, mes,dia)].append(item)

            # Calcular a soma de VALOR_LANCADO e imprimir os resultados
            
        estorno = defaultdict(list)
        
        tamanho = len(grupos_por_ano_mes)
        #print(tamanho)
        ##loop pra calcular o tamanho do estorno
        for (ano, mes,dia), items in sorted(grupos_por_ano_mes.items()):
                #print(grupos_por_ano_mes)
                for item in items:
                    if 'estorno' in item.get('HIS_LANCAMENTO', '').lower():
                                estorno_valor = item['VALOR_LANCADO']
                                estorno[(ano,mes,dia,item['VALOR_LANCADO'])].append(item)

        tamanho2 = len(estorno)
        #print(tamanho)
        #print(tamanho2)                    
        tamanho = tamanho-tamanho2                
        tabela = pegar_caminho(planilha)
        #print(tabela)
        estilo_conciliacoes_bancaria(tabela,tamanho,tamanho2)
       

        workb = openpyxl.load_workbook(tabela)
        worksheet333 = workb["Conciliação Bancária"]
        i = 16
        j=0
        estorno_valor = 0
        estorno_dia = []
        estorno_mes = []
        estorno_ano = []
        for (ano, mes,dia), items in sorted(grupos_por_ano_mes.items()):  
            #print(grupos_por_ano_mes)
            for item in items:
                if 'estorno' in item.get('HIS_LANCAMENTO', '').lower():
                            estorno_valor = item['VALOR_LANCADO']
                            estorno_dia = dia
                            estorno_mes = mes
                            estorno_ano = ano
                                
                                
                else:
                            valor_lancado = item['VALOR_LANCADO']


            anoss = {1:'jan',
                2:'fev',
                3:'mar',
                4:'abr',
                5:'mai',
                6:'jun',
                7:'jul',
                8:'ago',
                9:'sep',
                10:'out',
                11:'nov',
                12: 'dec'
                    
            }
            for a,b in anoss.items():
                #print(a)
                if mes == a :
                    mes = b
            cell_data = f'{dia}-{mes}-{ano}'
            # print(cell_data)
            # print(valor_lancado)
            if(valor_lancado != 0):
                worksheet333.cell(row=i, column=1, value=cell_data)
                worksheet333.cell(row=i,column=2,value=valor_lancado)
                worksheet333.cell(row=i,column=4,value= item['HIS_LANCAMENTO'])
            else:
                 i = i - 1
            # print(estorno_valor)
            # print(estorno_ano)
            # print(estorno_mes)
            # print(estorno_dia)

            if(estorno_valor != 0):       
                worksheet333.cell(row=16+tamanho+j+4, column=1, value=cell_data)
                worksheet333.cell(row=16+tamanho+j+4, column=2, value=estorno_valor)
               
                j = j +1
                    
            i = i + 1
            valor_lancado = 0
            estorno_valor = 0
        #print(j)
        workb.save(tabela)
        workb.close

# ##########################################Rendimento de Aplicação#########################################
def rendimentodeaplicacao(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Rendimento de Aplicação")
    workbook.save(tabela)
    workbook.close()
    tamanho = []
    categorized_data= separarporrubrica(codigo,data1,data2)
    #####pergar os dados do db e separar por mes e ano###################3
    
    grupos_por_ano_mes = defaultdict(list)
    if 3 not in categorized_data or not categorized_data[3]:
                print("Data not available or empty.")
                maior = 1
                maior2= 2
                tabela = pegar_caminho(planilha)
                estilo_rendimento_de_aplicacao(tabela,maior)
                return None  # or handle the case accordingly
    else:
        
        for item in categorized_data[3]:
            data_criacao_str = item['DATA_CRIACAO']
            
            # Converter a string de data para um objeto datetime
            data_criacao = datetime.strptime(data_criacao_str, '%d/%m/%Y')
            # Extrair o componente do ano e do mês
            ano = data_criacao.year
            mes = data_criacao.month
            dia = data_criacao.day
            # Adicionar o item ao grupo correspondente ao ano e mês
                
            grupos_por_ano_mes[(ano, mes,dia)].append(item)

            # Calcular a soma de VALOR_LANCADO e imprimir os resultados
            
        estorno = defaultdict(list)
        
        tamanho = len(grupos_por_ano_mes)     
        tabela = pegar_caminho(planilha)
        #print(tabela)
        estilo_rendimento_de_aplicacao(tabela,tamanho)
       

        workb = openpyxl.load_workbook(tabela)
        worksheet344 = workb["Rendimento de Aplicação"]
        i = 14
       
        for (ano, mes,dia), items in sorted(grupos_por_ano_mes.items()):  
            soma_valor_lancado = 0
            for item in items:
                soma_valor_lancado += item['VALOR_LANCADO']


            anoss = {1:'jan',
                2:'fev',
                3:'mar',
                4:'abr',
                5:'mai',
                6:'jun',
                7:'jul',
                8:'ago',
                9:'sep',
                10:'out',
                11:'nov',
                12: 'dec'
                    
            }
            for a,b in anoss.items():
                if mes == a :
                    mes = b
            cell_data = f'{mes}-{ano}'
            # print(cell_data)
            # print(valor_lancado)
            
            worksheet344.cell(row=i, column=1, value=cell_data)
            worksheet344.cell(row=i,column=8,value=soma_valor_lancado)
           
            i = i + 1
           
         
      
        workb.save(tabela)
        workb.close
   ##############################

def diaria(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Diárias"
    tituloStyle = "diarias"
    sheet2 = workbook.create_sheet(title="Diárias")
    workbook.save(tabela)
    workbook.close()
    maior = 20
    estiloGeral(tabela,maior,tituloStyle,nomeTabela)

def auxilio(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    nomeTabela ="Auxílio Financeiro Estudante"
    tituloStyle = "auxilioEstudante"
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Auxílio Financeiro Estudante")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloGeral(tabela,tamanho,tituloStyle,nomeTabela)

def bolsaExtensao(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Bolsa Extensão"
    tituloStyle = "bolsaExtensao"
    sheet2 = workbook.create_sheet(title="Bolsa Extensão")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloGeral(tabela,tamanho,tituloStyle,nomeTabela)

def estagiario(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Estagiário"
    tituloStyle = "estagiario"
    sheet2 = workbook.create_sheet(title="Estagiário")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloGeral(tabela,tamanho,tituloStyle,nomeTabela)

def custoIndireto(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Custos Indiretos - FUB"
    tituloStyle = "custosIndiretos"
    sheet2 = workbook.create_sheet(title="Custos Indiretos - FUB")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloGeral(tabela,tamanho,tituloStyle,nomeTabela)

def relacaodeBens(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Relação de Bens"
    tituloStyle = "relacaoBEns"
    sheet2 = workbook.create_sheet(title="Relação de Bens")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloRelacaoBens(tabela,tamanho,tituloStyle,nomeTabela)

def materialDeConsumo(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Material de Consumo"
    tituloStyle = "materialDeConsumo"
    sheet2 = workbook.create_sheet(title="Material de Consumo")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloGeral(tabela,tamanho,tituloStyle,nomeTabela)

def equipamentoMaterialPermanente(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    nomeTabela ="Equipamento Material Permanente"
    tituloStyle = "equipamentoMaterialPermanente"
    sheet2 = workbook.create_sheet(title="Equipamento Material Permanente")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estiloGeral(tabela,tamanho,tituloStyle,nomeTabela)

def demonstrativo(codigo,data1,data2,planilha):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Demonstrativo de Receita")
    workbook.save(tabela)
    workbook.close()
    tamanho = 20
    estilo_demonstrativoDeReceita(tabela,tamanho)

def preencher_fub_teste(codigo,data1,data2,keys,tabela):
    criaout(tabela,codigo,data1,data2)
    #preencherCapa(codigo,tabela)
    pessoaFisica(codigo,data1,data2,keys,tabela)
    pessoa_juridica(codigo,data1,data2,keys,tabela)
    #iss(codigo,data1,data2,keys,tabela)
    passagem_locomção(codigo,data1,data2,keys,tabela)
    #terclt(codigo,data1,data2,keys,tabela)
    obricacao_tributaria(codigo,data1,data2,keys,tabela)
    conciliacao_bancaria(codigo,data1,data2,tabela)
    rendimentodeaplicacao(codigo,data1,data2,tabela)
    diaria(codigo,data1,data2,tabela)
    auxilio(codigo,data1,data2,tabela)
    bolsaExtensao(codigo,data1,data2,tabela)
    #estagiario(codigo,data1,data2,tabela)
    custoIndireto(codigo,data1,data2,tabela)
    materialDeConsumo(codigo,data1,data2,tabela)
    equipamentoMaterialPermanente(codigo,data1,data2,tabela)
    demonstrativo(codigo,data1,data2,tabela)
    relacaodeBens(codigo,data1,data2,tabela)
    

# keys = ['NOME_FAVORECIDO','CNPJ_FAVORECIDO','TIPO_LANCAMENTO','HIS_LANCAMENTO','DATA_EMISSAO','DATA_PAGAMENTO', 'VALOR_PAGO']
# tabela = pegar_caminho("Modelo_Fub.xlsx")
# preencher_fub_teste(6411,'2020-01-01','2024-01-31',keys,tabela)

# pessoa_fisica(6858,'2022-09-09','2022-12-09',keys)
