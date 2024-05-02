import pyodbc
from datetime import datetime,date
import openpyxl
from openpyxl.styles import Font
import os
from collections import defaultdict
from .estiloFINEP import *
from .preencheFub import planilhaGeral,consultaNomeRubricaCodRubrica,consultaID,convert_datetime_to_string,convert_datetime_to_stringdt,formatar_data,formatarDataSemDia,formatar_cpf,check_format,pegar_caminho,pegar_pass
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.engine import URL
import numpy as np  
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re



#consulta

#ok
def consultaRelatorioExecFinanceiraA1(IDPROJETO, DATA1, DATA2):
    ''' Consulta que busca os valores executado no periodo, e os valores que foram executados até no periodo
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametros2 = [(IDPROJETO, DATA2)]
    parametros3 = IDPROJETO
    consultaComPeriodo =f"SELECT NomeRubrica, SUM(ValorPago) AS VALOR_TOTAL_PERIODO FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica"
    consultaAteAData =  f"SELECT NomeRubrica, SUM(ValorPago) AS VALOR_TOTAL_DATA FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento <= ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica "
    #consultaPrevisto = f"SELECT NomeRubrica, SUM(VALOR*Quantidade) AS VALOR_TOTAL_PREVISTO FROM [Conveniar].[dbo].[LisConvenioItemAprovado] WHERE CodConvenio = ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica"
    
    dfComPeriodo= pd.read_sql(consultaComPeriodo, engine, params=parametros)
    dfAteAData = pd.read_sql(consultaAteAData, engine, params=parametros2)
    #dfPrevisto = pd.read_sql(consultaPrevisto, engine, params=(IDPROJETO,))

    return dfComPeriodo,dfAteAData
#ok
def consultaDemonstrativoReceitaEDespesaA2(IDPROJETO,DATA1,DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametros2 = [(IDPROJETO, DATA1)]
    parametros4 = [(IDPROJETO, DATA2)]
    parametros3 = IDPROJETO
    consultaComPeriodo =f"SELECT NomeRubrica, SUM(ValorPago) AS VALOR_TOTAL_PERIODO FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento BETWEEN ? AND ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica"
    consultaAteAData =  f"SELECT NomeRubrica, SUM(ValorPago) AS VALOR_TOTAL_DATA FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento <= ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica "
    
    consultaPrevisto = f"SELECT NomeRubrica, SUM(VALOR*Quantidade) AS VALOR_TOTAL_PREVISTO FROM [Conveniar].[dbo].[LisConvenioItemAprovado] WHERE CodConvenio = ? GROUP BY NomeRubrica, CodRubrica Order by CodRubrica"
    consultaRubricaRecursoRecebidos =  f"SELECT sum(ValorPago) AS VALOR_TOTAL_DATA FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND DataPagamento <= ? "
    consultaRendimentoAplicacao = f"""
    SELECT 
        NomeTipoLancamento,
        --SUM(CASE WHEN NomeTipoLancamento = 'IRRF Pessoa Jurídica' THEN ValorPago ELSE 0 END) AS IRRF,
        SUM(CASE WHEN NomeTipoLancamento = 'Aplicação Financeira' THEN ValorPago ELSE 0 END) AS Aplicação
    FROM 
        [Conveniar].[dbo].[LisLancamentoConvenio] 
    WHERE 
        CodConvenio = ? 
        AND CodStatus = 27 
        AND CodRubrica = 3 
        AND DataPagamento BETWEEN ? AND ?
    GROUP BY 
        NomeTipoLancamento;
    """
   



    Soma = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
    dfComPeriodo= pd.read_sql(consultaComPeriodo, engine, params=parametros)
    dfAteAData = pd.read_sql(consultaAteAData, engine, params=parametros2)
    dfPrevisto = pd.read_sql(consultaPrevisto, engine, params=(IDPROJETO,))
    dfRubricaRecursoRecebidos = pd.read_sql(consultaRubricaRecursoRecebidos, engine, params=parametros4)



    return dfComPeriodo,dfAteAData,dfPrevisto,dfRubricaRecursoRecebidos,Soma
#ok
def consultaPagamentoPessoal(IDPROJETO,DATA1,DATA2):
    
    ''' Consulta dinamica do SQL relacionado a rubrica correspondente,cada pagina tem sua própria consulta correspondente a rubrica
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            codigoRubrica = código da rubrica 
    '''
    codigoRubrica = 87
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametrosComRubricaEstorno  = [(IDPROJETO, DATA1, DATA2,IDPROJETO,DATA1, DATA2)]
    parametrosPJ=[(IDPROJETO, DATA1, DATA2)]
    parametrosPJestorno=[(IDPROJETO, DATA1, DATA2,IDPROJETO, DATA1, DATA2)]
    queryConsultaComRubrica = f"""SELECT NomeFavorecido,
     CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     NomeTipoLancamento,
     [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     NumDocPago,
     DataEmissao,
     NumChequeDeposito,
     DataPagamento,
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
     WHERE [LisLancamentoConvenio].CodConvenio = ? 
     AND [LisLancamentoConvenio].CodStatus = 27
     AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ? 
     AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' 
     and [LisLancamentoConvenio].CodRubrica = 87 order by DataPagamento"""
    
    #SEMPRE TEM Q ADICIONAR DUAS COLUNAS DEPOIS
    queryConsultaComRubricaEstorno = f"""SELECT NomeFavorecido
    ,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-')
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     NomeTipoLancamento,   
    HisLancamento,
     NumChequeDeposito,
     DataPagamento, 
     ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado]
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado]
     WHERE 
     [LisLancamentoConvenio].CodConvenio = ? 
     AND CodStatus = 27 
     AND NomeTipoCreditoDebito = 'C' 
     AND DataPagamento BETWEEN ? AND ? 
     and [LisLancamentoConvenio].CodRubrica = 87 
     OR 
     CodStatus = 27
     AND [LisLancamentoConvenio].CodConvenio = ?  
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento)  LIKE '%estorno%'
     AND [LisLancamentoConvenio].CodRubrica = 87 
     
     order by DataPagamento """


    dfconsultaDadosPorRubrica = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)
    dfconsultaDadosPorRubricaComEstorno = pd.read_sql(queryConsultaComRubricaEstorno,engine, params=parametrosComRubricaEstorno)



    return dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno
#ok
def consultaestiloElementoDeDespesa1415Diarias(IDPROJETO,DATA1,DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)

    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametros2 = [(IDPROJETO, DATA1, DATA2,IDPROJETO, DATA1, DATA2)]
    
    queryConsultaSemEstorno = f"""
        SELECT [LisPessoa].[NomePessoa],

	    CASE WHEN LEN([LisPessoa].CPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF([LisPessoa].CPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
        WHEN LEN([LisPessoa].CPFCNPJ) = 11 THEN STUFF(STUFF(STUFF([LisPessoa].CPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE [LisPessoa].CPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
	    (SELECT TOP 1 [NomeCidadeOrigem]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Saída') + ' - ' +  
		(SELECT TOP 1 [NomeCidadeDestino]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Saída') 

		+ ' -> ' +
		 (SELECT TOP 1 [NomeCidadeOrigem]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Chegada') + ' - ' +  
		(SELECT TOP 1 [NomeCidadeDestino]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Chegada') AS Destino


        ,[LisPagamentoDiaria].[QuantDiaria]
        ,[LisPagamentoDiaria].[ObsPedido]
        ,[LisConvenioItemAprovado].[DescConvenioItemAprovado]
        ,[LisLancamentoConvenio]. NumDocPago
        ,[LisLancamentoConvenio]. DataEmissao
	    ,[LisLancamentoConvenio].NumChequeDeposito
	    ,[LisLancamentoConvenio].DataPagamento
        ,[LisLancamentoConvenio].ValorPago
        FROM [Conveniar].[dbo].[LisLancamentoConvenio]
        INNER JOIN [Conveniar].[dbo].[LisPagamentoDiaria] ON [LisLancamentoConvenio].[NumDocFinConvenio] = [LisPagamentoDiaria].[NumPedido]
        INNER JOIN [Conveniar].[dbo].[LisPessoa] ON [LisPagamentoDiaria].[CodPessoaFavorecida] = [LisPessoa].[CodPessoa]
        LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
        LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 


        WHERE LisLancamentoConvenio.CodConvenio = ? 
        AND LisLancamentoConvenio.CodStatus = 27
        AND LisLancamentoConvenio.DataPagamento BETWEEN ? AND ? 
        """

    queryConsultaComEstorno = f"""
    SELECT
     [LisPessoa].[NomePessoa]
	  ,

	    CASE WHEN LEN([LisPessoa].CPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF([LisPessoa].CPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
        WHEN LEN([LisPessoa].CPFCNPJ) = 11 THEN STUFF(STUFF(STUFF([LisPessoa].CPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE [LisPessoa].CPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
	    (SELECT TOP 1 [NomeCidadeOrigem]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Saída') + ' - ' +  
		(SELECT TOP 1 [NomeCidadeDestino]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Saída') 

		+ ' -> ' +
		 (SELECT TOP 1 [NomeCidadeOrigem]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Chegada') + ' - ' +  
		(SELECT TOP 1 [NomeCidadeDestino]
		FROM [Conveniar].[dbo].[LisPagamentoDiariaTrecho]
		WHERE [LisPagamentoDiariaTrecho].CodPedido = [LisPagamentoDiaria].[CodPedido]
		AND [NomeTipoDestino] = 'Chegada') AS Destino


    ,[LisPagamentoDiaria].[QuantDiaria]
    ,[LisPagamentoDiaria].[ObsPedido]
    ,[LisConvenioItemAprovado].[DescConvenioItemAprovado]
	,[LisLancamentoConvenio].NumChequeDeposito
	,[LisLancamentoConvenio]. NumDocPago
    ,[LisLancamentoConvenio]. DataEmissao
	,[LisLancamentoConvenio].NumChequeDeposito
	,[LisLancamentoConvenio].DataPagamento
    ,[LisLancamentoConvenio].ValorPago
    FROM [Conveniar].[dbo].[LisLancamentoConvenio]
    INNER JOIN [Conveniar].[dbo].[LisPagamentoDiaria] ON [LisLancamentoConvenio].[NumDocFinConvenio] = [LisPagamentoDiaria].[NumPedido]
    INNER JOIN [Conveniar].[dbo].[LisPessoa] ON [LisPagamentoDiaria].[CodPessoaFavorecida] = [LisPessoa].[CodPessoa]
    LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
    LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 


    WHERE  [LisLancamentoConvenio].CodConvenio = ?  
    AND [LisLancamentoConvenio].[CodStatus] = 27
    AND NomeTipoCreditoDebito = 'C' 
    AND LisLancamentoConvenio.DataPagamento BETWEEN ? AND ? 
    or
    [LisLancamentoConvenio].CodConvenio = ?  
    AND [LisLancamentoConvenio].[CodStatus] = 27
    AND LOWER(HisLancamento)  LIKE '%estorno%'
    AND LisLancamentoConvenio.DataPagamento BETWEEN ? AND ? 


    Order by LisLancamentoConvenio.DataPagamento"""


    dfConsultaDiaria = pd.read_sql(queryConsultaSemEstorno, engine, params=parametros)
    dfConsultaDiariaEstorno = pd.read_sql(queryConsultaComEstorno, engine, params=parametros2)


    return dfConsultaDiaria ,dfConsultaDiariaEstorno
#ok    
def consultaGeral30(IDPROJETO,DATA1,DATA2,codigoRubrica):

    ''' Consulta dinamica do SQL relacionado a rubrica correspondente,cada pagina tem sua própria consulta correspondente a rubrica
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            codigoRubrica = código da rubrica 
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2,codigoRubrica)]
    parametrosComRubricaEstorno  = [(IDPROJETO, DATA1, DATA2,codigoRubrica,IDPROJETO,DATA1, DATA2,codigoRubrica,)]
    parametrosPJ=[(IDPROJETO, DATA1, DATA2)]
    parametrosPJestorno=[(IDPROJETO, DATA1, DATA2,IDPROJETO, DATA1, DATA2)]

    queryConsultaComRubrica = f"""SELECT NomeFavorecido,
     CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     
     NumDocPago,
     DataEmissao,
     NumChequeDeposito,
     DataPagamento,
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
     WHERE [LisLancamentoConvenio].CodConvenio = ? 
     AND [LisLancamentoConvenio].CodStatus = 27
     AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ? 
     AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' 
     and [LisLancamentoConvenio].CodRubrica = ? order by DataPagamento"""
    
    queryConsultaComRubricaEstorno = f"""SELECT NomeFavorecido
    ,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-')
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     HisLancamento,
     NumChequeDeposito,
     DataPagamento, 
     ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado]
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado]
     WHERE 
     [LisLancamentoConvenio].CodConvenio = ? 
     AND CodStatus = 27 
     AND NomeTipoCreditoDebito = 'C' 
     AND DataPagamento BETWEEN ? AND ? 
     and [LisLancamentoConvenio].CodRubrica = ? 
     OR 
     CodStatus = 27
     AND [LisLancamentoConvenio].CodConvenio = ?  
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento)  LIKE '%estorno%'
     AND [LisLancamentoConvenio].CodRubrica = ? 
     
     order by DataPagamento """
    
    queryConsultaPJDOA = f"""SELECT NomeFavorecido,
     CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     
     NumDocPago,
     DataEmissao,
     NumChequeDeposito,
     DataPagamento,
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
     WHERE [LisLancamentoConvenio].CodConvenio = ? 
     AND [LisLancamentoConvenio].CodStatus = 27
     AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ? 
     AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' 
     and [LisLancamentoConvenio].CodRubrica IN (57,75,26) order by DataPagamento"""
    
    queryConsultaPJDOAEstorno = f"""SELECT NomeFavorecido
    ,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-')
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     HisLancamento,
     NumChequeDeposito,
     DataPagamento, 
     ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado]
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado]
     WHERE 
     [LisLancamentoConvenio].CodConvenio = ? 
     AND CodStatus = 27 
     AND NomeTipoCreditoDebito = 'C' 
     AND DataPagamento BETWEEN ? AND ? 
     and [LisLancamentoConvenio].CodRubrica IN (57,75,26) 
     OR 
     CodStatus = 27
     AND [LisLancamentoConvenio].CodConvenio = ?  
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento)  LIKE '%estorno%'
     AND [LisLancamentoConvenio].CodRubrica IN (57,75,26)
     
     order by DataPagamento """

    dfconsultaDadosPorRubrica = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)
    dfconsultaDadosPorRubricaComEstorno = pd.read_sql(queryConsultaComRubricaEstorno,engine, params=parametrosComRubricaEstorno)
    dfPJDOA = pd.read_sql(queryConsultaPJDOA, engine, params=parametrosPJ)
    dfPJDOAESTORNO = pd.read_sql(queryConsultaPJDOAEstorno,engine, params=parametrosPJestorno)
    
    return dfPJDOA,dfPJDOAESTORNO,dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno
#ok
def consultaestiloElementoDeDespesa33PassagemEDespesa(IDPROJETO,DATA1,DATA2):
     
    ''' Consulta dinamica do SQL relacionado a rubrica correspondente,cada pagina tem sua própria consulta correspondente a rubrica
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            codigoRubrica = código da rubrica 
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametrosComRubricaEstorno  = [(IDPROJETO, DATA1, DATA2,IDPROJETO,DATA1, DATA2,)]

    queryConsultaComRubrica = f"""SELECT NomeFavorecido,
     CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-') 
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
     [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     
     NumDocPago,
     DataEmissao,
     NumChequeDeposito,
     DataPagamento,
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado] 
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado] 
     WHERE [LisLancamentoConvenio].CodConvenio = ? 
     AND [LisLancamentoConvenio].CodStatus = 27
     AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ? 
     AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' 
     and [LisLancamentoConvenio].CodRubrica in (20,78,52)  order by [LisLancamentoConvenio].DataPagamento"""
    
    queryConsultaComRubricaEstorno = f"""SELECT NomeFavorecido
    ,CASE WHEN LEN(FavorecidoCPFCNPJ) > 11 THEN STUFF(STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 3, 0, '.'), 7, 0, '.'), 11, 0, '/'), 16, 0, '-')
     WHEN LEN(FavorecidoCPFCNPJ) = 11 THEN STUFF(STUFF(STUFF(FavorecidoCPFCNPJ, 4, 0, '.'), 8, 0, '.'), 12, 0, '-') ELSE FavorecidoCPFCNPJ END AS FormattedFavorecidoCPFCNPJ,
    [LisConvenioItemAprovado].[DescConvenioItemAprovado],
     NumChequeDeposito,
     DataPagamento, 
     ValorPago 
     FROM [Conveniar].[dbo].[LisLancamentoConvenio]
     LEFT JOIN [Conveniar].[dbo].[PlanoTrabalhoLancamento] ON [LisLancamentoConvenio].[CodLancamento] = [PlanoTrabalhoLancamento].[CodLancamentoGerado]
     LEFT JOIN [Conveniar].[dbo].[LisConvenioItemAprovado] ON [PlanoTrabalhoLancamento].[CodConvenioItemAprovado] = [LisConvenioItemAprovado].[CodConvenioItemAprovado]
     WHERE 
     [LisLancamentoConvenio].CodConvenio = ? 
     AND CodStatus = 27 
     AND NomeTipoCreditoDebito = 'C' 
     AND DataPagamento BETWEEN ? AND ? 
     and [LisLancamentoConvenio].CodRubrica in (20,78,52) 
     OR 
     CodStatus = 27
     AND [LisLancamentoConvenio].CodConvenio = ?  
     AND DataPagamento BETWEEN ? AND ? 
     AND LOWER(HisLancamento)  LIKE '%estorno%'
     AND [LisLancamentoConvenio].CodRubrica in (20,78,52)
     
     order by [LisLancamentoConvenio].DataPagamento """

    dfconsultaDadosPorRubrica = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)
    dfconsultaDadosPorRubricaComEstorno = pd.read_sql(queryConsultaComRubricaEstorno,engine, params=parametrosComRubricaEstorno)
  
    
    return dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno
#ok
def consultaBens(IDPROJETO,DATA1,DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passss.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    idprojetoComZero = f"0{IDPROJETO}"
    parametros = [(IDPROJETO,idprojetoComZero, DATA1, DATA2)]
    
    queryConsultaComRubrica = f"""
    SELECT [Descrição][descri],
    [Patrimônio][patri],
    CONVERT(varchar, [Data de Aquisição], 103) AS dataAqui,
    [Nº Nota][nota],
    [Localização][localiza]
   ,[Descrição],
    [Valor de Aquisição][valorAqui],
    [Valor de Aquisição][valorAqui2],
    [Responsável][responsavel] 
    FROM [SBO_FINATEC].[dbo].[VW_BENS_ADQUIRIDOS] 
    WHERE ([Cod Projeto] = ? or [Cod Projeto] = ? ) 
    AND [Status] = 'Imobilizado' 
    AND [Data de Aquisição] BETWEEN ? AND ? 
    Order by [Data de Aquisição]"""
    dfConsultaBens = pd.read_sql(queryConsultaComRubrica, engine, params=parametros)

   
    return dfConsultaBens   
#ok    
def consultaConciliacaoBancaria(IDPROJETO, DATA1, DATA2):
    ''' Função que vai pega os dados da Rubrica 9 Despesas Financeiras e transformalos em dataframe
    para poder popular a databela Despesas Financeiras
        Argumentos
            IDPROJETO = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
    '''
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    parametros2 = [(IDPROJETO, DATA1, DATA2, IDPROJETO, DATA1, DATA2)]
    #consultaSemEstorno = f"SELECT DISTINCT DataPagamento,ValorPago,NumChequeDeposito,HisLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) NOT LIKE '%estorno%' order by DataPagamento"
    #consultaComEstorno =  f"SELECT DISTINCT DataPagamento,ValorPago,NumChequeDeposito,HisLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 9 AND DataPagamento BETWEEN ? AND ? AND LOWER(HisLancamento) LIKE '%estorno%' order by DataPagamento"
    
    #consultaloucona assim por que precisa ser detalhado cada item na tabela
    consultaSemEstorno = f"""
    SELECT 
  	[LisLancamentoConvenio].HisLancamento, 
	CONVERT(varchar, CAST([LisLancamentoConvenio].DataPagamento AS datetime), 103) AS FormattedDate,
    [LisPagamentoDespesaConvenioAdministrativa].Valor 
    FROM [Conveniar].[dbo].[LisLancamentoConvenio]
    INNER JOIN  [Conveniar].[dbo].[LisDocumentoConvenio] ON [LisLancamentoConvenio].[CodDocFinConvenio] = [LisDocumentoConvenio].[CodDocFinConvenio]
    INNER JOIN  [Conveniar].[dbo].[DocFinConvPagDespesa] ON [LisDocumentoConvenio].[CodDocFinConvenio] = [DocFinConvPagDespesa].[CodDocFinConvenio]
    INNER JOIN  [Conveniar].[dbo].[LisPagamentoDespesaConvenio] ON [DocFinConvPagDespesa].[CodPedido] = [LisPagamentoDespesaConvenio].[CodPedido]
    INNER JOIN  [Conveniar].[dbo].[LisPagamentoDespesaConvenioAdministrativa] ON [LisPagamentoDespesaConvenio].CodDespesaConvenio = [LisPagamentoDespesaConvenioAdministrativa].CodDespesaConvenio
    AND [LisPagamentoDespesaConvenio].CodConvenio = [LisPagamentoDespesaConvenioAdministrativa].CodConvenio
    WHERE [LisLancamentoConvenio].CodConvenio = ? 
    AND [LisLancamentoConvenio].CodStatus = 27 
    AND [LisLancamentoConvenio].CodRubrica = 9 
    AND [LisLancamentoConvenio].DataPagamento BETWEEN ? AND ?
    AND LOWER([LisLancamentoConvenio].HisLancamento) NOT LIKE '%estorno%' 
    order by [LisLancamentoConvenio].DataPagamento"""

    consultaComEstorno = f"""SELECT 
    HisLancamento ,
    CONVERT(varchar, CAST(DataPagamento AS datetime), 103) AS FormattedDate,
    ValorPago
    
    FROM [Conveniar].[dbo].[LisLancamentoConvenio] 
    WHERE CodConvenio = ? 
    AND CodStatus = 27 
    AND CodRubrica = 9 
    AND DataPagamento BETWEEN ? AND ? 
    AND LOWER(HisLancamento)  LIKE '%estorno%'
    OR
    CodConvenio = ? 
    AND CodStatus = 27 
    AND CodRubrica = 9 
    AND NomeTipoCreditoDebito = 'C' 
    AND DataPagamento BETWEEN  ? AND ? 
    order by DataPagamento"""
    
    dfSemEstorno = pd.read_sql(consultaSemEstorno, engine, params=parametros)
    dfComEstorno = pd.read_sql(consultaComEstorno, engine, params=parametros2)
    

    return dfSemEstorno,dfComEstorno
#ok
def consultaRendimentosAplicacao(IDPROJETO,DATA1,DATA2):
    file_path = pegar_pass("passs.txt")
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    connection_url = URL.create("mssql+pyodbc", query={"odbc_connect": conStr})
    engine = create_engine(connection_url)
    parametros = [(IDPROJETO, DATA1, DATA2)]
    consultaRendimentoAplicacao = f"SELECT  DataPagamento,ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and NomeTipoLancamento = 'Aplicação Financeira'  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaRendimentoAplicacao = pd.read_sql(consultaRendimentoAplicacao, engine, params=parametros)
     
    consultaRendimentoEImposto =  f"SELECT DataPagamento,ValorPago,NomeTipoLancamento FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and (NomeTipoLancamento = 'Aplicação Financeira' or NomeTipoLancamento = 'IRRF Pessoa Jurídica')  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaRendimentoEImposto = pd.read_sql(consultaRendimentoEImposto, engine, params=parametros)
    
    consultaImposto =  f"SELECT  DataPagamento,ValorPago FROM [Conveniar].[dbo].[LisLancamentoConvenio] WHERE CodConvenio = ? AND CodStatus = 27 AND CodRubrica = 3 and NomeTipoLancamento = 'IRRF Pessoa Jurídica'  AND DataPagamento BETWEEN ? AND ?  order by DataPagamento"
    dfConsultaImposto = pd.read_sql(consultaImposto, engine, params=parametros)
     
   
    return dfConsultaRendimentoAplicacao,dfConsultaImposto,dfConsultaRendimentoEImposto




#preencher
#ok
def demostrativereceitaedepesaA2(codigo,data1,data2,planilha):
    #carrega planilha e colocca o estilo e retorna a localização da row brasilia
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="DEMOSTR. RECEITA E DESPESA A.2")
    workbook.save(tabela)
    workbook.close()
    #pega as 3 tabelas da o merge e retira o tamanho com base nas rubricas chaves que ja possuem um lugar no estilo
    dfComPeriodo,dfAteAData,dfPrevisto,dfRubricaRecursoRecebidos,Soma=consultaDemonstrativoReceitaEDespesaA2(codigo,data1,data2)
    merged_df = pd.merge(dfPrevisto, dfComPeriodo, on='NomeRubrica', how='outer')
    dfMerged = pd.merge(merged_df,dfAteAData, on = 'NomeRubrica', how = 'outer')
    tamanho = len(dfMerged)#tamanho para deixar dinamico para imprimir sa rubricas
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Receitas"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Rendimentos de Aplicações Financeiras"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Despesas Financeiras"]).any()
    if string_exists:
         tamanho = tamanho - 1
    
   
   
    rowBrasilia = estiloDEMOSTRRECEITEDESPESAA2(planilha,tamanho)
    
    #carregar planilha para preencher
    caminho = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['DEMOSTR. RECEITA E DESPESA A.2']
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None
    
    string_periodo = f"de {output_date_str} a {output_date_str2}"
    
    consulta_coordenador = consultaID(codigo)

    # o quarenta significa pra achar o lugar do coordenador dinamicamente mais o valor do tamanho da quantidade de rubricas
    tamanho_row_coordenador = tamanho + 40
    stringCoordenador= f'C{tamanho_row_coordenador}' # retorna lugar do coordanor
    stringTamanhoCPF = f'C{tamanho_row_coordenador+2}' # retorna lugar do coordanor
    sheet[stringCoordenador] = consulta_coordenador['NomePessoaResponsavel']
    sheet[stringTamanhoCPF] = formatar_cpf(consulta_coordenador['CPFCoordenador'])
    string_convenio = f"Convênio nº: {consulta_coordenador['SubProcesso']}"
    string_convenente= f"Fundação de Empreendimentos Científicos e Tecnológicos - FINATEC" 
    string_convenente_convenente= f"Convenente: "
    string_fonte_recursos = f"Fonte de Recursos:"
    string_participe = f"Partícipe (no caso de contrapartida):"
    string_periodo_relatorio = f"Período Abrangido por este Relatório: "
    stringPeriodoExececucao = f'Período de Execução do Convênio:'
   # Convert 'DataAssinatura' to "dd/mm/YYYY" format
    datetime_obj1 = consulta_coordenador['DataAssinatura']
    formatted_date1 = datetime_obj1.strftime("%d/%m/%Y")

    # Convert 'DataVigencia' to "dd/mm/YYYY" format
    datetime_obj2 = consulta_coordenador['DataVigencia']
    formatted_date2 = datetime_obj2.strftime("%d/%m/%Y")

# Create the string representing the period of execution
   
    string_periodo_convenio = f"de {formatted_date1} a {formatted_date2}"
    sheet['A4'] = string_convenio
    sheet['A5'] = string_convenente_convenente
    sheet['A6'] = stringPeriodoExececucao
    sheet['A7'] = string_periodo_relatorio
    sheet['A8'] = string_fonte_recursos
    sheet['A9'] = string_participe

    sheet['C5'] = string_convenente
    sheet['C6'] = string_periodo_convenio
    sheet['C7'] = string_periodo
    sheet['C8'] = f'RECURSOS FINEP/RECURSOS CONTRAPARTIDA'
    sheet['C9'] = F'XXX'
    
   
    meses_dict = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro"
    }   
    
    stringTamanho = f'A{tamanho + 37}' # retorna lugar de brasilia
    hoje = date.today()
    data_formatada = f"{hoje.day} de {meses_dict[hoje.month]} de {hoje.year}"
    sheet[stringTamanho] = f'Brasilia, {data_formatada}'


    #começar o preenchimento com o dataframe

    
    rowDespesasCapital = tamanho + 13 #vai ser o valro menos 1 para n bugar os codigos de tamanho abaixo

    #TOTAL DAS RECEITAS (B1+B2)
    #B.1
    #periodo receitas
    if dfMerged['NomeRubrica'].isin(["Receitas"]).any():
        stringReceitas = f'C{rowDespesasCapital + 11}'
        sheet[stringReceitas] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Receitas', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o perido anterior receitas
        stringReceitas = f'B{rowDespesasCapital + 11}'
        sheet[stringReceitas] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Receitas', 'VALOR_TOTAL_PREVISTO'].values[0]
    #B.2
    #periodo Rendimentos de Aplicações Financeiras
    if dfMerged['NomeRubrica'].isin(["Rendimentos de Aplicações Financeiras"]).any():
        stringRendimentosdeAplicaçõesFinanceiras = f'C{rowDespesasCapital + 12}'
        sheet[stringRendimentosdeAplicaçõesFinanceiras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Rendimentos de Aplicações Financeiras', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o perido anterior Rendimentos de Aplicações Financeiras
        stringRendimentosdeAplicaçõesFinanceiras = f'B{rowDespesasCapital + 12}'
        sheet[stringRendimentosdeAplicaçõesFinanceiras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Rendimentos de Aplicações Financeiras', 'VALOR_TOTAL_PREVISTO'].values[0]

   
      #Obras e Instalações
    #previsto
    string_exists = dfMerged['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
        
        stringObras = f'E{rowDespesasCapital + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PREVISTO'].values[0]
     
        #periodo
        stringObras = f'C{rowDespesasCapital + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 2}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_DATA'].values[0]

    string_exists = dfMerged['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
    #Materiais Equipamentos e Material Permanente   
        stringObras = f'E{rowDespesasCapital + 5}'
        
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PREVISTO'].values[0]
       
        #periodo
        stringObras = f'C{rowDespesasCapital + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_DATA'].values[0]

    #Materiais Equipamentos e Material nACIONAL
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
        stringObras = f'E{rowDespesasCapital + 6}'
        
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PREVISTO'].values[0]
       
        #periodo
        stringObras = f'C{rowDespesasCapital + 6}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 6}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_DATA'].values[0]

    #Materiais Equipamentos e Material iMPORTADO
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
        stringObras = f'E{rowDespesasCapital + 7}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PREVISTO'].values[0]
     
        #periodo
        stringObras = f'C{rowDespesasCapital + 7}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 7}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_DATA'].values[0]



    #remover essas linhas da tabela
    values_to_remove = ["Receitas", "Rendimentos de Aplicações Financeiras", "Despesas Financeiras",'Material Permanente e Equipamento Nacional','Material Permanente e Equipamento Importado','Devolução de Recursos','Obras e Instalações','Equipamentos e Material Permanente']

    # Use boolean indexing to drop rows based on the values in the first column
    dfMerged = dfMerged[~dfMerged['NomeRubrica'].isin(values_to_remove)]
    

    
    string_exists = dfMerged['NomeRubrica'].isin(["Despesas Operacionais e Administrativas - Finatec"]).any()
    if string_exists:
        # Extract the value from "Despesas Operacionais e Administrativas - Finatec"
        value_to_add = dfMerged.loc[dfMerged['NomeRubrica'] == 'Despesas Operacionais e Administrativas - Finatec'].iloc[0]
        
        string_exists = dfMerged['NomeRubrica'].isin(["Outros Serviços de Terceiros - Pessoa Jurídica "]).any()
        string_exists2 = dfMerged['NomeRubrica'].isin(["Serviços de Terceiros Pessoa Jurídica"]).any()
        string_exists3 = dfMerged['NomeRubrica'].isin(["Outros Serviços de Terceiros - Pessoa Jurídica"]).any()
        
        if string_exists or string_exists2 or string_exists3:
            if string_exists:
                
                
                dfMerged.loc[dfMerged['NomeRubrica'] == 'Outros Serviços de Terceiros - Pessoa Jurídica '] += value_to_add
             
                # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
                dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']
                
            if string_exists2:
               

                # Add the value to "Outros Serviços de Terceiros - Pessoa Jurídica"
                dfMerged.loc[dfMerged['NomeRubrica'] == 'Serviços de Terceiros Pessoa Jurídica'] += value_to_add

                # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
                dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']
                
            if string_exists3:
            # Find the index of "Outros Serviços de Terceiros - Pessoa Jurídica"
                
                index_to_update = dfMerged.loc[dfMerged['NomeRubrica'] == 'Outros Serviços de Terceiros - Pessoa Jurídica'].index[0]
                
                # Add the value to "Outros Serviços de Terceiros - Pessoa Jurídica"
                dfMerged.iloc[index_to_update] += value_to_add
                
                # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
                dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']   
                
            

          
    
    

    for row_num, row_data in enumerate(dfMerged.itertuples(index = False), start=14):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
              
                if col_num == 3:
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
                if col_num == 1:
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
                if col_num == 2:
                    col_num = 5
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
                    #sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value))
               
                if col_num == 4:
                    col_num = 2
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
         


    workbook.save(planilha)
    workbook.close()
   


    return rowBrasilia
#ok
def relatorioExecFinanceiraA1(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Relatório de Exec Financ A.1")
    workbook.save(tabela)
    workbook.close()
    dfComPeriodo,dfAteAData = consultaRelatorioExecFinanceiraA1(codigo, data1, data2)

    
    dfMerged = pd.merge(dfAteAData,dfComPeriodo, on = 'NomeRubrica', how = 'outer')
    tamanho = len(dfMerged)#tamanho para deixar dinamico para imprimir sa rubricas
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Receitas"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Rendimentos de Aplicações Financeiras"]).any()
    if string_exists:
         tamanho = tamanho - 1
    string_exists = dfMerged['NomeRubrica'].isin(["Despesas Financeiras"]).any()
    if string_exists:
         tamanho = tamanho - 1



    
    rowDespesasCapital = estiloRelatorioExecFinanceiroA1(tabela,tamanho,rowBrasilia)



    caminho = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook['Relatório de Exec Financ A.1']
    

 
    #Obras e Instalações
    #previsto
    string_exists = dfMerged['NomeRubrica'].isin(["Obras e Instalações"]).any()
    if string_exists:
        

     
        #periodo
        stringObras = f'F{rowDespesasCapital + 1}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 1}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Obras e Instalações', 'VALOR_TOTAL_DATA'].values[0]

    string_exists = dfMerged['NomeRubrica'].isin(["Equipamentos e Material Permanente"]).any()
    if string_exists:
    #Materiais Equipamentos e Material Permanente   
        
       
        #periodo
        stringObras = f'F{rowDespesasCapital + 4}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 4}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Equipamentos e Material Permanente', 'VALOR_TOTAL_DATA'].values[0]

    #Materiais Equipamentos e Material nACIONAL
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Nacional"]).any()
    if string_exists:
        
       
        #periodo
        stringObras = f'F{rowDespesasCapital + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 5}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Nacional', 'VALOR_TOTAL_DATA'].values[0]

    #Materiais Equipamentos e Material iMPORTADO
    string_exists = dfMerged['NomeRubrica'].isin(["Material Permanente e Equipamento Importado"]).any()
    if string_exists:
        
        #periodo
        stringObras = f'F{rowDespesasCapital + 6}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_PERIODO'].values[0]
        #Ate o momento
        stringObras = f'B{rowDespesasCapital + 6}'
        sheet[stringObras] = dfMerged.loc[dfMerged['NomeRubrica'] == 'Material Permanente e Equipamento Importado', 'VALOR_TOTAL_DATA'].values[0]



    #remover essas linhas da tabela
    values_to_remove = ["Receitas", "Rendimentos de Aplicações Financeiras", "Despesas Financeiras",'Material Permanente e Equipamento Nacional','Material Permanente e Equipamento Importado','Devolução de Recursos','Obras e Instalações','Equipamentos e Material Permanente']

    # Use boolean indexing to drop rows based on the values in the first column
    dfMerged = dfMerged[~dfMerged['NomeRubrica'].isin(values_to_remove)]
    

    
    string_exists = dfMerged['NomeRubrica'].isin(["Despesas Operacionais e Administrativas - Finatec"]).any()
    if string_exists:
        # Extract the value from "Despesas Operacionais e Administrativas - Finatec"
        value_to_add = dfMerged.loc[dfMerged['NomeRubrica'] == 'Despesas Operacionais e Administrativas - Finatec'].iloc[0]
        
        string_exists = dfMerged['NomeRubrica'].isin(["Outros Serviços de Terceiros - Pessoa Jurídica "]).any()
        string_exists2 = dfMerged['NomeRubrica'].isin(["Serviços de Terceiros Pessoa Jurídica"]).any()
        string_exists3 = dfMerged['NomeRubrica'].isin(["Outros Serviços de Terceiros - Pessoa Jurídica"]).any()
        
        if string_exists or string_exists2 or string_exists3:
            if string_exists:
                
                
                dfMerged.loc[dfMerged['NomeRubrica'] == 'Outros Serviços de Terceiros - Pessoa Jurídica '] += value_to_add
             
                # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
                dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']
                
            if string_exists2:
               

                # Add the value to "Outros Serviços de Terceiros - Pessoa Jurídica"
                dfMerged.loc[dfMerged['NomeRubrica'] == 'Serviços de Terceiros Pessoa Jurídica'] += value_to_add

                # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
                dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']
                
            if string_exists3:
            # Find the index of "Outros Serviços de Terceiros - Pessoa Jurídica"
                
                index_to_update = dfMerged.loc[dfMerged['NomeRubrica'] == 'Outros Serviços de Terceiros - Pessoa Jurídica'].index[0]
                
                # Add the value to "Outros Serviços de Terceiros - Pessoa Jurídica"
                dfMerged.iloc[index_to_update] += value_to_add
                
                # Drop the row for "Despesas Operacionais e Administrativas - Finatec"
                dfMerged = dfMerged[dfMerged['NomeRubrica'] != 'Despesas Operacionais e Administrativas - Finatec']   
                
            
    


    for row_num, row_data in enumerate(dfMerged.itertuples(index = False), start=12):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
              
           
                if col_num == 1:
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
                if col_num == 2:
                    col_num = 6
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
                    #sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value))
                if col_num == 3:
                    col_num = 2
                    sheet.cell(row=row_num, column=col_num, value=convert_datetime_to_string(value)).alignment=Alignment(horizontal="right",vertical="center",wrap_text=True)
               



    #PREENCHER CABECARIO
    input_date = []
    output_date_str = []
    input_date2  = []
    output_date_str2 = []
    if check_format(data1):
        input_date = datetime.strptime(data1, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str = input_date.strftime("%d/%m/%Y")
    else :
         return None
    if check_format(data2):
        input_date2 = datetime.strptime(data2, "%Y-%m-%d")
    # Format the datetime object to a string in dd/mm/yyyy format
        output_date_str2 = input_date2.strftime("%d/%m/%Y")
    else :
         return None
    
    string_periodo = f"de {output_date_str} a {output_date_str2}"


    consulta_coordenador = consultaID(codigo)

    # o quarenta significa pra achar o lugar do coordenador dinamicamente mais o valor do tamanho da quantidade de rubricas
    tamanho_row_coordenador = tamanho + 40
    stringCoordenador= f'C{tamanho_row_coordenador}' # retorna lugar do coordanor
    stringTamanhoCPF = f'C{tamanho_row_coordenador+2}' # retorna lugar do coordanor
    sheet[stringCoordenador] = consulta_coordenador['NomePessoaResponsavel']
    sheet[stringTamanhoCPF] = formatar_cpf(consulta_coordenador['CPFCoordenador'])
    string_convenio = f"Convênio nº: {consulta_coordenador['NomeConvenio']}"
    string_convenente= f"Fundação de Empreendimentos Científicos e Tecnológicos - FINATEC" 
    string_convenente_convenente= f"Convenente: "
    string_fonte_recursos = f"Fonte de Recursos:"
    string_participe = f"Partícipe (no caso de contrapartida):"
    string_periodo_relatorio = f"Período Abrangido por este Relatório: "
    stringPeriodoExececucao = f'Período de Execução do Convênio:'
   # Convert 'DataAssinatura' to "dd/mm/YYYY" format
    datetime_obj1 = consulta_coordenador['DataAssinatura']
    formatted_date1 = datetime_obj1.strftime("%d/%m/%Y")

    # Convert 'DataVigencia' to "dd/mm/YYYY" format
    datetime_obj2 = consulta_coordenador['DataVigencia']
    formatted_date2 = datetime_obj2.strftime("%d/%m/%Y")

# Create the string representing the period of execution
   
    string_periodo_convenio = f"de {formatted_date1} a {formatted_date2}"
    sheet['A4'] = string_convenio
    sheet['A5'] = string_convenente_convenente
    sheet['A6'] = stringPeriodoExececucao
    sheet['A7'] = string_periodo_relatorio
 

    sheet['C5'] = string_convenente
    sheet['C6'] = string_periodo_convenio
    sheet['C7'] = string_periodo
   
           


    workbook.save(planilha)
    workbook.close()
   

    
  
    return 0
#ok
def PagamentoDePessoal(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Pagamento de Pessoal")
    sheet2['B11'] = f'Pagamento de Pessoal (Vencimentos e Vantagens Fixas, Obrigações Patronais e Benefícios)'
    sheet2['B11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    workbook.save(tabela)
    workbook.close()

    #recebeOsDataFrames
    dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno = consultaPagamentoPessoal(codigo,data1,data2)
     
    tamanho = len(dfconsultaDadosPorRubrica)
    tamanhoestorno = len(dfconsultaDadosPorRubricaComEstorno)


    rowEstorno = estiloPagamentoPessoal(tabela,tamanho,rowBrasilia,tamanhoestorno)

    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook['Pagamento de Pessoal']

    dfconsultaDadosPorRubrica.index = dfconsultaDadosPorRubrica.index + 1
    for row_num, row_data in enumerate(dfconsultaDadosPorRubrica.itertuples(), start=16):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                value = convert_datetime_to_stringdt(value)
                sheet2.cell(row=row_num, column=col_num, value=value)
                    
                 # dfconsultaDadosPorRubricaComEstorno.index = dfconsultaDadosPorRubricaComEstorno.index + 1
    
    
    
    
                    #
    dfconsultaDadosPorRubricaComEstorno.insert(0, "col1", None)
    dfconsultaDadosPorRubricaComEstorno.insert(5, 'Col2', None)
    dfconsultaDadosPorRubricaComEstorno.insert(6, 'Col3', None)
    
     
    for row_num, row_data in enumerate(dfconsultaDadosPorRubricaComEstorno.itertuples(index=False), start=rowEstorno): #inicio linha
        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                        
           
            value = convert_datetime_to_stringdt(value)
            sheet2.cell(row=row_num, column=col_num, value=value)    
                    
    workbook.save(tabela)
    workbook.close()



    return 0
#ok
def ElementoDeDespesa1415Diarias(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Elemento de Despesa 14.15")
    sheet2['B11'] = f'14/15 - Diárias (Pessoal Civil/Militar)'
    sheet2['B11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    workbook.save(tabela)
    workbook.close()

    dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno = consultaestiloElementoDeDespesa1415Diarias(codigo,data1,data2)
     
    tamanho = len(dfconsultaDadosPorRubrica)
    tamanhoestorno = len(dfconsultaDadosPorRubricaComEstorno)


    rowEstorno = estiloElementoDeDespesa1415Diarias(tabela,tamanho,rowBrasilia,tamanhoestorno)


    
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook['Elemento de Despesa 14.15']

    #print(dfconsultaDadosPorRubrica.columns.values.tolist())

    dfconsultaDadosPorRubrica.index = dfconsultaDadosPorRubrica.index + 1
    for row_num, row_data in enumerate(dfconsultaDadosPorRubrica.itertuples(), start=16):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                value = convert_datetime_to_stringdt(value)
                sheet2.cell(row=row_num, column=col_num, value=value)
                    
                 # dfconsultaDadosPorRubricaComEstorno.index = dfconsultaDadosPorRubricaComEstorno.index + 1
    
    
    
    
                    #
    dfconsultaDadosPorRubricaComEstorno.insert(0, "col1", None)
    dfconsultaDadosPorRubricaComEstorno.insert(5, 'Col2', None)
    dfconsultaDadosPorRubricaComEstorno.insert(6, 'Col3', None)
    
     
    for row_num, row_data in enumerate(dfconsultaDadosPorRubricaComEstorno.itertuples(index=False), start=rowEstorno): #inicio linha
        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                        
           
            value = convert_datetime_to_stringdt(value)
            sheet2.cell(row=row_num, column=col_num, value=value)    
                    
    workbook.save(tabela)
    workbook.close()




    return 0
#ok
def geral30(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    dfNomeRubricaCodigoRubrica = consultaNomeRubricaCodRubrica(codigo, data1, data2)
    for index, values in dfNomeRubricaCodigoRubrica.iterrows():
        dfPJDOA,dfPJDOAESTORNO,dfConsultaProjeto ,dfconsultaDadosPorRubricaComEstorno= consultaGeral30(codigo,data1,data2,values['CodRubrica'])

        if values['NomeRubrica'] == "Obrigações Tributárias e contributivas":
            values['NomeRubrica'] = "Obrigações Tributárias"
        if values['NomeRubrica'] == "Material Permanente e Equipamento Nacional":
            values['NomeRubrica'] = "Equipamento Material Nacional"
        if values['NomeRubrica'] == "Serviços de Terceiros Pessoa Física":
            values['NomeRubrica'] = "Serviços de Terceiros PF"
        if values['NomeRubrica'] == f"Obrigações Tributárias e Contributivas - 20% de OST " :
            values['NomeRubrica'] = f"Obrigações Trib. - Encargos 20%"
        if values['NomeRubrica'] == f"Obrigações Tributárias e contributivas " :
            values['NomeRubrica'] = f"Obrigações Tributárias"
        if values['NomeRubrica'] == f"Outros Serviços de Terceiros - Pessoa Física" :
            values['NomeRubrica'] = f"Outros Serviços Terceiros - PF"
        if values['NomeRubrica'] == f"Outros Serviços de Terceiros - Pessoa Jurídica " :
            values['NomeRubrica'] = f"Outros Serviços Terceiros - PJ"
        if values['NomeRubrica'] == f"Passagens e Despesas com Locomoção" :
            values['NomeRubrica'] = f"Passagens e Desp. Locomoção"
        if values['NomeRubrica'] == f"Despesas Operacionais e Administrativas - Finatec" :
            values['NomeRubrica'] = f"Despesas Operacionais"
        if values['NomeRubrica'] == f"Despesas Operacionais e Administrativas - Finatec" :
            values['NomeRubrica'] = f"Despesas Operacionais"
        if values['NomeRubrica'] == f"Auxílio Financeiro a Pesquisador" :
            values['NomeRubrica'] = f"AuxFinanceiro Pesquisador"
        if values['NomeRubrica'] == f"Equipamentos e Material Permanente" :
            values['NomeRubrica'] = f"Equip e Mat Permanente"
        if values['NomeRubrica'] == f"Material Permanente e Equipamento Importado" :
            values['NomeRubrica'] = f"Equipamento Material Importado"
    
        

        

    

        if values['NomeRubrica'] == "Outros Serviços Terceiros - PJ" or values['NomeRubrica'] == "Serviços de Terceiros Pessoa Jurídica":
            values['NomeRubrica'] = "Outros Serviços Terceiros -PJ"
            nomeTabela = values['NomeRubrica']
            nomeTabelaElemento = f'Elemento de Despesa 39'
            tituloStyle = values['NomeRubrica']
            workbook = openpyxl.load_workbook(tabela)
            sheet2 = workbook.create_sheet(title=nomeTabelaElemento)
            sheet2['B11'] = values['NomeRubrica']
            sheet2['B11'].font = Font(name="Arial", size=12, color="000000",bold=True)
            workbook.save(tabela)
            workbook.close()

            
            tamanho = len(dfPJDOA)
            tamanhoRetorno = len(dfPJDOAESTORNO)
            
            rownovo = estiloG(tabela,tamanho,tituloStyle,nomeTabelaElemento,rowBrasilia,tamanhoRetorno)
            workbook = openpyxl.load_workbook(tabela)
            sheet2 = workbook[nomeTabelaElemento]
            dfPJDOA.index = dfPJDOA.index + 1
            for row_num, row_data in enumerate(dfPJDOA.itertuples(), start=16):#inicio linha
                for col_num, value in enumerate(row_data, start=1):#inicio coluna
                    value = convert_datetime_to_stringdt(value)
                    sheet2.cell(row=row_num, column=col_num, value=value)
            
            dfPJDOAESTORNO.insert(0, "col1", None)
            dfPJDOAESTORNO.insert(4, 'Col2', None)
            dfPJDOAESTORNO.insert(4, 'Col3', None)
            

            
            rownovo = rownovo 
            for row_num, row_data in enumerate(dfPJDOAESTORNO.itertuples(index=False), start=rownovo): #inicio linha
                for col_num, value in enumerate(row_data, start=1): #inicio coluna
                    if col_num == 5:
                        continue
                    value = convert_datetime_to_stringdt(value)
                    sheet2.cell(row=row_num, column=col_num, value=value)    
            
            workbook.save(tabela)
            workbook.close()
             
     
        else:
            excluded_values = ["Rendimentos de Aplicações Financeiras", 
                  "Despesas Financeiras", 
                  "Receitas", 
                  "Devolução de Recursos", 
                  "Outros Serviços Terceiros - PJ", 
                  "Despesas Operacionais",
                  'Passagens e Despesas com Locomoção',
                  "Diárias",
                  "Diárias - Celetistas",
                  "Diárias - Colaborador Eventual no País",
                  "Diárias - Pesquisadores",
                  "Diárias - Servidores Públicos",
                  "Diárias Internacional",
                  "Passagens e Desp. Locomoção",
                  'Pagamento de Pessoal',
                  "Diárias Nacionais"]
            
            if values['NomeRubrica'] not in excluded_values:

                    nomeTabela = values['NomeRubrica']
                    tituloStyle = values['NomeRubrica']
           
                    #switch case

                    #pessoa juridica
                    if values['NomeRubrica'] == f"Outros Serviços Terceiros - PF" :
                        nomeTabelaElementoGeral = f'Elemento de Despesa 36'
                    elif values['NomeRubrica'] == f"Obras e Instalações" :
                        nomeTabelaElementoGeral = f'Elemento de Despesa 51'
                    elif values['NomeRubrica'] == f"Equipamentos e Material Permanente" :
                        nomeTabelaElementoGeral = f'Elemento de Despesa 52'
                    elif values['NomeRubrica'] == f"Material de Consumo " :
                        nomeTabelaElementoGeral = f'Elemento de Despesa 30'
                    
                    else :
                        nomeTabelaElementoGeral = values['NomeRubrica']

                    workbook = openpyxl.load_workbook(tabela)
                    sheet2 = workbook.create_sheet(title=nomeTabelaElementoGeral)
                    sheet2['B11'] = values['NomeRubrica']
                    sheet2['B11'].font = Font(name="Arial", size=12, color="000000",bold=True)
                    workbook.save(tabela)
                    workbook.close()

                    tamanho = len(dfConsultaProjeto)
                    tamanhoRetorno = len(dfconsultaDadosPorRubricaComEstorno)
                    
                  
                    
                    rowEstorno = estiloG(tabela,tamanho,tituloStyle,nomeTabelaElementoGeral,rowBrasilia,tamanhoRetorno)
                    workbook = openpyxl.load_workbook(tabela)
                    sheet2 = workbook[nomeTabelaElementoGeral]
                    dfConsultaProjeto.index = dfConsultaProjeto.index + 1
                    for row_num, row_data in enumerate(dfConsultaProjeto.itertuples(), start=16):#inicio linha
                        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                            value = convert_datetime_to_stringdt(value)
                            sheet2.cell(row=row_num, column=col_num, value=value)
                    
                    # dfconsultaDadosPorRubricaComEstorno.index = dfconsultaDadosPorRubricaComEstorno.index + 1
                    rowEstorno = rowEstorno 
                    #
                    tamanhoDf = len(dfconsultaDadosPorRubricaComEstorno)
                    dfconsultaDadosPorRubricaComEstorno.insert(0, "col1", None)
                    dfconsultaDadosPorRubricaComEstorno.insert(4, 'Col2', None)
                    dfconsultaDadosPorRubricaComEstorno.insert(4, 'Col3', None)
                    
                    
                    
                    
                    
                    for row_num, row_data in enumerate(dfconsultaDadosPorRubricaComEstorno.itertuples(index=False), start=rowEstorno): #inicio linha
                        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                        
                            if col_num == 5:
                                continue
                            value = convert_datetime_to_stringdt(value)
                            sheet2.cell(row=row_num, column=col_num, value=value)    
                    
                    workbook.save(tabela)
                    workbook.close()
#ok
def PassagensEDespesa33(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Elemento de Despesa 33")
    sheet2['B11'] = f'33 - Passagens e Despesas com Locomoção'
    sheet2['B11'].font = Font(name="Arial", size=12, color="000000",bold=True)
    workbook.save(tabela)
    workbook.close()
    dfconsultaDadosPorRubrica,dfconsultaDadosPorRubricaComEstorno = consultaestiloElementoDeDespesa33PassagemEDespesa(codigo,data1,data2)
    tamanho = len(dfconsultaDadosPorRubrica)
    tamanhoestorno = len(dfconsultaDadosPorRubricaComEstorno)
    rowEstorno = estiloElementoDeDespesa33PassagensEDespesa(tabela,tamanho,rowBrasilia,tamanhoestorno)
     
    
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook['Elemento de Despesa 33']

    dfconsultaDadosPorRubrica.index = dfconsultaDadosPorRubrica.index + 1
    dfconsultaDadosPorRubrica.insert(2, "col1", None)
    dfconsultaDadosPorRubrica.insert(3, 'Col2', None)
    dfconsultaDadosPorRubrica.insert(4, 'Col3', None)
    for row_num, row_data in enumerate(dfconsultaDadosPorRubrica.itertuples(), start=16):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                value = convert_datetime_to_stringdt(value)
                sheet2.cell(row=row_num, column=col_num, value=value)
                    
                 # dfconsultaDadosPorRubricaComEstorno.index = dfconsultaDadosPorRubricaComEstorno.index + 1
    
    
    
    
                    #
    dfconsultaDadosPorRubricaComEstorno.insert(0, "col0", None)
    dfconsultaDadosPorRubricaComEstorno.insert(3, "col1", None)
    dfconsultaDadosPorRubricaComEstorno.insert(4, 'Col2', None)
    dfconsultaDadosPorRubricaComEstorno.insert(5, 'Col3', None)
    dfconsultaDadosPorRubricaComEstorno.insert(7, 'Col4', None)
    dfconsultaDadosPorRubricaComEstorno.insert(8, 'Col5', None)
     

    for row_num, row_data in enumerate(dfconsultaDadosPorRubricaComEstorno.itertuples(index=False), start=rowEstorno): #inicio linha
        for col_num, value in enumerate(row_data, start=1): #inicio coluna
                        
           
            value = convert_datetime_to_stringdt(value)
            sheet2.cell(row=row_num, column=col_num, value=value)    
                    
    workbook.save(tabela)
    workbook.close()



    return 0
#ok
def relacaoBensAdquiridosA5(codigo,data1,data2,planilha,rowBrasilia):
    #consult aprojeto 07318 monte de bens imobilizaos, 6995,7311
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Relação Bens Adquiridos A.5")
    workbook.save(tabela)
    workbook.close()
    dfConsultaBens = consultaBens(codigo,data1,data2)
    estiloRelacaoBens(tabela,len(dfConsultaBens),rowBrasilia)
    
    
    
    dfConsultaBens.insert(6, "col1", 1)

   

   

    workbook = openpyxl.load_workbook(tabela)
    sheet = workbook['Relação Bens Adquiridos A.5']
    for row_num, row_data in enumerate(dfConsultaBens.itertuples(), start=15):#inicio linha
         for col_num, value in enumerate(row_data, start=1):#inicio coluna
            value = re.sub("[^a-zA-ZÀ-ÿ0-9º+-//]", " ", str(value))
            value = convert_datetime_to_stringdt(value)
            
            sheet.cell(row=row_num, column=col_num, value=value)


    
    

    workbook.save(tabela)
    workbook.close()

    return 0
#ok
def rendimentoAplicacao(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Rendimento de Aplicação")
    workbook.save(tabela)
    workbook.close()

    dfConsultaRendimentoAplicacao,dfConsultaImposto,dfConsultaRendimentoEImposto = consultaRendimentosAplicacao(codigo,data1,data2)
    merged_df = pd.merge(dfConsultaRendimentoAplicacao, dfConsultaImposto, on='DataPagamento')
    estilo_rendimento_de_aplicacao(tabela,len(merged_df),rowBrasilia)
    
    workbook = openpyxl.load_workbook(tabela)
    sheet = workbook['Rendimento de Aplicação']

    merged_df['data_formatada'] = merged_df['DataPagamento'].apply(formatarDataSemDia)
    merged_df['DataPagamento'] = merged_df['data_formatada']
    merged_df = merged_df.drop('data_formatada', axis=1)

    for row_num, row_data in enumerate(merged_df.itertuples(index=False), start=15):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna 
            if col_num == 2:
                col_num = 5
            if col_num == 3:
                col_num = 6
            sheet.cell(row=row_num, column=col_num, value=value).number_format = 'R$ #,##0.00'


    workbook.save(tabela)
    workbook.close()
        
    return 0
#ok
def conciliacaoBancaria(codigo,data1,data2,planilha,rowBrasilia):
    tabela = pegar_caminho(planilha)
    workbook = openpyxl.load_workbook(tabela)
    sheet2 = workbook.create_sheet(title="Conciliação Bancária A.3")
    workbook.save(tabela)
    workbook.close()
    dfSemEstorno,dfComEstorno = consultaConciliacaoBancaria(codigo,data1,data2)
    estorno = estilo_conciliacoes_bancaria(tabela,len(dfSemEstorno)+1,len(dfComEstorno),rowBrasilia)
    
    workb = openpyxl.load_workbook(tabela)
    worksheet333 = workb["Conciliação Bancária A.3"]

    for row_num, row_data in enumerate(dfSemEstorno.itertuples(index=False), start=19):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                worksheet333.cell(row=row_num, column=col_num, value=value)
    
    for row_num, row_data in enumerate(dfComEstorno.itertuples(index=False), start=estorno):#inicio linha
        for col_num, value in enumerate(row_data, start=1):#inicio coluna
                worksheet333.cell(row=row_num, column=col_num, value=value)

    workb.save(tabela)
    workb.close            
    return 0
#ok
def preencheFinep(codigo,data1,data2,tabela):
    '''Preenche a planilha finep

        Argumentos: 
            codigo = CodConvenio na tabela nova, corresponde ao codigo do projeto
            DATA1 = Data Inicial Selecinado pelo Usuario
            DATA2 = Data Final Selecionado pelo Usuario
            tabela = tabela a ser preenchida  extensão xlsx
   '''
    planilhaGeral(tabela,codigo,data1,data2)
    rowBrasilia = demostrativereceitaedepesaA2(codigo,data1,data2,tabela)
    relatorioExecFinanceiraA1(codigo,data1,data2,tabela,rowBrasilia)
    PagamentoDePessoal(codigo,data1,data2,tabela,rowBrasilia)
    ElementoDeDespesa1415Diarias(codigo,data1,data2,tabela,rowBrasilia)
    PassagensEDespesa33(codigo,data1,data2,tabela,rowBrasilia)
    geral30(codigo,data1,data2,tabela,rowBrasilia)
    relacaoBensAdquiridosA5(codigo,data1,data2,tabela,rowBrasilia)
    rendimentoAplicacao(codigo,data1,data2,tabela,rowBrasilia)
    conciliacaoBancaria(codigo,data1,data2,tabela,rowBrasilia)
    
    