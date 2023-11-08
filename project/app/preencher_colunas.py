import oracledb
import datetime
import openpyxl
import os
#import estilo
from .estilo import estilo_fub_fisica_juridica
def pegar_caminho(nome_arquivo):

    # Obter o caminho absoluto do arquivo Python em execução
    caminho_script = os.path.abspath(__file__)

    # Obter o diretório da pasta onde o script está localizado
    pasta_script = os.path.dirname(caminho_script)

    # Combinar o caminho da pasta com o nome do arquivo Excel
    caminho = os.path.join(pasta_script, nome_arquivo)

    return caminho

def convert_datetime_to_string(value):
    if isinstance(value, datetime.datetime):
        return value.strftime('%d/%m/%Y')
    return value
#connection string in the format
#<username>/<password>@<dBhostAddress>:<dbPort>/<dbServiceName>


def getCollumNames(IDPROJETO, DATA1, DATA2):
    file_path = "/home/ubuntu/Desktop/devfront/devfull/pass.txt"
    conStr = ''
    with open(file_path, 'r') as file:
            conStr = file.readline().strip()

    conn = None
    conn = oracledb.connect(conStr)
    cur = conn.cursor()

    sql = """SELECT DISTINCT * FROM IDEA.FAT_LANCAMENTO_CONVENIAR 
             WHERE ID_PROJETO = :IDPROJETO 
             AND ID_STATUS = 27 
             AND DATA_PAGAMENTO BETWEEN TO_DATE(:DATA1, 'YYYY-MM-DD') 
             AND TO_DATE(:DATA2, 'YYYY-MM-DD') 
             ORDER BY NUM_DOC_FIN"""

    cur.execute(sql, {
        'IDPROJETO': IDPROJETO,
        'DATA1': DATA1,
        'DATA2': DATA2
    })
    return cur



def get_values_from_dict(keys,codigo,data1,data2):
    data1_str = f'{data1}'
    data2_str = f'{data2}'
    gete = getCollumNames(codigo,data1_str,data2_str)

    collums = []
    for i in gete.description:
        collums.append(i[0])
    
    #print(collums)

    value = []
    for i in gete:
        val = tuple(convert_datetime_to_string(item) for item in i)
        value.append(val)
    #print(value)
    list_of_dicts = [dict(zip(collums, values)) for values in value]
    #print(list_of_dicts)
    
    values = [d.get(key) for d in list_of_dicts for key in keys]
    
    #print(values)
    return values



def prenchimento_fub(tabela,keys,codigo,data1,data2):
     

   
    tamanho = []
    for j in keys:
        lj = [j]
        valores_dboracle = get_values_from_dict(lj,codigo,data1,data2)
        size = len(valores_dboracle)
        tamanho.append(size)
        #print(valores_dboracle)
   
    maior = max(tamanho)
    # print(maior)
    estilo_fub_fisica_juridica(tabela,maior)
    #estilo.estilo_fub_fisica_juridica('Modelo_Fub.xlsx',maior)
    coluna = 2
    caminho = pegar_caminho(tabela)
    #print(f'caminhozada{caminho}')
    #print(f'tabelazada{tabela}')
    workb = openpyxl.load_workbook(caminho)
    worksheet1 = workb['Pessoa Fisica']

    for i in range(1,maior+1):
        valor_coluna = 9 + i
        worksheet1.cell(row=valor_coluna, column=1, value=i)  # column index starts from 1

    #worksheet1.insert_rows(11, maior) 
    for i in keys:
        li = [i]
        valores_preenchimento = get_values_from_dict(li,codigo,data1,data2)
        #print(valores_preenchimento)
        n = len(valores_preenchimento)  

        for rowkek, cell_data in enumerate(valores_preenchimento, start=10):
           
            worksheet1.cell(row=rowkek, column=coluna, value=cell_data)  

        if coluna == 5 or coluna == 7 :
                coluna = coluna + 1  
        coluna = coluna + 1
        workb.save(tabela)
        workb.close



